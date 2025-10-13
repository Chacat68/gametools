#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel越南文扫描器
扫描整个文件夹下的所有Excel文件，检测越南文文本并输出结果到Excel文件
"""

import os
import re
from pathlib import Path
from typing import List, Dict, Tuple
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from core.localization_checker import VietnameseDetector


class ExcelVietnameseScanner:
    """Excel越南文扫描器"""
    
    def __init__(self):
        self.vietnamese_detector = VietnameseDetector()
        self.supported_extensions = {'.xlsx', '.xls'}
    
    def is_excel_file(self, file_path: Path) -> bool:
        """
        检查文件是否为支持的Excel格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是Excel格式返回True
        """
        return file_path.suffix.lower() in self.supported_extensions
    
    def scan_excel_file(self, file_path: Path) -> List[Dict]:
        """
        扫描单个Excel文件中的越南文
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            List[Dict]: 包含越南文的位置信息列表
        """
        results = []
        
        try:
            # 读取Excel文件的所有工作表
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    # 读取工作表
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # 扫描每个单元格
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.notna(value) and self.vietnamese_detector.contains_vietnamese(str(value)):
                                # 获取实际检测到的单元格内容
                                content = str(value)
                                # 基于实际检测到的内容判断语言类型
                                language_type = self.vietnamese_detector.detect_language_type(content)
                                results.append({
                                    'excel_file': file_path.name,
                                    'sheet_name': sheet_name,
                                    'row': row_idx + 2,  # +2 因为pandas从0开始，且Excel有标题行
                                    'col': col_idx + 1,  # +1 因为pandas从0开始
                                    'column_name': df.columns[col_idx] if col_idx < len(df.columns) else f'Column_{col_idx + 1}',
                                    'content': content,
                                    'language_type': language_type,  # 基于实际检测内容判断的语言类型
                                    'position': f"第{row_idx + 2}行第{col_idx + 1}列"
                                })
                
                except Exception as e:
                    print(f"读取工作表 '{sheet_name}' 时出错: {e}")
                    continue
        
        except Exception as e:
            print(f"读取Excel文件 {file_path} 时出错: {e}")
        
        return results
    
    def scan_directory(self, directory_path: str) -> List[Dict]:
        """
        扫描目录下的所有Excel文件
        
        Args:
            directory_path: 要扫描的目录路径
            
        Returns:
            List[Dict]: 所有Excel文件中越南文的位置信息
        """
        directory = Path(directory_path)
        
        if not directory.exists():
            print(f"错误: 目录 {directory_path} 不存在")
            return []
        
        if not directory.is_dir():
            print(f"错误: {directory_path} 不是一个目录")
            return []
        
        all_results = []
        excel_files = []
        
        # 收集所有Excel文件
        for file_path in directory.rglob('*'):
            if file_path.is_file() and self.is_excel_file(file_path):
                excel_files.append(file_path)
        
        print(f"找到 {len(excel_files)} 个Excel文件")
        
        # 扫描每个Excel文件
        for i, file_path in enumerate(excel_files, 1):
            print(f"正在扫描 ({i}/{len(excel_files)}): {file_path.name}")
            
            file_results = self.scan_excel_file(file_path)
            all_results.extend(file_results)
            
            if file_results:
                print(f"  - 找到 {len(file_results)} 个越南文位置")
            else:
                print(f"  - 未找到越南文")
        
        return all_results
    
    def create_output_excel(self, results: List[Dict], output_path: str) -> bool:
        """
        创建输出Excel文件
        
        Args:
            results: 扫描结果列表
            output_path: 输出文件路径
            
        Returns:
            bool: 创建成功返回True
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "越南文检测结果"
            
            # 设置标题行
            headers = ['序号', 'Excel文件名', '位置', '越南文内容', '语言类型']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加数据
            for row_idx, result in enumerate(results, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
                ws.cell(row=row_idx, column=2, value=result['excel_file'])
                ws.cell(row=row_idx, column=3, value=result['position'])
                ws.cell(row=row_idx, column=4, value=result['content'])
                ws.cell(row=row_idx, column=5, value=result['language_type'])
            
            # 设置列宽
            column_widths = [8, 25, 15, 50, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 保存文件
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"创建输出Excel文件时出错: {e}")
            return False
    
    def scan_and_export(self, directory_path: str, output_path: str) -> Dict:
        """
        扫描目录并导出结果
        
        Args:
            directory_path: 要扫描的目录路径
            output_path: 输出Excel文件路径
            
        Returns:
            Dict: 包含扫描统计信息的字典
        """
        print("开始扫描Excel文件中的越南文...")
        print("=" * 50)
        
        # 扫描目录
        results = self.scan_directory(directory_path)
        
        # 统计信息
        stats = {
            'total_files_scanned': len(list(Path(directory_path).rglob('*.xlsx'))) + len(list(Path(directory_path).rglob('*.xls'))),
            'files_with_vietnamese': len(set(result['excel_file'] for result in results)),
            'total_vietnamese_locations': len(results),
            'results': results
        }
        
        print("\n" + "=" * 50)
        print("扫描完成！")
        print(f"扫描的Excel文件总数: {stats['total_files_scanned']}")
        print(f"包含越南文的文件数: {stats['files_with_vietnamese']}")
        print(f"越南文位置总数: {stats['total_vietnamese_locations']}")
        
        if results:
            # 创建输出Excel文件
            print(f"\n正在创建输出文件: {output_path}")
            if self.create_output_excel(results, output_path):
                print("输出文件创建成功！")
                stats['output_success'] = True
            else:
                print("输出文件创建失败！")
                stats['output_success'] = False
        else:
            print("未找到任何越南文内容，不创建输出文件。")
            stats['output_success'] = False
        
        return stats


def main():
    """主函数 - 命令行版本"""
    import sys
    
    print("Excel越南文扫描器")
    print("=" * 50)
    
    # 获取用户输入
    if len(sys.argv) > 1:
        directory_path = sys.argv[1]
    else:
        directory_path = input("请输入要扫描的目录路径: ").strip()
    
    if not directory_path:
        print("错误: 未提供目录路径")
        return
    
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    else:
        output_path = input("请输入输出Excel文件路径: ").strip()
    
    if not output_path:
        print("错误: 未提供输出文件路径")
        return
    
    # 创建扫描器并执行扫描
    scanner = ExcelVietnameseScanner()
    stats = scanner.scan_and_export(directory_path, output_path)
    
    print("\n按任意键退出...")
    input()


if __name__ == "__main__":
    main()
