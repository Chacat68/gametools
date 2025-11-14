#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
输出格式模块
提供多种输出格式支持（Excel、CSV、JSON、HTML、Markdown）
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class OutputFormat:
    """输出格式枚举"""
    EXCEL = "xlsx"
    CSV = "csv"
    JSON = "json"
    HTML = "html"
    MARKDOWN = "md"
    TEXT = "txt"


class ResultExporter:
    """结果导出器"""
    
    def __init__(self):
        """初始化导出器"""
        self.supported_formats = [
            OutputFormat.EXCEL,
            OutputFormat.CSV,
            OutputFormat.JSON,
            OutputFormat.HTML,
            OutputFormat.MARKDOWN,
            OutputFormat.TEXT
        ]
    
    def export(self, data: List[Dict], output_path: str, 
               format_type: Optional[str] = None,
               metadata: Optional[Dict] = None,
               **options) -> bool:
        """
        导出数据到指定格式
        
        Args:
            data: 要导出的数据
            output_path: 输出路径
            format_type: 格式类型（None则根据文件扩展名判断）
            metadata: 元数据（标题、描述等）
            **options: 格式特定选项
            
        Returns:
            是否导出成功
        """
        try:
            # 确定格式
            if format_type is None:
                format_type = Path(output_path).suffix.lstrip('.')
            
            format_type = format_type.lower()
            
            # 调用对应的导出方法
            if format_type == OutputFormat.EXCEL:
                return self._export_excel(data, output_path, metadata, **options)
            elif format_type == OutputFormat.CSV:
                return self._export_csv(data, output_path, **options)
            elif format_type == OutputFormat.JSON:
                return self._export_json(data, output_path, metadata, **options)
            elif format_type == OutputFormat.HTML:
                return self._export_html(data, output_path, metadata, **options)
            elif format_type == OutputFormat.MARKDOWN:
                return self._export_markdown(data, output_path, metadata, **options)
            elif format_type == OutputFormat.TEXT:
                return self._export_text(data, output_path, **options)
            else:
                logger.error(f"不支持的输出格式: {format_type}")
                return False
                
        except Exception as e:
            logger.error(f"导出失败: {e}")
            return False
    
    def _export_excel(self, data: List[Dict], output_path: str,
                     metadata: Optional[Dict] = None,
                     **options) -> bool:
        """
        导出为Excel格式
        
        Args:
            data: 数据
            output_path: 输出路径
            metadata: 元数据
            **options: 选项（sheet_name, auto_filter, freeze_panes等）
            
        Returns:
            是否成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 创建Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                sheet_name = options.get('sheet_name', '结果')
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # 样式设置
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置表头样式
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # 设置数据行样式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 自动筛选
                if options.get('auto_filter', True):
                    worksheet.auto_filter.ref = worksheet.dimensions
                
                # 冻结首行
                if options.get('freeze_panes', True):
                    worksheet.freeze_panes = 'A2'
                
                # 添加元数据（作为另一个工作表）
                if metadata:
                    meta_sheet = workbook.create_sheet('元数据')
                    meta_sheet.append(['键', '值'])
                    for key, value in metadata.items():
                        meta_sheet.append([key, str(value)])
            
            logger.info(f"Excel文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return False
    
    def _export_csv(self, data: List[Dict], output_path: str, **options) -> bool:
        """
        导出为CSV格式
        
        Args:
            data: 数据
            output_path: 输出路径
            **options: 选项（encoding, delimiter等）
            
        Returns:
            是否成功
        """
        try:
            encoding = options.get('encoding', 'utf-8-sig')  # 使用BOM以便Excel正确识别
            delimiter = options.get('delimiter', ',')
            
            df = pd.DataFrame(data)
            df.to_csv(output_path, index=False, encoding=encoding, sep=delimiter)
            
            logger.info(f"CSV文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"CSV导出失败: {e}")
            return False
    
    def _export_json(self, data: List[Dict], output_path: str,
                    metadata: Optional[Dict] = None,
                    **options) -> bool:
        """
        导出为JSON格式
        
        Args:
            data: 数据
            output_path: 输出路径
            metadata: 元数据
            **options: 选项（indent, ensure_ascii等）
            
        Returns:
            是否成功
        """
        try:
            indent = options.get('indent', 2)
            ensure_ascii = options.get('ensure_ascii', False)
            
            output_data = {
                'metadata': metadata or {},
                'data': data,
                'export_time': datetime.now().isoformat(),
                'total_records': len(data)
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=indent, ensure_ascii=ensure_ascii)
            
            logger.info(f"JSON文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
            return False
    
    def _export_html(self, data: List[Dict], output_path: str,
                    metadata: Optional[Dict] = None,
                    **options) -> bool:
        """
        导出为HTML格式
        
        Args:
            data: 数据
            output_path: 输出路径
            metadata: 元数据
            **options: 选项（title, style等）
            
        Returns:
            是否成功
        """
        try:
            title = options.get('title', '检测结果')
            
            df = pd.DataFrame(data)
            
            # HTML模板
            html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #366092;
            padding-bottom: 10px;
        }}
        .metadata {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .footer {{
            margin-top: 20px;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        
        <div class="metadata">
            <strong>导出时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
            <strong>记录数量:</strong> {len(data)}
        </div>
        
        {df.to_html(index=False, classes='data-table', border=0)}
        
        <div class="footer">
            Generated by GameTools v1.23.0
        </div>
    </div>
</body>
</html>
"""
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_template)
            
            logger.info(f"HTML文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"HTML导出失败: {e}")
            return False
    
    def _export_markdown(self, data: List[Dict], output_path: str,
                        metadata: Optional[Dict] = None,
                        **options) -> bool:
        """
        导出为Markdown格式
        
        Args:
            data: 数据
            output_path: 输出路径
            metadata: 元数据
            **options: 选项（title等）
            
        Returns:
            是否成功
        """
        try:
            title = options.get('title', '检测结果')
            
            df = pd.DataFrame(data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                # 标题
                f.write(f"# {title}\n\n")
                
                # 元数据
                if metadata:
                    f.write("## 元数据\n\n")
                    for key, value in metadata.items():
                        f.write(f"- **{key}**: {value}\n")
                    f.write("\n")
                
                # 统计信息
                f.write("## 统计信息\n\n")
                f.write(f"- 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"- 记录数量: {len(data)}\n\n")
                
                # 数据表格
                f.write("## 数据\n\n")
                f.write(df.to_markdown(index=False))
                f.write("\n\n---\n")
                f.write("*Generated by GameTools v1.23.0*\n")
            
            logger.info(f"Markdown文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Markdown导出失败: {e}")
            return False
    
    def _export_text(self, data: List[Dict], output_path: str, **options) -> bool:
        """
        导出为纯文本格式
        
        Args:
            data: 数据
            output_path: 输出路径
            **options: 选项
            
        Returns:
            是否成功
        """
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(f"检测结果\n")
                f.write(f"=" * 60 + "\n\n")
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"记录数量: {len(data)}\n\n")
                f.write("=" * 60 + "\n\n")
                
                for i, item in enumerate(data, 1):
                    f.write(f"记录 #{i}\n")
                    f.write("-" * 40 + "\n")
                    for key, value in item.items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")
            
            logger.info(f"文本文件已导出: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"文本导出失败: {e}")
            return False


# 便捷函数

def export_results(data: List[Dict], output_path: str, 
                  format_type: Optional[str] = None,
                  **options) -> bool:
    """
    导出结果的便捷函数
    
    Args:
        data: 数据
        output_path: 输出路径
        format_type: 格式类型
        **options: 格式选项
        
    Returns:
        是否成功
    """
    exporter = ResultExporter()
    return exporter.export(data, output_path, format_type, **options)
