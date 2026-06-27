#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量Excel修改器
根据映射表和JSON配置文件，批量修改指定文件夹中的Excel文件
"""

import os
import re
import json
import copy
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional
import pandas as pd
import logging

# 导入统一的常量
from core.constants import (
    SUPPORTED_LANGUAGES, SUPPORTED_EXCEL_EXTENSIONS, SUPPORTED_MAPPING_FORMATS,
    FIELD_NAME_ROW, DATA_START_ROW,
)
from core.text_patterns import parse_cell_reference, is_translatable_text

# 尝试导入xlwings
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    xw = None

# 尝试导入openpyxl（备用引擎）
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # 备用函数
    def get_column_letter(col_num: int) -> str:
        """将列号（从1开始）转换为Excel列字母"""
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def column_index_from_string(col_letter: str) -> int:
        """将Excel列字母转换为列号（从1开始）"""
        col_letter = col_letter.upper()
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

# 别名，保持向后兼容
get_column_number = column_index_from_string

logger = logging.getLogger(__name__)


class BatchExcelModifier:
    """批量Excel修改器（使用 xlwings 引擎）"""
    
    # 使用统一的语言配置
    SUPPORTED_LANGUAGES = SUPPORTED_LANGUAGES
    
    def __init__(self):
        """
        初始化批量修改器
        
        使用 xlwings 库进行修改，需要安装 Microsoft Excel
        xlwings 可以完全保留 Excel 文件的原有结构（包括批注、格式等）
        """
        self.supported_extensions = SUPPORTED_EXCEL_EXTENSIONS
        self.supported_mapping_formats = SUPPORTED_MAPPING_FORMATS
        
        # 使用 xlwings 引擎
        self.use_xlwings = True
        
        # 处理统计
        self.processing_stats = {
            'total_rows': 0,           # 映射表总行数
            'processed_rows': 0,       # 已处理行数
            'modified_files': 0,       # 修改的文件数
            'modified_cells': 0,       # 修改的单元格数
            'skipped_rows': 0,         # 跳过的行数
            'errors': 0                # 错误数
        }
        
        # 错误日志
        self.error_logs = []
        
        # 修改日志
        self.modification_logs = []
        
        # JSON配置中的字段信息
        self.field_config = {}
        
        # JSON中的语言标记
        self.json_language = None
        
        # 进度回调
        self.progress_callback = None
        
        # xlwings Excel 应用实例（延迟初始化）
        self._excel_app = None
    
    def set_progress_callback(self, callback):
        """设置进度回调函数"""
        self.progress_callback = callback
    
    def _report_progress(self, message: str, percentage: float = None):
        """报告进度"""
        if self.progress_callback:
            self.progress_callback(message, percentage)
        logger.info(message)
    
    def _convert_csv_format_if_needed(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        检测CSV格式并转换（如果需要）
        
        支持两种格式：
        1. 标准格式：Table,Classification,ID,VN,TH...
        2. 翻译提取格式：Table,Sheet,Field,Type,Position,ZH,VN,TH,EN...
        
        如果是翻译提取格式，自动转换为标准格式
        
        Args:
            df: 原始DataFrame
            
        Returns:
            pd.DataFrame: 转换后的DataFrame
        """
        columns = df.columns.tolist()
        
        # 检测是否为翻译提取格式
        is_translation_format = ('Position' in columns and 'Field' in columns and 'Sheet' in columns)
        
        if not is_translation_format:
            logger.info("检测到标准批量改表格式")
            return df
        
        logger.info("检测到翻译提取格式，开始转换...")
        
        try:
            # 保留Position列用于直接定位单元格
            # Position格式如 "B7", "E24" 等，包含了完整的列和行信息
            
            # Field列作为Classification
            df['Classification'] = df['Field']
            
            # 创建一个虚拟的ID列（使用行索引），虽然实际定位会用Position
            df['ID'] = range(1, len(df) + 1)
            
            # 删除不需要的列（保留Position和Sheet）
            # Sheet列用于定位三维表的工作表
            columns_to_drop = ['Field', 'Type', 'ZH']
            df = df.drop(columns=[col for col in columns_to_drop if col in df.columns], errors='ignore')
            
            # 重新排列列顺序：Table, Sheet, Classification, ID, Position, 然后是语言列
            base_cols = ['Table', 'Classification', 'ID', 'Position']
            if 'Sheet' in df.columns:
                base_cols = ['Table', 'Sheet', 'Classification', 'ID', 'Position']
            lang_cols = [col for col in df.columns if col not in base_cols]
            new_columns = base_cols + lang_cols
            df = df[new_columns]
            
            logger.info(f"格式转换完成！")
            logger.info(f"  - 转换后列名: {df.columns.tolist()}")
            logger.info(f"  - 转换后行数: {len(df)}")
            
            return df
            
        except Exception as e:
            logger.error(f"格式转换失败: {e}")
            logger.warning("将使用原始格式继续处理")
            return df
    
    def get_available_languages_from_json(self, json_path: str) -> List[Dict[str, str]]:
        """
        获取JSON配置文件中所有可用的语言
        
        Args:
            json_path: JSON配置文件路径
            
        Returns:
            List[Dict]: 语言列表，每项包含 {'code': 'vn', 'name': '越南语', 'key': 'VN'}
        """
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            languages = []
            lang_code_keys = ['ZH', 'VN', 'TH', 'EN', 'JP', 'KR', 'TW', 'CN', 
                             'zh', 'vn', 'th', 'en', 'jp', 'kr', 'tw', 'cn']
            lang_names = {
                'zh': '中文', 'cn': '中文', 'vn': '越南语', 'th': '泰语',
                'en': '英语', 'jp': '日语', 'kr': '韩语', 'tw': '繁体中文'
            }
            
            # 检测格式3（多个语言作为顶层key）
            for key in config.keys():
                if key.upper() in [k.upper() for k in lang_code_keys] and isinstance(config.get(key), dict):
                    lang_code = key.lower()
                    languages.append({
                        'code': lang_code,
                        'name': lang_names.get(lang_code, key),
                        'key': key
                    })
            
            # 如果找到了语言key，返回
            if languages:
                return languages
            
            # 格式1或2：检查language字段
            if 'language' in config and isinstance(config['language'], dict):
                lang_code = config['language'].get('code', '')
                lang_name = config['language'].get('name', '')
                if lang_code:
                    return [{'code': lang_code, 'name': lang_name, 'key': None}]
            
            return []
        except Exception as e:
            logger.error(f"获取JSON语言列表失败: {e}")
            return []
    
    def load_json_config(self, json_path: str, target_lang_code: str = None) -> Dict:
        """
        加载JSON配置文件，提取字段信息
        
        支持三种JSON格式：
        
        格式1（传统格式）:
        {
            "language": {"code": "vn", "name": "越南语"},
            "text_tables": [...]
        }
        
        格式2（新格式 - 按语言组织字段）:
        {
            "language": {"code": "vn", "name": "越南语"},
            "text_tables": [
                {
                    "table_name": "armor_ancient.xlsx",
                    "fields_by_language": {"zh": [...], "vn": [...]}
                }
            ]
        }
        
        格式3（语言代码作为顶层key）:
        {
            "ZH": {"text_tables": [...]},
            "VN": {"text_tables": [...]},
            "TH": {"text_tables": [...]}
        }
        
        Args:
            json_path: JSON配置文件路径
            target_lang_code: 目标语言代码（如'vn', 'zh'），用于格式3
            
        Returns:
            Dict: 表名到字段信息的映射
        """
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # 检测JSON格式类型
            # 格式3：语言代码作为顶层key（如 "ZH", "VN", "TH"）
            lang_code_keys = ['ZH', 'VN', 'TH', 'EN', 'JP', 'KR', 'TW', 'CN', 
                             'zh', 'vn', 'th', 'en', 'jp', 'kr', 'tw', 'cn']
            available_lang_keys = []
            for key in config.keys():
                if key.upper() in [k.upper() for k in lang_code_keys] and isinstance(config.get(key), dict):
                    available_lang_keys.append(key)
            
            detected_lang_key = None
            if available_lang_keys:
                # 格式3：有多个语言key
                if target_lang_code:
                    # 根据用户指定的语言代码选择
                    for key in available_lang_keys:
                        if key.lower() == target_lang_code.lower():
                            detected_lang_key = key
                            break
                
                # 如果没有指定或没找到，使用第一个
                if not detected_lang_key:
                    detected_lang_key = available_lang_keys[0]
                    if target_lang_code:
                        logger.warning(f"未找到语言 '{target_lang_code}'，使用 '{detected_lang_key}'")
            
            if detected_lang_key:
                # 格式3：提取语言配置
                lang_code = detected_lang_key.lower()
                lang_names = {
                    'zh': '中文', 'cn': '中文', 'vn': '越南语', 'th': '泰语',
                    'en': '英语', 'jp': '日语', 'kr': '韩语', 'tw': '繁体中文'
                }
                lang_name = lang_names.get(lang_code, detected_lang_key)
                
                self.json_language = {'code': lang_code, 'name': lang_name}
                logger.info(f"  - 检测到格式3（语言代码顶层key）: {lang_name} ({lang_code})")
                if len(available_lang_keys) > 1:
                    logger.info(f"  - 可用语言: {', '.join(available_lang_keys)}")
                
                # 提取该语言下的配置
                lang_config = config[detected_lang_key]
                text_tables = lang_config.get('text_tables', [])
                no_text_tables = lang_config.get('no_text_tables', [])
            else:
                # 格式1或格式2：传统格式
                if 'language' in config:
                    self.json_language = config['language']
                    logger.info(f"  - 语言标记: {self.json_language.get('name', '')} ({self.json_language.get('code', '')})")
                else:
                    self.json_language = None
                
                text_tables = config.get('text_tables', [])
                no_text_tables = config.get('no_text_tables', [])
            
            # 获取当前语言代码
            current_lang_code = self.get_json_language()
            
            # 构建表名到字段的映射
            field_config = {}
            
            # 处理text_tables
            for table_info in text_tables:
                table_name = table_info.get('table_name', '')
                sheet_name = table_info.get('sheet_name', '')
                fields = table_info.get('fields', [])
                fields_with_examples = table_info.get('fields_with_examples', [])
                fields_by_language = table_info.get('fields_by_language', {})
                
                # 解析fields_with_examples获取字段类型
                field_types = {}
                all_fields_from_examples = []
                for field_str in fields_with_examples:
                    if ',' in field_str:
                        parts = field_str.split(',', 1)
                        field_name = parts[0].strip()
                        field_type = parts[1].strip() if len(parts) > 1 else ''
                        field_types[field_name] = field_type
                        all_fields_from_examples.append(field_name)
                    else:
                        all_fields_from_examples.append(field_str.strip())
                
                # 合并所有字段来源
                all_fields = set(fields)
                all_fields.update(all_fields_from_examples)
                
                # 如果有按语言组织的字段，也加入
                for lang_fields in fields_by_language.values():
                    all_fields.update(lang_fields)
                
                # 使用完整表名作为key
                field_config[table_name] = {
                    'table_name': table_name,
                    'sheet_name': sheet_name,
                    'fields': list(all_fields),
                    'fields_with_examples': fields_with_examples,
                    'field_types': field_types,
                    'fields_by_language': fields_by_language
                }
                
                # 同时用不含扩展名的表名作为key（兼容）
                table_key = Path(table_name).stem
                field_config[table_key] = field_config[table_name]
            
            # 也处理no_text_tables（用于识别哪些表不需要处理）
            for table_info in no_text_tables:
                table_name = table_info.get('table_name', '')
                sheet_name = table_info.get('sheet_name', '')
                
                # 标记为无文本表
                field_config[table_name] = {
                    'table_name': table_name,
                    'sheet_name': sheet_name,
                    'fields': [],
                    'fields_with_examples': [],
                    'field_types': {},
                    'fields_by_language': {},
                    'is_no_text_table': True
                }
                
                table_key = Path(table_name).stem
                field_config[table_key] = field_config[table_name]
            
            self.field_config = field_config
            logger.info(f"成功加载JSON配置: {json_path}")
            logger.info(f"  - 包含 {len(text_tables)} 个有文本表, {len(no_text_tables)} 个无文本表")
            if current_lang_code:
                logger.info(f"  - 当前语言代码: {current_lang_code}")
            
            return field_config
            
            return field_config
        
        except Exception as e:
            error_msg = f"加载JSON配置失败: {e}"
            logger.error(error_msg)
            self.error_logs.append(error_msg)
            return {}
    
    def get_json_language(self) -> Optional[str]:
        """
        获取JSON配置中的语言代码
        
        Returns:
            str or None: 语言代码 ('zh', 'vn', 'th') 或 None
        """
        if self.json_language and isinstance(self.json_language, dict):
            return self.json_language.get('code')
        return None
    
    def get_json_language_name(self) -> str:
        """
        获取JSON配置中的语言名称
        
        Returns:
            str: 语言名称，如 '中文'、'越南语'、'泰语'
        """
        if self.json_language and isinstance(self.json_language, dict):
            return self.json_language.get('name', '')
        return ''
    
    def match_directory_by_language(self, lang_dirs: Dict[str, str]) -> Optional[str]:
        """
        根据JSON中的语言标记匹配对应的目录
        
        Args:
            lang_dirs: 语言目录映射 {'zh': '/path/to/zh', 'vn': '/path/to/vn', 'th': '/path/to/th'}
            
        Returns:
            str or None: 匹配的目录路径
        """
        lang_code = self.get_json_language()
        if lang_code and lang_code in lang_dirs:
            return lang_dirs[lang_code]
        return None
    
    def get_language_suffix_patterns(self, lang_code: str) -> List[str]:
        """
        获取语言代码对应的字段后缀模式列表
        
        Args:
            lang_code: 语言代码 ('zh', 'vn', 'th', 'en' 等)
            
        Returns:
            List[str]: 可能的字段后缀列表
        """
        # 语言代码到可能的后缀映射
        suffix_patterns = {
            'zh': ['_zh', '_cn', '_ch', '_chinese', 'cn', 'ch', 'zh'],
            'vn': ['_vn', '_vi', '_vietnamese', 'vn', 'vi'],
            'th': ['_th', '_thai', 'th'],
            'en': ['_en', '_english', 'en'],
            'jp': ['_jp', '_ja', '_japanese', 'jp', 'ja'],
            'kr': ['_kr', '_ko', '_korean', 'kr', 'ko'],
            'tw': ['_tw', '_tc', '_traditional', 'tw', 'tc'],
            'ru': ['_ru', '_russian', 'ru'],
            'de': ['_de', '_german', 'de'],
            'fr': ['_fr', '_french', 'fr'],
            'es': ['_es', '_spanish', 'es'],
            'pt': ['_pt', '_portuguese', 'pt'],
        }
        
        lang_lower = lang_code.lower()
        patterns = suffix_patterns.get(lang_lower, [f'_{lang_lower}', lang_lower])
        return patterns
    
    def detect_language_from_mapping_columns(self, mapping_columns: List[str]) -> Optional[str]:
        """
        从映射表的列名中检测语言代码
        
        Args:
            mapping_columns: 映射表的列名列表
            
        Returns:
            Optional[str]: 检测到的语言代码（如'vn', 'zh', 'th'），未检测到返回None
        """
        # 定义所有支持的语言及其可能的列名
        language_patterns = {
            'zh': ['Support-CH', 'Polish-CH', 'CH', 'CN', 'ZH', 'Chinese', '中文', 'ch'],
            'vn': ['VN', 'VI', 'Vietnamese', '越南语', '越南', 'VN.1'],
            'th': ['TH', 'Thai', '泰语', '泰文'],
            'en': ['EN', 'English', '英语', '英文'],
            'jp': ['JP', 'JA', 'Japanese', '日语', '日文'],
            'kr': ['KR', 'KO', 'Korean', '韩语', '韩文'],
            'tw': ['TW', 'TC', 'Traditional', '繁体', '繁体中文'],
        }
        
        # 按优先级检测（优先检测越南语、泰语等常用语言）
        priority_order = ['vn', 'th', 'zh', 'en', 'jp', 'kr', 'tw']
        
        for lang_code in priority_order:
            patterns = language_patterns.get(lang_code, [])
            for pattern in patterns:
                # 不区分大小写匹配
                for col in mapping_columns:
                    if col.strip().upper() == pattern.upper():
                        logger.info(f"  - 从映射表列名检测到语言: {lang_code} (列名: {col})")
                        return lang_code
        
        # 如果没有匹配到，尝试模糊匹配
        for col in mapping_columns:
            col_upper = col.strip().upper()
            # 检查是否包含语言标识
            if 'VN' in col_upper or 'VIET' in col_upper:
                logger.info(f"  - 从映射表列名模糊匹配到语言: vn (列名: {col})")
                return 'vn'
            elif 'TH' in col_upper or 'THAI' in col_upper:
                logger.info(f"  - 从映射表列名模糊匹配到语言: th (列名: {col})")
                return 'th'
            elif ('CH' in col_upper or 'ZH' in col_upper or 'CN' in col_upper) and 'CLASSIFICATION' not in col_upper:
                logger.info(f"  - 从映射表列名模糊匹配到语言: zh (列名: {col})")
                return 'zh'
            elif col_upper == 'EN' or 'ENGLISH' in col_upper:
                logger.info(f"  - 从映射表列名模糊匹配到语言: en (列名: {col})")
                return 'en'
        
        logger.warning("  - 未能从映射表列名检测到语言")
        return None
    
    def get_mapping_column_for_language(self, lang_code: str) -> List[str]:
        """
        获取语言代码对应的映射表列名候选列表
        
        Args:
            lang_code: 语言代码 ('zh', 'vn', 'th', 'en' 等)
            
        Returns:
            List[str]: 可能的映射表列名列表
        """
        # 语言代码到映射表列名的映射
        column_patterns = {
            'zh': ['Support-CH', 'Polish-CH', 'CH', 'CN', 'ZH', 'Chinese', 'cn', 'zh', 'ch', '中文'],
            'vn': ['VN', 'VI', 'Vietnamese', 'vn', 'vi', 'VN.1', '越南语', '越南'],
            'th': ['TH', 'Thai', 'th', '泰语', '泰文'],
            'en': ['EN', 'English', 'en', '英语', '英文'],
            'jp': ['JP', 'JA', 'Japanese', 'jp', 'ja', '日语', '日文'],
            'kr': ['KR', 'KO', 'Korean', 'kr', 'ko', '韩语', '韩文'],
            'tw': ['TW', 'TC', 'Traditional', 'tw', 'tc', '繁体', '繁体中文'],
        }
        
        lang_lower = lang_code.lower()
        columns = column_patterns.get(lang_lower, [lang_code.upper(), lang_code.lower()])
        return columns
    
    def get_table_fields_by_language(self, table_name: str, lang_code: str = None) -> List[str]:
        """
        根据语言代码获取指定表的对应语言字段列表
        
        优先级:
        1. 如果有 fields_by_language，直接使用对应语言的字段列表
        2. 否则从所有字段中筛选包含语言后缀的字段
        3. 如果都没有，返回所有字段
        
        Args:
            table_name: 表名（可以带或不带扩展名）
            lang_code: 语言代码（如 'zh', 'vn', 'th'），为 None 时使用 JSON 中的语言代码
            
        Returns:
            List[str]: 匹配该语言的字段名列表
        """
        # 获取语言代码
        if not lang_code:
            lang_code = self.get_json_language()
        
        if not lang_code:
            # 没有语言代码，返回所有字段
            return self.get_table_fields(table_name)
        
        table_config = None
        
        # 尝试直接匹配
        if table_name in self.field_config:
            table_config = self.field_config[table_name]
        else:
            # 尝试不带扩展名匹配
            table_key = Path(table_name).stem
            if table_key in self.field_config:
                table_config = self.field_config[table_key]
        
        if not table_config:
            return []
        
        # 优先使用 fields_by_language
        fields_by_language = table_config.get('fields_by_language', {})
        if lang_code in fields_by_language:
            return fields_by_language[lang_code]
        
        # 也尝试小写匹配
        lang_lower = lang_code.lower()
        for key, fields in fields_by_language.items():
            if key.lower() == lang_lower:
                return fields
        
        # 从所有字段中筛选包含语言后缀的字段
        all_fields = self.get_table_fields(table_name)
        suffix_patterns = self.get_language_suffix_patterns(lang_code)
        
        matched_fields = []
        for field in all_fields:
            field_lower = field.lower()
            # 精确匹配：字段名以语言后缀结尾
            for pattern in suffix_patterns:
                # 优先匹配下划线开头的后缀（如 _vn, _zh）
                if pattern.startswith('_') and field_lower.endswith(pattern):
                    matched_fields.append(field)
                    break
                # 其次匹配非下划线后缀（如 vn, zh），但要确保是完整单词
                elif not pattern.startswith('_'):
                    # 检查是否以后缀结尾且前面有分隔符或大写字母
                    if field_lower.endswith('_' + pattern) or field_lower.endswith(pattern):
                        # 确保不是部分匹配（如 avnx 不应匹配 vn）
                        if field_lower.endswith('_' + pattern):
                            matched_fields.append(field)
                            break
                        # 检查是否有驼峰命名（如 nameVn）
                        elif len(field) >= len(pattern) + 1:
                            before_pattern = field[-(len(pattern)+1):-len(pattern)]
                            if before_pattern.isupper() or before_pattern == '_':
                                matched_fields.append(field)
                                break
        
        # 如果匹配到了字段，返回匹配结果
        if matched_fields:
            logger.info(f"为表 {table_name} 匹配到 {len(matched_fields)} 个 {lang_code} 语言字段: {matched_fields[:5]}...")
            return matched_fields
        
        # 如果没有匹配到任何字段，记录警告并返回所有字段
        logger.warning(f"表 {table_name} 未找到 {lang_code} 语言的字段，返回所有字段")
        return all_fields

    def get_table_fields(self, table_name: str) -> List[str]:
        """
        获取指定表的字段列表
        
        Args:
            table_name: 表名（可以带或不带扩展名）
            
        Returns:
            List[str]: 字段名列表（合并 fields 和 fields_with_examples）
        """
        table_config = None
        
        # 尝试直接匹配
        if table_name in self.field_config:
            table_config = self.field_config[table_name]
        else:
            # 尝试不带扩展名匹配
            table_key = Path(table_name).stem
            if table_key in self.field_config:
                table_config = self.field_config[table_key]
        
        if not table_config:
            return []
        
        # 合并 fields 和 fields_with_examples 中的字段名
        all_fields = set()
        
        # 从 fields 列表获取
        fields = table_config.get('fields', [])
        all_fields.update(fields)
        
        # 从 fields_with_examples 列表获取（格式可能是 "字段名,类型" 或纯字段名）
        fields_with_examples = table_config.get('fields_with_examples', [])
        for field_entry in fields_with_examples:
            if isinstance(field_entry, str):
                # 取第一个逗号前的部分作为字段名
                field_name = field_entry.split(',')[0].strip()
                if field_name:
                    all_fields.add(field_name)
        
        return list(all_fields)
    
    def load_mapping_table(self, mapping_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, List[str]]:
        """
        加载映射表文件（支持Excel和CSV）
        
        支持两种CSV格式：
        1. 标准批量改表格式：
           Table,Classification,ID,VN,TH,EN...
        2. 翻译提取格式（自动转换）：
           Table,Sheet,Field,Type,Position,ZH,VN,TH,EN...
        
        Args:
            mapping_path: 映射表文件路径（.xlsx/.xls/.csv）
            sheet_name: 工作表名称（仅Excel文件有效，CSV文件忽略此参数）
            
        Returns:
            Tuple[pd.DataFrame, List[str]]: (DataFrame, 列名列表)
        """
        try:
            # 检查文件扩展名
            file_ext = os.path.splitext(mapping_path)[1].lower()
            
            if file_ext == '.csv':
                # 读取CSV文件
                # 尝试多种编码
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        df = pd.read_csv(mapping_path, header=0, encoding=encoding)
                        logger.info(f"使用编码 {encoding} 成功读取CSV文件")
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    # 如果所有编码都失败，使用默认编码并忽略错误
                    df = pd.read_csv(mapping_path, header=0, encoding='utf-8', errors='ignore')
                    logger.warning(f"使用默认编码读取CSV，可能存在字符问题")
                
                # 检测CSV格式类型并转换
                df = self._convert_csv_format_if_needed(df)
            else:
                # 读取Excel文件
                if sheet_name:
                    df = pd.read_excel(mapping_path, sheet_name=sheet_name, header=0)
                else:
                    df = pd.read_excel(mapping_path, header=0)
            
            columns = df.columns.tolist()
            
            logger.info(f"成功加载映射表: {mapping_path}")
            logger.info(f"  - 文件格式: {file_ext}")
            logger.info(f"  - 行数: {len(df)}")
            logger.info(f"  - 列名: {columns}")
            
            self.processing_stats['total_rows'] = len(df)
            
            return df, columns
        
        except Exception as e:
            error_msg = f"加载映射表失败: {e}"
            logger.error(error_msg)
            self.error_logs.append(error_msg)
            return pd.DataFrame(), []

    def _get_csv_mapping_table_names(self, df: pd.DataFrame) -> List[str]:
        """从 CSV 映射表中提取唯一的目标表名，并统一去除扩展名。"""
        if 'Table' not in df.columns:
            return []

        sheet_names = []
        seen = set()
        for table_value in df['Table'].dropna().astype(str):
            normalized_name = Path(table_value).stem.strip()
            normalized_key = normalized_name.lower()
            if normalized_name and normalized_key not in seen:
                seen.add(normalized_key)
                sheet_names.append(normalized_name)

        return sheet_names

    def _filter_csv_mapping_rows(self, df: pd.DataFrame, table_name: str) -> pd.DataFrame:
        """按表名筛选 CSV 映射数据，兼容扩展名、大小写和带路径的表名。"""
        if df.empty or 'Table' not in df.columns:
            return df.copy()

        normalized_table_name = table_name.strip().lower()
        table_values = df['Table'].fillna('').astype(str).str.strip()
        normalized_exact_names = {
            normalized_table_name,
            f"{normalized_table_name}.xlsx",
            f"{normalized_table_name}.xls",
        }

        normalized_values = table_values.str.lower()
        exact_mask = normalized_values.isin(normalized_exact_names)
        if exact_mask.any():
            return df[exact_mask].copy()

        stem_mask = table_values.apply(
            lambda value: Path(value).stem.strip().lower() == normalized_table_name if value else False
        )
        if stem_mask.any():
            return df[stem_mask].copy()

        return df.iloc[0:0].copy()
    
    def get_mapping_sheets(self, mapping_path: str) -> List[str]:
        """
        获取映射表中的所有工作表名称（仅Excel文件有效）
        
        Args:
            mapping_path: 映射表文件路径（.xlsx/.xls/.csv）
            
        Returns:
            List[str]: 工作表名称列表（CSV文件返回空列表）
        """
        try:
            # 检查文件扩展名
            file_ext = os.path.splitext(mapping_path)[1].lower()
            
            if file_ext == '.csv':
                # CSV文件没有工作表概念，返回空列表
                return []
            else:
                # Excel文件，返回工作表列表
                with pd.ExcelFile(mapping_path) as xl:
                    return xl.sheet_names
        except Exception as e:
            logger.error(f"获取工作表列表失败: {e}")
            return []

    def close(self):
        """显式释放批量改表过程中持有的 Excel 应用实例。"""
        self._close_excel_app()

    def __del__(self):
        try:
            self.close()
        except Exception:
            pass
    
    def parse_mapping_row(self, row: pd.Series, columns: List[str], 
                         table_col: str, id_col: str, 
                         modify_cols: List[str]) -> Dict:
        """
        解析映射表中的一行数据
        
        Args:
            row: DataFrame行
            columns: 列名列表
            table_col: 表名列名
            id_col: ID列名
            modify_cols: 要修改的列名列表
            
        Returns:
            Dict: 解析后的行信息
        """
        try:
            table_name = str(row[table_col]).strip() if pd.notna(row[table_col]) else ''
            id_value = row[id_col] if pd.notna(row[id_col]) else ''
            
            # 获取要修改的值
            modify_values = {}
            for col in modify_cols:
                if col in row.index:
                    value = row[col]
                    if pd.notna(value):
                        modify_values[col] = str(value).strip()
            
            return {
                'table_name': table_name,
                'id': id_value,
                'modify_values': modify_values
            }
        
        except Exception as e:
            logger.error(f"解析行数据失败: {e}")
            return {}
    
    def get_field_name_for_column(self, mapping_col: str, table_key: str) -> Optional[str]:
        """
        根据映射表列名和表配置，获取对应的Excel字段名
        
        这个方法根据JSON配置中的字段列表，找到与映射表列名对应的字段
        
        Args:
            mapping_col: 映射表中的列名（如 "VN", "Support-CH"等）
            table_key: 表名key
            
        Returns:
            Optional[str]: Excel中的字段名，未找到返回None
        """
        # 获取表的字段配置
        if table_key not in self.field_config:
            return None
        
        fields = self.field_config[table_key].get('fields', [])
        
        # 常见映射规则
        column_field_mapping = {
            'VN': ['vn', 'VN', 'vietnamese', 'name_vn'],
            'EN': ['en', 'EN', 'english', 'name_en'],
            'Support-CH': ['support_ch', 'Support-CH', 'chinese', 'name_ch', 'zh'],
            'Polish-CH': ['polish_ch', 'Polish-CH'],
            'VN.1': ['vn_1', 'VN.1', 'vn1'],
            'TH': ['th', 'TH', 'thai', 'name_th'],
        }
        
        # 查找匹配的字段
        possible_names = column_field_mapping.get(mapping_col, [mapping_col.lower()])
        
        for field in fields:
            field_lower = field.lower()
            for possible in possible_names:
                if possible.lower() == field_lower or field_lower.endswith(possible.lower()):
                    return field
        
        # 如果没找到精确匹配，返回列名本身（让用户在GUI中指定映射）
        return None
    
    def _get_excel_app(self):
        """获取或创建 xlwings Excel 应用实例"""
        if self._excel_app is None and XLWINGS_AVAILABLE:
            self._excel_app = xw.App(visible=False, add_book=False)
            self._excel_app.display_alerts = False
            self._excel_app.screen_updating = False
        return self._excel_app
    
    def _close_excel_app(self):
        """关闭 xlwings Excel 应用实例"""
        if self._excel_app is not None:
            try:
                self._excel_app.quit()
            except Exception:
                pass
            self._excel_app = None
    
    def modify_excel_file(self, excel_path: str, modifications: List[Dict], 
                         field_mapping: Dict[str, str] = None,
                         id_col: int = 1, 
                         field_row: int = FIELD_NAME_ROW,
                         data_start_row: int = DATA_START_ROW,
                         use_position: bool = False,
                         sheet_name: str = None) -> Tuple[int, List[str], int]:
        """
        使用 xlwings 修改单个Excel文件（完全保留原文件结构）
        
        两种定位模式：
        1. Position模式（use_position=True）：使用Position列（如"B7"）直接定位单元格
        2. 行号模式（use_position=False）：使用ID列的值直接作为Excel行号
        
        Args:
            excel_path: Excel文件路径
            modifications: 修改列表
                - Position模式：{position: "B7", modify_values: {...}}
                - 行号模式：{id: 7, modify_values: {...}}  # id直接作为行号
            field_mapping: 列名到字段名的映射 {映射表列名: Excel字段名}
            id_col: 保留参数，不再使用
            field_row: 字段名所在行
            data_start_row: 数据起始行（保护表头）
            use_position: 是否使用Position直接定位
            sheet_name: 目标工作表名称（三维表支持），为None时使用第一个工作表
            
        Returns:
            Tuple[int, List[str], int]: (修改的单元格数, 错误列表, 跳过的相同值数量)
        """
        return self._modify_excel_file_xlwings(
            excel_path, modifications, field_mapping, 
            id_col, field_row, data_start_row, use_position, sheet_name
        )
    
    def _modify_excel_file_xlwings(self, excel_path: str, modifications: List[Dict], 
                                   field_mapping: Dict[str, str] = None,
                                   id_col: int = 1, 
                                   field_row: int = FIELD_NAME_ROW,
                                   data_start_row: int = DATA_START_ROW,
                                   use_position: bool = False,
                                   sheet_name: str = None) -> Tuple[int, List[str], int]:
        """
        使用 xlwings 修改单个Excel文件（完全保留原文件结构）
        
        定位逻辑：
        - Position模式：从"B7"提取列B和行7，直接定位单元格
        - 行号模式：ID值直接作为Excel行号（如ID=7表示第7行）
        
        修改范围限制：
        - Position模式：检查行号 >= data_start_row
        - 行号模式：检查行号 >= data_start_row
        - 避免修改表头区域（data_start_row 之前的行）
        
        Args:
            excel_path: Excel文件路径
            modifications: 修改列表
            field_mapping: 列名到字段名的映射
            id_col: 保留参数
            field_row: 字段名所在行
            data_start_row: 数据起始行（修改范围的最小行号，小于此行号的将被跳过）
            use_position: True=Position模式，False=行号模式
            sheet_name: 目标工作表名称（三维表支持），为None时使用第一个工作表
            
        Returns:
            Tuple[int, List[str], int]: (修改的单元格数, 错误列表, 跳过的相同值数量)
        """
        modified_count = 0
        skipped_same_value = 0  # 跳过的相同值数量
        errors = []
        
        wb = None
        try:
            # 使用 xlwings 打开文件
            app = self._get_excel_app()
            wb = app.books.open(excel_path)
            
            # 选择工作表（三维表支持）
            if sheet_name:
                # 按名称查找工作表
                sheet_names_in_wb = [s.name for s in wb.sheets]
                if sheet_name in sheet_names_in_wb:
                    ws = wb.sheets[sheet_name]
                    logger.debug(f"使用指定工作表: {sheet_name}")
                else:
                    # 工作表不存在，使用第一个工作表并记录警告
                    ws = wb.sheets[0]
                    error_msg = f"工作表 '{sheet_name}' 不存在，使用第一个工作表 '{ws.name}'"
                    errors.append(error_msg)
                    logger.warning(error_msg)
            else:
                # 默认使用第一个工作表
                ws = wb.sheets[0]
            
            # 获取数据范围
            used_range = ws.used_range
            max_row = used_range.last_cell.row
            max_col = used_range.last_cell.column
            
            # 如果使用Position模式，不需要构建ID映射和字段映射
            if use_position:
                # Position模式：直接使用Position定位单元格
                for mod in modifications:
                    position = mod.get('position')
                    modify_values = mod.get('modify_values', {})
                    
                    if not position or not modify_values:
                        continue
                    
                    # 解析Position（如"B7"）
                    position_str = str(position).strip()
                    match = re.match(r'([A-Z]+)(\d+)', position_str.upper())
                    if not match:
                        error_msg = f"无效的Position格式: {position}"
                        errors.append(error_msg)
                        continue
                    
                    col_letter = match.group(1)
                    target_row = int(match.group(2))
                    col_num = get_column_number(col_letter)
                    
                    # 检查行号是否在允许的修改范围内（防止修改表头）
                    if target_row < data_start_row:
                        error_msg = f"跳过表头区域: Position {position_str} (行号{target_row} < 数据起始行{data_start_row})"
                        errors.append(error_msg)
                        logger.warning(error_msg)
                        continue
                    
                    # modify_values应该只有一个值（对应Position所在的单元格）
                    for col_name, new_value in modify_values.items():
                        # 读取旧值
                        cell = ws.range((target_row, col_num))
                        old_value = cell.value
                        
                        # 比较新旧值，避免不必要的修改
                        old_str = str(old_value).strip() if old_value is not None else ''
                        new_str = str(new_value).strip() if new_value is not None else ''
                        
                        if old_str == new_str:
                            # 值相同，跳过修改
                            skipped_same_value += 1
                            continue
                        
                        # 修改单元格
                        cell.value = new_value
                        
                        # 记录修改日志
                        self.modification_logs.append({
                            'file': os.path.basename(excel_path),
                            'position': position_str,
                            'field': col_name,
                            'old_value': old_value,
                            'new_value': new_value
                        })
                        
                        modified_count += 1
            else:
                # 行号直接定位模式：ID直接作为Excel行号使用
                # 读取字段行（用于查找列号）
                field_row_values = ws.range((field_row, 1), (field_row, max_col)).value
                if not isinstance(field_row_values, list):
                    field_row_values = [field_row_values]
                
                # 构建字段名到列号的映射
                field_to_col = {}
                for col_idx, field_val in enumerate(field_row_values, start=1):
                    if field_val:
                        field_to_col[str(field_val).strip()] = col_idx
                
                for mod in modifications:
                    id_value = mod.get('id')
                    modify_values = mod.get('modify_values', {})
                
                    if not id_value or not modify_values:
                        continue
                    
                    # 直接使用ID作为行号
                    try:
                        target_row = int(float(id_value))
                    except (ValueError, TypeError):
                        error_msg = f"无效的行号: {id_value}"
                        errors.append(error_msg)
                        continue
                    
                    # 检查行号是否在有效范围内（防止修改表头）
                    if target_row < data_start_row:
                        error_msg = f"跳过表头区域: 行号 {target_row} < 数据起始行 {data_start_row}"
                        errors.append(error_msg)
                        logger.warning(error_msg)
                        continue
                    
                    if target_row > max_row:
                        error_msg = f"行号超出范围: {target_row} (最大: {max_row})"
                        errors.append(error_msg)
                        continue
                    
                    # 修改每个指定的列
                    for col_name, new_value in modify_values.items():
                        # 获取字段名
                        field_name = field_mapping.get(col_name) if field_mapping else col_name
                        
                        if not field_name:
                            continue
                        
                        # 查找字段对应的列
                        col_num = field_to_col.get(field_name)
                        
                        if col_num is None:
                            error_msg = f"未找到字段: {field_name}"
                            if error_msg not in errors:
                                errors.append(error_msg)
                            continue
                        
                        # 读取旧值
                        cell = ws.range((target_row, col_num))
                        old_value = cell.value
                        
                        # 比较新旧值，避免不必要的修改
                        old_str = str(old_value).strip() if old_value is not None else ''
                        new_str = str(new_value).strip() if new_value is not None else ''
                        
                        if old_str == new_str:
                            # 值相同，跳过修改
                            skipped_same_value += 1
                            continue
                        
                        # 修改单元格
                        cell.value = new_value
                        
                        # 记录修改日志
                        self.modification_logs.append({
                            'file': os.path.basename(excel_path),
                            'row': target_row,
                            'field': field_name,
                            'position': f"{get_column_letter(col_num)}{target_row}",
                            'old_value': old_value,
                            'new_value': new_value
                        })
                        
                        modified_count += 1
            
            # 保存文件 - 只有实际发生修改时才保存
            if modified_count > 0:
                wb.save()
                logger.info(f"[xlwings] 已修改并保存: {excel_path} ({modified_count} 处修改, {skipped_same_value} 处跳过-值相同)")
            else:
                logger.info(f"[xlwings] 无需修改: {excel_path} (数据未变化, {skipped_same_value} 处跳过-值相同)")
            
        except Exception as e:
            error_msg = f"修改文件失败 {excel_path}: {e}"
            errors.append(error_msg)
            logger.error(error_msg)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        
        return modified_count, errors, skipped_same_value
    
    def process_batch_modification_with_json(self, mapping_path: str, 
                                             excel_directory: str,
                                             table_col: str,
                                             id_col: str,
                                             mapping_sheet: str = None,
                                             backup: bool = True) -> Dict:
        """
        根据JSON配置执行批量修改（自动匹配字段）
        
        工作流程:
        1. 读取映射表中的每一行
        2. 根据表名从JSON配置中获取该表的字段列表
        3. 检查映射表中是否有同名的列
        4. 如果有，则用映射表中的值更新目标Excel
        
        Args:
            mapping_path: 映射表路径
            excel_directory: Excel文件所在目录
            table_col: 映射表中的表名列
            id_col: 映射表中的ID列
            mapping_sheet: 映射表工作表名称
            backup: 是否创建备份
            
        Returns:
            Dict: 处理结果统计
        """
        # 重置统计
        self.processing_stats = {
            'total_rows': 0,
            'processed_rows': 0,
            'modified_files': 0,
            'modified_cells': 0,
            'skipped_rows': 0,
            'skipped_no_config': 0,
            'errors': 0
        }
        self.error_logs = []
        self.modification_logs = []
        
        if not self.field_config:
            self._report_progress("错误: 未加载JSON配置文件")
            return self.processing_stats
        
        self._report_progress("开始加载映射表...")
        
        # 加载映射表
        df, mapping_columns = self.load_mapping_table(mapping_path, mapping_sheet)
        if df.empty:
            return self.processing_stats
        
        self._report_progress(f"映射表加载完成，共 {len(df)} 行数据")
        self._report_progress(f"映射表列: {mapping_columns}")
        
        # 验证列名是否存在
        if table_col not in mapping_columns:
            error_msg = f"错误: 映射表中不存在表名列 '{table_col}'，可用列: {mapping_columns}"
            self._report_progress(error_msg)
            self.error_logs.append(error_msg)
            return self.processing_stats
        
        # 检测是否有Position列（翻译提取格式）
        use_position_mode = 'Position' in mapping_columns
        
        if not use_position_mode:
            # 传统模式需要ID列
            if id_col not in mapping_columns:
                error_msg = f"错误: 映射表中不存在ID列 '{id_col}'，可用列: {mapping_columns}"
                self._report_progress(error_msg)
                self.error_logs.append(error_msg)
                return self.processing_stats
        else:
            self._report_progress("检测到Position列，使用Position直接定位模式")
        
        # 检测是否有Sheet列（三维表支持）
        has_sheet_column = 'Sheet' in mapping_columns
        if has_sheet_column:
            self._report_progress("检测到Sheet列，启用三维表支持")
        
        # 按表名和工作表分组修改
        # key: (table_name, sheet_name), value: list of modifications
        grouped_modifications = {}
        
        for idx, row in df.iterrows():
            # 获取表名
            table_name = str(row[table_col]).strip() if pd.notna(row[table_col]) else ''
            
            if not table_name:
                self.processing_stats['skipped_rows'] += 1
                continue
            
            # 获取工作表名（三维表支持）
            sheet_name = None
            if has_sheet_column:
                sheet_val = row.get('Sheet')
                if pd.notna(sheet_val) and str(sheet_val).strip():
                    sheet_name = str(sheet_val).strip()
            
            if use_position_mode:
                # Position模式
                position = row.get('Position')
                if pd.isna(position) or not str(position).strip():
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 从JSON配置中获取该表的字段列表
                table_fields = self.get_table_fields(table_name)
                
                if not table_fields:
                    # 该表不在JSON配置中，跳过
                    self.processing_stats['skipped_no_config'] += 1
                    continue
                
                # 找出映射表中与JSON字段匹配的列
                modify_values = {}
                for field in table_fields:
                    if field in mapping_columns and field in row.index:
                        value = row[field]
                        if pd.notna(value) and str(value).strip():
                            modify_values[field] = str(value).strip()
                
                if not modify_values:
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 使用 (table_name, sheet_name) 作为key
                group_key = (table_name, sheet_name)
                if group_key not in grouped_modifications:
                    grouped_modifications[group_key] = []
                
                grouped_modifications[group_key].append({
                    'position': position,
                    'modify_values': modify_values
                })
            else:
                # 传统ID模式
                id_value = row[id_col] if pd.notna(row[id_col]) else ''
                
                if not id_value:
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 从JSON配置中获取该表的字段列表
                table_fields = self.get_table_fields(table_name)
                
                if not table_fields:
                    # 该表不在JSON配置中，跳过
                    self.processing_stats['skipped_no_config'] += 1
                    continue
                
                # 找出映射表中与JSON字段匹配的列
                modify_values = {}
                for field in table_fields:
                    if field in mapping_columns and field in row.index:
                        value = row[field]
                        if pd.notna(value) and str(value).strip():
                            modify_values[field] = str(value).strip()
                
                if not modify_values:
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 使用 (table_name, sheet_name) 作为key
                group_key = (table_name, sheet_name)
                if group_key not in grouped_modifications:
                    grouped_modifications[group_key] = []
                
                grouped_modifications[group_key].append({
                    'id': id_value,
                    'modify_values': modify_values
                })
        
        self._report_progress(f"需要修改 {len(grouped_modifications)} 个文件/工作表组合")
        
        # 处理每个文件/工作表组合
        total_files = len(grouped_modifications)
        if total_files == 0:
            self._report_progress("没有找到需要修改的文件")
            return self.processing_stats
        
        processed_files = 0
        
        for (table_name, sheet_name), modifications in grouped_modifications.items():
            # 构建文件路径
            excel_path = os.path.join(excel_directory, table_name)
            
            if not os.path.exists(excel_path):
                error_msg = f"文件不存在: {excel_path}"
                self.error_logs.append(error_msg)
                self.processing_stats['errors'] += 1
                logger.warning(error_msg)
                continue
            
            # 创建备份
            if backup:
                backup_path = excel_path + '.bak'
                try:
                    import shutil
                    shutil.copy2(excel_path, backup_path)
                except Exception as e:
                    logger.warning(f"创建备份失败: {e}")
            
            # 获取该表需要修改的字段（用于显示）
            first_mod = modifications[0] if modifications else {}
            fields_to_modify = list(first_mod.get('modify_values', {}).keys())
            sheet_info = f" [工作表: {sheet_name}]" if sheet_name else ""
            self._report_progress(f"处理: {table_name}{sheet_info} (字段: {', '.join(fields_to_modify)})")
            
            # 修改文件 - 传递use_position参数、data_start_row保护表头和sheet_name支持三维表
            modified_count, errors, skipped_same = self.modify_excel_file(
                excel_path, 
                modifications, 
                field_mapping=None,  # 字段名已经是正确的
                data_start_row=DATA_START_ROW,  # 默认数据起始行，保护表头
                use_position=use_position_mode,  # 使用检测到的模式
                sheet_name=sheet_name  # 三维表支持
            )
            
            if modified_count > 0:
                self.processing_stats['modified_files'] += 1
                self.processing_stats['modified_cells'] += modified_count
            
            if skipped_same > 0:
                self.processing_stats['skipped_same_value'] = self.processing_stats.get('skipped_same_value', 0) + skipped_same
            
            self.processing_stats['processed_rows'] += len(modifications)
            
            if errors:
                self.error_logs.extend(errors)
                self.processing_stats['errors'] += len(errors)
            
            processed_files += 1
            progress = (processed_files / total_files) * 100
            self._report_progress(f"已处理: {processed_files}/{total_files} 文件", progress)
        
        # 关闭 xlwings Excel 应用
        self._close_excel_app()
        
        self._report_progress("批量修改完成！", 100)
        
        return self.processing_stats

    def process_batch_modification_by_json_language(self, mapping_path: str, 
                                                     excel_directory: str,
                                                     id_col: str = None,
                                                     field_col: str = None,
                                                     target_language: str = None,
                                                     backup: bool = True) -> Dict:
        """
        根据JSON中的语言代码自动匹配并执行批量修改
        
        工作流程:
        1. 从JSON配置中获取语言代码（如 'vn', 'zh', 'th'）
        2. 根据语言代码自动匹配映射表中的语言列（如 'VN', 'Support-CH', 'TH'）
        3. 根据语言代码自动匹配Excel中的目标字段（如 'name_vn', 'desc_cn'）
        4. 如果JSON中有 fields_by_language，直接使用对应语言的字段列表
        
        映射表结构:
        - 每个工作表名 = 目标Excel文件名（不含扩展名）
        - Classification列: 字段名（可选，如果没有则自动匹配）
        - ID列: 数据ID
        - 语言列: 根据JSON语言代码自动匹配
        
        Args:
            mapping_path: 映射表路径
            excel_directory: Excel文件所在目录
            id_col: 映射表中的ID列名（可选，自动检测）
            field_col: 字段名列名（可选，自动检测）
            target_language: 目标语言列名（可选，从JSON语言代码自动匹配）
            backup: 是否创建备份
            
        Returns:
            Dict: 处理结果统计
        """
        # 重置统计
        self.processing_stats = {
            'total_rows': 0,
            'processed_rows': 0,
            'modified_files': 0,
            'modified_cells': 0,
            'skipped_rows': 0,
            'skipped_no_config': 0,
            'skipped_no_file': 0,
            'skipped_same_value': 0,
            'errors': 0
        }
        self.error_logs = []
        self.modification_logs = []
        
        # 获取JSON中的语言代码
        json_lang_code = self.get_json_language()
        json_lang_name = self.get_json_language_name()
        
        if not json_lang_code:
            error_msg = "JSON配置中未找到语言标记（language字段）"
            self._report_progress(f"错误: {error_msg}")
            self.error_logs.append(error_msg)
            return self.processing_stats
        
        self._report_progress(f"JSON语言标记: {json_lang_name} ({json_lang_code})")
        
        # 获取所有工作表（CSV文件需要从Table列提取表名）
        file_ext = os.path.splitext(mapping_path)[1].lower()
        csv_mapping_df = None
        try:
            if file_ext == '.csv':
                # CSV文件，从Table列提取所有唯一的表名
                csv_mapping_df, _ = self.load_mapping_table(mapping_path)
                if csv_mapping_df is None or csv_mapping_df.empty:
                    error_msg = "CSV文件为空或加载失败"
                    self._report_progress(error_msg)
                    self.error_logs.append(error_msg)
                    return self.processing_stats
                
                if 'Table' not in csv_mapping_df.columns:
                    error_msg = "CSV文件缺少'Table'列"
                    self._report_progress(error_msg)
                    self.error_logs.append(error_msg)
                    return self.processing_stats
                
                # 提取唯一的表名（去除.xlsx/.xls扩展名）
                sheet_names = self._get_csv_mapping_table_names(csv_mapping_df)
                self._report_progress(f"映射表为CSV格式，包含 {len(sheet_names)} 个表文件")
            else:
                with pd.ExcelFile(mapping_path) as xl:
                    sheet_names = xl.sheet_names
                self._report_progress(f"映射表包含 {len(sheet_names)} 个工作表")
        except Exception as e:
            error_msg = f"无法读取映射表: {e}"
            self._report_progress(error_msg)
            self.error_logs.append(error_msg)
            return self.processing_stats
        
        # 自动检测映射表中的语言列（如果没有指定）
        detected_lang_column = target_language
        if not detected_lang_column:
            possible_columns = self.get_mapping_column_for_language(json_lang_code)
            
            # 读取第一个数据工作表来检测列
            skip_sheets = ['汇总信息', '汇总', 'Summary', 'summary', '说明', 'Info']
            
            for sheet in sheet_names:
                if sheet not in skip_sheets:
                    try:
                        if file_ext == '.csv':
                            if csv_mapping_df is None:
                                continue
                            columns = csv_mapping_df.columns.tolist()
                        else:
                            df_sample = pd.read_excel(mapping_path, sheet_name=sheet, nrows=0)
                            columns = df_sample.columns.tolist()
                        
                        for col in possible_columns:
                            if col in columns:
                                detected_lang_column = col
                                break
                        
                        if detected_lang_column:
                            break
                    except Exception:
                        continue
            
            if detected_lang_column:
                self._report_progress(f"自动检测到语言列: {detected_lang_column}")
            else:
                error_msg = f"无法在映射表中找到语言 '{json_lang_code}' 对应的列"
                self._report_progress(f"错误: {error_msg}")
                self.error_logs.append(error_msg)
                return self.processing_stats
        
        # 跳过汇总信息等非数据工作表
        skip_sheets = ['汇总信息', '汇总', 'Summary', 'summary', '说明', 'Info']
        data_sheets = [s for s in sheet_names if s not in skip_sheets]
        
        self._report_progress(f"将处理 {len(data_sheets)} 个数据工作表")
        
        total_sheets = len(data_sheets)
        processed_sheets = 0
        
        for sheet_name in data_sheets:
            # 工作表名就是目标Excel文件名
            table_name = sheet_name
            excel_filename = f"{table_name}.xlsx"
            excel_path = os.path.join(excel_directory, excel_filename)
            
            # 检查文件是否存在
            if not os.path.exists(excel_path):
                # 尝试其他扩展名
                excel_path_xls = os.path.join(excel_directory, f"{table_name}.xls")
                if os.path.exists(excel_path_xls):
                    excel_path = excel_path_xls
                else:
                    self.processing_stats['skipped_no_file'] += 1
                    continue
            
            # 获取该表对应语言的字段列表
            lang_fields = self.get_table_fields_by_language(table_name, json_lang_code)
            if not lang_fields:
                # 也尝试带扩展名
                lang_fields = self.get_table_fields_by_language(excel_filename, json_lang_code)
            
            if not lang_fields:
                self.processing_stats['skipped_no_config'] += 1
                continue
            
            # 读取该工作表的数据
            try:
                if file_ext == '.csv':
                    if csv_mapping_df is None or csv_mapping_df.empty:
                        continue
                    df = self._filter_csv_mapping_rows(csv_mapping_df, table_name)
                else:
                    df = pd.read_excel(mapping_path, sheet_name=sheet_name, header=0)
            except Exception as e:
                error_msg = f"读取工作表 {sheet_name} 失败: {e}"
                self.error_logs.append(error_msg)
                continue
            
            if df.empty:
                continue
            
            columns = df.columns.tolist()
            self.processing_stats['total_rows'] += len(df)
            
            # 检测是否使用Position模式（翻译提取CSV格式）
            use_position_mode = 'Position' in columns
            if use_position_mode:
                logger.info(f"工作表 {sheet_name} 检测到Position列，将使用Position直接定位模式")
            
            # 自动检测ID列
            actual_id_col = id_col
            if not actual_id_col or actual_id_col not in columns:
                for possible_id in ['ID', 'id', 'Id', 'ID列', '编号']:
                    if possible_id in columns:
                        actual_id_col = possible_id
                        break
            
            # 验证必要的列存在
            if actual_id_col not in columns:
                error_msg = f"工作表 {sheet_name} 中未找到ID列"
                self.error_logs.append(error_msg)
                continue
            
            if detected_lang_column not in columns:
                # 跳过没有该语言列的工作表
                continue
            
            # 确定字段列（Classification 或类似）
            actual_field_col = field_col
            if not actual_field_col:
                # 自动检测字段列
                for possible_col in ['Classification', 'classification', 'Field', 'field', '字段', '字段名']:
                    if possible_col in columns:
                        actual_field_col = possible_col
                        break
            
            # 收集该工作表的所有修改
            modifications = []
            skipped_no_field = 0
            
            for idx, row in df.iterrows():
                id_value = row[actual_id_col] if pd.notna(row[actual_id_col]) else ''
                lang_value = row[detected_lang_column] if pd.notna(row[detected_lang_column]) else ''
                
                if not id_value or not lang_value or not is_translatable_text(lang_value):
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 确定目标字段
                target_field = None
                
                # 如果有Classification列，使用它
                if actual_field_col and actual_field_col in columns:
                    field_from_mapping = str(row[actual_field_col]).strip() if pd.notna(row[actual_field_col]) else None
                    if field_from_mapping:
                        # 先直接检查该字段是否在语言字段列表中
                        if field_from_mapping in lang_fields:
                            target_field = field_from_mapping
                        else:
                            # 尝试匹配带语言后缀的字段
                            suffix_patterns = self.get_language_suffix_patterns(json_lang_code)
                            for lang_field in lang_fields:
                                field_base = field_from_mapping.lower()
                                lang_field_lower = lang_field.lower()
                                # 检查是否是同一个字段的语言版本
                                # 例如：Classification='des', lang_field='des_zh'
                                for pattern in suffix_patterns:
                                    if lang_field_lower == f"{field_base}{pattern}" or \
                                       lang_field_lower == f"{field_base.replace('_', '')}{pattern}":
                                        target_field = lang_field
                                        break
                                if target_field:
                                    break
                            
                            # 如果还没找到，尝试去掉语言后缀后匹配
                            if not target_field:
                                for lang_field in lang_fields:
                                    # 从lang_field中去掉语言后缀，看是否匹配field_from_mapping
                                    for pattern in suffix_patterns:
                                        if pattern.startswith('_') and lang_field.lower().endswith(pattern):
                                            base_name = lang_field[:-len(pattern)]
                                            if base_name.lower() == field_from_mapping.lower():
                                                target_field = lang_field
                                                break
                                    if target_field:
                                        break
                            
                            # 如果还是没找到，直接使用Classification中的字段名
                            if not target_field:
                                target_field = field_from_mapping
                                logger.debug(f"使用Classification字段名: {field_from_mapping}")
                else:
                    # 没有Classification列，尝试智能匹配
                    # 如果只有一个语言字段，直接使用
                    if len(lang_fields) == 1:
                        target_field = lang_fields[0]
                    elif len(lang_fields) > 1:
                        # 多个字段时，尝试根据映射表列名推断
                        # 例如：映射表列名是'VN'，可能对应'name_vn'或'desc_vn'
                        # 优先选择包含常见名称字段（name, title, text等）
                        common_name_fields = ['name', 'title', 'text', 'label', 'des', 'desc']
                        for common in common_name_fields:
                            for field in lang_fields:
                                if common in field.lower():
                                    target_field = field
                                    break
                            if target_field:
                                break
                        
                        # 如果还没找到，使用第一个字段
                        if not target_field:
                            target_field = lang_fields[0]
                            logger.debug(f"没有Classification列，使用第一个语言字段: {target_field}")
                
                if target_field:
                    modification_item = {
                        'id': id_value,
                        'modify_values': {target_field: str(lang_value).strip()}
                    }
                    # 如果是Position模式，添加Position信息
                    if use_position_mode and 'Position' in columns:
                        position_value = row['Position'] if pd.notna(row['Position']) else ''
                        if position_value:
                            modification_item['position'] = str(position_value).strip()
                    modifications.append(modification_item)
                else:
                    skipped_no_field += 1
            
            if skipped_no_field > 0:
                logger.warning(f"工作表 {sheet_name}: 有 {skipped_no_field} 行因未找到目标字段而跳过")
            
            if not modifications:
                continue
            
            # 创建备份
            if backup:
                backup_path = excel_path + '.bak'
                try:
                    import shutil
                    shutil.copy2(excel_path, backup_path)
                except Exception as e:
                    logger.warning(f"创建备份失败: {e}")
            
            self._report_progress(f"处理: {table_name} ({len(modifications)} 条修改，语言字段: {lang_fields[:3]}...)")
            
            # 修改文件（添加data_start_row保护表头）
            modified_count, errors, skipped_same = self.modify_excel_file(
                excel_path, 
                modifications, 
                field_mapping=None,
                data_start_row=DATA_START_ROW,  # 默认数据起始行，保护表头
                use_position=use_position_mode  # 使用检测到的模式
            )
            
            if modified_count > 0:
                self.processing_stats['modified_files'] += 1
                self.processing_stats['modified_cells'] += modified_count
            
            if skipped_same > 0:
                self.processing_stats['skipped_same_value'] = self.processing_stats.get('skipped_same_value', 0) + skipped_same
            
            self.processing_stats['processed_rows'] += len(modifications)
            
            if errors:
                self.error_logs.extend(errors)
                self.processing_stats['errors'] += len(errors)
            
            processed_sheets += 1
            progress = (processed_sheets / total_sheets) * 100
            self._report_progress(f"已处理: {processed_sheets}/{total_sheets} 工作表", progress)
        
        # 关闭 xlwings Excel 应用
        self._close_excel_app()
        
        self._report_progress("批量修改完成！", 100)
        
        return self.processing_stats
    def process_batch_modification(self, mapping_path: str, 
                                  excel_directory: str,
                                  table_col: str,
                                  id_col: str,
                                  modify_cols: List[str],
                                  field_mapping: Dict[str, str] = None,
                                  mapping_sheet: str = None,
                                  backup: bool = True) -> Dict:
        """
        执行批量修改
        
        Args:
            mapping_path: 映射表路径
            excel_directory: Excel文件所在目录
            table_col: 映射表中的表名列
            id_col: 映射表中的ID列
            modify_cols: 映射表中要用于修改的列名列表
            field_mapping: 映射表列名到Excel字段名的映射
            mapping_sheet: 映射表工作表名称
            backup: 是否创建备份
            
        Returns:
            Dict: 处理结果统计
        """
        # 重置统计
        self.processing_stats = {
            'total_rows': 0,
            'processed_rows': 0,
            'modified_files': 0,
            'modified_cells': 0,
            'skipped_rows': 0,
            'errors': 0
        }
        self.error_logs = []
        self.modification_logs = []
        
        self._report_progress("开始加载映射表...")
        
        # 加载映射表
        df, columns = self.load_mapping_table(mapping_path, mapping_sheet)
        if df.empty:
            return self.processing_stats
        
        self._report_progress(f"映射表加载完成，共 {len(df)} 行数据")
        
        # 按表名分组修改
        grouped_modifications = {}
        
        for idx, row in df.iterrows():
            parsed = self.parse_mapping_row(row, columns, table_col, id_col, modify_cols)
            
            if not parsed or not parsed.get('table_name') or not parsed.get('modify_values'):
                self.processing_stats['skipped_rows'] += 1
                continue
            
            table_name = parsed['table_name']
            
            if table_name not in grouped_modifications:
                grouped_modifications[table_name] = []
            
            grouped_modifications[table_name].append({
                'id': parsed['id'],
                'modify_values': parsed['modify_values']
            })
        
        self._report_progress(f"需要修改 {len(grouped_modifications)} 个文件")
        
        # 处理每个文件
        total_files = len(grouped_modifications)
        processed_files = 0
        
        for table_name, modifications in grouped_modifications.items():
            # 构建文件路径
            excel_path = os.path.join(excel_directory, table_name)
            
            if not os.path.exists(excel_path):
                error_msg = f"文件不存在: {excel_path}"
                self.error_logs.append(error_msg)
                self.processing_stats['errors'] += 1
                logger.warning(error_msg)
                continue
            
            # 创建备份
            if backup:
                backup_path = excel_path + '.bak'
                try:
                    import shutil
                    shutil.copy2(excel_path, backup_path)
                except Exception as e:
                    logger.warning(f"创建备份失败: {e}")
            
            # 修改文件（添加data_start_row保护表头）
            modified_count, errors, skipped_same = self.modify_excel_file(
                excel_path, 
                modifications, 
                field_mapping,
                data_start_row=DATA_START_ROW  # 默认数据起始行，保护表头
            )
            
            if modified_count > 0:
                self.processing_stats['modified_files'] += 1
                self.processing_stats['modified_cells'] += modified_count
            
            if skipped_same > 0:
                self.processing_stats['skipped_same_value'] = self.processing_stats.get('skipped_same_value', 0) + skipped_same
            
            self.processing_stats['processed_rows'] += len(modifications)
            
            if errors:
                self.error_logs.extend(errors)
                self.processing_stats['errors'] += len(errors)
            
            processed_files += 1
            progress = (processed_files / total_files) * 100
            self._report_progress(f"已处理: {processed_files}/{total_files} 文件", progress)
        
        # 关闭 xlwings Excel 应用
        self._close_excel_app()
        
        self._report_progress("批量修改完成！", 100)
        
        return self.processing_stats
    
    def generate_modification_report(self, output_path: str) -> bool:
        """
        生成修改报告
        
        Args:
            output_path: 输出Excel路径
            
        Returns:
            bool: 成功返回True
        """
        try:
            if not self.modification_logs:
                logger.warning("没有修改记录可导出")
                return False
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "修改记录"
            
            # 设置表头
            headers = ['文件名', 'ID', '字段名', 'Excel位置', '原值', '新值']
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for log in self.modification_logs:
                ws.append([
                    log.get('file', ''),
                    log.get('id', ''),
                    log.get('field', ''),
                    log.get('position', ''),
                    log.get('old_value', ''),
                    log.get('new_value', '')
                ])
            
            # 调整列宽
            for col_idx, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # 添加统计页
            ws_stats = wb.create_sheet(title="统计信息")
            stats_data = [
                ['统计项', '数值'],
                ['映射表总行数', self.processing_stats['total_rows']],
                ['已处理行数', self.processing_stats['processed_rows']],
                ['跳过行数', self.processing_stats['skipped_rows']],
                ['修改的文件数', self.processing_stats['modified_files']],
                ['修改的单元格数', self.processing_stats['modified_cells']],
                ['错误数', self.processing_stats['errors']]
            ]
            
            for row_data in stats_data:
                ws_stats.append(row_data)
            
            # 设置统计页表头样式
            for col_idx in range(1, 3):
                cell = ws_stats.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
            
            # 如果有错误，添加错误日志页
            if self.error_logs:
                ws_errors = wb.create_sheet(title="错误日志")
                ws_errors.append(['错误信息'])
                ws_errors.cell(row=1, column=1).fill = PatternFill(
                    start_color='FF0000', end_color='FF0000', fill_type='solid')
                ws_errors.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF')
                
                for error in self.error_logs:
                    ws_errors.append([error])
            
            wb.save(output_path)
            logger.info(f"修改报告已保存: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成报告失败: {e}")
            return False
    
    def get_stats_summary(self) -> str:
        """
        获取统计摘要文本
        
        Returns:
            str: 统计摘要
        """
        return f"""批量修改完成！

统计信息:
- 映射表总行数: {self.processing_stats['total_rows']}
- 已处理行数: {self.processing_stats['processed_rows']}
- 跳过行数: {self.processing_stats['skipped_rows']}
- 修改的文件数: {self.processing_stats['modified_files']}
- 修改的单元格数: {self.processing_stats['modified_cells']}
- 错误数: {self.processing_stats['errors']}

修改记录数: {len(self.modification_logs)} 条"""

    def get_mapping_file_languages(self, mapping_path: str) -> List[str]:
        """
        获取映射表中可用的语言列（支持Excel和CSV）
        
        Args:
            mapping_path: 映射表文件路径（.xlsx/.xls/.csv）
            
        Returns:
            List[str]: 语言列名列表
        """
        # 常见的语言列名
        language_columns = ['VN', 'EN', 'TH', 'Support-CH', 'Polish-CH', 'VN.1', 
                           'CN', 'JP', 'KR', 'TW', 'RU', 'DE', 'FR', 'ES', 'PT']
        
        # 需要排除的列名
        exclude_columns = ['Classification', 'classification', 'ID', 'id', 'Field', 'field',
                          '字段', '字段名', '表名', 'Table', 'table', '项目', '值', 'Name', 'name']
        
        try:
            # 检查文件扩展名
            file_ext = os.path.splitext(mapping_path)[1].lower()
            
            if file_ext == '.csv':
                # 读取CSV文件（只读取列名）
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        df = pd.read_csv(mapping_path, nrows=0, encoding=encoding)
                        columns = df.columns.tolist()
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    df = pd.read_csv(mapping_path, nrows=0, encoding='utf-8', errors='ignore')
                    columns = df.columns.tolist()
            else:
                # Excel文件
                with pd.ExcelFile(mapping_path) as xl:
                    # 跳过汇总信息等非数据工作表
                    skip_sheets = ['汇总信息', '汇总', 'Summary', 'summary', '说明', 'Info']
                    data_sheet = None
                    for sheet in xl.sheet_names:
                        if sheet not in skip_sheets:
                            data_sheet = sheet
                            break

                    if not data_sheet:
                        data_sheet = xl.sheet_names[0] if xl.sheet_names else None

                    if data_sheet:
                        df = pd.read_excel(xl, sheet_name=data_sheet, nrows=0)
                        columns = df.columns.tolist()
                    else:
                        return []
            
            # 过滤出语言列
            if columns:
                
                # 过滤出语言列
                available_langs = []
                for col in columns:
                    # 排除非语言列
                    if col in exclude_columns:
                        continue
                    
                    col_upper = str(col).upper()
                    # 匹配预定义语言列
                    if col in language_columns or col_upper in [l.upper() for l in language_columns]:
                        available_langs.append(col)
                    # 也检查是否包含语言相关关键词
                    elif any(lang in col_upper for lang in ['VN', 'EN', 'TH', 'CH', 'CN', 'JP', 'KR']):
                        available_langs.append(col)
                
                return available_langs
        except Exception as e:
            logger.error(f"获取语言列失败: {e}")
        
        return []

    def process_batch_modification_by_language(self, mapping_path: str, 
                                               excel_directory: str,
                                               id_col: str,
                                               target_language: str,
                                               field_col: str = None,
                                               backup: bool = True,
                                               field_row: int = FIELD_NAME_ROW,
                                               data_start_row: int = DATA_START_ROW) -> Dict:
        """
        按语言执行批量修改（每个工作表对应一个Excel文件）
        
        映射表结构:
        - 每个工作表名 = 目标Excel文件名（不含扩展名）
        - Classification列: 字段名
        - ID列: 数据ID
        - 语言列（如 VN, EN, TH）: 要修改的内容
        
        Args:
            mapping_path: 映射表路径
            excel_directory: Excel文件所在目录
            id_col: 映射表中的ID列名
            target_language: 目标语言列名（如 'VN', 'EN', 'TH'）
            field_col: 字段名列名（如 'Classification'），可选
            backup: 是否创建备份
            field_row: 目标Excel的字段名所在行（默认 FIELD_NAME_ROW）
            data_start_row: 数据起始行，小于此行号的将被跳过（默认 DATA_START_ROW，保护表头）
            
        Returns:
            Dict: 处理结果统计
        """
        # 重置统计
        self.processing_stats = {
            'total_rows': 0,
            'processed_rows': 0,
            'modified_files': 0,
            'modified_cells': 0,
            'skipped_rows': 0,
            'skipped_no_config': 0,
            'skipped_no_file': 0,
            'skipped_field_mismatch': 0,  # CSV字段不在JSON配置中
            'skipped_same_value': 0,  # 值相同跳过的数量
            'errors': 0
        }
        self.error_logs = []
        self.modification_logs = []
        
        self._report_progress(f"开始按语言批量修改，目标语言: {target_language}")
        
        # 获取所有工作表（CSV文件需要从Table列提取表名）
        file_ext = os.path.splitext(mapping_path)[1].lower()
        csv_mapping_df = None
        try:
            if file_ext == '.csv':
                # CSV文件，从Table列提取所有唯一的表名
                csv_mapping_df, _ = self.load_mapping_table(mapping_path)
                if csv_mapping_df is None or csv_mapping_df.empty:
                    error_msg = "CSV文件为空或加载失败"
                    self._report_progress(error_msg)
                    self.error_logs.append(error_msg)
                    return self.processing_stats
                
                if 'Table' not in csv_mapping_df.columns:
                    error_msg = "CSV文件缺少'Table'列"
                    self._report_progress(error_msg)
                    self.error_logs.append(error_msg)
                    return self.processing_stats
                
                # 提取唯一的表名（去除.xlsx/.xls扩展名）
                sheet_names = self._get_csv_mapping_table_names(csv_mapping_df)
                self._report_progress(f"映射表为CSV格式，包含 {len(sheet_names)} 个表文件")
            else:
                with pd.ExcelFile(mapping_path) as xl:
                    sheet_names = xl.sheet_names
                self._report_progress(f"映射表包含 {len(sheet_names)} 个工作表")
        except Exception as e:
            error_msg = f"无法读取映射表: {e}"
            self._report_progress(error_msg)
            self.error_logs.append(error_msg)
            return self.processing_stats
        
        # 跳过汇总信息等非数据工作表
        skip_sheets = ['汇总信息', '汇总', 'Summary', 'summary', '说明', 'Info']
        data_sheets = [s for s in sheet_names if s not in skip_sheets]
        
        self._report_progress(f"将处理 {len(data_sheets)} 个数据工作表")
        
        total_sheets = len(data_sheets)
        processed_sheets = 0
        
        for sheet_name in data_sheets:
            # 工作表名就是目标Excel文件名
            table_name = sheet_name
            excel_filename = f"{table_name}.xlsx"
            excel_path = os.path.join(excel_directory, excel_filename)
            
            # 检查文件是否存在
            if not os.path.exists(excel_path):
                # 尝试其他扩展名
                excel_path_xls = os.path.join(excel_directory, f"{table_name}.xls")
                if os.path.exists(excel_path_xls):
                    excel_path = excel_path_xls
                else:
                    self.processing_stats['skipped_no_file'] += 1
                    self._report_progress(f"  跳过: 文件不存在 - {excel_filename}")
                    continue
            
            # 检查JSON配置中是否有该表
            table_fields = self.get_table_fields(table_name)
            if not table_fields:
                # 也尝试带扩展名
                table_fields = self.get_table_fields(excel_filename)
            
            if not table_fields:
                self.processing_stats['skipped_no_config'] += 1
                continue
            
            # 读取该工作表的数据（支持CSV）
            try:
                if file_ext == '.csv':
                    if csv_mapping_df is None or csv_mapping_df.empty:
                        continue
                    df = self._filter_csv_mapping_rows(csv_mapping_df, table_name)
                    if df.empty:
                        self._report_progress(f"  跳过: CSV中无 {table_name} 的数据")
                        continue
                else:
                    # Excel文件，读取指定工作表
                    df = pd.read_excel(mapping_path, sheet_name=sheet_name, header=0)
            except Exception as e:
                error_msg = f"读取{'CSV文件' if file_ext == '.csv' else '工作表 ' + sheet_name} 失败: {e}"
                self.error_logs.append(error_msg)
                continue
            
            if df.empty:
                continue
            
            columns = df.columns.tolist()
            self.processing_stats['total_rows'] += len(df)
            
            # 检测是否使用Position模式（翻译提取CSV格式）
            use_position_mode = 'Position' in columns
            if use_position_mode:
                logger.info(f"工作表 {sheet_name} 检测到Position列，将使用Position直接定位模式")
            
            # 自动检测ID列
            actual_id_col = id_col
            if not actual_id_col or actual_id_col not in columns:
                for possible_id in ['ID', 'id', 'Id', 'ID列', '编号']:
                    if possible_id in columns:
                        actual_id_col = possible_id
                        break
            
            # 验证必要的列存在
            if actual_id_col not in columns:
                error_msg = f"工作表 {sheet_name} 中未找到ID列"
                self.error_logs.append(error_msg)
                continue
            
            if target_language not in columns:
                error_msg = f"工作表 {sheet_name} 中不存在语言列 '{target_language}'"
                self.error_logs.append(error_msg)
                continue
            
            # 确定字段列（Classification 或类似）
            actual_field_col = field_col
            if not actual_field_col:
                # 自动检测字段列
                for possible_col in ['Classification', 'classification', 'Field', 'field', '字段', '字段名']:
                    if possible_col in columns:
                        actual_field_col = possible_col
                        break
            
            # 收集该工作表的所有修改
            modifications = []
            skipped_field_mismatch = 0  # 记录因字段不匹配而跳过的行数
            
            for idx, row in df.iterrows():
                id_value = row[actual_id_col] if pd.notna(row[actual_id_col]) else ''
                lang_value = row[target_language] if pd.notna(row[target_language]) else ''
                
                if not id_value or not lang_value or not is_translatable_text(lang_value):
                    self.processing_stats['skipped_rows'] += 1
                    continue
                
                # 获取字段名
                field_name = None
                if actual_field_col and actual_field_col in columns:
                    field_name = str(row[actual_field_col]).strip() if pd.notna(row[actual_field_col]) else None
                
                # 如果没有字段列，尝试从JSON配置匹配
                # 在JSON配置的字段列表中查找与目标语言对应的字段
                target_field = None
                if field_name:
                    # 有Classification列，校验字段是否在JSON配置中
                    if table_fields and field_name not in table_fields:
                        # 字段不在JSON配置中，跳过此行
                        skipped_field_mismatch += 1
                        self.processing_stats['skipped_rows'] += 1
                        self.processing_stats['skipped_field_mismatch'] += 1
                        continue
                    target_field = field_name
                else:
                    # 没有Classification列，尝试匹配语言字段
                    lang_lower = target_language.lower().replace('-', '_').replace('.', '_')
                    for field in table_fields:
                        field_lower = field.lower()
                        if lang_lower in field_lower or field_lower.endswith(f'_{lang_lower}'):
                            target_field = field
                            break
                    
                    # 如果还没找到，使用语言列名本身
                    if not target_field:
                        target_field = target_language
                
                if target_field:
                    modification_item = {
                        'id': id_value,
                        'modify_values': {target_field: str(lang_value).strip()}
                    }
                    # 如果是Position模式，添加Position信息
                    if use_position_mode and 'Position' in columns:
                        position_value = row['Position'] if pd.notna(row['Position']) else ''
                        if position_value:
                            modification_item['position'] = str(position_value).strip()
                    modifications.append(modification_item)
            
            # 输出字段不匹配的警告
            if skipped_field_mismatch > 0:
                warn_msg = f"  ⚠️ {table_name}: 跳过 {skipped_field_mismatch} 行（CSV字段不在JSON配置中）"
                self._report_progress(warn_msg)
                logger.warning(warn_msg)
            
            if not modifications:
                continue
            
            # 创建备份
            if backup:
                backup_path = excel_path + '.bak'
                try:
                    import shutil
                    shutil.copy2(excel_path, backup_path)
                except Exception as e:
                    logger.warning(f"创建备份失败: {e}")

            self._report_progress(f"处理: {table_name} ({len(modifications)} 条修改){'[Position模式]' if use_position_mode else ''}")

            # 修改文件（传递field_row和data_start_row以保护表头）
            modified_count, errors, skipped_same = self.modify_excel_file(
                excel_path,
                modifications,
                field_mapping=None,
                field_row=field_row,
                data_start_row=data_start_row,
                use_position=use_position_mode  # 使用检测到的模式
            )
            
            if modified_count > 0:
                self.processing_stats['modified_files'] += 1
                self.processing_stats['modified_cells'] += modified_count
            
            if skipped_same > 0:
                self.processing_stats['skipped_same_value'] += skipped_same
            
            self.processing_stats['processed_rows'] += len(modifications)
            
            if errors:
                self.error_logs.extend(errors)
                self.processing_stats['errors'] += len(errors)
            
            processed_sheets += 1
            progress = (processed_sheets / total_sheets) * 100
            self._report_progress(f"已处理: {processed_sheets}/{total_sheets} 工作表", progress)
        
        # 关闭 xlwings Excel 应用
        self._close_excel_app()
        
        self._report_progress("批量修改完成！", 100)
        
        return self.processing_stats
