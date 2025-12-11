#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel配置同步器
将一个目录的Excel文件配置同步到其他目录的同名Excel文件上
"""

import os
import json
import shutil
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional, Callable
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelConfigSync:
    """Excel配置同步器 - 将源目录的Excel配置同步到目标目录"""
    
    def __init__(self):
        """初始化配置同步器"""
        self.supported_extensions = {'.xlsx', '.xls'}
        
        # 处理统计
        self.processing_stats = {
            'source_files': 0,          # 源目录文件数
            'target1_synced': 0,        # 目标目录1同步的文件数
            'target2_synced': 0,        # 目标目录2同步的文件数
            'target1_skipped': 0,       # 目标目录1跳过的文件数
            'target2_skipped': 0,       # 目标目录2跳过的文件数
            'total_cells_synced': 0,    # 同步的单元格数
            'errors': 0                  # 错误数
        }
        
        # 同步日志
        self.sync_logs = []
        
        # 错误日志
        self.error_logs = []
        
        # JSON配置（仅用于参考，不做修改）
        self.json_config = {}
        self.json_path = ""
        
        # 字段过滤配置
        self.filter_config = {}
        self.filter_path = ""
        self.skip_fields = {}  # {表名: [字段名列表]}
        
        # 进度回调
        self.progress_callback = None
        
        # 配置选项
        self.sync_options = {
            'sync_values': True,          # 同步单元格值
            'sync_formulas': True,        # 同步公式
            'sync_styles': False,         # 同步样式（可选）
            'sync_column_widths': False,  # 同步列宽（可选）
            'backup_before_sync': True,   # 同步前备份
            'skip_first_n_rows': 0,       # 跳过前N行（表头）
        }
    
    def set_progress_callback(self, callback: Callable[[str, float], None]):
        """设置进度回调函数"""
        self.progress_callback = callback
    
    def _report_progress(self, message: str, percentage: float = None):
        """报告进度"""
        if self.progress_callback:
            self.progress_callback(message, percentage)
        logger.info(message)
    
    def reset_stats(self):
        """重置统计信息"""
        self.processing_stats = {
            'source_files': 0,
            'target1_synced': 0,
            'target2_synced': 0,
            'target1_skipped': 0,
            'target2_skipped': 0,
            'total_cells_synced': 0,
            'total_cells_skipped': 0,
            'errors': 0
        }
        self.sync_logs = []
        self.error_logs = []
    
    def load_json_config(self, json_path: str) -> Dict:
        """
        加载JSON配置文件（仅用于参考，不做修改）
        
        Args:
            json_path: JSON配置文件路径
            
        Returns:
            Dict: JSON配置内容
        """
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                self.json_config = json.load(f)
            self.json_path = json_path
            logger.info(f"成功加载JSON配置: {json_path}")
            return self.json_config
        except Exception as e:
            error_msg = f"加载JSON配置失败: {e}"
            logger.error(error_msg)
            self.error_logs.append(error_msg)
            return {}
    
    def load_filter_config(self, filter_path: str) -> Dict:
        """
        加载字段过滤配置文件
        
        配置格式示例：
        {
            "skip_fields": {
                "table1.xlsx": ["field1", "field2"],
                "table2.xlsx": ["field3"]
            }
        }
        
        或者:
        {
            "text_tables": [
                {
                    "table_name": "table1.xlsx",
                    "skip_fields": ["field1", "field2"]
                }
            ]
        }
        
        Args:
            filter_path: 过滤配置文件路径
            
        Returns:
            Dict: 过滤配置内容
        """
        try:
            with open(filter_path, 'r', encoding='utf-8') as f:
                self.filter_config = json.load(f)
            self.filter_path = filter_path
            
            # 解析过滤配置
            self.skip_fields = {}
            
            # 格式1: 直接的 skip_fields 字典
            if 'skip_fields' in self.filter_config:
                self.skip_fields = self.filter_config['skip_fields']
            
            # 格式2: text_tables 数组格式
            elif 'text_tables' in self.filter_config:
                for table_info in self.filter_config['text_tables']:
                    table_name = table_info.get('table_name', '')
                    skip_fields = table_info.get('skip_fields', [])
                    if table_name and skip_fields:
                        self.skip_fields[table_name] = skip_fields
                        # 同时支持不带扩展名的表名
                        table_key = Path(table_name).stem
                        self.skip_fields[table_key] = skip_fields
            
            logger.info(f"成功加载过滤配置: {filter_path}")
            logger.info(f"  - 包含 {len(self.skip_fields)} 个表的过滤字段配置")
            return self.filter_config
        
        except Exception as e:
            error_msg = f"加载过滤配置失败: {e}"
            logger.error(error_msg)
            self.error_logs.append(error_msg)
            return {}
    
    def get_skip_fields_for_table(self, table_name: str) -> List[str]:
        """
        获取指定表需要跳过的字段列表
        
        Args:
            table_name: 表名（可以带或不带扩展名）
            
        Returns:
            List[str]: 需要跳过的字段名列表
        """
        # 尝试直接匹配
        if table_name in self.skip_fields:
            return self.skip_fields[table_name]
        
        # 尝试不带扩展名匹配
        table_key = Path(table_name).stem
        if table_key in self.skip_fields:
            return self.skip_fields[table_key]
        
        return []
    
    def get_excel_files(self, directory: str) -> List[str]:
        """
        获取目录中的所有Excel文件
        
        Args:
            directory: 目录路径
            
        Returns:
            List[str]: Excel文件名列表（不含路径）
        """
        files = []
        try:
            for file in os.listdir(directory):
                if Path(file).suffix.lower() in self.supported_extensions:
                    files.append(file)
        except Exception as e:
            logger.error(f"读取目录失败: {e}")
        return files
    
    def find_matching_files(self, source_dir: str, target_dirs: List[str]) -> Dict[str, Dict]:
        """
        查找源目录和目标目录中匹配的文件
        
        Args:
            source_dir: 源目录路径
            target_dirs: 目标目录路径列表
            
        Returns:
            Dict: 文件匹配信息 {文件名: {'source': 源路径, 'targets': [目标路径列表]}}
        """
        source_files = set(self.get_excel_files(source_dir))
        self.processing_stats['source_files'] = len(source_files)
        
        matching = {}
        
        for filename in source_files:
            source_path = os.path.join(source_dir, filename)
            targets = []
            
            for target_dir in target_dirs:
                target_path = os.path.join(target_dir, filename)
                if os.path.exists(target_path):
                    targets.append(target_path)
            
            if targets:
                matching[filename] = {
                    'source': source_path,
                    'targets': targets
                }
        
        return matching
    
    def sync_excel_file(self, source_path: str, target_path: str, 
                        sync_sheets: List[str] = None) -> Tuple[int, List[str]]:
        """
        将源Excel文件的配置同步到目标Excel文件
        
        Args:
            source_path: 源Excel文件路径
            target_path: 目标Excel文件路径
            sync_sheets: 要同步的工作表列表（None表示全部）
            
        Returns:
            Tuple[int, List[str]]: (同步的单元格数, 错误信息列表)
        """
        errors = []
        cells_synced = 0
        
        try:
            # 创建备份
            if self.sync_options['backup_before_sync']:
                backup_path = target_path + '.bak'
                try:
                    shutil.copy2(target_path, backup_path)
                except Exception as e:
                    logger.warning(f"创建备份失败: {e}")
            
            # 加载工作簿，保留所有原有结构（VBA、链接、样式、批注等）
            source_wb = load_workbook(source_path, data_only=False, keep_vba=True, keep_links=True, rich_text=True)
            target_wb = load_workbook(target_path, data_only=False, keep_vba=True, keep_links=True, rich_text=True)
            
            # 确定要同步的工作表
            if sync_sheets is None:
                sheets_to_sync = source_wb.sheetnames
            else:
                sheets_to_sync = [s for s in sync_sheets if s in source_wb.sheetnames]
            
            # 获取该表的过滤字段
            table_name = os.path.basename(source_path)
            skip_fields = self.get_skip_fields_for_table(table_name)
            
            for sheet_name in sheets_to_sync:
                if sheet_name not in target_wb.sheetnames:
                    # 目标文件中没有该工作表，跳过
                    continue
                
                source_ws = source_wb[sheet_name]
                target_ws = target_wb[sheet_name]
                
                # 如果有过滤字段，构建字段行到列索引的映射
                skip_cols = set()
                if skip_fields:
                    # 假设字段名在第5行（可配置）
                    field_row = 5
                    for col in range(1, source_ws.max_column + 1):
                        cell_value = source_ws.cell(row=field_row, column=col).value
                        if cell_value and str(cell_value).strip() in skip_fields:
                            skip_cols.add(col)
                            logger.debug(f"跳过字段列: {cell_value} (列 {col})")
                
                # 同步单元格内容
                skip_rows = self.sync_options['skip_first_n_rows']
                cells_skipped = 0
                
                for row in range(skip_rows + 1, source_ws.max_row + 1):
                    for col in range(1, source_ws.max_column + 1):
                        # 检查是否需要跳过该列
                        if col in skip_cols:
                            cells_skipped += 1
                            continue
                        
                        source_cell = source_ws.cell(row=row, column=col)
                        target_cell = target_ws.cell(row=row, column=col)
                        
                        # 同步值或公式
                        if self.sync_options['sync_formulas'] and source_cell.data_type == 'f':
                            # 同步公式
                            target_cell.value = source_cell.value
                        elif self.sync_options['sync_values']:
                            target_cell.value = source_cell.value
                        
                        # 同步样式（可选）
                        if self.sync_options['sync_styles']:
                            target_cell.font = source_cell.font.copy() if source_cell.font else None
                            target_cell.fill = source_cell.fill.copy() if source_cell.fill else None
                            target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else None
                            target_cell.border = source_cell.border.copy() if source_cell.border else None
                        
                        cells_synced += 1
                
                # 更新跳过的单元格统计
                self.processing_stats['total_cells_skipped'] += cells_skipped
                
                # 同步列宽（可选）
                if self.sync_options['sync_column_widths']:
                    for col_letter, col_dim in source_ws.column_dimensions.items():
                        if col_dim.width:
                            target_ws.column_dimensions[col_letter].width = col_dim.width
                
                # 记录同步日志
                self.sync_logs.append({
                    'source': source_path,
                    'target': target_path,
                    'sheet': sheet_name,
                    'cells_synced': cells_synced,
                    'status': 'success'
                })
            
            # 保存目标文件
            target_wb.save(target_path)
            source_wb.close()
            target_wb.close()
            
        except Exception as e:
            error_msg = f"同步文件失败 [{os.path.basename(source_path)} -> {os.path.basename(target_path)}]: {e}"
            errors.append(error_msg)
            logger.error(error_msg)
        
        return cells_synced, errors
    
    def sync_directories(self, source_dir: str, target_dir1: str, target_dir2: str = None,
                        file_filter: List[str] = None) -> Dict:
        """
        将源目录的Excel配置同步到目标目录
        
        Args:
            source_dir: 源目录路径
            target_dir1: 目标目录1路径
            target_dir2: 目标目录2路径（可选）
            file_filter: 文件名过滤列表（仅同步这些文件）
            
        Returns:
            Dict: 处理统计信息
        """
        self.reset_stats()
        
        # 构建目标目录列表
        target_dirs = [target_dir1]
        if target_dir2:
            target_dirs.append(target_dir2)
        
        # 查找匹配的文件
        self._report_progress("正在扫描目录...", 0)
        matching_files = self.find_matching_files(source_dir, target_dirs)
        
        if file_filter:
            # 应用文件过滤
            matching_files = {k: v for k, v in matching_files.items() 
                            if k in file_filter or Path(k).stem in file_filter}
        
        total_files = len(matching_files)
        if total_files == 0:
            self._report_progress("没有找到匹配的文件", 100)
            return self.processing_stats
        
        self._report_progress(f"找到 {total_files} 个匹配的文件", 5)
        
        # 同步每个文件
        processed = 0
        for filename, file_info in matching_files.items():
            source_path = file_info['source']
            
            for idx, target_path in enumerate(file_info['targets']):
                target_dir = os.path.dirname(target_path)
                target_num = idx + 1
                
                self._report_progress(
                    f"正在同步: {filename} -> 目标目录{target_num}",
                    5 + (processed / total_files) * 90
                )
                
                cells_synced, errors = self.sync_excel_file(source_path, target_path)
                
                if errors:
                    self.error_logs.extend(errors)
                    self.processing_stats['errors'] += len(errors)
                else:
                    self.processing_stats['total_cells_synced'] += cells_synced
                    if target_num == 1:
                        self.processing_stats['target1_synced'] += 1
                    else:
                        self.processing_stats['target2_synced'] += 1
            
            # 记录未找到的目标文件
            if target_dir1 and not any(target_dir1 in t for t in file_info['targets']):
                self.processing_stats['target1_skipped'] += 1
            if target_dir2 and not any(target_dir2 in t for t in file_info['targets']):
                self.processing_stats['target2_skipped'] += 1
            
            processed += 1
        
        self._report_progress("同步完成！", 100)
        return self.processing_stats
    
    def generate_sync_report(self, output_path: str) -> bool:
        """
        生成同步报告
        
        Args:
            output_path: 输出Excel路径
            
        Returns:
            bool: 成功返回True
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "同步记录"
            
            # 设置表头
            headers = ['源文件', '目标文件', '工作表', '同步单元格数', '状态']
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入同步记录
            for log in self.sync_logs:
                ws.append([
                    os.path.basename(log.get('source', '')),
                    os.path.basename(log.get('target', '')),
                    log.get('sheet', ''),
                    log.get('cells_synced', 0),
                    log.get('status', '')
                ])
            
            # 调整列宽
            for col_idx, _ in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            
            # 添加统计页
            ws_stats = wb.create_sheet(title="统计信息")
            stats_data = [
                ['统计项', '数值'],
                ['源目录文件数', self.processing_stats['source_files']],
                ['目标目录1同步文件数', self.processing_stats['target1_synced']],
                ['目标目录2同步文件数', self.processing_stats['target2_synced']],
                ['目标目录1跳过文件数', self.processing_stats['target1_skipped']],
                ['目标目录2跳过文件数', self.processing_stats['target2_skipped']],
                ['同步的单元格总数', self.processing_stats['total_cells_synced']],
                ['跳过的单元格总数', self.processing_stats['total_cells_skipped']],
                ['错误数', self.processing_stats['errors']]
            ]
            
            for row_data in stats_data:
                ws_stats.append(row_data)
            
            # 设置统计页样式
            for col_idx in range(1, 3):
                cell = ws_stats.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
            
            # 如果有错误，添加错误日志页
            if self.error_logs:
                ws_errors = wb.create_sheet(title="错误日志")
                ws_errors.append(['错误信息'])
                ws_errors.cell(row=1, column=1).fill = PatternFill(
                    start_color='FF0000', end_color='FF0000', fill_type='solid')
                ws_errors.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF')
                
                for error in self.error_logs:
                    ws_errors.append([error])
            
            wb.save(output_path)
            logger.info(f"同步报告已保存: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成报告失败: {e}")
            return False
    
    def get_stats_summary(self) -> str:
        """
        获取统计摘要文本
        
        Returns:
            str: 统计摘要
        """
        return f"""配置同步完成！

统计信息:
- 源目录文件数: {self.processing_stats['source_files']}
- 目标目录1同步文件数: {self.processing_stats['target1_synced']}
- 目标目录2同步文件数: {self.processing_stats['target2_synced']}
- 目标目录1跳过文件数: {self.processing_stats['target1_skipped']}
- 目标目录2跳过文件数: {self.processing_stats['target2_skipped']}
- 同步的单元格总数: {self.processing_stats['total_cells_synced']}
- 跳过的单元格总数: {self.processing_stats['total_cells_skipped']}
- 错误数: {self.processing_stats['errors']}

同步记录数: {len(self.sync_logs)} 条"""


# 命令行接口
if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel配置同步工具')
    parser.add_argument('--source', '-s', required=True, help='源目录路径')
    parser.add_argument('--target1', '-t1', required=True, help='目标目录1路径')
    parser.add_argument('--target2', '-t2', help='目标目录2路径（可选）')
    parser.add_argument('--json', '-j', help='JSON配置文件路径（可选，仅用于参考）')
    parser.add_argument('--filter', '-f', help='字段过滤配置文件路径（可选）')
    parser.add_argument('--report', '-r', help='报告输出路径')
    parser.add_argument('--no-backup', action='store_true', help='不创建备份')
    
    args = parser.parse_args()
    
    syncer = ExcelConfigSync()
    
    if args.no_backup:
        syncer.sync_options['backup_before_sync'] = False
    
    if args.json:
        syncer.load_json_config(args.json)
    
    if args.filter:
        syncer.load_filter_config(args.filter)
    
    stats = syncer.sync_directories(
        source_dir=args.source,
        target_dir1=args.target1,
        target_dir2=args.target2
    )
    
    print(syncer.get_stats_summary())
    
    if args.report:
        syncer.generate_sync_report(args.report)
