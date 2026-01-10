#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多语言翻译提取器
根据字段导出的JSON配置文件，智能提取多语言翻译内容
只提取前端、后端、前后端的字段，忽略策划字段
生成多语言翻译总表，支持中文、越南文、泰文
"""

import os
import json
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

# 导入统一的常量和模式
from core.constants import (
    SUPPORTED_LANGUAGES, SUPPORTED_EXCEL_EXTENSIONS, EXPORTABLE_FIELD_TYPES
)
from core.text_patterns import (
    CHINESE_PATTERN, VIETNAMESE_PATTERN, THAI_PATTERN,
    is_filterable_content, contains_chinese, contains_vietnamese, contains_thai
)

logger = logging.getLogger(__name__)


class TableRangeTranslator:
    """多语言翻译提取器"""
    
    # 使用统一的语言配置
    SUPPORTED_LANGUAGES = SUPPORTED_LANGUAGES
    
    def __init__(self):
        """初始化翻译提取器"""
        self.supported_extensions = SUPPORTED_EXCEL_EXTENSIONS
        
        # 需要导出的字段类型（忽略"策划"）
        self.exportable_field_types = EXPORTABLE_FIELD_TYPES
        
        # 结果存储
        self.translation_results = []
        self.processing_stats = {
            'total_tables': 0,
            'processed_tables': 0,
            'skipped_tables': 0,
            'total_fields': 0,
            'exported_fields': 0,
            'skipped_fields': 0,
            'total_rows': 0
        }
        
        self.error_logs = []
    
    def load_json_config(self, json_path: str, target_language: str = None) -> Dict:
        """
        加载JSON配置文件
        
        Args:
            json_path: JSON配置文件路径
            target_language: 目标语言代码 ('zh', 'vn', 'th')，用于从合并的JSON中提取对应语言配置
            
        Returns:
            Dict: JSON配置内容（包含language字段信息）
        """
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            logger.info(f"成功加载JSON配置: {json_path}")
            
            # 检查是否为合并的JSON结构（包含大写语言代码键）
            if target_language:
                lang_key = target_language.upper()
                if lang_key in config:
                    logger.info(f"检测到合并JSON结构，提取 {lang_key} 语言配置")
                    config = config[lang_key]
            
            # 读取语言标记
            if 'language' in config:
                lang_info = config['language']
                logger.info(f"  - 语言标记: {lang_info.get('name', '')} ({lang_info.get('code', '')})")
            
            logger.info(f"  - no_text_tables: {len(config.get('no_text_tables', []))} 个")
            logger.info(f"  - text_tables: {len(config.get('text_tables', []))} 个")
            
            return config
        
        except Exception as e:
            logger.error(f"加载JSON配置失败: {e}")
            return {}
    
    def get_json_language(self, config: Dict) -> Optional[str]:
        """
        从JSON配置中获取语言代码
        
        Args:
            config: JSON配置字典
            
        Returns:
            str or None: 语言代码 ('zh', 'vn', 'th') 或 None
        """
        if 'language' in config and isinstance(config['language'], dict):
            return config['language'].get('code')
        return None
    
    def match_directory_by_language(self, config: Dict, lang_dirs: Dict[str, str]) -> Optional[str]:
        """
        根据JSON中的语言标记匹配对应的目录
        
        Args:
            config: JSON配置字典
            lang_dirs: 语言目录映射 {'zh': '/path/to/zh', 'vn': '/path/to/vn', 'th': '/path/to/th'}
            
        Returns:
            str or None: 匹配的目录路径
        """
        lang_code = self.get_json_language(config)
        if lang_code and lang_code in lang_dirs:
            return lang_dirs[lang_code]
        return None
    
    def parse_field_with_type(self, field_str: str) -> Tuple[str, str, Optional[str]]:
        """
        解析字段字符串，提取字段名、字段类型和列号
        格式: "field_name,field_type" 或 "field_name,field_type,col_letter"
        
        Args:
            field_str: 字段字符串
            
        Returns:
            Tuple[str, str, Optional[str]]: (字段名, 字段类型, 列号字母)
        """
        parts = [p.strip() for p in field_str.split(',')]
        field_name = parts[0]
        field_type = parts[1] if len(parts) > 1 else ''
        col_letter = parts[2] if len(parts) > 2 else None
        
        return field_name, field_type, col_letter
    
    def load_merged_json_config(self, json_path: str) -> Dict:
        """
        加载合并的JSON配置文件（包含多语言配置）
        
        Args:
            json_path: 合并的JSON配置文件路径
            
        Returns:
            Dict: 完整的JSON配置（包含ZH/VN/TH等语言键）
        """
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            logger.info(f"成功加载合并JSON配置: {json_path}")
            
            # 检测包含的语言
            detected_langs = []
            for lang_key in ['ZH', 'VN', 'TH']:
                if lang_key in config:
                    lang_data = config[lang_key]
                    text_count = len(lang_data.get('text_tables', []))
                    no_text_count = len(lang_data.get('no_text_tables', []))
                    detected_langs.append(f"{lang_key}({text_count}表)")
                    logger.info(f"  - {lang_key}: {text_count}个文本表, {no_text_count}个无文本表")
            
            logger.info(f"检测到语言: {', '.join(detected_langs)}")
            
            return config
        
        except Exception as e:
            logger.error(f"加载合并JSON配置失败: {e}")
            return {}
    
    def process_with_merged_json(self, json_path: str, lang_dirs: Dict[str, str],
                                 progress_callback=None) -> List[Dict]:
        """
        根据合并的JSON配置文件处理多语言Excel文件
        JSON结构: {"ZH": {...}, "VN": {...}, "TH": {...}}
        每个语言配置包含 text_tables 列表，其中有 field_column_letters 用于精确定位
        
        Args:
            json_path: 合并的JSON配置文件路径
            lang_dirs: 语言目录字典 {'zh': 'path', 'vn': 'path', 'th': 'path'}
            progress_callback: 进度回调函数
            
        Returns:
            List[Dict]: 所有提取的数据
        """
        def log_progress(msg):
            logger.info(msg)
            if progress_callback:
                progress_callback(msg)
        
        try:
            lang_names = {'zh': '中文', 'vn': '越南语', 'th': '泰语'}
            
            # 加载合并的JSON配置
            merged_config = self.load_merged_json_config(json_path)
            if not merged_config:
                log_progress("✗ 加载JSON配置失败")
                return []
            
            # 识别JSON中的语言配置
            lang_configs = {}
            for lang_key in ['ZH', 'VN', 'TH']:
                if lang_key in merged_config:
                    lang_code = lang_key.lower()
                    lang_configs[lang_code] = merged_config[lang_key]
                    log_progress(f"✓ 检测到{lang_names.get(lang_code, lang_code)}配置")
            
            if not lang_configs:
                log_progress("✗ JSON中未找到有效的语言配置（ZH/VN/TH）")
                return []
            
            # 合并所有语言的表格配置
            # 以表名+工作表名为键，合并字段信息（包含列字母）
            table_key_to_info: Dict[Tuple[str, str], Dict] = {}
            
            for lang_code, config in lang_configs.items():
                text_tables = config.get('text_tables', [])
                for table_info in text_tables:
                    table_name = table_info.get('table_name', '')
                    sheet_name = table_info.get('sheet_name', '')
                    if not table_name or not sheet_name:
                        continue
                    
                    key = (table_name, sheet_name)
                    
                    if key not in table_key_to_info:
                        table_key_to_info[key] = {
                            'table_name': table_name,
                            'sheet_name': sheet_name,
                            'fields_by_lang': {},
                            'columns_by_lang': {}
                        }
                    
                    # 存储该语言的字段信息和列字母
                    fields = table_info.get('fields_with_examples', [])
                    columns = table_info.get('field_column_letters', [])
                    table_key_to_info[key]['fields_by_lang'][lang_code] = fields
                    table_key_to_info[key]['columns_by_lang'][lang_code] = columns
            
            # 转为列表并排序
            all_tables = sorted(table_key_to_info.values(), 
                               key=lambda x: (x['table_name'], x['sheet_name']))
            self.processing_stats['total_tables'] = len(all_tables)
            
            log_progress(f"\n📊 开始处理 {len(all_tables)} 个表格")
            active_langs = [k for k in lang_dirs.keys() if k in lang_configs]
            log_progress(f"🌐 活跃语言版本: {', '.join([lang_names.get(k, k) for k in active_langs])}")
            log_progress("")
            
            all_data = []
            
            for idx, table_info in enumerate(all_tables, 1):
                table_name = table_info['table_name']
                sheet_name = table_info['sheet_name']
                fields_by_lang = table_info['fields_by_lang']
                columns_by_lang = table_info['columns_by_lang']
                
                log_progress(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
                log_progress(f"📋 [{idx}/{len(all_tables)}] 处理表格: {table_name}")
                log_progress(f"   工作表: {sheet_name}")
                
                # 合并所有语言的字段列表（使用第一个有效语言的字段作为基准）
                base_lang = None
                base_fields = []
                base_columns = []
                for lang in ['zh', 'vn', 'th']:
                    if lang in fields_by_lang and fields_by_lang[lang]:
                        base_lang = lang
                        base_fields = fields_by_lang[lang]
                        base_columns = columns_by_lang.get(lang, [])
                        break
                
                if not base_fields:
                    log_progress(f"   ⚠️ 没有可用的字段配置，跳过")
                    continue
                
                # 解析字段信息，过滤出需要导出的字段
                exportable_fields = []
                skipped_count = 0
                for i, field_str in enumerate(base_fields):
                    field_name, field_type, col_letter = self.parse_field_with_type(field_str)
                    # 如果JSON没有列字母在field_str中，尝试从columns列表获取
                    if not col_letter and i < len(base_columns):
                        col_letter = base_columns[i]
                    
                    if self.is_exportable_field(field_type):
                        exportable_fields.append((field_name, field_type, col_letter))
                    else:
                        self.processing_stats['skipped_fields'] += 1
                        skipped_count += 1
                
                if not exportable_fields:
                    log_progress(f"   ⚠️ 没有需要导出的字段，跳过")
                    continue
                
                log_progress(f"   ✓ 可导出字段: {len(exportable_fields)} 个")
                if skipped_count > 0:
                    log_progress(f"   • 跳过策划字段: {skipped_count} 个")
                
                self.processing_stats['exported_fields'] += len(exportable_fields)
                
                # 从各语言目录读取Excel数据
                log_progress("   📁 读取语言版本文件:")
                table_data_by_lang = {}
                for lang, lang_dir in lang_dirs.items():
                    if lang not in lang_configs:
                        continue
                    
                    excel_path = os.path.join(lang_dir, table_name)
                    lang_name = lang_names.get(lang, lang)
                    
                    if not os.path.exists(excel_path):
                        log_progress(f"      ✗ {lang_name}: 文件不存在")
                        continue
                    
                    try:
                        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                        if len(df) < 7:
                            log_progress(f"      ✗ {lang_name}: 数据行不足")
                            continue
                        table_data_by_lang[lang] = df
                        log_progress(f"      ✓ {lang_name}: {len(df)-6} 行数据")
                    except Exception as e:
                        log_progress(f"      ✗ {lang_name}: 读取失败 - {e}")
                        continue
                
                if not table_data_by_lang:
                    log_progress(f"   ⚠️ 所有语言版本都不可用，跳过")
                    self.processing_stats['skipped_tables'] += 1
                    continue
                
                # 提取并合并数据
                log_progress(f"   🔍 提取字段数据:")
                table_extracted_count = 0
                
                # 构建ID映射
                id_to_row_by_lang = {lang: self._build_id_to_row_index(df) 
                                    for lang, df in table_data_by_lang.items()}
                
                for field_name, field_type, col_letter in exportable_fields:
                    # 使用列字母精确定位（交叉验证）
                    col_idx_by_lang: Dict[str, Optional[int]] = {}
                    
                    for lang, df in table_data_by_lang.items():
                        col_idx = None
                        
                        # 优先使用列字母定位
                        if col_letter:
                            try:
                                col_idx = self.column_letter_to_index(col_letter)
                                # 验证列索引有效性
                                if col_idx >= len(df.columns):
                                    col_idx = None
                                else:
                                    # 交叉验证：检查第5行的字段名是否匹配
                                    if len(df) >= 5:
                                        actual_field = df.iloc[4, col_idx]
                                        if pd.notna(actual_field) and str(actual_field).strip() != field_name:
                                            log_progress(f"      ⚠️ {lang_names.get(lang)}: 列{col_letter}字段名不匹配 (期望:{field_name}, 实际:{actual_field})")
                                            # 回退到按名称查找
                                            col_idx = self.find_column_index_by_name(df, field_name)
                            except Exception:
                                col_idx = None
                        
                        # 如果列字母定位失败，回退到按名称查找
                        if col_idx is None:
                            col_idx = self.find_column_index_by_name(df, field_name)
                        
                        col_idx_by_lang[lang] = col_idx
                    
                    if all(v is None for v in col_idx_by_lang.values()):
                        log_progress(f"      ⚠️ {field_name} ({field_type}): 所有语言版本均未找到列")
                        continue
                    
                    # 选择锚点语言
                    anchor_lang = 'zh' if 'zh' in table_data_by_lang else list(table_data_by_lang.keys())[0]
                    anchor_df = table_data_by_lang[anchor_lang]
                    anchor_col_idx = col_idx_by_lang.get(anchor_lang)
                    
                    if anchor_col_idx is None:
                        for lang, idx_val in col_idx_by_lang.items():
                            if idx_val is not None:
                                anchor_lang = lang
                                anchor_df = table_data_by_lang[lang]
                                anchor_col_idx = idx_val
                                break
                    
                    excel_col = self.column_index_to_letter(anchor_col_idx)
                    
                    # 遍历数据行
                    for anchor_row_idx in range(6, len(anchor_df)):
                        anchor_id = self._normalize_id_value(anchor_df.iloc[anchor_row_idx, 0])
                        
                        lang_contents: Dict[str, str] = {}
                        has_content = False
                        
                        for lang, df in table_data_by_lang.items():
                            col_idx = col_idx_by_lang.get(lang)
                            if col_idx is None:
                                continue
                            
                            # 按ID对齐行
                            target_row_idx = None
                            if anchor_id:
                                target_row_idx = id_to_row_by_lang.get(lang, {}).get(anchor_id)
                            if target_row_idx is None:
                                target_row_idx = anchor_row_idx if anchor_row_idx < len(df) else None
                            
                            if target_row_idx is None:
                                continue
                            
                            cell_value = df.iloc[target_row_idx, col_idx]
                            if pd.notna(cell_value) and str(cell_value).strip():
                                lang_contents[lang] = str(cell_value).strip()
                                has_content = True
                        
                        if not has_content:
                            continue
                        
                        # 过滤掉非文本内容（{}、数组、纯数字）
                        lang_contents = self.filter_lang_contents(lang_contents)
                        if not lang_contents:
                            continue
                        
                        excel_row = anchor_row_idx + 1
                        excel_position = f"{excel_col}{excel_row}"
                        
                        row_data = {
                            'table_name': table_name,
                            'sheet_name': sheet_name,
                            'field_name': field_name,
                            'field_type': field_type,
                            'excel_position': excel_position,
                            'chinese': lang_contents.get('zh', ''),
                            'vietnamese': lang_contents.get('vn', ''),
                            'thai': lang_contents.get('th', '')
                        }
                        
                        all_data.append(row_data)
                        self.processing_stats['total_rows'] += 1
                        table_extracted_count += 1
                
                log_progress(f"      ✓ 共提取 {table_extracted_count} 条数据")
                log_progress("")
                
                self.processing_stats['processed_tables'] += 1
            
            self.translation_results = all_data
            
            log_progress("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            log_progress(f"✅ 处理完成！")
            log_progress(f"   • 处理表格: {self.processing_stats['processed_tables']}/{self.processing_stats['total_tables']}")
            log_progress(f"   • 导出字段: {self.processing_stats['exported_fields']} 个")
            log_progress(f"   • 提取数据: {len(all_data)} 条")
            log_progress("")
            
            return all_data
        
        except Exception as e:
            logger.error(f"处理失败: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def is_exportable_field(self, field_type: str) -> bool:
        """
        判断字段类型是否需要导出
        
        Args:
            field_type: 字段类型（前端、后端、前后端、策划等）
            
        Returns:
            bool: True表示需要导出
        """
        return field_type in self.exportable_field_types
    
    def is_valid_translation_text(self, text: str) -> bool:
        """
        判断文本是否为有效的翻译文本（非结构化数据）
        过滤掉：空花括号{}、数组格式[2,99]/{2,99}、纯数字
        
        Args:
            text: 待检测文本
            
        Returns:
            bool: True表示是有效的翻译文本，False表示应该被过滤
        """
        if pd.isna(text):
            return False
        
        text_str = str(text).strip()
        if not text_str:
            return False
        
        # 使用统一的过滤函数
        return not is_filterable_content(text_str)
    
    def filter_lang_contents(self, lang_contents: Dict[str, str]) -> Dict[str, str]:
        """
        过滤语言内容字典，移除非文本值
        
        Args:
            lang_contents: 语言内容字典 {'zh': 'xxx', 'vn': 'xxx', 'th': 'xxx'}
            
        Returns:
            Dict[str, str]: 过滤后的语言内容字典
        """
        return {lang: text for lang, text in lang_contents.items() 
                if self.is_valid_translation_text(text)}
    
    def detect_language_type(self, text: str) -> str:
        """
        检测文本语言类型
        
        Args:
            text: 待检测文本
            
        Returns:
            str: 语言类型（中文、越南文、泰文、中越混合等）
        """
        if pd.isna(text) or not str(text).strip():
            return "空"
        
        text_str = str(text).strip()
        
        # 使用统一的语言检测函数
        has_chinese = contains_chinese(text_str)
        has_vietnamese = contains_vietnamese(text_str)
        has_thai = contains_thai(text_str)
        
        # 判断语言类型
        if has_chinese and has_vietnamese:
            return "中越混合"
        elif has_chinese and has_thai:
            return "中泰混合"
        elif has_vietnamese and has_thai:
            return "越泰混合"
        elif has_chinese:
            return "中文"
        elif has_vietnamese:
            return "越南文"
        elif has_thai:
            return "泰文"
        else:
            return "其他"
    
    def column_index_to_letter(self, col_idx: int) -> str:
        """
        将列索引转换为Excel列字母（A, B, C, ..., Z, AA, AB, ...）
        
        Args:
            col_idx: 列索引（从0开始）
            
        Returns:
            str: Excel列字母
        """
        result = ""
        col_num = col_idx + 1  # Excel列从1开始
        
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        
        return result

    def column_letter_to_index(self, col_letter: str) -> int:
        """
        将Excel列字母转换为列索引（A->0, B->1, ...）
        
        Args:
            col_letter: Excel列字母
            
        Returns:
            int: 列索引（从0开始）
        """
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def _normalize_id_value(self, value) -> str:
        """规范化ID值，用于跨语言版本按ID对齐行。"""
        if pd.isna(value):
            return ""
        value_str = str(value).strip()
        if not value_str:
            return ""
        # pandas 读取数字时可能变成 123.0，这里尽量还原
        try:
            if isinstance(value, (int, float)) and float(value).is_integer():
                return str(int(value))
        except Exception:
            pass
        if value_str.endswith('.0'):
            prefix = value_str[:-2]
            if prefix.isdigit():
                return prefix
        return value_str

    def _build_id_to_row_index(self, df: pd.DataFrame) -> Dict[str, int]:
        """为一个语言版本的表构建 ID -> 行索引 映射（数据区从第7行开始，索引6）。"""
        id_to_row: Dict[str, int] = {}
        if df is None or len(df) < 7 or len(df.columns) == 0:
            return id_to_row

        for row_idx in range(6, len(df)):
            id_key = self._normalize_id_value(df.iloc[row_idx, 0])
            if not id_key:
                continue
            # 如果有重复ID，保留第一次出现的位置
            if id_key not in id_to_row:
                id_to_row[id_key] = row_idx

        return id_to_row
    
    def find_column_index_by_name(self, df: pd.DataFrame, field_name: str) -> Optional[int]:
        """
        在DataFrame的第5行（索引4）中查找字段名对应的列索引
        
        Args:
            df: pandas DataFrame
            field_name: 要查找的字段名
            
        Returns:
            Optional[int]: 列索引，未找到返回None
        """
        if len(df) < 5:
            return None
        
        # 第5行是字段名行（索引为4）
        field_row = df.iloc[4]
        
        for col_idx, cell_value in enumerate(field_row):
            if pd.notna(cell_value) and str(cell_value).strip() == field_name:
                return col_idx
        
        return None
    
    def extract_table_data(self, excel_path: str, table_info: Dict) -> List[Dict]:
        """
        从Excel表格中提取指定字段的数据
        
        Args:
            excel_path: Excel文件路径
            table_info: 表格信息（包含 table_name, sheet_name, fields_with_examples）
            
        Returns:
            List[Dict]: 提取的数据行列表
        """
        try:
            table_name = table_info.get('table_name', '')
            sheet_name = table_info.get('sheet_name', '')
            fields_with_examples = table_info.get('fields_with_examples', [])
            
            logger.info(f"开始提取表格: {table_name} - {sheet_name}")
            
            # 预先计算可导出字段数量用于统计
            exportable_count = 0
            for field_str in fields_with_examples:
                _, field_type, _ = self.parse_field_with_type(field_str)
                if self.is_exportable_field(field_type):
                    exportable_count += 1
            
            if exportable_count == 0:
                logger.info(f"表格 {table_name} 没有需要导出的字段，跳过")
                return []
            
            logger.info(f"可导出字段: {exportable_count} 个")
            self.processing_stats['exported_fields'] += exportable_count
            
            # 读取Excel文件
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            
            # 从第7行开始提取数据（索引6）
            if len(df) < 7:
                logger.warning(f"表格 {table_name} 数据行不足，跳过")
                return []
            
            extracted_rows = []
            
            # 遍历每个需要导出的字段
            for field_str in fields_with_examples:
                field_name, field_type, col_letter_hint = self.parse_field_with_type(field_str)
                
                if not self.is_exportable_field(field_type):
                    self.processing_stats['skipped_fields'] += 1
                    continue

                # 优先使用列号定位
                col_idx = None
                if col_letter_hint:
                    try:
                        col_idx = self.column_letter_to_index(col_letter_hint)
                        # 验证列索引是否有效
                        if col_idx >= len(df.columns):
                            logger.warning(f"字段 {field_name} 的列号 {col_letter_hint} 超出范围，尝试按名称查找")
                            col_idx = None
                    except Exception:
                        logger.warning(f"字段 {field_name} 的列号 {col_letter_hint} 解析失败，尝试按名称查找")
                        col_idx = None
                
                # 如果没有列号或列号无效，回退到按名称查找
                if col_idx is None:
                    col_idx = self.find_column_index_by_name(df, field_name)
                
                if col_idx is None:
                    logger.warning(f"未找到字段: {field_name}")
                    continue
                
                # 从第7行开始提取数据
                for row_idx in range(6, len(df)):  # 从索引6开始（第7行）
                    cell_value = df.iloc[row_idx, col_idx]
                    
                    # 跳过空值
                    if pd.isna(cell_value) or not str(cell_value).strip():
                        continue
                    
                    # 检测语言类型
                    lang_type = self.detect_language_type(cell_value)
                    
                    # 只提取包含中文、越南文、泰文的内容
                    if lang_type not in ["中文", "越南文", "泰文", "中越混合", "中泰混合", "越泰混合"]:
                        continue
                    
                    # 提取A列的ID值
                    id_value = df.iloc[row_idx, 0] if len(df.columns) > 0 else ''
                    
                    # 生成Excel物理位置（如F8）
                    current_col_letter = self.column_index_to_letter(col_idx)
                    excel_position = f"{current_col_letter}{row_idx + 1}"
                    
                    extracted_rows.append({
                        'table_name': table_name,
                        'sheet_name': sheet_name,
                        'field_name': field_name,
                        'field_type': field_type,
                        'excel_position': excel_position,  # Excel物理位置（如F8）
                        'id': id_value,
                        'content': str(cell_value).strip(),
                        'language_type': lang_type
                    })
                    
                    self.processing_stats['total_rows'] += 1
            
            logger.info(f"提取完成: {len(extracted_rows)} 行数据")
            return extracted_rows
        
        except Exception as e:
            error_msg = f"提取表格数据失败 {table_name}: {e}"
            logger.error(error_msg)
            self.error_logs.append(error_msg)
            return []
    
    def process_with_json_config(self, json_path: str, excel_directory: str) -> List[Dict]:
        """
        根据JSON配置处理Excel文件
        
        Args:
            json_path: JSON配置文件路径
            excel_directory: Excel文件所在目录
            
        Returns:
            List[Dict]: 所有提取的数据
        """
        try:
            # 加载JSON配置
            config = self.load_json_config(json_path)
            if not config:
                return []
            
            # 获取text_tables列表（跳过no_text_tables）
            text_tables = config.get('text_tables', [])
            self.processing_stats['total_tables'] = len(text_tables)
            
            logger.info(f"开始处理 {len(text_tables)} 个包含文本的表格")
            
            all_data = []
            
            # 处理每个表格
            for table_info in text_tables:
                table_name = table_info.get('table_name', '')
                
                # 构建Excel文件路径
                excel_path = os.path.join(excel_directory, table_name)
                
                if not os.path.exists(excel_path):
                    logger.warning(f"文件不存在: {excel_path}")
                    self.processing_stats['skipped_tables'] += 1
                    continue
                
                # 提取表格数据
                table_data = self.extract_table_data(excel_path, table_info)
                all_data.extend(table_data)
                
                self.processing_stats['processed_tables'] += 1
            
            self.translation_results = all_data
            logger.info(f"处理完成，共提取 {len(all_data)} 条数据")
            
            return all_data
        
        except Exception as e:
            logger.error(f"处理失败: {e}")
            return []
    
    def generate_translation_master_table(self, output_path: str, 
                                          chinese_dir: str = None,
                                          vietnamese_dir: str = None,
                                          thai_dir: str = None):
        """
        生成多语言翻译总表
        每个表格文件对应一个工作表标签
        列格式: 字段名 | 字段类型 | Excel位置 | ZH | VN | TH
        
        Args:
            output_path: 输出Excel文件路径
            chinese_dir: 中文版Excel文件目录
            vietnamese_dir: 越南文版Excel文件目录
            thai_dir: 泰文版Excel文件目录
        """
        try:
            if not self.translation_results:
                logger.warning("没有数据可导出")
                return False
            
            logger.info(f"开始生成翻译总表: {output_path}")
            
            # 按表格名称分组
            tables_data = {}
            for row in self.translation_results:
                table_name = row['table_name']
                if table_name not in tables_data:
                    tables_data[table_name] = []
                tables_data[table_name].append(row)
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 为每个表格创建一个工作表
            for table_name, rows in tables_data.items():
                # 工作表名称（去除扩展名）
                sheet_name = Path(table_name).stem
                # Excel工作表名称限制31字符
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + '...'
                
                ws = wb.create_sheet(title=sheet_name)
                
                # 设置表头
                headers = ['Field', 'Type', 'Position', 'ZH', 'VN', 'TH']
                ws.append(headers)
                
                # 设置表头样式
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 写入数据行
                for row_data in rows:
                    ws.append([
                        row_data.get('field_name', ''),
                        row_data.get('field_type', ''),
                        row_data.get('excel_position', ''),  # Excel物理位置
                        row_data.get('content', '') if row_data.get('language_type', '') in ['中文', '中越混合', '中泰混合'] else '',
                        '',  # 越南文列（待填充）
                        ''   # 泰文列（待填充）
                    ])
                
                # 设置列宽
                column_widths = [20, 12, 12, 40, 40, 40]
                for col_idx, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 添加边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1:  # 数据行
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # 保存工作簿
            wb.save(output_path)
            logger.info(f"翻译总表已生成: {output_path}")
            logger.info(f"  - 工作表数量: {len(tables_data)}")
            logger.info(f"  - 总数据行数: {len(self.translation_results)}")
            
            return True
        
        except Exception as e:
            logger.error(f"生成翻译总表失败: {e}")
            return False
    
    def process_with_multi_json_configs(self, json_configs: Dict[str, str], lang_dirs: Dict[str, str],
                                        progress_callback=None) -> List[Dict]:
        """
        根据多个JSON配置（带语言标记）和对应的语言目录处理Excel文件
        每个JSON配置中包含language字段，自动匹配对应的目录
        
        Args:
            json_configs: 语言JSON配置字典 {'zh': 'path/to/zh.json', 'vn': 'path/to/vn.json', 'th': 'path/to/th.json'}
            lang_dirs: 语言目录字典 {'zh': 'path', 'vn': 'path', 'th': 'path'}
            progress_callback: 进度回调函数，接收进度消息字符串
            
        Returns:
            List[Dict]: 所有提取的数据（包含多语言内容）
        """
        def log_progress(msg):
            """输出进度信息"""
            logger.info(msg)
            if progress_callback:
                progress_callback(msg)
        
        try:
            lang_names = {'zh': '中文', 'vn': '越南语', 'th': '泰语'}
            
            # 加载所有JSON配置
            all_configs = {}
            for lang_code, json_path in json_configs.items():
                config = self.load_json_config(json_path)
                if config:
                    # 验证JSON中的语言标记
                    json_lang = self.get_json_language(config)
                    if json_lang and json_lang != lang_code:
                        log_progress(f"⚠️ 警告: {lang_names.get(lang_code, lang_code)}JSON的语言标记为{json_lang}，与预期不符")
                    all_configs[lang_code] = config
                    log_progress(f"✓ 加载{lang_names.get(lang_code, lang_code)}JSON配置成功")
            
            if not all_configs:
                log_progress("✗ 没有成功加载任何JSON配置")
                return []
            
            # 合并所有JSON配置的表格列表（不同语言版本可能存在表/字段差异）
            table_key_to_infos: Dict[Tuple[str, str], List[Dict]] = {}
            for cfg in all_configs.values():
                for table_info in cfg.get('text_tables', []):
                    table_name = table_info.get('table_name', '')
                    sheet_name = table_info.get('sheet_name', '')
                    if not table_name or not sheet_name:
                        continue
                    key = (table_name, sheet_name)
                    table_key_to_infos.setdefault(key, []).append(table_info)

            text_tables = []
            for (table_name, sheet_name), infos in table_key_to_infos.items():
                # 合并字段列表（去重）
                merged_fields: List[str] = []
                seen: Set[str] = set()
                for info in infos:
                    for f in info.get('fields_with_examples', []) or []:
                        f_str = str(f).strip()
                        if not f_str or f_str in seen:
                            continue
                        seen.add(f_str)
                        merged_fields.append(f_str)

                text_tables.append({
                    'table_name': table_name,
                    'sheet_name': sheet_name,
                    'fields_with_examples': merged_fields
                })

            # 保持处理顺序稳定
            text_tables.sort(key=lambda x: (x.get('table_name', ''), x.get('sheet_name', '')))
            self.processing_stats['total_tables'] = len(text_tables)
            
            log_progress(f"\n📊 开始处理 {len(text_tables)} 个表格")
            log_progress(f"🌐 语言版本: {', '.join([lang_names.get(k, k) for k in lang_dirs.keys()])}")
            log_progress("")
            
            all_data = []
            
            # 处理每个表格
            for idx, table_info in enumerate(text_tables, 1):
                table_name = table_info.get('table_name', '')
                sheet_name = table_info.get('sheet_name', '')
                fields_with_examples = table_info.get('fields_with_examples', [])
                
                log_progress(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
                log_progress(f"📋 [{idx}/{len(text_tables)}] 处理表格: {table_name}")
                log_progress(f"   工作表: {sheet_name}")
                
                # 解析字段信息，过滤出需要导出的字段
                exportable_fields = []
                skipped_count = 0
                for field_str in fields_with_examples:
                    field_name, field_type = self.parse_field_with_type(field_str)
                    if self.is_exportable_field(field_type):
                        exportable_fields.append((field_name, field_type))
                    else:
                        self.processing_stats['skipped_fields'] += 1
                        skipped_count += 1
                
                if not exportable_fields:
                    log_progress(f"   ⚠️  没有需要导出的字段，跳过")
                    log_progress("")
                    continue
                
                log_progress(f"   ✓ 可导出字段: {len(exportable_fields)} 个")
                if skipped_count > 0:
                    log_progress(f"   • 跳过策划字段: {skipped_count} 个")
                
                self.processing_stats['exported_fields'] += len(exportable_fields)
                
                # 从各语言目录读取数据
                log_progress("   ")
                log_progress("   📁 读取语言版本文件:")
                table_data_by_lang = {}
                for lang, lang_dir in lang_dirs.items():
                    excel_path = os.path.join(lang_dir, table_name)
                    lang_name = lang_names.get(lang, lang)
                    
                    if not os.path.exists(excel_path):
                        log_progress(f"      ✗ {lang_name}: 文件不存在")
                        continue
                    
                    # 读取Excel
                    try:
                        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                        if len(df) < 7:
                            log_progress(f"      ✗ {lang_name}: 数据行不足")
                            continue
                        table_data_by_lang[lang] = df
                        log_progress(f"      ✓ {lang_name}: {len(df)-6} 行数据")
                    except Exception as e:
                        log_progress(f"      ✗ {lang_name}: 读取失败 - {e}")
                        continue
                
                if not table_data_by_lang:
                    log_progress(f"   ⚠️  所有语言版本都不可用，跳过")
                    log_progress("")
                    self.processing_stats['skipped_tables'] += 1
                    continue
                
                # 提取并合并数据
                log_progress("   ")
                log_progress(f"   🔍 提取字段数据:")
                table_extracted_count = 0
                
                for field_name, field_type in exportable_fields:
                    # 每个语言版本分别按字段名定位列（不同语言表结构可能不同）
                    col_idx_by_lang: Dict[str, Optional[int]] = {}
                    for lang, df in table_data_by_lang.items():
                        col_idx_by_lang[lang] = self.find_column_index_by_name(df, field_name)

                    if all(v is None for v in col_idx_by_lang.values()):
                        log_progress(f"      ⚠️  {field_name} ({field_type}): 所有语言版本均未找到列")
                        continue

                    # 选择一个锚点语言用于生成Position与遍历行（优先中文）
                    anchor_lang = 'zh' if 'zh' in table_data_by_lang else list(table_data_by_lang.keys())[0]
                    anchor_df = table_data_by_lang[anchor_lang]
                    anchor_col_idx = col_idx_by_lang.get(anchor_lang)
                    if anchor_col_idx is None:
                        # 如果锚点语言缺列，退化为任意一个有列的语言
                        for lang, idx_val in col_idx_by_lang.items():
                            if idx_val is not None:
                                anchor_lang = lang
                                anchor_df = table_data_by_lang[lang]
                                anchor_col_idx = idx_val
                                break

                    excel_col = self.column_index_to_letter(anchor_col_idx)

                    # 为每个语言版本构建ID -> 行索引映射，支持按ID对齐行
                    id_to_row_by_lang = {lang: self._build_id_to_row_index(df) for lang, df in table_data_by_lang.items()}

                    for anchor_row_idx in range(6, len(anchor_df)):
                        anchor_id = self._normalize_id_value(anchor_df.iloc[anchor_row_idx, 0])

                        lang_contents: Dict[str, str] = {}
                        has_content = False

                        for lang, df in table_data_by_lang.items():
                            col_idx = col_idx_by_lang.get(lang)
                            if col_idx is None:
                                continue

                            # 优先用ID在该语言版本中定位行；若ID为空则回退按行号
                            target_row_idx: Optional[int] = None
                            if anchor_id:
                                target_row_idx = id_to_row_by_lang.get(lang, {}).get(anchor_id)
                            if target_row_idx is None:
                                target_row_idx = anchor_row_idx if anchor_row_idx < len(df) else None

                            if target_row_idx is None:
                                continue

                            cell_value = df.iloc[target_row_idx, col_idx]
                            if pd.notna(cell_value) and str(cell_value).strip():
                                lang_contents[lang] = str(cell_value).strip()
                                has_content = True

                        if not has_content:
                            continue

                        # 过滤掉非文本内容（{}、数组、纯数字）
                        lang_contents = self.filter_lang_contents(lang_contents)
                        if not lang_contents:
                            continue

                        excel_row = anchor_row_idx + 1
                        excel_position = f"{excel_col}{excel_row}"

                        row_data = {
                            'table_name': table_name,
                            'sheet_name': sheet_name,
                            'field_name': field_name,
                            'field_type': field_type,
                            'excel_position': excel_position,
                            'chinese': lang_contents.get('zh', ''),
                            'vietnamese': lang_contents.get('vn', ''),
                            'thai': lang_contents.get('th', '')
                        }

                        all_data.append(row_data)
                        self.processing_stats['total_rows'] += 1
                        table_extracted_count += 1
                
                log_progress(f"      ✓ 共提取 {table_extracted_count} 条数据")
                log_progress("")
                
                self.processing_stats['processed_tables'] += 1
            
            self.translation_results = all_data
            
            log_progress("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            log_progress(f"✅ 处理完成！")
            log_progress(f"   • 处理表格: {self.processing_stats['processed_tables']}/{self.processing_stats['total_tables']}")
            log_progress(f"   • 导出字段: {self.processing_stats['exported_fields']} 个")
            log_progress(f"   • 提取数据: {len(all_data)} 条")
            log_progress("")
            
            return all_data
        
        except Exception as e:
            logger.error(f"处理失败: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def process_with_json_config_multi_lang(self, json_path: str, lang_dirs: Dict[str, str], 
                                            progress_callback=None) -> List[Dict]:
        """
        根据JSON配置和多语言目录处理Excel文件
        
        Args:
            json_path: JSON配置文件路径
            lang_dirs: 语言目录字典 {'vn': 'path', 'zh': 'path', 'th': 'path'}
            progress_callback: 进度回调函数，接收进度消息字符串
            
        Returns:
            List[Dict]: 所有提取的数据（包含多语言内容）
        """
        def log_progress(msg):
            """输出进度信息"""
            logger.info(msg)
            if progress_callback:
                progress_callback(msg)
        
        try:
            # 加载JSON配置
            config = self.load_json_config(json_path)
            if not config:
                return []
            
            # 获取text_tables列表
            text_tables = config.get('text_tables', [])
            self.processing_stats['total_tables'] = len(text_tables)
            
            lang_names = {'vn': '越南文', 'zh': '中文', 'th': '泰文'}
            lang_list = ', '.join([lang_names.get(k, k) for k in lang_dirs.keys()])
            
            log_progress(f"📊 开始处理 {len(text_tables)} 个表格")
            log_progress(f"🌐 语言版本: {lang_list}")
            log_progress("")
            
            all_data = []
            
            # 处理每个表格
            for idx, table_info in enumerate(text_tables, 1):
                table_name = table_info.get('table_name', '')
                sheet_name = table_info.get('sheet_name', '')
                fields_with_examples = table_info.get('fields_with_examples', [])
                
                log_progress(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
                log_progress(f"📋 [{idx}/{len(text_tables)}] 处理表格: {table_name}")
                log_progress(f"   工作表: {sheet_name}")
                
                # 解析字段信息，过滤出需要导出的字段
                exportable_fields = []
                skipped_count = 0
                for field_str in fields_with_examples:
                    field_name, field_type = self.parse_field_with_type(field_str)
                    if self.is_exportable_field(field_type):
                        exportable_fields.append((field_name, field_type))
                    else:
                        self.processing_stats['skipped_fields'] += 1
                        skipped_count += 1
                
                if not exportable_fields:
                    log_progress(f"   ⚠️  没有需要导出的字段，跳过")
                    log_progress("")
                    continue
                
                log_progress(f"   ✓ 可导出字段: {len(exportable_fields)} 个")
                if skipped_count > 0:
                    log_progress(f"   • 跳过策划字段: {skipped_count} 个")
                
                self.processing_stats['exported_fields'] += len(exportable_fields)
                
                # 从各语言目录读取数据
                log_progress("   ")
                log_progress("   📁 读取语言版本文件:")
                table_data_by_lang = {}
                for lang, lang_dir in lang_dirs.items():
                    excel_path = os.path.join(lang_dir, table_name)
                    lang_name = lang_names.get(lang, lang)
                    
                    if not os.path.exists(excel_path):
                        log_progress(f"      ✗ {lang_name}: 文件不存在")
                        continue
                    
                    # 读取Excel
                    try:
                        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                        if len(df) < 7:
                            log_progress(f"      ✗ {lang_name}: 数据行不足")
                            continue
                        table_data_by_lang[lang] = df
                        log_progress(f"      ✓ {lang_name}: {len(df)-6} 行数据")
                    except Exception as e:
                        log_progress(f"      ✗ {lang_name}: 读取失败 - {e}")
                        continue
                
                if not table_data_by_lang:
                    log_progress(f"   ⚠️  所有语言版本都不可用，跳过")
                    log_progress("")
                    self.processing_stats['skipped_tables'] += 1
                    continue
                
                # 提取并合并数据
                log_progress("   ")
                log_progress(f"   🔍 提取字段数据:")
                table_extracted_count = 0
                
                for field_name, field_type in exportable_fields:
                    # 每个语言版本分别按字段名定位列（不同语言表结构可能不同）
                    col_idx_by_lang: Dict[str, Optional[int]] = {}
                    for lang, df in table_data_by_lang.items():
                        col_idx_by_lang[lang] = self.find_column_index_by_name(df, field_name)

                    if all(v is None for v in col_idx_by_lang.values()):
                        log_progress(f"      ⚠️  {field_name} ({field_type}): 所有语言版本均未找到列")
                        continue

                    anchor_lang = 'zh' if 'zh' in table_data_by_lang else list(table_data_by_lang.keys())[0]
                    anchor_df = table_data_by_lang[anchor_lang]
                    anchor_col_idx = col_idx_by_lang.get(anchor_lang)
                    if anchor_col_idx is None:
                        for lang, idx_val in col_idx_by_lang.items():
                            if idx_val is not None:
                                anchor_lang = lang
                                anchor_df = table_data_by_lang[lang]
                                anchor_col_idx = idx_val
                                break

                    excel_col = self.column_index_to_letter(anchor_col_idx)
                    id_to_row_by_lang = {lang: self._build_id_to_row_index(df) for lang, df in table_data_by_lang.items()}

                    for anchor_row_idx in range(6, len(anchor_df)):
                        anchor_id = self._normalize_id_value(anchor_df.iloc[anchor_row_idx, 0])

                        lang_contents: Dict[str, str] = {}
                        has_content = False

                        for lang, df in table_data_by_lang.items():
                            col_idx = col_idx_by_lang.get(lang)
                            if col_idx is None:
                                continue

                            target_row_idx: Optional[int] = None
                            if anchor_id:
                                target_row_idx = id_to_row_by_lang.get(lang, {}).get(anchor_id)
                            if target_row_idx is None:
                                target_row_idx = anchor_row_idx if anchor_row_idx < len(df) else None

                            if target_row_idx is None:
                                continue

                            cell_value = df.iloc[target_row_idx, col_idx]
                            if pd.notna(cell_value) and str(cell_value).strip():
                                lang_contents[lang] = str(cell_value).strip()
                                has_content = True

                        if not has_content:
                            continue

                        # 过滤掉非文本内容（{}、数组、纯数字）
                        lang_contents = self.filter_lang_contents(lang_contents)
                        if not lang_contents:
                            continue

                        excel_row = anchor_row_idx + 1
                        excel_position = f"{excel_col}{excel_row}"

                        row_data = {
                            'table_name': table_name,
                            'sheet_name': sheet_name,
                            'field_name': field_name,
                            'field_type': field_type,
                            'excel_position': excel_position,
                            'chinese': lang_contents.get('zh', ''),
                            'vietnamese': lang_contents.get('vn', ''),
                            'thai': lang_contents.get('th', '')
                        }

                        all_data.append(row_data)
                        self.processing_stats['total_rows'] += 1
                        table_extracted_count += 1
                
                log_progress(f"      ✓ 共提取 {table_extracted_count} 条数据")
                log_progress("")
                
                self.processing_stats['processed_tables'] += 1
            
            self.translation_results = all_data
            
            log_progress("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            log_progress(f"✅ 处理完成！")
            log_progress(f"   • 处理表格: {self.processing_stats['processed_tables']}/{self.processing_stats['total_tables']}")
            log_progress(f"   • 导出字段: {self.processing_stats['exported_fields']} 个")
            log_progress(f"   • 提取数据: {len(all_data)} 条")
            log_progress("")
            
            return all_data
        
        except Exception as e:
            logger.error(f"处理失败: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def generate_translation_master_table_multi_lang(self, output_path: str) -> bool:
        """
        生成多语言翻译总表（基于多语言目录提取的数据）
        每个表格文件对应一个工作表标签
        列格式: 字段名 | 字段类型 | Excel位置 | ZH | VN | TH
        
        Args:
            output_path: 输出Excel文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            if not self.translation_results:
                logger.warning("没有数据可导出")
                return False
            
            logger.info(f"开始生成翻译总表: {output_path}")
            
            # 按表格名称分组
            tables_data = {}
            for row in self.translation_results:
                table_name = row['table_name']
                if table_name not in tables_data:
                    tables_data[table_name] = []
                tables_data[table_name].append(row)
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 为每个表格创建一个工作表
            for table_name, rows in tables_data.items():
                # 工作表名称（去除扩展名）
                sheet_name = Path(table_name).stem
                # Excel工作表名称限制31字符
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + '...'
                
                ws = wb.create_sheet(title=sheet_name)
                
                # 设置表头
                headers = ['Field', 'Type', 'Position', 'ZH', 'VN', 'TH']
                ws.append(headers)
                
                # 设置表头样式
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 写入数据行
                for row_data in rows:
                    ws.append([
                        row_data.get('field_name', ''),
                        row_data.get('field_type', ''),
                        row_data.get('excel_position', ''),
                        row_data.get('chinese', ''),
                        row_data.get('vietnamese', ''),
                        row_data.get('thai', '')
                    ])
                
                # 设置列宽
                column_widths = [20, 12, 12, 40, 40, 40]
                for col_idx, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 添加边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1:  # 数据行
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # 保存工作簿
            wb.save(output_path)
            logger.info(f"翻译总表已生成: {output_path}")
            logger.info(f"  - 工作表数量: {len(tables_data)}")
            logger.info(f"  - 总数据行数: {len(self.translation_results)}")
            
            return True
        
        except Exception as e:
            logger.error(f"生成翻译总表失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def generate_translation_csv(self, output_path: str) -> bool:
        """
        生成多语言翻译CSV文件
        列格式: 表名 | 工作表 | 字段名 | 字段类型 | Excel位置 | ZH | VN | TH
        
        Args:
            output_path: 输出CSV文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            if not self.translation_results:
                logger.warning("没有数据可导出")
                return False
            
            logger.info(f"开始生成翻译CSV: {output_path}")
            
            # 创建DataFrame
            df = pd.DataFrame(self.translation_results)
            
            # 重新排列列顺序
            columns = ['table_name', 'sheet_name', 'field_name', 'field_type', 
                      'excel_position', 'chinese', 'vietnamese', 'thai']
            
            # 确保所有列都存在
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[columns]
            
            # 重命名列标题
            df.columns = ['Table', 'Sheet', 'Field', 'Type', 'Position', 'ZH', 'VN', 'TH']
            
            # 保存CSV（使用UTF-8 BOM编码以便Excel正确识别中文）
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            logger.info(f"翻译CSV已生成: {output_path}")
            logger.info(f"  - 总数据行数: {len(self.translation_results)}")
            
            return True
        
        except Exception as e:
            logger.error(f"生成翻译CSV失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def generate_output_filename(output_dir: str = None) -> str:
        """
        自动生成输出文件名
        格式: 翻译提取_YYYYMMDD_HHMMSS.csv
        
        Args:
            output_dir: 输出目录，如果为None则使用当前目录
            
        Returns:
            str: 完整的输出文件路径
        """
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"翻译提取_{timestamp}.csv"
        
        if output_dir:
            return os.path.join(output_dir, filename)
        return filename
    
    def get_processing_report(self) -> str:
        """
        获取处理报告
        
        Returns:
            str: 处理报告文本
        """
        report_lines = [
            "=" * 70,
            "多语言翻译提取处理报告",
            "=" * 70,
            f"总表格数: {self.processing_stats['total_tables']}",
            f"已处理: {self.processing_stats['processed_tables']}",
            f"已跳过: {self.processing_stats['skipped_tables']}",
            "",
            f"总字段数: {self.processing_stats['total_fields']}",
            f"已导出字段: {self.processing_stats['exported_fields']}",
            f"已跳过字段（策划）: {self.processing_stats['skipped_fields']}",
            "",
            f"提取数据行数: {self.processing_stats['total_rows']}",
            "=" * 70
        ]
        
        if self.error_logs:
            report_lines.append("")
            report_lines.append("错误日志:")
            for error in self.error_logs:
                report_lines.append(f"  - {error}")
            report_lines.append("=" * 70)
        
        return "\n".join(report_lines)


def main():
    """命令行测试"""
    import argparse
    
    parser = argparse.ArgumentParser(description="多语言翻译提取器")
    parser.add_argument("json_config", help="JSON配置文件路径")
    parser.add_argument("excel_directory", help="Excel文件目录")
    parser.add_argument("--output", default="翻译总表.xlsx", help="输出文件路径")
    
    args = parser.parse_args()
    
    # 创建提取器
    translator = TableRangeTranslator()
    
    # 处理数据
    results = translator.process_with_json_config(args.json_config, args.excel_directory)
    
    if results:
        # 生成翻译总表
        translator.generate_translation_master_table(args.output)
        
        # 显示报告
        print(translator.get_processing_report())
    else:
        print("处理失败，没有数据")


if __name__ == "__main__":
    main()
