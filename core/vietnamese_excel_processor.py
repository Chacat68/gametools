#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
越南文Excel处理器
合并越南文检测和Excel扫描导出功能，支持文件夹输出
支持并行处理和流式读取大文件
"""

import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Union, Optional
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
import multiprocessing
from functools import partial

# 添加当前目录到路径
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

from localization_checker import VietnameseDetector
from error_handler import (
    FileProcessingError, DirectoryError, ExcelReadError, CSVReadError,
    validate_file_path, validate_directory, handle_errors, safe_execute,
    log_error_with_context
)
from log_manager import get_logger

# 使用增强的日志系统
logger = get_logger(__name__)


class VietnameseExcelProcessor:
    """越南文Excel处理器 - 合并检测和导出功能"""
    
    def __init__(self, max_workers: Optional[int] = None, enable_parallel: bool = True, chunk_size: int = 10000):
        """
        初始化处理器
        
        Args:
            max_workers: 最大并行工作进程数，None表示使用CPU核心数
            enable_parallel: 是否启用并行处理
            chunk_size: 大文件分块读取的行数
        """
        self.vietnamese_detector = VietnameseDetector()
        self.supported_extensions = {'.xlsx', '.xls', '.csv', '.tsv'}
        self.max_workers = max_workers or max(1, multiprocessing.cpu_count() - 1)
        self.enable_parallel = enable_parallel
        self.chunk_size = chunk_size
    
    def _get_excel_cell_reference(self, row: int, col: int) -> str:
        """
        将行列号转换为Excel单元格引用格式
        
        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）
            
        Returns:
            str: Excel单元格引用格式（如"C5"）
        """
        # 将列号转换为Excel列字母
        col_letter = ""
        while col > 0:
            col -= 1
            col_letter = chr(ord('A') + col % 26) + col_letter
            col //= 26
        
        return f"{col_letter}{row}"
    
    def is_supported_file(self, file_path: Path) -> bool:
        """
        检查文件是否为支持的格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是支持的格式返回True
        """
        return file_path.suffix.lower() in self.supported_extensions
    
    @handle_errors(default_return=[])
    def scan_excel_file(self, file_path: Path, use_chunks: bool = False) -> List[Dict]:
        """
        扫描单个Excel文件中的越南文（支持大文件流式处理）
        
        Args:
            file_path: Excel文件路径
            use_chunks: 是否使用分块读取（适用于大文件）
            
        Returns:
            List[Dict]: 包含越南文的位置信息列表
        """
        results = []
        
        try:
            # 验证文件
            validate_file_path(str(file_path), must_exist=True, extensions=['.xlsx', '.xls'])
            
            # 获取文件大小以决定是否使用分块
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            should_use_chunks = use_chunks or file_size_mb > 50  # 大于50MB使用分块
            
            if should_use_chunks:
                logger.info(f"文件大小 {file_size_mb:.2f}MB，使用分块读取模式")
            
            # 读取Excel文件的所有工作表
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    if should_use_chunks:
                        # 大文件使用分块读取
                        results.extend(self._scan_sheet_chunked(
                            excel_file, sheet_name, file_path
                        ))
                    else:
                        # 小文件一次性读取
                        df = excel_file.parse(sheet_name=sheet_name)
                        results.extend(self._scan_dataframe(
                            df, sheet_name, file_path
                        ))
                
                except Exception as e:
                    error = ExcelReadError(str(file_path), sheet_name, e)
                    log_error_with_context(error, {'sheet': sheet_name}, logger)
                    continue
        
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(str(file_path), original_error=e)
        
        return results
    
    def _scan_dataframe(self, df: pd.DataFrame, sheet_name: str, 
                       file_path: Path) -> List[Dict]:
        """
        扫描DataFrame中的越南文
        
        Args:
            df: 要扫描的DataFrame
            sheet_name: 工作表名
            file_path: 文件路径
            
        Returns:
            List[Dict]: 扫描结果列表
        """
        results = []
        
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                if pd.notna(value) and self.vietnamese_detector.contains_vietnamese(str(value)):
                    # 获取实际检测到的单元格内容
                    content = str(value)
                    # 基于实际检测到的内容判断语言类型
                    language_type = self.vietnamese_detector.detect_language_type(content)
                    results.append({
                        'excel_file': file_path.name,
                        'sheet_name': sheet_name,
                        'row': row_idx + 2,  # +2 因为pandas从0开始，且Excel有标题行
                        'col': col_idx + 1,  # +1 因为pandas从0开始
                        'column_name': df.columns[col_idx] if col_idx < len(df.columns) else f'Column_{col_idx + 1}',
                        'content': content,
                        'language_type': language_type,
                        'position': self._get_excel_cell_reference(row_idx + 2, col_idx + 1),
                        'file_path': str(file_path)
                    })
        
        return results
    
    def _scan_sheet_chunked(self, excel_file: pd.ExcelFile, sheet_name: str, 
                           file_path: Path) -> List[Dict]:
        """
        分块扫描工作表（适用于大文件）
        
        Args:
            excel_file: Excel文件对象
            sheet_name: 工作表名
            file_path: 文件路径
            
        Returns:
            List[Dict]: 扫描结果列表
        """
        results = []
        
        try:
            # 读取工作表以获取列信息
            df_sample = excel_file.parse(sheet_name=sheet_name, nrows=1)
            columns = df_sample.columns
            
            # 分块读取
            chunk_iter = pd.read_excel(
                file_path, 
                sheet_name=sheet_name, 
                chunksize=self.chunk_size
            )
            
            row_offset = 0
            for chunk_idx, df_chunk in enumerate(chunk_iter):
                for local_row_idx, row in df_chunk.iterrows():
                    actual_row_idx = row_offset + local_row_idx
                    
                    for col_idx, value in enumerate(row):
                        if pd.notna(value) and self.vietnamese_detector.contains_vietnamese(str(value)):
                            content = str(value)
                            language_type = self.vietnamese_detector.detect_language_type(content)
                            results.append({
                                'excel_file': file_path.name,
                                'sheet_name': sheet_name,
                                'row': actual_row_idx + 2,
                                'col': col_idx + 1,
                                'column_name': columns[col_idx] if col_idx < len(columns) else f'Column_{col_idx + 1}',
                                'content': content,
                                'language_type': language_type,
                                'position': self._get_excel_cell_reference(actual_row_idx + 2, col_idx + 1),
                                'file_path': str(file_path)
                            })
                
                row_offset += len(df_chunk)
                
                # 每处理一个块输出进度（可选）
                if chunk_idx % 10 == 0 and chunk_idx > 0:
                    print(f"  - 已处理 {row_offset} 行...")
        
        except Exception as e:
            print(f"分块读取工作表 '{sheet_name}' 时出错: {e}")
        
        return results
    
    def scan_csv_file(self, file_path: Path) -> List[Dict]:
        """
        扫描单个CSV文件中的越南文
        
        Args:
            file_path: CSV文件路径
            
        Returns:
            List[Dict]: 包含越南文的位置信息列表
        """
        results = []
        
        try:
            # 尝试不同的编码
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    
                    # 扫描每个单元格
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.notna(value) and self.vietnamese_detector.contains_vietnamese(str(value)):
                                # 获取实际检测到的单元格内容
                                content = str(value)
                                # 基于实际检测到的内容判断语言类型
                                language_type = self.vietnamese_detector.detect_language_type(content)
                                results.append({
                                    'excel_file': file_path.name,
                                    'sheet_name': 'CSV数据',
                                    'row': row_idx + 2,  # +2 因为pandas从0开始，且CSV有标题行
                                    'col': col_idx + 1,  # +1 因为pandas从0开始
                                    'column_name': df.columns[col_idx] if col_idx < len(df.columns) else f'Column_{col_idx + 1}',
                                    'content': content,
                                    'language_type': language_type,  # 基于实际检测内容判断的语言类型
                                    'position': self._get_excel_cell_reference(row_idx + 2, col_idx + 1),
                                    'file_path': str(file_path)
                                })
                    break  # 成功读取后跳出循环
                    
                except UnicodeDecodeError:
                    continue
            
        except Exception as e:
            print(f"读取CSV文件 {file_path} 时出错: {e}")
        
        return results
    
    def scan_single_file(self, file_path: Path) -> List[Dict]:
        """
        扫描单个文件中的越南文
        
        Args:
            file_path: 文件路径
            
        Returns:
            List[Dict]: 包含越南文的位置信息列表
        """
        if not self.is_supported_file(file_path):
            return []
        
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            return self.scan_excel_file(file_path)
        elif file_path.suffix.lower() in ['.csv', '.tsv']:
            return self.scan_csv_file(file_path)
        
        return []
    
    def scan_directory(
        self,
        directory_path: str,
        recursive: bool = True,
        return_files: bool = False
    ) -> Union[List[Dict], Tuple[List[Dict], List[Path]]]:
        """
        扫描目录下的所有支持文件（支持并行处理）
        
        Args:
            directory_path: 要扫描的目录路径
            recursive: 是否递归扫描子目录
            return_files: 是否返回文件列表，供统计使用
            
        Returns:
            List[Dict] 或 (List[Dict], List[Path]): 越南文位置信息及可选的文件列表
        """
        try:
            # 验证目录
            directory = validate_directory(directory_path, must_exist=True)
        except DirectoryError as e:
            logger.error(str(e))
            return ([], []) if return_files else []
        
        all_results = []
        supported_files = []
        
        # 收集所有支持的文件
        if recursive:
            for file_path in directory.rglob('*'):
                if file_path.is_file() and self.is_supported_file(file_path):
                    supported_files.append(file_path)
        else:
            for file_path in directory.iterdir():
                if file_path.is_file() and self.is_supported_file(file_path):
                    supported_files.append(file_path)
        
        print(f"找到 {len(supported_files)} 个支持的文件")
        
        # 根据配置选择串行或并行处理
        if self.enable_parallel and len(supported_files) > 3:
            all_results = self._scan_directory_parallel(supported_files)
        else:
            all_results = self._scan_directory_serial(supported_files)
        
        if return_files:
            return all_results, supported_files
        return all_results
    
    def _scan_directory_serial(self, file_paths: List[Path]) -> List[Dict]:
        """
        串行扫描文件列表
        
        Args:
            file_paths: 文件路径列表
            
        Returns:
            List[Dict]: 扫描结果列表
        """
        all_results = []
        for i, file_path in enumerate(file_paths, 1):
            print(f"正在扫描 ({i}/{len(file_paths)}): {file_path.name}")
            
            file_results = self.scan_single_file(file_path)
            all_results.extend(file_results)
            
            if file_results:
                print(f"  - 找到 {len(file_results)} 个越南文位置")
            else:
                print(f"  - 未找到越南文")
        
        return all_results
    
    def _scan_directory_parallel(self, file_paths: List[Path]) -> List[Dict]:
        """
        并行扫描文件列表
        
        Args:
            file_paths: 文件路径列表
            
        Returns:
            List[Dict]: 扫描结果列表
        """
        all_results = []
        total_files = len(file_paths)
        
        print(f"使用并行处理模式，工作进程数: {self.max_workers}")
        
        # 使用ThreadPoolExecutor以避免序列化问题
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_file = {executor.submit(self.scan_single_file, fp): fp for fp in file_paths}
            
            # 处理完成的任务
            completed = 0
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                completed += 1
                
                try:
                    file_results = future.result()
                    all_results.extend(file_results)
                    
                    status = f"找到 {len(file_results)} 个越南文位置" if file_results else "未找到越南文"
                    print(f"[{completed}/{total_files}] {file_path.name}: {status}")
                    
                except Exception as e:
                    print(f"[{completed}/{total_files}] {file_path.name}: 处理失败 - {str(e)}")
        
        return all_results
    
    def create_output_excel(self, results: List[Dict], output_folder: str, filename: str = "越南文检测结果.xlsx") -> str:
        """
        创建输出Excel文件
        
        Args:
            results: 扫描结果列表
            output_folder: 输出文件夹路径
            filename: 输出文件名
            
        Returns:
            str: 输出文件的完整路径
        """
        try:
            # 确保输出文件夹存在
            output_path = Path(output_folder)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # 构建完整的输出文件路径
            full_output_path = output_path / filename
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "越南文检测结果"
            
            # 设置标题行
            headers = ['序号', '文件名', '位置', '越南文内容', '语言类型']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加数据
            for row_idx, result in enumerate(results, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
                ws.cell(row=row_idx, column=2, value=result['excel_file'])  # 文件名
                ws.cell(row=row_idx, column=3, value=result['position'])  # 位置
                ws.cell(row=row_idx, column=4, value=result['content'])  # 越南文内容
                ws.cell(row=row_idx, column=5, value=result['language_type'])  # 语言类型
            
            # 设置列宽
            column_widths = [8, 30, 20, 60, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 保存文件
            wb.save(full_output_path)
            return str(full_output_path)
            
        except Exception as e:
            print(f"创建输出Excel文件时出错: {e}")
            return ""
    
    
    def process_directory(self, directory_path: str, output_folder: str, recursive: bool = True, 
                         create_excel: bool = True, create_report: bool = False) -> Dict:
        """
        处理目录并导出结果
        
        Args:
            directory_path: 要扫描的目录路径
            output_folder: 输出文件夹路径
            recursive: 是否递归扫描子目录
            create_excel: 是否创建Excel结果文件
            create_report: 是否创建汇总报告（已废弃，始终为False）
            
        Returns:
            Dict: 包含处理统计信息的字典
        """
        print("开始扫描文件中的越南文...")
        print("=" * 50)
        
        # 扫描目录，复用已收集的文件列表用于统计
        results, supported_files = self.scan_directory(directory_path, recursive, return_files=True)
        
        # 统计信息
        stats = {
            'total_files_scanned': len(supported_files),
            'files_with_vietnamese': len(set(result['excel_file'] for result in results)),
            'total_vietnamese_locations': len(results),
            'results': results,
            'output_files': []
        }
        
        print("\n" + "=" * 50)
        print("扫描完成！")
        print(f"扫描的文件总数: {stats['total_files_scanned']}")
        print(f"包含越南文的文件数: {stats['files_with_vietnamese']}")
        print(f"越南文位置总数: {stats['total_vietnamese_locations']}")
        
        # 创建输出文件
        if results:
            print(f"\n正在创建输出文件到: {output_folder}")
            
            if create_excel:
                excel_path = self.create_output_excel(results, output_folder)
                if excel_path:
                    print(f"Excel结果文件创建成功: {excel_path}")
                    stats['output_files'].append(excel_path)
                    stats['excel_success'] = True
                else:
                    print("Excel结果文件创建失败！")
                    stats['excel_success'] = False
            
        else:
            print("未找到任何越南文内容，不创建输出文件。")
            stats['excel_success'] = False
        
        return stats


def main():
    """主函数 - 命令行版本"""
    import sys
    
    print("越南文Excel处理器")
    print("=" * 50)
    
    # 获取用户输入
    if len(sys.argv) > 1:
        directory_path = sys.argv[1]
    else:
        directory_path = input("请输入要扫描的目录路径: ").strip()
    
    if not directory_path:
        print("错误: 未提供目录路径")
        return
    
    if len(sys.argv) > 2:
        output_folder = sys.argv[2]
    else:
        output_folder = input("请输入输出文件夹路径: ").strip()
    
    if not output_folder:
        print("错误: 未提供输出文件夹路径")
        return
    
    # 创建处理器并执行处理
    processor = VietnameseExcelProcessor()
    stats = processor.process_directory(directory_path, output_folder)
    
    print("\n按任意键退出...")
    input()


if __name__ == "__main__":
    main()
