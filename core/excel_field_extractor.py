#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel表字段导出器
扫描Excel文件，检测包含本地化文本内容（中文、越南文、泰文）的列，
从 FIELD_NAME_ROW 提取字段名，从 FIELD_TYPE_ROW 提取字段类型（策划、前端、后端、前后端）。
忽略纯数字、英文代码和配置项。
"""

import os
import json
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 导入统一的常量和模式
from core.constants import (
    SUPPORTED_LANGUAGES,
    SUPPORTED_EXCEL_EXTENSIONS,
    COLUMN_MARKER,
    FIELD_EXTRACTION_MERGED_JSON_NAME,
    FIELD_NAME_ROW,
    FIELD_TYPE_ROW,
    DATA_START_ROW,
    ROW_BOUNDARY_KEYWORD,
)
from core.text_patterns import (
    TEXT_PATTERN,
    is_filterable_content,
    contains_localized_text,
    contains_cjk_vietnamese_or_thai,
)


class ExcelFieldExtractor:
    """Excel表字段导出器"""
    
    # 使用统一的语言配置
    SUPPORTED_LANGUAGES = SUPPORTED_LANGUAGES
    
    def __init__(self):
        self.supported_extensions = SUPPORTED_EXCEL_EXTENSIONS
        # 要过滤的字段名列表（不区分大小写）
        # 注意：仅在“该列不包含本地化字符”时才会过滤，避免误伤真实的中文/越南文/泰文 name 列。
        self.excluded_field_names = {
            'name', 'model', 'id', 'code', 'type'
        }
        # 错误日志列表
        self.error_logs = []
        self.extraction_warnings = []  # 提取警告（例如第6行数据为空）
        
        # 行下限边界关键字（与 core.constants.ROW_BOUNDARY_KEYWORD 一致）
        self.boundary_keyword = ROW_BOUNDARY_KEYWORD        
        # 进度回调函数
        self.progress_callback = None        
        # 警告输出控制：避免大量警告导致卡顿
        self.verbose_warnings = False  # 是否输出每条警告到控制台
        self.max_printed_warnings = 10  # 最多打印的警告数量
        self._printed_warning_count = 0  # 已打印的警告计数
        
        # 警告级别控制
        # 字段类型为空会直接影响后续翻译提取与日志排障，默认开启。
        self.warn_empty_field_type = True  # 是否警告字段类型为空
        self.warn_missing_c_marker = False  # 是否警告缺少 c_ 标记（默认关闭，会自动全表扫描）
    
    def _add_warning(self, warning_msg: str):
        """
        添加警告信息（控制输出频率避免卡顿）
        
        Args:
            warning_msg: 警告消息
        """
        self.extraction_warnings.append(warning_msg)
        
        # 控制打印输出，避免大量警告导致卡顿
        if self.verbose_warnings or self._printed_warning_count < self.max_printed_warnings:
            print(warning_msg)
            self._printed_warning_count += 1
        elif self._printed_warning_count == self.max_printed_warnings:
            print(f"... 后续警告已省略，共 {len(self.extraction_warnings)} 条警告将记录在日志中")
            self._printed_warning_count += 1
    
    def set_progress_callback(self, callback):
        """
        设置进度回调函数
        
        Args:
            callback: 回调函数，签名为 callback(message: str, percentage: float = None)
        """
        self.progress_callback = callback
    
    def _report_progress(self, message: str, percentage: float = None):
        """
        报告处理进度
        
        Args:
            message: 进度消息
            percentage: 进度百分比 (0-100)
        """
        if self.progress_callback:
            self.progress_callback(message, percentage)    
    def _check_row_boundary(self, sheet, row_idx: int) -> bool:
        """
        检查当前行是否为「数据区下限边界行」。
        若该行任意单元格 strip 后小写等于 ``self.boundary_keyword``（默认 ``ROW_BOUNDARY_KEYWORD``），
        则本行及以下不再作为数据遍历，调用方应立即 ``break`` 结束行循环。
        
        Args:
            sheet: openpyxl工作表对象
            row_idx: 行号（从1开始）
            
        Returns:
            bool: True 表示已到达边界行，应停止向下遍历
        """
        if row_idx > sheet.max_row:
            return False
        
        for cell in sheet[row_idx]:
            if cell.value is not None:
                cell_str = str(cell.value).strip().lower()
                if cell_str == self.boundary_keyword:
                    return True
        return False
    
    def _find_boundary_row(self, sheet, start_row: int = DATA_START_ROW) -> int:
        """
        从 start_row 起向下查找第一个边界行（任意格等于行边界关键字）。
        返回该行的物理行号；未找到则返回 ``sheet.max_row + 1``。
        数据区遍历应使用 ``range(data_start_row, 返回值)``，不包含边界行本身。
        
        Args:
            sheet: openpyxl工作表对象
            start_row: 开始搜索的行号（默认从 DATA_START_ROW 开始）
            
        Returns:
            int: 边界行的物理行号；未找到则为 sheet.max_row + 1
        """
        for row_idx in range(start_row, sheet.max_row + 1):
            if self._check_row_boundary(sheet, row_idx):
                return row_idx
        return sheet.max_row + 1
    
    def is_excel_file(self, file_path: Path) -> bool:
        """
        检查文件是否为支持的Excel格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是Excel格式返回True
        """
        return file_path.suffix.lower() in self.supported_extensions
    
    def contains_text(self, value) -> bool:
        """
        检查值是否包含本地化文本内容（仅限中文、越南文、泰文）
        排除：纯数字、英文代码、配置项、数组、对象等
        
        Args:
            value: 要检查的值
            
        Returns:
            bool: 包含本地化文本返回True
        """
        if pd.isna(value):
            return False
        
        value_str = str(value).strip()
        
        # 排除空值和应被过滤的内容
        if not value_str or is_filterable_content(value_str):
            return False
        
        # 检查是否包含中文、越南文或泰文字符
        return contains_localized_text(value_str)
    
    def find_column_range_between_markers(self, sheet, marker: Optional[str] = None):
        """
        查找两个标记之间的列范围（不包含标记本身）
        
        Args:
            sheet: openpyxl工作表对象
            marker: 要查找的标记文本（默认 COLUMN_MARKER）
            
        Returns:
            Tuple[int, int] or None: (起始列号, 结束列号)，如果未找到2个标记返回 None
        """
        if marker is None:
            marker = COLUMN_MARKER
        marker_columns = []
        
        # 在字段名行查找标记
        field_row = FIELD_NAME_ROW
        if sheet.max_row >= field_row:
            for cell in sheet[field_row]:
                if cell.value is None:
                    continue

                # 只将“值严格等于 marker”的列视为边界标记。
                # 注意：像 c_story、c_classic_battle_story 这类以 c_ 开头的真实字段不应被当成标记。
                cell_value = str(cell.value).strip()
                if cell_value == marker:
                    marker_columns.append(cell.column)
        
        # 如果找到至少两个标记，返回它们之间的范围（不包含标记列本身）
        if len(marker_columns) >= 2:
            start_col = marker_columns[0] + 1  # 第一个标记的下一列
            end_col = marker_columns[1] - 1    # 第二个标记的上一列
            # 确保范围有效
            if start_col <= end_col:
                return (start_col, end_col)
        
        # 如果没有找到两个标记，返回 None
        return None
    
    def extract_fields_from_excel(self, file_path: Path) -> List[Dict]:
        """
        从Excel文件中提取包含文本内容的列的字段信息
        仅扫描两个 c_ 标记之间的列范围
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            List[Dict]: 每个工作表的字段信息列表
        """
        results = []
        
        try:
            # 使用openpyxl读取，以便精确访问物理行
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]
                    data_start_row = DATA_START_ROW
                    boundary_row = self._find_boundary_row(sheet, data_start_row)

                    def scan_text_columns_in_range(start_col: int, end_col: int) -> Tuple[Set[int], Dict[int, int], int]:
                        """在指定列范围内扫描数据行，返回文本列集合与统计信息。"""
                        text_column_counts: Dict[int, int] = {}
                        total_data_rows = 0

                        if boundary_row > data_start_row:
                            total_data_rows = boundary_row - data_start_row
                            for row in sheet.iter_rows(
                                min_row=data_start_row,
                                max_row=boundary_row - 1,  # 不包含边界行
                                min_col=start_col,
                                max_col=end_col,
                            ):
                                for cell in row:
                                    if cell.value is not None and self.contains_text(cell.value):
                                        col_num = cell.column
                                        text_column_counts[col_num] = text_column_counts.get(col_num, 0) + 1

                        min_text_rows = 3  # 最少文本行数
                        min_text_ratio = 0.10  # 最小文本占比 10%

                        text_columns: Set[int] = set()
                        for col_num, count in text_column_counts.items():
                            if count >= min_text_rows or (
                                total_data_rows > 0 and count / total_data_rows >= min_text_ratio
                            ):
                                text_columns.add(col_num)

                        return text_columns, text_column_counts, total_data_rows

                    # 1) 优先按两个 c_ 标记之间的范围扫描（如果存在）
                    # 2) 若该范围未检测到任何文本列，则自动回退为全表扫描，避免漏掉文本列在标记外的表
                    column_range = self.find_column_range_between_markers(sheet)

                    if column_range is None:
                        if self.warn_missing_c_marker:
                            warning_msg = (
                                f"⚠️ 未找到两个 c_ 标记，改为全表扫描 | "
                                f"文件: {file_path.name} | "
                                f"工作表: {sheet_name}"
                            )
                            self._add_warning(warning_msg)
                        start_col, end_col = 1, sheet.max_column
                        text_columns, _, _ = scan_text_columns_in_range(start_col, end_col)
                    else:
                        start_col, end_col = column_range
                        text_columns, _, _ = scan_text_columns_in_range(start_col, end_col)

                        if not text_columns:
                            if self.warn_missing_c_marker:
                                warning_msg = (
                                    f"⚠️ c_ 标记范围未检测到文本列，改为全表扫描 | "
                                    f"文件: {file_path.name} | "
                                    f"工作表: {sheet_name}"
                                )
                                self._add_warning(warning_msg)
                            start_col, end_col = 1, sheet.max_column
                            text_columns, _, _ = scan_text_columns_in_range(start_col, end_col)
                    
                    if not text_columns:
                        # 未检测到本地化文本列：仍然加入结果，标记 has_text=False，字段为空
                        results.append({
                            'excel_file': file_path.name,
                            'sheet_name': sheet_name,
                            'fields': [],
                            'fields_with_examples': [],
                            'field_column_letters': [],
                            'field_count': 0,
                            'text_columns': [],
                            'has_text': False
                        })
                        continue
                    
                    # 从第5行提取字段名，从第6行提取字段类型
                    fields = []
                    field_with_types = []  # 字段名+字段类型+列字母的组合列表
                    field_column_letters = []  # 列字母列表
                    field_row = FIELD_NAME_ROW
                    type_row = FIELD_TYPE_ROW
                    
                    if sheet.max_row >= field_row:
                        for col_num in sorted(text_columns):
                            # 提取字段名
                            field_cell = sheet.cell(row=field_row, column=col_num)
                            field_name = str(field_cell.value) if field_cell.value is not None else f"列{col_num}"

                            # 字段名过滤（内容检测优先）：
                            # - 若字段名在 excluded_field_names 中，但该列数据区抽样【不含】中文/越南文/泰文，则视为代码列过滤掉。
                            # - 若该列含上述脚本字符，则保留，避免过滤真实本地化 name 列。
                            if field_name and field_name.strip().lower() in self.excluded_field_names:
                                has_localized = False
                                if boundary_row > data_start_row:
                                    # 仅抽样检查，避免整列扫描带来的性能成本
                                    checks = 0
                                    max_checks = 50
                                    for row_idx in range(data_start_row, boundary_row):  # 不超过边界行
                                        # 检查是否到达边界
                                        if self._check_row_boundary(sheet, row_idx):
                                            break
                                        cell_val = sheet.cell(row=row_idx, column=col_num).value
                                        if cell_val is None:
                                            continue
                                        val_str = str(cell_val).strip()
                                        if not val_str or is_filterable_content(val_str):
                                            continue
                                        if contains_cjk_vietnamese_or_thai(val_str):
                                            has_localized = True
                                            break
                                        checks += 1
                                        if checks >= max_checks:
                                            break

                                if not has_localized:
                                    # 过滤代码列/ID列等
                                    continue
                            
                            # 获取列字母（如 A, B, F 等）
                            col_letter = get_column_letter(col_num)
                            
                            fields.append(field_name)
                            field_column_letters.append(col_letter)
                            
                            # 提取字段类型行的单元格
                            if sheet.max_row >= type_row:
                                type_cell = sheet.cell(row=type_row, column=col_num)
                                field_type = str(type_cell.value) if type_cell.value is not None else ""
                                
                                # 检查字段类型是否为空
                                if self.warn_empty_field_type and (not field_type or field_type.strip() == ""):
                                    warning_msg = (
                                        f"⚠️ 字段类型为空 | "
                                        f"文件: {file_path.name} | "
                                        f"工作表: {sheet_name} | "
                                        f"字段: {field_name} | "
                                        f"位置: 第{type_row}行,第{col_num}列({col_letter}{type_row})"
                                    )
                                    self._add_warning(warning_msg)
                                
                                # 格式：字段名,字段类型,列字母
                                field_with_type = f"{field_name},{field_type},{col_letter}"
                            else:
                                # 表格行数不足 FIELD_TYPE_ROW
                                warning_msg = (
                                    f"⚠️ 表格行数不足 | "
                                    f"文件: {file_path.name} | "
                                    f"工作表: {sheet_name} | "
                                    f"字段: {field_name} | "
                                    f"当前行数: {sheet.max_row} (需要至少{FIELD_TYPE_ROW}行)"
                                )
                                self._add_warning(warning_msg)
                                field_with_type = f"{field_name},,{col_letter}"
                            
                            field_with_types.append(field_with_type)
                    else:
                        # 如果表格行数不足5行，使用列号
                        for col_num in sorted(text_columns):
                            col_letter = get_column_letter(col_num)
                            fields.append(f"列{col_num}")
                            field_column_letters.append(col_letter)
                            field_with_types.append(f"列{col_num},,{col_letter}")
                    
                    if fields:
                        results.append({
                            'excel_file': file_path.name,
                            'sheet_name': sheet_name,
                            'fields': fields,
                            'fields_with_examples': field_with_types,  # 字段名+字段类型+列字母列表
                            'field_column_letters': field_column_letters,  # 列字母列表
                            'field_count': len(fields),
                            'text_columns': sorted(text_columns),
                            'has_text': True
                        })
                    elif text_columns:
                        # 曾检测到文本列，但经字段名/类型过滤后无可导出字段 → 视为无文本表
                        results.append({
                            'excel_file': file_path.name,
                            'sheet_name': sheet_name,
                            'fields': [],
                            'fields_with_examples': [],
                            'field_column_letters': [],
                            'field_count': 0,
                            'text_columns': sorted(text_columns),
                            'has_text': False
                        })
                
                except Exception as e:
                    error_msg = f"❌ 读取工作表失败 | 文件: {file_path.name} | 工作表: {sheet_name} | 错误: {str(e)}"
                    self.error_logs.append(error_msg)
                    print(error_msg)
            
            wb.close()
        
        except Exception as e:
            error_msg = f"❌ 读取文件失败 | 文件: {file_path} | 错误: {str(e)}"
            self.error_logs.append(error_msg)
            print(error_msg)
        
        return results
    
    def scan_directory(self, directory: Path, recursive: bool = True, 
                        lang_name: str = None, lang_progress_base: float = 0,
                        lang_progress_range: float = 100) -> List[Dict]:
        """
        扫描目录中的所有Excel文件
        
        Args:
            directory: 要扫描的目录
            recursive: 是否递归扫描子目录
            lang_name: 当前处理的语言名称（用于进度显示）
            lang_progress_base: 当前语言的进度基准值
            lang_progress_range: 当前语言的进度范围
            
        Returns:
            List[Dict]: 所有Excel文件的字段信息
        """
        all_results = []
        
        # 重置警告打印计数
        self._printed_warning_count = 0
        
        if not directory.exists():
            print(f"目录不存在: {directory}")
            return all_results
        
        # 报告开始扫描
        lang_prefix = f"[{lang_name}] " if lang_name else ""
        self._report_progress(f"{lang_prefix}正在扫描目录...", lang_progress_base)
        
        # 获取所有Excel文件
        if recursive:
            excel_files = [f for f in directory.rglob("*") if self.is_excel_file(f)]
        else:
            excel_files = [f for f in directory.glob("*") if self.is_excel_file(f)]
        
        total_files = len(excel_files)
        print(f"找到 {total_files} 个Excel文件")
        self._report_progress(f"{lang_prefix}找到 {total_files} 个Excel文件", lang_progress_base)
        
        for idx, file_path in enumerate(excel_files, 1):
            # 计算当前进度
            file_progress = lang_progress_base + (idx / total_files) * lang_progress_range if total_files > 0 else lang_progress_base
            
            # 报告文件处理进度
            self._report_progress(
                f"{lang_prefix}处理 ({idx}/{total_files}): {file_path.name}", 
                file_progress
            )
            print(f"处理文件 {idx}/{total_files}: {file_path.name}")
            
            results = self.extract_fields_from_excel(file_path)
            all_results.extend(results)
        
        return all_results
    
    def _format_json_data(self, results: List[Dict], language: str = None) -> Dict:
        """
        格式化JSON数据
        
        Args:
            results: 字段信息列表
            language: 语言代码
            
        Returns:
            Dict: 格式化后的JSON数据
        """
        # 分组：无文本表与有文本表
        no_text_results = [r for r in results if not r.get('has_text', True)]
        text_results = [r for r in results if r.get('has_text', True)]
        
        json_output = {
            "no_text_tables": [],
            "text_tables": []
        }
        
        # 添加语言标记（如果指定了语言）
        if language and language in self.SUPPORTED_LANGUAGES:
            lang_info = self.SUPPORTED_LANGUAGES[language]
            json_output["language"] = {
                "code": lang_info['code'],
                "name": lang_info['name']
            }
        
        # 无文本表：只包含表名信息
        for result in no_text_results:
            json_output["no_text_tables"].append({
                "table_name": result['excel_file'],
                "sheet_name": result['sheet_name']
            })
        
        # 有文本表：包含完整字段结构
        for result in text_results:
            table_entry = {
                "table_name": result['excel_file'],
                "sheet_name": result['sheet_name'],
                "fields_with_examples": result.get('fields_with_examples', []),
                "field_column_letters": result.get('field_column_letters', []),
                "field_count": result['field_count']
            }
            json_output["text_tables"].append(table_entry)
            
        return json_output

    def export_to_json(self, results: List[Dict], output_file: Path, language: str = None):
        """
        导出结果到JSON文件
        分为两部分：
        1) no_text_tables: 无文本内容的表格（只显示表名和工作表名）
        2) text_tables: 包含文本内容的表格（显示字段名+字段类型）
        
        Args:
            results: 字段信息列表（已按 has_text 排序）
            output_file: 输出文件路径
            language: 语言代码（如 'zh', 'vn', 'th', 'en'），如果指定则添加语言标记
        """
        try:
            # 构建JSON结构
            json_output = self._format_json_data(results, language)
            
            # 写入JSON文件
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, ensure_ascii=False, indent=2)
            
            lang_str = f" [{self.SUPPORTED_LANGUAGES[language]['name']}]" if language else ""
            print(f"结果已导出到{lang_str}: {output_file}")
            
            # 统计数量
            no_text_count = len(json_output["no_text_tables"])
            text_count = len(json_output["text_tables"])
            print(f"  - 无文本表格: {no_text_count} 个")
            print(f"  - 有文本表格: {text_count} 个")
        
        except Exception as e:
            print(f"导出到JSON时出错: {e}")
    
    def export_to_csv(self, results: List[Dict], output_file: Path):
        """
        导出结果到CSV文件
        分为两部分：1) 无文本内容的表格（只显示表名）；2) 有文本内容的表格（显示完整字段结构）
        格式：表名,字段1,字段2,...
        
        Args:
            results: 字段信息列表（已按 has_text 排序）
            output_file: 输出文件路径
        """
        try:
            # 分组：无文本表与有文本表
            no_text_results = [r for r in results if not r.get('has_text', True)]
            text_results = [r for r in results if r.get('has_text', True)]
            
            with open(output_file, 'w', encoding='utf-8-sig') as f:
                # 写入无文本表部分
                if no_text_results:
                    f.write("# 无文本内容的表格\n")
                    for result in no_text_results:
                        table_name = f"{result['excel_file']}#{result['sheet_name']}"
                        f.write(f"{table_name}\n")
                    f.write("\n")
                
                # 写入有文本表部分
                if text_results:
                    f.write("# 包含文本内容的表格\n")
                    for result in text_results:
                        table_name = f"{result['excel_file']}#{result['sheet_name']}"
                        fields_with_examples = result.get('fields_with_examples', result['fields'])
                        if isinstance(fields_with_examples, list) and len(fields_with_examples) > 0:
                            if ',' in str(fields_with_examples[0]):
                                fields_str = ','.join(fields_with_examples)
                            else:
                                fields_str = ','.join(result['fields'])
                        else:
                            fields_str = ','.join(result['fields'])
                        f.write(f"{table_name},{fields_str}\n")
            
            print(f"结果已导出到: {output_file}")
            print(f"  - 无文本表格: {len(no_text_results)} 个")
            print(f"  - 有文本表格: {len(text_results)} 个")
        
        except Exception as e:
            print(f"导出到CSV时出错: {e}")
    
    def export_to_excel(self, results: List[Dict], output_file: Path):
        """
        导出结果到Excel文件，带格式
        
        Args:
            results: 字段信息列表
            output_file: 输出文件路径
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "字段导出结果"
            
            # 设置样式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            ws['A1'] = '表名'
            ws['B1'] = '工作表'
            ws['C1'] = '字段数量'
            ws['D1'] = '字段+示例'
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}1']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 设置列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 100
            
            # 分组：无文本表与有文本表
            no_text_results = [r for r in results if not r.get('has_text', True)]
            text_results = [r for r in results if r.get('has_text', True)]

            current_row = 2
            # 写入无文本表部分（只显示表名）
            if no_text_results:
                # 添加分组标题
                ws[f'A{current_row}'] = '无文本内容的表格'
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                title_cell = ws[f'A{current_row}']
                title_cell.font = Font(bold=True, size=12, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
                
                for result in no_text_results:
                    ws[f'A{current_row}'] = result['excel_file']
                    ws[f'B{current_row}'] = result['sheet_name']
                    ws[f'C{current_row}'] = 0
                    ws[f'D{current_row}'] = ''
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{current_row}'].border = border
                    ws[f'C{current_row}'].alignment = Alignment(horizontal="center")
                    current_row += 1
                
                # 空行作为分隔
                current_row += 1

            # 写入有文本表部分（显示完整字段结构）
            if text_results:
                # 添加分组标题
                ws[f'A{current_row}'] = '包含文本内容的表格'
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                title_cell = ws[f'A{current_row}']
                title_cell.font = Font(bold=True, size=12, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
                
                for result in text_results:
                    ws[f'A{current_row}'] = result['excel_file']
                    ws[f'B{current_row}'] = result['sheet_name']
                    ws[f'C{current_row}'] = result['field_count']
                    fields_with_examples = result.get('fields_with_examples', [])
                    ws[f'D{current_row}'] = ', '.join(fields_with_examples) if fields_with_examples else ''
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{current_row}'].border = border
                    ws[f'C{current_row}'].alignment = Alignment(horizontal="center")
                    current_row += 1
            
            wb.save(output_file)
            print(f"结果已导出到: {output_file}")
            print(f"  - 无文本表格: {len(no_text_results)} 个")
            print(f"  - 有文本表格: {len(text_results)} 个")
        
        except Exception as e:
            print(f"导出到Excel时出错: {e}")
    
    def process_directory(self, 
                         directory_path: str, 
                         output_folder: str = None,
                         output_format: str = 'json',
                         recursive: bool = True,
                         language: str = None,
                         write_output: bool = True,
                         lang_progress_base: float = 0,
                         lang_progress_range: float = 100) -> Dict:
        """
        处理目录并导出结果
        
        Args:
            directory_path: 要扫描的目录路径
            output_folder: 输出文件夹（如果为None，使用扫描目录）
            output_format: 输出格式 ('json', 'csv' 或 'excel')
            recursive: 是否递归扫描
            language: 语言代码（如 'zh', 'vn', 'th', 'en'），用于输出文件命名和JSON语言标记
            write_output: 是否写入输出文件
            lang_progress_base: 当前语言的进度基准值
            lang_progress_range: 当前语言的进度范围
            
        Returns:
            Dict: 处理统计信息，包含results数据
        """
        directory = Path(directory_path)
        
        if output_folder:
            output_dir = Path(output_folder)
        else:
            output_dir = directory
        
        if write_output:
            output_dir.mkdir(parents=True, exist_ok=True)
        
        # 扫描目录
        lang_name = self.SUPPORTED_LANGUAGES[language]['name'] if language and language in self.SUPPORTED_LANGUAGES else None
        lang_str = f" [{lang_name}]" if lang_name else ""
        print(f"开始扫描目录{lang_str}: {directory}")
        self._report_progress(f"{lang_str} 开始扫描...", lang_progress_base)
        
        results = self.scan_directory(
            directory, 
            recursive=recursive,
            lang_name=lang_name,
            lang_progress_base=lang_progress_base,
            lang_progress_range=lang_progress_range * 0.9  # 90% 用于扫描
        )

        # 重新排序：无文本表优先，其次按文件名与工作表名
        results.sort(key=lambda r: (0 if not r.get('has_text', True) else 1, r['excel_file'], r['sheet_name']))
        
        output_file = None
        if not results:
            print("未找到包含文本内容的Excel表格")
            return {
                'total_files': 0,
                'total_sheets': 0,
                'total_fields': 0,
                'output_file': None,
                'results': [],
                'language': language
            }
        
        # 生成输出文件名（带语言后缀）
        lang_suffix = self.SUPPORTED_LANGUAGES[language]['suffix'] if language and language in self.SUPPORTED_LANGUAGES else ""
        
        # 导出结果
        if write_output:
            if output_format == 'json':
                output_file = output_dir / f"field_extraction_result{lang_suffix}.json"
                self.export_to_json(results, output_file, language=language)
            elif output_format == 'csv':
                output_file = output_dir / f"field_extraction_result{lang_suffix}.csv"
                self.export_to_csv(results, output_file)
            else:
                output_file = output_dir / f"field_extraction_result{lang_suffix}.xlsx"
                self.export_to_excel(results, output_file)
        
        # 统计信息
        total_files = len(set(r['excel_file'] for r in results))
        total_sheets = len(results)
        total_fields = sum(r['field_count'] for r in results)
        
        stats = {
            'total_files': total_files,
            'total_sheets': total_sheets,
            'total_fields': total_fields,
            'output_file': str(output_file) if output_file else None,
            'results': results,
            'language': language
        }
        
        print(f"\n处理完成!")
        print(f"总文件数: {total_files}")
        print(f"总工作表数: {total_sheets}")
        print(f"总字段数: {total_fields}")
        
        # 显示日志统计
        if self.error_logs:
            print(f"错误日志数: {len(self.error_logs)}")
        if self.extraction_warnings:
            print(f"警告日志数: {len(self.extraction_warnings)}")
        
        return stats
    
    def process_multi_language_directories(self, 
                                           directories: Dict[str, str],
                                           output_folder: str = None,
                                           output_format: str = 'json',
                                           recursive: bool = True) -> Dict:
        """
        批量处理多语言目录
        
        Args:
            directories: 语言目录映射，格式为 {'zh': '...', 'vn': '...', 'th': '...', 'en': '...'} 等
            output_folder: 输出文件夹（如果为None，使用第一个有效目录）
            output_format: 输出格式 ('json', 'csv' 或 'excel')
            recursive: 是否递归扫描
            
        Returns:
            Dict: 包含各语言处理统计信息的字典
        """
        all_stats = {
            'languages': {},
            'total_files': 0,
            'total_sheets': 0,
            'total_fields': 0,
            'output_files': []
        }
        
        # 确定输出目录
        if not output_folder:
            for lang, dir_path in directories.items():
                if dir_path and Path(dir_path).exists():
                    output_folder = dir_path
                    break
        
        if not output_folder:
            print("错误: 未指定有效的输出目录")
            return all_stats
            
        output_dir_path = Path(output_folder)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        
        merged_json_data = {}
        
        # 计算有效的语言目录数量，用于进度计算
        valid_languages = []
        for lang_code, dir_path in directories.items():
            if dir_path and Path(dir_path).exists() and lang_code in self.SUPPORTED_LANGUAGES:
                valid_languages.append((lang_code, dir_path))
        
        total_languages = len(valid_languages)
        if total_languages == 0:
            self._report_progress("没有有效的语言目录", 100)
            return all_stats
        
        # 每个语言占用的进度范围
        progress_per_lang = 100.0 / total_languages
        
        # 依次处理每个语言目录
        for lang_idx, (lang_code, dir_path) in enumerate(valid_languages):
            lang_name = self.SUPPORTED_LANGUAGES[lang_code]['name']
            
            # 计算当前语言的进度范围
            lang_progress_base = lang_idx * progress_per_lang
            
            print(f"\n{'='*60}")
            print(f"开始处理 {lang_name} 目录...")
            print(f"{'='*60}")
            
            self._report_progress(f"正在处理 {lang_name} ({lang_idx + 1}/{total_languages})...", lang_progress_base)
            
            # 如果是JSON格式，暂时不写入单个文件，而是收集数据合并
            should_write = (output_format != 'json')
            
            # 处理该语言的目录
            stats = self.process_directory(
                directory_path=dir_path,
                output_folder=output_folder,
                output_format=output_format,
                recursive=recursive,
                language=lang_code,
                write_output=should_write,
                lang_progress_base=lang_progress_base,
                lang_progress_range=progress_per_lang
            )
            
            # 如果是JSON格式，收集数据
            if output_format == 'json':
                formatted_data = self._format_json_data(stats['results'], lang_code)
                # 使用大写语言代码作为Key
                merged_json_data[lang_code.upper()] = formatted_data
            
            # 汇总统计
            all_stats['languages'][lang_code] = {
                'name': lang_name,
                'directory': dir_path,
                'stats': stats
            }
            all_stats['total_files'] += stats['total_files']
            all_stats['total_sheets'] += stats['total_sheets']
            all_stats['total_fields'] += stats['total_fields']
            if stats.get('output_file'):
                all_stats['output_files'].append(stats['output_file'])
        
        # 如果是JSON格式，写入合并后的文件
        if output_format == 'json' and merged_json_data:
            merged_output_file = output_dir_path / FIELD_EXTRACTION_MERGED_JSON_NAME
            try:
                with open(merged_output_file, 'w', encoding='utf-8') as f:
                    json.dump(merged_json_data, f, ensure_ascii=False, indent=2)
                print(f"\n合并结果已导出到: {merged_output_file}")
                all_stats['output_files'].append(str(merged_output_file))
            except Exception as e:
                print(f"导出合并JSON时出错: {e}")
        
        # 输出总结
        print(f"\n{'='*60}")
        print("多语言处理完成汇总")
        print(f"{'='*60}")
        print(f"处理语言数: {len(all_stats['languages'])}")
        print(f"总文件数: {all_stats['total_files']}")
        print(f"总工作表数: {all_stats['total_sheets']}")
        print(f"总字段数: {all_stats['total_fields']}")
        print(f"输出文件: {all_stats['output_files']}")
        
        return all_stats
    
    def get_error_logs(self) -> List[str]:
        """
        获取所有错误日志
        
        Returns:
            List[str]: 错误日志列表
        """
        return self.error_logs.copy()
    
    def get_warning_logs(self) -> List[str]:
        """
        获取所有警告日志（提取失败警告）
        
        Returns:
            List[str]: 警告日志列表
        """
        return self.extraction_warnings.copy()
    
    def get_all_logs(self) -> Dict[str, List[str]]:
        """
        获取所有日志（错误+警告）
        
        Returns:
            Dict: 包含errors和warnings的字典
        """
        return {
            'errors': self.error_logs.copy(),
            'warnings': self.extraction_warnings.copy()
        }
    
    def clear_logs(self):
        """清除所有日志"""
        self.error_logs.clear()
        self.extraction_warnings.clear()
        self._printed_warning_count = 0  # 重置警告打印计数
    
    def save_logs_to_file(self, output_file: Path):
        """
        保存日志到文件
        
        Args:
            output_file: 输出文件路径
        """
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("="*70 + "\n")
                f.write("Excel字段提取 - 错误与警告日志\n")
                f.write("="*70 + "\n\n")
                
                if self.error_logs:
                    f.write(f"【错误日志】 共 {len(self.error_logs)} 条\n")
                    f.write("-"*70 + "\n")
                    for i, log in enumerate(self.error_logs, 1):
                        f.write(f"{i}. {log}\n")
                    f.write("\n")
                else:
                    f.write("【错误日志】 无错误\n\n")
                
                if self.extraction_warnings:
                    f.write(f"【警告日志】 共 {len(self.extraction_warnings)} 条\n")
                    f.write("-"*70 + "\n")
                    for i, log in enumerate(self.extraction_warnings, 1):
                        f.write(f"{i}. {log}\n")
                    f.write("\n")
                else:
                    f.write("【警告日志】 无警告\n\n")
                
                f.write("="*70 + "\n")
                f.write(f"总计: {len(self.error_logs)} 个错误, {len(self.extraction_warnings)} 个警告\n")
            
            print(f"日志已保存到: {output_file}")
            return True
        
        except Exception as e:
            print(f"保存日志文件失败: {e}")
            return False


if __name__ == "__main__":
    # 测试代码
    extractor = ExcelFieldExtractor()
    
    # 示例用法（与 test/create_test_data.py --field 生成目录一致）
    _repo_root = Path(__file__).resolve().parent.parent
    test_dir = _repo_root / "test" / "_runtime" / "generated" / "test_excel_files"
    if test_dir.exists():
        stats = extractor.process_directory(
            directory_path=str(test_dir),
            output_format='excel',
            recursive=True
        )
        print(f"\n统计信息: {stats}")
    else:
        print(f"测试目录不存在: {test_dir}")
