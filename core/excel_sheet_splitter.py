#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel分页拆分器
根据第一列的文件名，将同一文件名的行数据迁移到新表格的对应分页中
"""

import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


class ExcelSheetSplitter:
    """Excel分页拆分器 - 根据第一列文件名创建分页"""
    
    def __init__(self):
        """初始化拆分器"""
        self.supported_extensions = {'.xlsx', '.xls', '.csv'}
        self.split_data = {}
        self.original_columns = []
        self.source_file = None
        self.total_rows = 0
        self.processed_rows = 0
    
    def read_source_file(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        读取源Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，None表示读取第一个工作表
            
        Returns:
            pandas DataFrame对象
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            file_ext = Path(file_path).suffix.lower()
            if file_ext not in self.supported_extensions:
                raise ValueError(f"不支持的文件格式: {file_ext}，支持的格式: {', '.join(self.supported_extensions)}")
            
            logger.info(f"正在读取文件: {file_path}")
            self.source_file = file_path
            
            # 根据文件格式读取
            if file_ext == '.csv':
                # 尝试不同编码读取CSV
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise ValueError(f"无法读取CSV文件，编码不支持: {file_path}")
            else:
                # 读取Excel文件
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(file_path)
            
            self.total_rows = len(df)
            self.original_columns = list(df.columns)
            
            logger.info(f"成功读取文件，共 {self.total_rows} 行，{len(df.columns)} 列")
            logger.info(f"列名: {self.original_columns}")
            
            return df
            
        except Exception as e:
            logger.error(f"读取文件失败: {str(e)}")
            raise
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        获取Excel文件中的所有工作表名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            工作表名称列表
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            if file_ext == '.csv':
                return ['Sheet1']  # CSV只有一个工作表
            
            excel_file = pd.ExcelFile(file_path)
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"获取工作表名称失败: {str(e)}")
            return []
    
    def split_by_first_column(self, df: pd.DataFrame, 
                               group_column: Optional[str] = None,
                               extract_filename: bool = True) -> Dict[str, pd.DataFrame]:
        """
        根据第一列（或指定列）的文件名进行数据拆分
        
        Args:
            df: 输入的DataFrame
            group_column: 分组列名，默认为第一列
            extract_filename: 是否从路径中提取文件名（不含扩展名）
            
        Returns:
            字典，键为分组值（文件名），值为对应的DataFrame
        """
        try:
            if df.empty:
                logger.warning("输入数据为空")
                return {}
            
            # 确定分组列
            if group_column is None:
                group_column = df.columns[0]  # 使用第一列
            
            if group_column not in df.columns:
                raise ValueError(f"分组列 '{group_column}' 不存在于数据中")
            
            logger.info(f"使用列 '{group_column}' 进行分组拆分")
            
            # 处理分组列的值
            df_copy = df.copy()
            
            if extract_filename:
                # 从路径中提取文件名
                df_copy['_group_key'] = df_copy[group_column].apply(self._extract_filename)
            else:
                df_copy['_group_key'] = df_copy[group_column].astype(str)
            
            # 按分组键进行分组
            grouped = df_copy.groupby('_group_key')
            
            # 创建拆分后的数据字典
            self.split_data = {}
            
            for group_key, group_df in grouped:
                # 移除临时分组键列
                result_df = group_df.drop(columns=['_group_key']).reset_index(drop=True)
                
                # 确保分组键有效
                if pd.isna(group_key) or str(group_key).strip() == '':
                    group_key = "未分类"
                else:
                    group_key = str(group_key).strip()
                
                self.split_data[group_key] = result_df
                logger.info(f"分组 '{group_key}': {len(result_df)} 行数据")
            
            logger.info(f"数据拆分完成，共 {len(self.split_data)} 个分组")
            self.processed_rows = sum(len(df) for df in self.split_data.values())
            
            return self.split_data
            
        except Exception as e:
            logger.error(f"数据拆分失败: {str(e)}")
            raise
    
    def _extract_filename(self, value) -> str:
        """
        从路径或值中提取文件名（不含扩展名）
        
        Args:
            value: 原始值（可能是路径或文件名）
            
        Returns:
            提取的文件名
        """
        if pd.isna(value):
            return "未分类"
        
        value_str = str(value).strip()
        if not value_str:
            return "未分类"
        
        # 尝试从路径中提取文件名
        try:
            # 处理Windows和Unix路径
            path = Path(value_str)
            filename = path.stem  # 获取不含扩展名的文件名
            
            # 如果提取结果为空，使用原值
            if not filename:
                filename = value_str
            
            return filename
        except Exception:
            return value_str
    
    def _clean_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        清理工作表名称，使其符合Excel规范
        
        Args:
            name: 原始名称
            max_length: 最大长度（Excel限制为31）
            
        Returns:
            清理后的名称
        """
        # Excel工作表名称不能包含的字符
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']', "'"]
        
        clean_name = name
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        # 移除首尾空格和引号
        clean_name = clean_name.strip().strip("'\"")
        
        # 如果为空，使用默认名称
        if not clean_name:
            clean_name = "数据"
        
        # 截断到最大长度
        if len(clean_name) > max_length:
            clean_name = clean_name[:max_length - 3] + "..."
        
        return clean_name
    
    def create_split_excel(self, output_path: str, 
                           include_summary: bool = True,
                           apply_formatting: bool = True,
                           remove_first_column: bool = False) -> bool:
        """
        创建拆分后的Excel文件，每个分组作为一个工作表
        
        Args:
            output_path: 输出文件路径
            include_summary: 是否包含汇总工作表
            apply_formatting: 是否应用格式化样式
            remove_first_column: 是否在输出中移除第一列（分组列）
            
        Returns:
            是否成功创建文件
        """
        try:
            if not self.split_data:
                logger.error("没有可拆分的数据，请先调用 split_by_first_column 方法")
                return False
            
            logger.info(f"正在创建拆分后的Excel文件: {output_path}")
            
            # 确保输出目录存在
            output_dir = Path(output_path).parent
            if not output_dir.exists():
                output_dir.mkdir(parents=True, exist_ok=True)
            
            # 创建工作簿
            wb = Workbook()
            
            # 移除默认工作表
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # 用于处理重复工作表名称
            sheet_name_counter = {}
            
            # 按分组名称排序
            sorted_groups = sorted(self.split_data.keys())
            
            for group_key in sorted_groups:
                group_df = self.split_data[group_key]
                
                # 是否移除第一列
                if remove_first_column and len(group_df.columns) > 1:
                    output_df = group_df.iloc[:, 1:].copy()
                else:
                    output_df = group_df.copy()
                
                # 清理工作表名称
                clean_name = self._clean_sheet_name(group_key)
                
                # 处理重复名称
                if clean_name in sheet_name_counter:
                    sheet_name_counter[clean_name] += 1
                    clean_name = f"{clean_name[:27]}_{sheet_name_counter[clean_name]:03d}"
                else:
                    sheet_name_counter[clean_name] = 0
                
                # 创建工作表
                ws = wb.create_sheet(title=clean_name)
                
                # 写入数据
                for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 应用格式化
                if apply_formatting:
                    self._apply_sheet_formatting(ws, len(output_df.columns))
            
            # 添加汇总工作表
            if include_summary:
                self._create_summary_sheet(wb)
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"Excel文件创建成功: {output_path}")
            logger.info(f"共创建 {len(self.split_data)} 个数据工作表")
            
            return True
            
        except Exception as e:
            logger.error(f"创建Excel文件失败: {str(e)}")
            return False
    
    def _apply_sheet_formatting(self, ws, col_count: int):
        """
        应用工作表格式化
        
        Args:
            ws: 工作表对象
            col_count: 列数
        """
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头格式
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽（自动调整）
        for col in range(1, col_count + 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col).column_letter
            
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
            
            # 设置列宽（最大50，最小10）
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, wb: Workbook):
        """
        创建汇总工作表
        
        Args:
            wb: 工作簿对象
        """
        ws = wb.create_sheet(title="汇总信息", index=0)
        
        # 写入汇总数据
        summary_data = [
            ["项目", "值"],
            ["源文件", self.source_file or "未知"],
            ["总行数", self.total_rows],
            ["处理行数", self.processed_rows],
            ["分组数量", len(self.split_data)],
            ["原始列数", len(self.original_columns)],
            ["", ""],
            ["分组详情", ""],
        ]
        
        # 添加各分组统计
        for group_key in sorted(self.split_data.keys()):
            row_count = len(self.split_data[group_key])
            summary_data.append([f"  {group_key}", f"{row_count} 行"])
        
        # 写入数据
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 应用格式化
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # 设置列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
    
    def get_split_report(self) -> str:
        """
        获取拆分报告
        
        Returns:
            拆分报告字符串
        """
        if not self.split_data:
            return "没有拆分数据"
        
        report_lines = []
        report_lines.append("=" * 60)
        report_lines.append("Excel分页拆分报告")
        report_lines.append("=" * 60)
        report_lines.append(f"源文件: {self.source_file or '未知'}")
        report_lines.append(f"总行数: {self.total_rows}")
        report_lines.append(f"处理行数: {self.processed_rows}")
        report_lines.append(f"分组数量: {len(self.split_data)}")
        report_lines.append(f"原始列数: {len(self.original_columns)}")
        report_lines.append("")
        report_lines.append("分组详情:")
        report_lines.append("-" * 40)
        
        for group_key in sorted(self.split_data.keys()):
            row_count = len(self.split_data[group_key])
            report_lines.append(f"  {group_key}: {row_count} 行")
        
        report_lines.append("=" * 60)
        
        return "\n".join(report_lines)
    
    def process_file(self, input_path: str, output_path: str,
                     sheet_name: Optional[str] = None,
                     group_column: Optional[str] = None,
                     extract_filename: bool = True,
                     include_summary: bool = True,
                     apply_formatting: bool = True,
                     remove_first_column: bool = False) -> Tuple[bool, str]:
        """
        一键处理文件：读取、拆分、导出
        
        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径
            sheet_name: 要读取的工作表名称
            group_column: 分组列名
            extract_filename: 是否从路径中提取文件名
            include_summary: 是否包含汇总工作表
            apply_formatting: 是否应用格式化
            remove_first_column: 是否移除第一列
            
        Returns:
            (成功标志, 报告/错误信息)
        """
        try:
            # 1. 读取源文件
            df = self.read_source_file(input_path, sheet_name)
            
            # 2. 按第一列拆分
            self.split_by_first_column(df, group_column, extract_filename)
            
            # 3. 创建输出文件
            success = self.create_split_excel(
                output_path,
                include_summary=include_summary,
                apply_formatting=apply_formatting,
                remove_first_column=remove_first_column
            )
            
            if success:
                return True, self.get_split_report()
            else:
                return False, "创建输出文件失败"
                
        except Exception as e:
            error_msg = f"处理失败: {str(e)}"
            logger.error(error_msg)
            return False, error_msg


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Excel分页拆分工具 - 根据第一列文件名创建分页',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python excel_sheet_splitter.py input.xlsx output.xlsx
  python excel_sheet_splitter.py input.xlsx output.xlsx --no-summary
  python excel_sheet_splitter.py input.xlsx output.xlsx --remove-first-column
  python excel_sheet_splitter.py input.xlsx output.xlsx --sheet "Sheet2" --column "文件名"
        """
    )
    
    parser.add_argument('input', help='输入Excel文件路径')
    parser.add_argument('output', help='输出Excel文件路径')
    parser.add_argument('--sheet', '-s', help='指定要读取的工作表名称')
    parser.add_argument('--column', '-c', help='指定分组列名（默认为第一列）')
    parser.add_argument('--no-extract', action='store_true', help='不从路径中提取文件名')
    parser.add_argument('--no-summary', action='store_true', help='不生成汇总工作表')
    parser.add_argument('--no-format', action='store_true', help='不应用格式化样式')
    parser.add_argument('--remove-first-column', '-r', action='store_true', 
                        help='在输出中移除第一列（分组列）')
    
    args = parser.parse_args()
    
    # 创建处理器并执行
    splitter = ExcelSheetSplitter()
    success, report = splitter.process_file(
        input_path=args.input,
        output_path=args.output,
        sheet_name=args.sheet,
        group_column=args.column,
        extract_filename=not args.no_extract,
        include_summary=not args.no_summary,
        apply_formatting=not args.no_format,
        remove_first_column=args.remove_first_column
    )
    
    print(report)
    
    if success:
        print(f"\n✓ 文件已成功保存到: {args.output}")
        return 0
    else:
        print(f"\n✗ 处理失败")
        return 1


if __name__ == '__main__':
    sys.exit(main())
