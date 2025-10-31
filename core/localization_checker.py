#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
越南文检测工具
用于检测目录下包含越南文的表格文件
"""

import os
import re
import sys
from pathlib import Path
from typing import List, Set
import pandas as pd
import openpyxl
from openpyxl import load_workbook


class VietnameseDetector:
    """越南文检测器"""
    
    def __init__(self):
        # 越南文字符范围
        self.vietnamese_patterns = [
            # 基本拉丁字母 + 越南语声调符号
            r'[àáạảãâầấậẩẫăằắặẳẵ]',  # a with tones
            r'[èéẹẻẽêềếệểễ]',        # e with tones
            r'[ìíịỉĩ]',              # i with tones
            r'[òóọỏõôồốộổỗơờớợởỡ]',  # o with tones
            r'[ùúụủũưừứựửữ]',        # u with tones
            r'[ỳýỵỷỹ]',              # y with tones
            r'[đ]',                  # d with stroke
            # 大写字母
            r'[ÀÁẠẢÃÂẦẤẬẨẪĂẰẮẶẲẴ]',
            r'[ÈÉẸẺẼÊỀẾỆỂỄ]',
            r'[ÌÍỊỈĨ]',
            r'[ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ]',
            r'[ÙÚỤỦŨƯỪỨỰỬỮ]',
            r'[ỲÝỴỶỸ]',
            r'[Đ]'
        ]
        
        # 编译正则表达式
        self.compiled_patterns = [re.compile(pattern, re.IGNORECASE) for pattern in self.vietnamese_patterns]
        # 合并后的高性能正则（用于向量化/批量匹配）
        self.combined_pattern = re.compile("|".join(self.vietnamese_patterns), re.IGNORECASE)
        
        # 中文字符范围
        self.chinese_patterns = [
            r'[\u4e00-\u9fff]',  # 基本中文字符
            r'[\u3400-\u4dbf]',  # 扩展A区
            r'[\uf900-\ufaff]',   # 兼容汉字
        ]
        
        # 编译中文字符正则表达式
        self.chinese_compiled_patterns = [re.compile(pattern) for pattern in self.chinese_patterns]
    
    def contains_vietnamese(self, text: str) -> bool:
        """
        检测文本中是否包含越南文字符
        
        Args:
            text: 要检测的文本
            
        Returns:
            bool: 如果包含越南文返回True，否则返回False
        """
        if not isinstance(text, str):
            return False
            
        # 先用合并正则快速判断
        return bool(self.combined_pattern.search(text))
    
    def contains_chinese(self, text: str) -> bool:
        """
        检测文本中是否包含中文字符
        
        Args:
            text: 要检测的文本
            
        Returns:
            bool: 如果包含中文返回True，否则返回False
        """
        if not isinstance(text, str):
            return False
            
        # 检查所有中文模式
        for pattern in self.chinese_compiled_patterns:
            if pattern.search(text):
                return True
        return False
    
    def contains_english(self, text: str) -> bool:
        """
        检测文本中是否包含英文字符
        
        Args:
            text: 要检测的文本
            
        Returns:
            bool: 如果包含英文返回True，否则返回False
        """
        if not isinstance(text, str):
            return False
        
        # 英文字符范围：基本拉丁字母
        english_pattern = re.compile(r'[a-zA-Z]')
        return bool(english_pattern.search(text))
    
    def detect_language_type(self, text: str) -> str:
        """
        检测文本的语言类型
        
        Args:
            text: 要检测的文本
            
        Returns:
            str: 语言类型 - "英文", "越南文", "中文", "中越混合", "越英混合", "其他"
        """
        if not isinstance(text, str) or not text.strip():
            return "其他"
        
        has_vietnamese = self.contains_vietnamese(text)
        has_chinese = self.contains_chinese(text)
        has_english = self.contains_english(text)
        
        # 统计各种字符的数量
        vietnamese_count = sum(1 for pattern in self.compiled_patterns for _ in pattern.finditer(text))
        chinese_count = sum(1 for pattern in self.chinese_compiled_patterns for _ in pattern.finditer(text))
        english_count = len(re.findall(r'[a-zA-Z]', text))
        
        # 计算总字符数（排除空格和标点）
        total_chars = len(re.findall(r'[^\s\W]', text))
        
        # 如果总字符数很少，使用简单判断
        if total_chars <= 3:
            if has_vietnamese and has_chinese:
                return "中越混合"
            elif has_vietnamese and has_english:
                return "越英混合"
            elif has_vietnamese:
                return "越南文"
            elif has_chinese:
                return "中文"
            elif has_english:
                return "英文"
            else:
                return "其他"
        
        # 计算各种字符的比例
        vietnamese_ratio = vietnamese_count / total_chars if total_chars > 0 else 0
        chinese_ratio = chinese_count / total_chars if total_chars > 0 else 0
        english_ratio = english_count / total_chars if total_chars > 0 else 0
        
        # 优化判断逻辑：
        # 1. 如果包含越南文声调符号，优先考虑越南文
        # 2. 如果越南文字符比例 > 20%，且没有中文，判断为越南文
        # 3. 如果同时有越南文和中文，判断为混合
        # 4. 如果只有英文，判断为英文
        
        if has_vietnamese and has_chinese:
            return "中越混合"
        elif has_vietnamese:
            # 如果越南文字符比例 > 20%，判断为越南文
            if vietnamese_ratio > 0.2:
                return "越南文"
            else:
                # 如果越南文字符很少，但有英文，可能是越英混合
                if has_english and english_ratio > 0.5:
                    return "越英混合"
                else:
                    return "越南文"
        elif has_chinese:
            return "中文"
        elif has_english:
            return "英文"
        else:
            return "其他"


class TableChecker:
    """表格检测器"""
    
    def __init__(self):
        self.vietnamese_detector = VietnameseDetector()
        self.supported_extensions = {'.xlsx', '.xls', '.csv', '.tsv'}
    
    def is_table_file(self, file_path: Path) -> bool:
        """
        检查文件是否为支持的表格格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是支持的表格格式返回True
        """
        return file_path.suffix.lower() in self.supported_extensions
    
    def read_excel_file(self, file_path: Path) -> List[str]:
        """
        读取Excel文件内容（保留旧接口，仍返回所有文本）。
        注意：为节省内存，扫描场景推荐使用流式检测而非一次性返回。
        """
        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            all_text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            all_text.append(str(cell.value))
            return all_text
        except Exception as e:
            print(f"读取Excel文件 {file_path} 时出错: {e}")
            return []
    
    def read_csv_file(self, file_path: Path) -> List[str]:
        """
        读取CSV/TSV文件内容（保留旧接口）。
        注意：为节省内存，扫描场景推荐使用分块流式检测而非一次性返回。
        """
        try:
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            sep = '\t' if file_path.suffix.lower() == '.tsv' else None
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding, sep=sep, dtype=str, engine='python')
                    all_text = []
                    for column in df.columns:
                        for value in df[column].dropna():
                            all_text.append(str(value))
                    return all_text
                except UnicodeDecodeError:
                    continue
            print(f"无法读取CSV文件 {file_path}，尝试了多种编码")
            return []
        except Exception as e:
            print(f"读取CSV文件 {file_path} 时出错: {e}")
            return []

    def _excel_contains_vietnamese_stream(self, file_path: Path) -> bool:
        """
        以只读流式方式扫描Excel，检测到即提前返回。
        """
        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True, keep_links=False)
            detector = self.vietnamese_detector
            combined = detector.combined_pattern
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # 直接返回原始值，避免创建单元格对象，速度更快
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if value is None:
                            continue
                        # 仅处理字符串，避免无意义的类型转换
                        if isinstance(value, str):
                            # 先做快速非ASCII过滤，减少正则调用
                            has_non_ascii = False
                            for ch in value:
                                if ord(ch) > 127:
                                    has_non_ascii = True
                                    break
                            if not has_non_ascii:
                                continue
                            if combined.search(value):
                                return True
            return False
        except Exception as e:
            print(f"读取Excel文件 {file_path} 时出错: {e}")
            return False

        
    def _csv_contains_vietnamese_stream(self, file_path: Path) -> bool:
        """
        分块读取CSV/TSV并向量化检测，检测到即提前返回。
        """
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
        sep = '\t' if file_path.suffix.lower() == '.tsv' else None
        pattern = self.vietnamese_detector.combined_pattern
        for encoding in encodings:
            try:
                for chunk in pd.read_csv(
                    file_path,
                    encoding=encoding,
                    sep=sep,
                    dtype=str,
                    engine='python',
                    chunksize=10000,
                    low_memory=True,
                    on_bad_lines='skip'
                ):
                    for column in chunk.columns:
                        series = chunk[column].astype(str)
                        # 使用向量化正则匹配加速
                        if series.str.contains(pattern, regex=True, na=False).any():
                            return True
                # 该编码读取完成仍未发现
            except UnicodeDecodeError:
                continue
            except Exception as e:
                # 其他读取异常，继续尝试下个编码
                print(f"读取CSV文件 {file_path} 使用编码 {encoding} 时出错: {e}")
                continue
        return False
    
    def check_table_has_vietnamese(self, file_path: Path) -> bool:
        """
        检查表格文件是否包含越南文
        
        Args:
            file_path: 表格文件路径
            
        Returns:
            bool: 如果包含越南文返回True
        """
        if not self.is_table_file(file_path):
            return False
        
        suffix = file_path.suffix.lower()
        # 流式快速路径：避免一次性加载全部内容
        if suffix in ['.xlsx', '.xls']:
            return self._excel_contains_vietnamese_stream(file_path)
        if suffix in ['.csv', '.tsv']:
            return self._csv_contains_vietnamese_stream(file_path)
        return False


class LocalizationChecker:
    """本地化检测主类"""
    
    def __init__(self):
        self.table_checker = TableChecker()
    
    def scan_directory(self, directory_path: str, recursive: bool = False) -> List[str]:
        """
        扫描目录下的所有表格文件，检测包含越南文的文件
        
        Args:
            directory_path: 要扫描的目录路径
            recursive: 是否递归扫描子目录，默认为False
            
        Returns:
            List[str]: 包含越南文的表格文件名列表
        """
        directory = Path(directory_path)
        
        if not directory.exists():
            print(f"错误: 目录 {directory_path} 不存在")
            return []
        
        if not directory.is_dir():
            print(f"错误: {directory_path} 不是一个目录")
            return []
        
        valid_tables = []
        
        print(f"Scanning directory: {directory_path}")
        print(f"Recursive scan: {'Yes' if recursive else 'No'}")
        print("Supported formats: .xlsx, .xls, .csv, .tsv")
        print("-" * 50)
        
        # 根据recursive参数决定扫描方式
        if recursive:
            # 递归扫描所有文件
            for file_path in directory.rglob('*'):
                if file_path.is_file() and self.table_checker.is_table_file(file_path):
                    print(f"Checking file: {file_path.name}...", end=" ")
                    
                    if self.table_checker.check_table_has_vietnamese(file_path):
                        valid_tables.append(file_path.name)
                        print("YES - Contains Vietnamese")
                    else:
                        print("NO - No Vietnamese")
        else:
            # 只扫描当前目录
            for file_path in directory.iterdir():
                if file_path.is_file() and self.table_checker.is_table_file(file_path):
                    print(f"Checking file: {file_path.name}...", end=" ")
                    
                    if self.table_checker.check_table_has_vietnamese(file_path):
                        valid_tables.append(file_path.name)
                        print("YES - Contains Vietnamese")
                    else:
                        print("NO - No Vietnamese")
        
        return valid_tables
    
    def print_results(self, valid_tables: List[str]):
        """
        打印检测结果
        
        Args:
            valid_tables: 有效表格文件名列表
        """
        print("\n" + "=" * 50)
        print("Scan Results:")
        print("=" * 50)
        
        if valid_tables:
            print(f"Found {len(valid_tables)} table files containing Vietnamese:")
            for i, table_name in enumerate(valid_tables, 1):
                print(f"{i:2d}. {table_name}")
        else:
            print("No table files containing Vietnamese found")
        
        print("=" * 50)


def main():
    """主函数"""
    print("Localization Checker - Vietnamese Table Detector")
    print("=" * 50)
    
    # 获取用户输入的目录路径
    if len(sys.argv) > 1:
        directory_path = sys.argv[1]
    else:
        directory_path = input("Enter directory path to scan: ").strip()
    
    if not directory_path:
        print("Error: No directory path provided")
        return
    
    # 创建检测器并开始扫描
    checker = LocalizationChecker()
    valid_tables = checker.scan_directory(directory_path)
    
    # 打印结果
    checker.print_results(valid_tables)


if __name__ == "__main__":
    main()
