#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gametools - 统一用户界面
集成越南文检测和JSON格式检测工具
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import sys
from pathlib import Path
import subprocess
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 添加模块路径
sys.path.append(str(Path(__file__).parent.parent))

from core.localization_checker import LocalizationChecker
from core.excel_vietnamese_scanner import ExcelVietnameseScanner
from core.vietnamese_excel_processor import VietnameseExcelProcessor
from core.cross_project_translator import CrossProjectTranslator
from core.excel_field_extractor import ExcelFieldExtractor
from core.table_range_translator import TableRangeTranslator
from tools.json_error_detector.json_error_detector import JSONErrorDetector
from tools.excel_data_processor import ExcelDataProcessor
from tools.excel_text_extractor import ExcelTextExtractor
from version import get_version, format_version_string, get_description, get_latest_changes


class GameToolsUnified:
    """gametools统一界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title(f"gametools - 游戏工具集 v{get_version()}")
        self.root.geometry("1200x900")
        self.root.minsize(1000, 800)
        
        # 设置窗口图标
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass
        
        # 设置样式
        self.setup_styles()
        
        # 创建界面
        self.create_widgets()
        
        # 初始化检测器
        self.localization_checker = LocalizationChecker()
        self.excel_scanner = ExcelVietnameseScanner()
        self.vietnamese_processor = VietnameseExcelProcessor()
        self.cross_project_translator = CrossProjectTranslator()
        self.json_detector = JSONErrorDetector()
        self.excel_processor = ExcelDataProcessor()
        self.text_extractor = ExcelTextExtractor()
        self.field_extractor = ExcelFieldExtractor()
        self.table_range_translator = TableRangeTranslator()
        
        # 扫描状态
        self.is_scanning = False
        
        # 结果存储字典
        self.results_storage = {
            'vietnamese_processor': '',
            'cross_project_translator': '',
            'json_detector': '',
            'excel_processor': '',
            'text_extractor': '',
            'field_extractor': '',
            'table_range_translator': ''
        }
        
        # 字段提取结果数据
        self.field_extraction_results = None
    
    def setup_styles(self):
        """设置界面样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置样式
        style.configure('Title.TLabel', font=('Microsoft YaHei', 18, 'bold'))
        style.configure('Heading.TLabel', font=('Microsoft YaHei', 12, 'bold'))
        style.configure('Info.TLabel', font=('Microsoft YaHei', 10))
        style.configure('Success.TLabel', font=('Microsoft YaHei', 10), foreground='green')
        style.configure('Error.TLabel', font=('Microsoft YaHei', 10), foreground='red')
        style.configure('Accent.TButton', font=('Microsoft YaHei', 10, 'bold'))
    
    def create_widgets(self):
        """创建界面控件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # 标题区域（隐藏）
        # title_frame = ttk.Frame(main_frame)
        # title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        # title_frame.columnconfigure(0, weight=1)
        
        # # 主标题
        # title_label = ttk.Label(title_frame, text="gametools - 游戏工具集", 
        #                        style='Title.TLabel')
        # title_label.grid(row=0, column=0, pady=(0, 5))
        
        # # 副标题
        # subtitle_label = ttk.Label(title_frame, text="集成策划本地化、JSON检测、Excel处理、翻译提取等功能", 
        #                           style='Info.TLabel')
        # subtitle_label.grid(row=1, column=0)
        
        # 创建笔记本控件（页签）
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 5))
        
        # 创建各个功能页签
        self.create_vietnamese_processor_tab()
        self.create_cross_project_translator_tab()
        self.create_json_detector_tab()
        self.create_excel_data_processor_tab()
        self.create_excel_text_extractor_tab()
        self.create_field_extractor_tab()
        self.create_table_range_translator_tab()
        self.create_about_tab()
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, 
                              relief=tk.SUNKEN, anchor=tk.W, padding="3")
        status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(2, 0))
    
    def create_vietnamese_processor_tab(self):
        """创建越南文检测和导出页签"""
        # 越南文处理器框架
        processor_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(processor_frame, text="越南文检测导出")
        
        # 配置网格
        processor_frame.columnconfigure(0, weight=1)
        processor_frame.rowconfigure(2, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(processor_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="越南文检测和导出工具", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, text="检测Excel和CSV文件中的越南文内容，并导出详细结果到Excel文件", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(processor_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 目录选择区域
        dir_frame = ttk.LabelFrame(control_frame, text="扫描设置", padding="12")
        dir_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        dir_frame.columnconfigure(1, weight=1)
        
        # 扫描目录
        ttk.Label(dir_frame, text="扫描目录:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.vp_scan_dir_var = tk.StringVar()
        self.vp_scan_dir_entry = ttk.Entry(dir_frame, textvariable=self.vp_scan_dir_var, 
                                          font=("Microsoft YaHei", 9))
        self.vp_scan_dir_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.vp_scan_browse_button = ttk.Button(dir_frame, text="浏览目录", 
                                               command=self.browse_vp_scan_directory)
        self.vp_scan_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        # 输出文件夹
        ttk.Label(dir_frame, text="输出文件夹:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.vp_output_folder_var = tk.StringVar()
        self.vp_output_folder_entry = ttk.Entry(dir_frame, textvariable=self.vp_output_folder_var, 
                                               font=("Microsoft YaHei", 9))
        self.vp_output_folder_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(5, 0))
        
        self.vp_output_browse_button = ttk.Button(dir_frame, text="选择输出文件夹", 
                                                 command=self.browse_vp_output_folder)
        self.vp_output_browse_button.grid(row=1, column=2, pady=(5, 0))
        
        # 选项设置区域
        options_frame = ttk.LabelFrame(control_frame, text="处理选项", padding="12")
        options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 递归扫描选项
        self.vp_recursive_var = tk.BooleanVar(value=False)
        self.vp_recursive_check = ttk.Checkbutton(options_frame, text="递归扫描子目录", 
                                                 variable=self.vp_recursive_var)
        self.vp_recursive_check.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        # 输出文件选项
        output_options_frame = ttk.Frame(options_frame)
        output_options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        self.vp_create_excel_var = tk.BooleanVar(value=True)
        self.vp_create_excel_check = ttk.Checkbutton(output_options_frame, text="创建Excel结果文件", 
                                                    variable=self.vp_create_excel_var)
        self.vp_create_excel_check.pack(side=tk.LEFT)
        
        # 操作按钮区域
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 主要操作按钮
        self.vp_process_button = ttk.Button(button_frame, text="🔍 开始检测导出", 
                                           command=self.start_vietnamese_processing, 
                                           style='Accent.TButton')
        self.vp_process_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.vp_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                         command=self.clear_vp_results)
        self.vp_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.vp_demo_button = ttk.Button(button_frame, text="📁 创建演示文件", 
                                        command=self.create_demo_files)
        self.vp_demo_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.vp_view_results_button = ttk.Button(button_frame, text="👁️ 查看结果", 
                                                command=lambda: self.show_results_dialog('vietnamese_processor'))
        self.vp_view_results_button.pack(side=tk.LEFT)
        
        # 进度条
        progress_frame = ttk.Frame(control_frame)
        progress_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.vp_progress_var = tk.StringVar(value="就绪")
        self.vp_progress_label = ttk.Label(progress_frame, textvariable=self.vp_progress_var)
        self.vp_progress_label.grid(row=0, column=0, sticky=tk.W)
        
        self.vp_progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.vp_progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
    
    def create_cross_project_translator_tab(self):
        """创建跨项目翻译对应页签"""
        # 跨项目翻译对应框架
        translator_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(translator_frame, text="跨项目翻译对应")
        
        # 配置网格
        translator_frame.columnconfigure(0, weight=1)
        translator_frame.rowconfigure(2, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(translator_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="跨项目翻译对应工具", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, text="根据Excel表格中的文件名列和位置列查找对应的内容", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(translator_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(control_frame, text="文件选择", padding="12")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # 映射文件选择
        ttk.Label(file_frame, text="映射文件:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.cpt_mapping_file_var = tk.StringVar()
        self.cpt_mapping_file_entry = ttk.Entry(file_frame, textvariable=self.cpt_mapping_file_var, 
                                               font=("Microsoft YaHei", 9))
        self.cpt_mapping_file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.cpt_mapping_browse_button = ttk.Button(file_frame, text="浏览文件", 
                                                   command=self.browse_cpt_mapping_file)
        self.cpt_mapping_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        # 项目目录选择
        ttk.Label(file_frame, text="项目目录:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.cpt_project_dir_var = tk.StringVar()
        self.cpt_project_dir_entry = ttk.Entry(file_frame, textvariable=self.cpt_project_dir_var, 
                                              font=("Microsoft YaHei", 9))
        self.cpt_project_dir_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(5, 0))
        
        self.cpt_project_browse_button = ttk.Button(file_frame, text="浏览目录", 
                                                   command=self.browse_cpt_project_directory)
        self.cpt_project_browse_button.grid(row=1, column=2, pady=(5, 0))
        
        # 输出设置区域
        output_frame = ttk.LabelFrame(control_frame, text="输出设置", padding="12")
        output_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)
        
        # 输出文件选择
        ttk.Label(output_frame, text="输出文件:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.cpt_output_file_var = tk.StringVar()
        self.cpt_output_file_entry = ttk.Entry(output_frame, textvariable=self.cpt_output_file_var, 
                                              font=("Microsoft YaHei", 9))
        self.cpt_output_file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.cpt_output_browse_button = ttk.Button(output_frame, text="选择文件", 
                                                  command=self.browse_cpt_output_file)
        self.cpt_output_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        # 操作按钮区域
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        # 主要操作按钮
        self.cpt_process_button = ttk.Button(button_frame, text="🔍 开始对应", 
                                            command=self.start_cross_project_translation, 
                                            style='Accent.TButton')
        self.cpt_process_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.cpt_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                          command=self.clear_cpt_results)
        self.cpt_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.cpt_export_button = ttk.Button(button_frame, text="💾 导出结果", 
                                           command=self.export_cpt_results,
                                           state="disabled")
        self.cpt_export_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.cpt_view_results_button = ttk.Button(button_frame, text="👁️ 查看结果", 
                                                 command=lambda: self.show_results_dialog('cross_project_translator'))
        self.cpt_view_results_button.pack(side=tk.LEFT)
    
    
    def create_json_detector_tab(self):
        """创建JSON错误检测工具页签"""
        # JSON检测工具框架
        json_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(json_frame, text="JSON错误检测工具")
        
        # 配置网格
        json_frame.columnconfigure(0, weight=1)
        json_frame.rowconfigure(2, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(json_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="JSON错误检测器", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, text="检测JSON文件中的语法错误、结构错误、数据类型错误、编码错误和性能问题", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(json_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 路径选择区域
        path_frame = ttk.LabelFrame(control_frame, text="路径选择", padding="12")
        path_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        path_frame.columnconfigure(1, weight=1)
        
        # 路径输入
        ttk.Label(path_frame, text="路径:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.json_path_var = tk.StringVar()
        self.json_path_entry = ttk.Entry(path_frame, textvariable=self.json_path_var, 
                                       font=("Microsoft YaHei", 9))
        self.json_path_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.json_browse_button = ttk.Button(path_frame, text="浏览文件夹", 
                                            command=self.browse_json_folder)
        self.json_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        
        # 操作按钮区域
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # 主要操作按钮
        self.json_detect_button = ttk.Button(button_frame, text="🔍 开始检测", 
                                            command=self.start_json_detection, 
                                            style='Accent.TButton')
        self.json_detect_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.json_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                           command=self.clear_json_results)
        self.json_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.json_save_button = ttk.Button(button_frame, text="💾 保存报告", 
                                          command=self.save_json_report, 
                                          state="disabled")
        self.json_save_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.json_view_results_button = ttk.Button(button_frame, text="👁️ 查看结果", 
                                                  command=lambda: self.show_results_dialog('json_detector'))
        self.json_view_results_button.pack(side=tk.LEFT)
    
    def create_excel_data_processor_tab(self):
        """创建Excel数据处理工具页签"""
        # Excel数据处理工具框架
        excel_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(excel_frame, text="Excel数据处理工具")
        
        # 配置网格
        excel_frame.columnconfigure(0, weight=1)
        excel_frame.rowconfigure(2, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(excel_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="Excel数据处理工具", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, text="根据指定列对Excel数据进行分组处理，支持多工作表输出", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(excel_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(control_frame, text="文件选择", padding="12")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # 输入文件
        ttk.Label(file_frame, text="输入文件:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.excel_input_var = tk.StringVar()
        self.excel_input_entry = ttk.Entry(file_frame, textvariable=self.excel_input_var, 
                                         font=("Microsoft YaHei", 9))
        self.excel_input_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.excel_input_browse_button = ttk.Button(file_frame, text="浏览文件", 
                                                    command=self.browse_excel_input_file)
        self.excel_input_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        # 输出设置
        output_frame = ttk.LabelFrame(control_frame, text="输出设置", padding="12")
        output_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)
        
        # 输出文件夹
        ttk.Label(output_frame, text="输出文件夹:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 5))
        self.excel_output_folder_var = tk.StringVar()
        self.excel_output_folder_entry = ttk.Entry(output_frame, textvariable=self.excel_output_folder_var, 
                                                 font=("Microsoft YaHei", 9))
        self.excel_output_folder_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 5))
        
        self.excel_output_browse_button = ttk.Button(output_frame, text="浏览文件夹", 
                                                     command=self.browse_excel_output_folder)
        self.excel_output_browse_button.grid(row=0, column=2, pady=(0, 5))
        
        # 输出文件名
        ttk.Label(output_frame, text="输出文件名:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.excel_output_filename_var = tk.StringVar(value="整合结果.xlsx")
        self.excel_output_filename_entry = ttk.Entry(output_frame, textvariable=self.excel_output_filename_var, 
                                                   width=25, font=("Microsoft YaHei", 9))
        self.excel_output_filename_entry.grid(row=1, column=1, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        
        # 处理选项区域
        options_frame = ttk.LabelFrame(control_frame, text="处理选项", padding="12")
        options_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        options_frame.columnconfigure(1, weight=1)
        
        # 分组列设置
        ttk.Label(options_frame, text="分组列:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.excel_group_column_var = tk.StringVar()
        self.excel_group_column_entry = ttk.Entry(options_frame, textvariable=self.excel_group_column_var, 
                                                width=15, font=("Microsoft YaHei", 9))
        self.excel_group_column_entry.grid(row=0, column=1, sticky=tk.W, padx=(0, 10))
        ttk.Label(options_frame, text="(留空使用第一列)", style='Info.TLabel').grid(row=0, column=2, sticky=tk.W)
        
        # 工作表前缀
        ttk.Label(options_frame, text="工作表前缀:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.excel_sheet_prefix_var = tk.StringVar()
        self.excel_sheet_prefix_entry = ttk.Entry(options_frame, textvariable=self.excel_sheet_prefix_var, 
                                                width=15, font=("Microsoft YaHei", 9))
        self.excel_sheet_prefix_entry.grid(row=1, column=1, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        
        # 包含汇总信息选项
        self.excel_include_summary_var = tk.BooleanVar(value=True)
        self.excel_include_summary_check = ttk.Checkbutton(options_frame, text="包含汇总信息工作表", 
                                                          variable=self.excel_include_summary_var)
        self.excel_include_summary_check.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # 操作按钮区域
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=3, column=0, sticky=(tk.W, tk.E))
        
        # 主要操作按钮
        self.excel_process_button = ttk.Button(button_frame, text="⚙️ 开始整合", 
                                               command=self.start_excel_consolidation, 
                                               style='Accent.TButton')
        self.excel_process_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.excel_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                             command=self.clear_excel_results)
        self.excel_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.excel_preview_button = ttk.Button(button_frame, text="👁️ 预览数据", 
                                               command=self.preview_excel_data,
                                               state="disabled")
        self.excel_preview_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.excel_view_results_button = ttk.Button(button_frame, text="📊 查看结果", 
                                                   command=lambda: self.show_results_dialog('excel_processor'))
        self.excel_view_results_button.pack(side=tk.LEFT)
    
    def create_excel_text_extractor_tab(self):
        """创建Excel文本提取器页签 - 多语言版本"""
        # Excel文本提取器框架
        extractor_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(extractor_frame, text="翻译提取")
        
        # 配置网格
        extractor_frame.columnconfigure(0, weight=1)
        extractor_frame.rowconfigure(3, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(extractor_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="翻译提取工具 - 多语言版本", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, text="💡 为每种语言版本指定独立的Excel文件路径，系统将分别提取文本内容", 
                              style='Info.TLabel', foreground='blue')
        desc_label.grid(row=1, column=0)
        
        # 语言路径配置字典（中文为准，默认越南文，额外支持英文、泰文）
        self.extractor_language_paths = {
            '中文版': tk.StringVar(),
            '越南文版': tk.StringVar(),
            '英文版': tk.StringVar(),
            '泰文版': tk.StringVar(),
        }
        
        # 控制面板
        control_frame = ttk.Frame(extractor_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        control_frame.columnconfigure(0, weight=1)
        
        # 多语言路径选择区域
        lang_frame = ttk.LabelFrame(control_frame, text="多语言文件路径配置", padding="10")
        lang_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        lang_frame.columnconfigure(1, weight=1)
        
        # 为每种语言创建路径选择行
        row_idx = 0
        for lang_name, lang_var in self.extractor_language_paths.items():
            # 语言标签
            lang_label = ttk.Label(lang_frame, text=f"{lang_name}:", style='Info.TLabel')
            lang_label.grid(row=row_idx, column=0, sticky=tk.W, padx=(0, 8), pady=(3, 3))
            
            # 路径输入框
            lang_entry = ttk.Entry(lang_frame, textvariable=lang_var, font=("Microsoft YaHei", 9))
            lang_entry.grid(row=row_idx, column=1, sticky=(tk.W, tk.E), padx=(0, 8), pady=(3, 3))
            
            # 浏览按钮
            browse_btn = ttk.Button(lang_frame, text="浏览", 
                                   command=lambda l=lang_name: self.browse_extractor_language_path(l))
            browse_btn.grid(row=row_idx, column=2, pady=(3, 3))
            
            # 清除按钮
            clear_btn = ttk.Button(lang_frame, text="✕", width=3,
                                  command=lambda l=lang_name: self.clear_extractor_language_path(l))
            clear_btn.grid(row=row_idx, column=3, padx=(5, 0), pady=(3, 3))
            
            row_idx += 1
        
        # 通用输出目录区域
        output_frame = ttk.LabelFrame(control_frame, text="输出配置", padding="10")
        output_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)
        
        # 输出目录
        ttk.Label(output_frame, text="输出目录:", style='Info.TLabel').grid(row=0, column=0, sticky=tk.W, padx=(0, 8))
        self.extractor_output_var = tk.StringVar()
        self.extractor_output_entry = ttk.Entry(output_frame, textvariable=self.extractor_output_var, 
                                              font=("Microsoft YaHei", 9))
        self.extractor_output_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 8))
        
        self.extractor_output_browse_button = ttk.Button(output_frame, text="浏览", 
                                                        command=self.browse_extractor_output_directory)
        self.extractor_output_browse_button.grid(row=0, column=2)
        
        # 输出说明
        output_info = ttk.Label(output_frame, 
                               text="提示：每种语言的提取结果将保存在输出目录的对应子文件夹中", 
                               style='Info.TLabel', foreground='gray')
        output_info.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # 提取选项区域
        options_frame = ttk.LabelFrame(control_frame, text="提取选项", padding="10")
        options_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        options_frame.columnconfigure(1, weight=1)
        
        # 递归扫描选项
        self.extractor_recursive_var = tk.BooleanVar(value=False)
        self.extractor_recursive_check = ttk.Checkbutton(options_frame, text="递归扫描子目录", 
                                                         variable=self.extractor_recursive_var)
        self.extractor_recursive_check.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        
        # 文本类型过滤
        ttk.Label(options_frame, text="文本类型:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.extractor_text_type_var = tk.StringVar(value="全部")
        text_type_combo = ttk.Combobox(options_frame, textvariable=self.extractor_text_type_var, 
                                      values=["全部", "中文", "英文", "中英混合"], state="readonly", 
                                      width=15, font=("Microsoft YaHei", 9))
        text_type_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        
        # 策划检测说明
        planner_info = ttk.Label(options_frame, 
                                text="💡 自动检测：如果Excel文件第6行包含'策划'，将跳过该文件的文本提取", 
                                style='Info.TLabel', foreground='blue')
        planner_info.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(8, 0))
        
        # 语言支持说明
        language_info = ttk.Label(options_frame, 
                                 text="🌐 支持语言：中文、越南文（跳过纯英文），从第7行开始检测，同时提取A列内容", 
                                 style='Info.TLabel', foreground='green')
        language_info.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(3, 0))
        
        # 进度显示框架
        progress_frame = ttk.LabelFrame(extractor_frame, text="处理进度", padding="10")
        progress_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        # 进度条
        self.extractor_progress_var = tk.DoubleVar()
        self.extractor_progress_bar = ttk.Progressbar(progress_frame, variable=self.extractor_progress_var, 
                                                      maximum=100)
        self.extractor_progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # 进度文本
        self.extractor_progress_text_var = tk.StringVar(value="就绪")
        self.extractor_progress_text = ttk.Label(progress_frame, textvariable=self.extractor_progress_text_var, 
                                                style='Info.TLabel')
        self.extractor_progress_text.grid(row=1, column=0, sticky=tk.W)
        
        # 操作按钮区域
        button_frame = ttk.Frame(extractor_frame)
        button_frame.grid(row=3, column=0, pady=(0, 10))
        
        # 主要操作按钮
        self.extractor_process_button = ttk.Button(button_frame, text="📄 开始提取", 
                                                   command=self.start_text_extraction, 
                                                   style='Accent.TButton')
        self.extractor_process_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.extractor_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                                command=self.clear_extractor_results)
        self.extractor_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.extractor_preview_button = ttk.Button(button_frame, text="👁️ 预览文件", 
                                                  command=self.preview_extractor_files)
        self.extractor_preview_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.extractor_view_results_button = ttk.Button(button_frame, text="📝 查看结果", 
                                                       command=lambda: self.show_results_dialog('text_extractor'))
        self.extractor_view_results_button.pack(side=tk.LEFT)
    
    def create_field_extractor_tab(self):
        """创建表字段导出页签"""
        # 字段导出器框架
        field_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(field_frame, text="表字段导出")
        
        # 配置网格
        field_frame.columnconfigure(0, weight=1)
        field_frame.rowconfigure(3, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(field_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="表字段导出工具", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, 
                              text="扫描Excel文件，检测包含文本内容的列，从物理行第5行提取字段名", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(field_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 目录选择区域
        dir_frame = ttk.LabelFrame(control_frame, text="扫描设置", padding="12")
        dir_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        dir_frame.columnconfigure(1, weight=1)
        
        # 扫描目录
        ttk.Label(dir_frame, text="扫描目录:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 8))
        self.field_scan_dir_var = tk.StringVar()
        self.field_scan_dir_entry = ttk.Entry(dir_frame, textvariable=self.field_scan_dir_var, 
                                             font=("Microsoft YaHei", 9))
        self.field_scan_dir_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        
        self.field_scan_browse_button = ttk.Button(dir_frame, text="浏览目录", 
                                                  command=self.browse_field_scan_directory)
        self.field_scan_browse_button.grid(row=0, column=2, pady=(0, 8))
        
        # 输出文件夹
        ttk.Label(dir_frame, text="输出目录:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10))
        self.field_output_dir_var = tk.StringVar()
        self.field_output_dir_entry = ttk.Entry(dir_frame, textvariable=self.field_output_dir_var, 
                                               font=("Microsoft YaHei", 9))
        self.field_output_dir_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        self.field_output_browse_button = ttk.Button(dir_frame, text="选择输出目录", 
                                                    command=self.browse_field_output_directory)
        self.field_output_browse_button.grid(row=1, column=2)
        
        # 选项设置区域
        options_frame = ttk.LabelFrame(control_frame, text="处理选项", padding="12")
        options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 递归扫描选项
        self.field_recursive_var = tk.BooleanVar(value=False)
        self.field_recursive_check = ttk.Checkbutton(options_frame, text="递归扫描子目录", 
                                                    variable=self.field_recursive_var)
        self.field_recursive_check.grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        
        # 输出格式选择
        format_frame = ttk.Frame(options_frame)
        format_frame.grid(row=1, column=0, sticky=tk.W)
        
        ttk.Label(format_frame, text="输出格式:").pack(side=tk.LEFT, padx=(0, 10))
        self.field_output_format_var = tk.StringVar(value="json")
        ttk.Radiobutton(format_frame, text="JSON格式", 
                       variable=self.field_output_format_var, 
                       value="json").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(format_frame, text="CSV格式", 
                       variable=self.field_output_format_var, 
                       value="csv").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(format_frame, text="Excel格式", 
                       variable=self.field_output_format_var, 
                       value="excel").pack(side=tk.LEFT)
        
        # 说明信息
        info_label = ttk.Label(options_frame, 
                              text="💡 输出标准JSON格式，方便其他工具读取", 
                              style='Info.TLabel', foreground='blue')
        info_label.grid(row=2, column=0, sticky=tk.W, pady=(8, 0))
        
        # 操作按钮区域
        button_frame = ttk.Frame(field_frame)
        button_frame.grid(row=2, column=0, pady=(0, 15))
        
        self.field_extract_button = ttk.Button(button_frame, text="📊 开始提取", 
                                              command=self.start_field_extraction, 
                                              style='Accent.TButton')
        self.field_extract_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.field_copy_button = ttk.Button(button_frame, text="📋 复制JSON", 
                                           command=self.copy_field_json_result)
        self.field_copy_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.field_error_log_button = ttk.Button(button_frame, text="⚠️ 错误日志", 
                                                command=self.show_field_error_logs)
        self.field_error_log_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.field_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                            command=self.clear_field_results)
        self.field_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        self.field_view_results_button = ttk.Button(button_frame, text="📝 查看结果", 
                                                   command=lambda: self.show_results_dialog('field_extractor'))
        self.field_view_results_button.pack(side=tk.LEFT)
        
        # 结果显示区域
        results_frame = ttk.LabelFrame(field_frame, text="提取结果", padding="10")
        results_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # 结果文本框
        self.field_results_text = scrolledtext.ScrolledText(results_frame, 
                                                           wrap=tk.WORD,
                                                           font=("Consolas", 9))
        self.field_results_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    def create_table_range_translator_tab(self):
        """创建多语言翻译提取页签"""
        # 多语言翻译提取器框架
        trt_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(trt_frame, text="多语言翻译提取")
        
        # 配置网格
        trt_frame.columnconfigure(0, weight=1)
        trt_frame.rowconfigure(2, weight=1)
        
        # 标题和描述
        header_frame = ttk.Frame(trt_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(header_frame, text="多语言翻译提取工具", 
                               style='Heading.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        desc_label = ttk.Label(header_frame, 
                              text="根据字段导出的JSON配置，智能提取多语言翻译内容（只导出前端、后端、前后端字段）", 
                              style='Info.TLabel')
        desc_label.grid(row=1, column=0)
        
        # 控制面板
        control_frame = ttk.Frame(trt_frame)
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        control_frame.columnconfigure(0, weight=1)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(control_frame, text="文件配置", padding="12")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # JSON配置文件
        ttk.Label(file_frame, text="JSON配置:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 8))
        self.trt_json_var = tk.StringVar()
        self.trt_json_entry = ttk.Entry(file_frame, textvariable=self.trt_json_var, 
                                       font=("Microsoft YaHei", 9))
        self.trt_json_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        
        self.trt_json_browse_button = ttk.Button(file_frame, text="浏览JSON", 
                                                command=self.browse_trt_json_file)
        self.trt_json_browse_button.grid(row=0, column=2, pady=(0, 8))
        
        # 越南文目录（默认，无后缀）
        ttk.Label(file_frame, text="越南文目录:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 8))
        self.trt_vn_dir_var = tk.StringVar()
        self.trt_vn_dir_entry = ttk.Entry(file_frame, textvariable=self.trt_vn_dir_var, 
                                         font=("Microsoft YaHei", 9))
        self.trt_vn_dir_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        
        self.trt_vn_browse_button = ttk.Button(file_frame, text="浏览目录", 
                                              command=self.browse_trt_vn_directory)
        self.trt_vn_browse_button.grid(row=1, column=2, pady=(0, 8))
        
        # 中文目录（_zh后缀）
        ttk.Label(file_frame, text="中文目录:").grid(row=2, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 8))
        self.trt_zh_dir_var = tk.StringVar()
        self.trt_zh_dir_entry = ttk.Entry(file_frame, textvariable=self.trt_zh_dir_var, 
                                         font=("Microsoft YaHei", 9))
        self.trt_zh_dir_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        
        self.trt_zh_browse_button = ttk.Button(file_frame, text="浏览目录", 
                                              command=self.browse_trt_zh_directory)
        self.trt_zh_browse_button.grid(row=2, column=2, pady=(0, 8))
        
        # 泰文目录（_th后缀）
        ttk.Label(file_frame, text="泰文目录:").grid(row=3, column=0, sticky=tk.W, padx=(0, 10), pady=(0, 8))
        self.trt_th_dir_var = tk.StringVar()
        self.trt_th_dir_entry = ttk.Entry(file_frame, textvariable=self.trt_th_dir_var, 
                                         font=("Microsoft YaHei", 9))
        self.trt_th_dir_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        
        self.trt_th_browse_button = ttk.Button(file_frame, text="浏览目录", 
                                              command=self.browse_trt_th_directory)
        self.trt_th_browse_button.grid(row=3, column=2, pady=(0, 8))
        
        # 输出文件
        ttk.Label(file_frame, text="输出文件:").grid(row=4, column=0, sticky=tk.W, padx=(0, 10))
        self.trt_output_var = tk.StringVar()
        self.trt_output_entry = ttk.Entry(file_frame, textvariable=self.trt_output_var, 
                                         font=("Microsoft YaHei", 9))
        self.trt_output_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        self.trt_output_browse_button = ttk.Button(file_frame, text="选择保存位置", 
                                                  command=self.browse_trt_output_file)
        self.trt_output_browse_button.grid(row=4, column=2)
        
        # 说明信息区域
        info_frame = ttk.LabelFrame(control_frame, text="功能说明", padding="12")
        info_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        info_text = """✓ 跳过 no_text_tables 中的表格
✓ 处理 text_tables，只导出前端、后端、前后端字段（忽略策划字段）
✓ 多语言目录：越南文（无后缀）、中文（_zh）、泰文（_th）
✓ 生成翻译总表，每个表格文件对应一个工作表标签
✓ 列格式：字段名 | 字段类型 | Excel位置 | 中文内容 | 越南文 | 泰文"""
        
        info_label = ttk.Label(info_frame, text=info_text, 
                              font=("Microsoft YaHei", 9), 
                              justify=tk.LEFT, foreground='blue')
        info_label.pack(anchor=tk.W)
        
        # 操作按钮区域
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        # 主要操作按钮
        self.trt_process_button = ttk.Button(button_frame, text="🚀 开始提取", 
                                            command=self.start_table_range_translation, 
                                            style='Accent.TButton')
        self.trt_process_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 辅助操作按钮
        self.trt_clear_button = ttk.Button(button_frame, text="🗑️ 清空结果", 
                                          command=self.clear_trt_results)
        self.trt_clear_button.pack(side=tk.LEFT, padx=(0, 8))
        
        # 查看结果按钮
        self.trt_view_results_button = ttk.Button(button_frame, text="👁️ 查看结果", 
                                                 command=lambda: self.show_results_dialog('table_range_translator'))
        self.trt_view_results_button.pack(side=tk.LEFT)
    
    def create_about_tab(self):
        """创建关于页签"""
        about_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(about_frame, text="关于")
        
        # 配置网格
        about_frame.columnconfigure(0, weight=1)
        about_frame.rowconfigure(1, weight=1)
        about_frame.rowconfigure(2, weight=0)  # 底部信息不扩展
        
        # 标题区域
        title_frame = ttk.Frame(about_frame)
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        title_frame.columnconfigure(0, weight=1)
        
        # 主标题
        title_label = ttk.Label(title_frame, text="gametools - 游戏工具集", 
                               style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # 版本信息
        version_label = ttk.Label(title_frame, text=format_version_string(), 
                                 style='Info.TLabel')
        version_label.grid(row=1, column=0, pady=(0, 20))
        
        # 内容区域
        content_frame = ttk.Frame(about_frame)
        content_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        content_frame.columnconfigure(0, weight=1)
        content_frame.columnconfigure(1, weight=1)
        
        # 左侧：功能模块
        left_frame = ttk.LabelFrame(content_frame, text="功能模块", padding="15")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        features_text = f"""🔍 越南文检测导出
   检测Excel和CSV文件中的越南文内容并导出详细结果

📊 JSON格式检测工具  
   检测JSON文件中text字段的格式一致性

📈 Excel数据处理工具
   根据指定列对Excel数据进行分组和处理

📄 翻译提取工具
   批量提取Excel文件中的文本内容

📋 表字段导出工具
   扫描Excel文件，提取包含文本的列的字段信息

🌐 多语言翻译提取工具
   根据字段导出的JSON配置，智能提取多语言翻译内容

📋 版本信息
   当前版本: v{get_version()}
   项目描述: {get_description()}"""
        
        features_label = ttk.Label(left_frame, text=features_text, 
                                  font=("Microsoft YaHei", 10), 
                                  justify=tk.LEFT)
        features_label.pack(anchor=tk.W)
        
        # 右侧：技术信息
        right_frame = ttk.LabelFrame(content_frame, text="技术信息", padding="15")
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        
        # 获取最新更新内容
        latest_changes = get_latest_changes()
        changes_text = "\n".join([f"• {change}" for change in latest_changes])
        
        tech_text = f"""🛠️ 技术栈:
• Python 3.7+
• Tkinter (GUI界面)
• pandas (数据处理)
• openpyxl (Excel文件处理)

✨ 主要特性:
• 支持多种文件格式
• 图形化界面，操作简单
• 多线程处理，界面响应流畅
• 支持exe文件打包和分发

🆕 最新更新 (v{get_version()}):
{changes_text}

⚠️ 注意事项:
• 确保文件格式正确
• 大文件处理可能需要较长时间
• 建议在检测前备份重要文件"""
        
        tech_label = ttk.Label(right_frame, text=tech_text, 
                              font=("Microsoft YaHei", 10), 
                              justify=tk.LEFT)
        tech_label.pack(anchor=tk.W)
        
        # 底部信息
        bottom_frame = ttk.Frame(about_frame)
        bottom_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(20, 0))
        bottom_frame.columnconfigure(0, weight=1)
        
        # 使用方法
        usage_text = "📖 使用方法: 选择相应的功能页签 → 按照界面提示操作 → 查看检测结果"
        usage_label = ttk.Label(bottom_frame, text=usage_text, 
                               font=("Microsoft YaHei", 10), 
                               style='Info.TLabel')
        usage_label.grid(row=0, column=0, pady=(0, 10))
        
        # 版权信息
        copyright_text = "💬 技术支持: 如有问题或建议，请联系开发团队\n© 2024 gametools - 版权所有"
        copyright_label = ttk.Label(bottom_frame, text=copyright_text, 
                                   font=("Microsoft YaHei", 9), 
                                   style='Info.TLabel')
        copyright_label.grid(row=1, column=0)
    
    # 结果存储辅助函数
    def append_result(self, result_type, text):
        """追加文本到结果存储"""
        self.results_storage[result_type] += text
    
    def clear_result(self, result_type):
        """清空结果存储"""
        self.results_storage[result_type] = ''
    
    def get_result(self, result_type):
        """获取结果存储内容"""
        return self.results_storage.get(result_type, '')
    
    # 统一的结果查看对话框
    def show_results_dialog(self, result_type):
        """显示结果查看对话框（二级菜单）"""
        # 获取对应的结果内容
        result_content = self.results_storage.get(result_type, '')
        
        if not result_content.strip():
            messagebox.showinfo("提示", "暂无处理结果")
            return
        
        # 创建对话框窗口
        dialog = tk.Toplevel(self.root)
        dialog.title("查看处理结果")
        dialog.geometry("900x700")
        dialog.minsize(700, 500)
        
        # 结果标题映射
        title_map = {
            'vietnamese_processor': '越南文检测结果',
            'cross_project_translator': '跨项目翻译对应结果',
            'json_detector': 'JSON错误检测结果',
            'excel_processor': 'Excel数据处理结果',
            'text_extractor': '翻译提取结果',
            'field_extractor': '表字段导出结果',
            'table_range_translator': '多语言翻译提取结果'
        }
        
        # 标题
        title_frame = ttk.Frame(dialog, padding="10")
        title_frame.pack(fill=tk.X)
        
        title_label = ttk.Label(title_frame, 
                               text=title_map.get(result_type, '处理结果'),
                               style='Heading.TLabel')
        title_label.pack()
        
        # 结果显示区域
        result_frame = ttk.Frame(dialog, padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        result_text = scrolledtext.ScrolledText(result_frame, 
                                               wrap=tk.WORD, 
                                               font=("Consolas", 9))
        result_text.pack(fill=tk.BOTH, expand=True)
        result_text.insert(tk.END, result_content)
        result_text.config(state='disabled')  # 只读
        
        # 按钮区域
        button_frame = ttk.Frame(dialog, padding="10")
        button_frame.pack(fill=tk.X)
        
        # 复制按钮
        def copy_to_clipboard():
            dialog.clipboard_clear()
            dialog.clipboard_append(result_content)
            messagebox.showinfo("成功", "结果已复制到剪贴板")
        
        copy_button = ttk.Button(button_frame, text="📋 复制到剪贴板", 
                                command=copy_to_clipboard)
        copy_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # 保存按钮
        def save_to_file():
            file_path = filedialog.asksaveasfilename(
                title="保存结果",
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(result_content)
                    messagebox.showinfo("成功", f"结果已保存到: {file_path}")
                except Exception as e:
                    messagebox.showerror("错误", f"保存失败: {str(e)}")
        
        save_button = ttk.Button(button_frame, text="💾 保存到文件", 
                                command=save_to_file)
        save_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # 关闭按钮
        close_button = ttk.Button(button_frame, text="关闭", 
                                 command=dialog.destroy)
        close_button.pack(side=tk.RIGHT)
        
        # 设置对话框为模态
        dialog.transient(self.root)
        dialog.grab_set()
    
    # 越南文处理器相关方法
    def browse_vp_scan_directory(self):
        """浏览越南文处理器扫描目录"""
        directory = filedialog.askdirectory(title="选择要扫描的目录")
        if directory:
            self.vp_scan_dir_var.set(directory)
            # 自动设置输出文件夹为扫描目录下的子文件夹
            if not self.vp_output_folder_var.get():
                output_folder = os.path.join(directory, "越南文检测结果")
                self.vp_output_folder_var.set(output_folder)
    
    def browse_vp_output_folder(self):
        """浏览越南文处理器输出文件夹"""
        folder = filedialog.askdirectory(title="选择输出文件夹")
        if folder:
            self.vp_output_folder_var.set(folder)
    
    def start_vietnamese_processing(self):
        """开始越南文处理"""
        scan_dir = self.vp_scan_dir_var.get().strip()
        output_folder = self.vp_output_folder_var.get().strip()
        
        if not scan_dir:
            messagebox.showerror("错误", "请选择要扫描的目录")
            return
        
        if not output_folder:
            messagebox.showerror("错误", "请选择输出文件夹")
            return
        
        if not os.path.exists(scan_dir):
            messagebox.showerror("错误", "扫描目录不存在")
            return
        
        # 在新线程中执行处理
        self.vp_process_button.config(state="disabled")
        self.vp_progress_bar.start()
        self.vp_progress_var.set("正在处理...")
        self.status_var.set("正在处理越南文检测...")
        
        thread = threading.Thread(target=self._vietnamese_processing_thread, 
                                 args=(scan_dir, output_folder))
        thread.daemon = True
        thread.start()
    
    def _vietnamese_processing_thread(self, scan_dir, output_folder):
        """越南文处理线程"""
        try:
            # 清空结果
            self.root.after(0, self.clear_vp_results)
            
            # 开始处理
            self.root.after(0, lambda: self.append_result('vietnamese_processor', 
                f"开始扫描目录: {scan_dir}\n"))
            self.root.after(0, lambda: self.append_result('vietnamese_processor', 
                f"输出文件夹: {output_folder}\n"))
            self.root.after(0, lambda: self.append_result('vietnamese_processor', 
                f"递归扫描: {'是' if self.vp_recursive_var.get() else '否'}\n"))
            self.root.after(0, lambda: self.append_result('vietnamese_processor', 
                "支持的格式: .xlsx, .xls, .csv, .tsv\n"))
            self.root.after(0, lambda: self.append_result('vietnamese_processor', 
                "-" * 50 + "\n"))
            
            # 执行处理
            stats = self.vietnamese_processor.process_directory(
                directory_path=scan_dir,
                output_folder=output_folder,
                recursive=self.vp_recursive_var.get(),
                create_excel=self.vp_create_excel_var.get(),
                create_report=False
            )
            
            # 显示结果
            self.root.after(0, self._show_vp_result, stats)
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}"
            self.root.after(0, self._show_vp_error, error_msg)
        finally:
            # 恢复界面状态
            self.root.after(0, self._vp_finished)
    
    def _show_vp_result(self, stats):
        """显示越南文处理结果"""
        self.append_result('vietnamese_processor', "\n" + "=" * 50 + "\n")
        self.append_result('vietnamese_processor', "处理完成！\n")
        self.append_result('vietnamese_processor', "=" * 50 + "\n")
        self.append_result('vietnamese_processor', f"扫描的文件总数: {stats['total_files_scanned']}\n")
        self.append_result('vietnamese_processor', f"包含越南文的文件数: {stats['files_with_vietnamese']}\n")
        self.append_result('vietnamese_processor', f"越南文位置总数: {stats['total_vietnamese_locations']}\n")
        
        if stats['output_files']:
            self.append_result('vietnamese_processor', "\n✓ 输出文件创建成功！\n")
            self.append_result('vietnamese_processor', "生成的文件:\n")
            for output_file in stats['output_files']:
                self.append_result('vietnamese_processor', f"  - {output_file}\n")
        else:
            self.append_result('vietnamese_processor', "\n✗ 未找到越南文内容，未创建输出文件\n")
        
        # 显示成功消息
        if stats['output_files']:
            messagebox.showinfo("成功", f"越南文检测完成！\n找到 {stats['total_vietnamese_locations']} 个越南文位置\n已生成 {len(stats['output_files'])} 个输出文件")
        else:
            messagebox.showinfo("完成", "扫描完成，未发现越南文内容")
    
    def _show_vp_error(self, error_msg):
        """显示越南文处理错误"""
        self.append_result('vietnamese_processor', "\n" + "=" * 50 + "\n")
        self.append_result('vietnamese_processor', f"错误: {error_msg}\n")
        self.append_result('vietnamese_processor', "=" * 50 + "\n")
        messagebox.showerror("错误", error_msg)
    
    def _vp_finished(self):
        """越南文处理完成后的界面恢复"""
        self.vp_process_button.config(state="normal")
        self.vp_progress_bar.stop()
        self.vp_progress_var.set("处理完成")
        self.status_var.set("就绪")
    
    def clear_vp_results(self):
        """清空越南文处理结果"""
        self.clear_result('vietnamese_processor')
        self.vp_progress_var.set("就绪")
    
    def create_demo_files(self):
        """创建演示文件"""
        try:
            # 运行演示脚本
            result = subprocess.run([sys.executable, "tools/demo.py"], 
                                  capture_output=True, text=True, encoding='utf-8')
            
            if result.returncode == 0:
                self.append_result('vietnamese_processor', "演示文件创建成功！\n")
                self.append_result('vietnamese_processor', "文件位置: demo_tables/\n")
                self.append_result('vietnamese_processor', "现在可以使用批量扫描功能测试这些文件。\n")
                self.status_var.set("演示文件创建成功")
                messagebox.showinfo("成功", "演示文件创建成功！\n文件位置: demo_tables/")
            else:
                self.append_result('vietnamese_processor', f"创建演示文件失败: {result.stderr}\n")
                self.status_var.set("演示文件创建失败")
        except Exception as e:
            self.append_result('vietnamese_processor', f"创建演示文件时发生错误: {str(e)}\n")
            self.status_var.set("演示文件创建失败")
    
    # JSON格式检测工具相关方法
    def browse_json_folder(self):
        """浏览JSON文件夹"""
        folder_path = filedialog.askdirectory(
            title="选择包含JSON文件的文件夹"
        )
        if folder_path:
            self.json_path_var.set(folder_path)
    
    def start_json_detection(self):
        """开始JSON错误检测"""
        path = self.json_path_var.get().strip()
        
        if not path:
            messagebox.showerror("错误", "请选择路径")
            return
        
        if not os.path.exists(path):
            messagebox.showerror("错误", "路径不存在")
            return
        
        # 在新线程中执行检测
        self.json_detect_button.config(state="disabled")
        self.status_var.set("正在检测...")
        
        thread = threading.Thread(target=self._json_detection, 
                                 args=(path,))
        thread.daemon = True
        thread.start()
    
    def _json_detection(self, path):
        """JSON错误检测（后台线程）"""
        try:
            # 自动检测：如果是文件夹则检测文件夹，如果是文件则检测单个文件
            if os.path.isdir(path):
                report = self.json_detector.detect_errors_in_folder(path)
            else:
                report = self.json_detector.detect_errors(path)
            
            self.root.after(0, self._update_json_results, report)
        except Exception as e:
            error_msg = f"检测过程中发生错误: {str(e)}"
            self.root.after(0, self._show_json_error, error_msg)
    
    def _update_json_results(self, report):
        """更新JSON错误检测结果"""
        self.clear_result('json_detector')
        self.append_result('json_detector', report)
        
        self.json_detect_button.config(state="normal")
        self.json_save_button.config(state="normal")
        self.status_var.set("检测完成")
        messagebox.showinfo("完成", "JSON检测完成！请点击查看结果按钮查看详细报告")
    
    def _show_json_error(self, error_msg):
        """显示JSON错误检测错误"""
        self.clear_result('json_detector')
        self.append_result('json_detector', error_msg)
        
        self.json_detect_button.config(state="normal")
        self.status_var.set("检测失败")
        messagebox.showerror("错误", error_msg)
    
    def clear_json_results(self):
        """清空JSON检测结果"""
        self.clear_result('json_detector')
        self.json_save_button.config(state="disabled")
    
    def save_json_report(self):
        """保存JSON检测报告"""
        content = self.get_result('json_detector').strip()
        if not content:
            messagebox.showwarning("警告", "没有可保存的内容")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存检测报告",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("成功", f"报告已保存到: {file_path}")
                self.status_var.set(f"报告已保存: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    # Excel数据处理工具相关方法
    def browse_excel_input_file(self):
        """浏览Excel输入文件"""
        file_path = filedialog.askopenfilename(
            title="选择输入Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.excel_input_var.set(file_path)
            # 自动设置输出文件名
            if not self.excel_output_var.get():
                input_path = Path(file_path)
                output_path = input_path.parent / f"{input_path.stem}_整合{input_path.suffix}"
                self.excel_output_var.set(str(output_path))
    
    def browse_excel_output_folder(self):
        """浏览Excel输出文件夹"""
        folder_path = filedialog.askdirectory(title="选择输出文件夹")
        if folder_path:
            self.excel_output_folder_var.set(folder_path)
            # 自动设置输出文件名
            if not self.excel_output_filename_var.get():
                self.excel_output_filename_var.set("整合结果.xlsx")
    
    def start_excel_consolidation(self):
        """开始Excel数据整合"""
        input_file = self.excel_input_var.get().strip()
        output_folder = self.excel_output_folder_var.get().strip()
        output_filename = self.excel_output_filename_var.get().strip()
        
        if not input_file:
            messagebox.showerror("错误", "请选择输入文件")
            return
        
        if not output_folder:
            messagebox.showerror("错误", "请选择输出文件夹")
            return
        
        if not output_filename:
            messagebox.showerror("错误", "请输入输出文件名")
            return
        
        if not os.path.exists(input_file):
            messagebox.showerror("错误", "输入文件不存在")
            return
        
        if not os.path.exists(output_folder):
            messagebox.showerror("错误", "输出文件夹不存在")
            return
        
        # 构建完整的输出文件路径
        output_file = os.path.join(output_folder, output_filename)
        
        # 在新线程中执行整合
        self.excel_process_button.config(state="disabled")
        self.excel_preview_button.config(state="disabled")
        self.status_var.set("正在处理Excel数据...")
        
        thread = threading.Thread(target=self._excel_consolidation_process, 
                                 args=(input_file, output_file))
        thread.daemon = True
        thread.start()
    
    def _excel_consolidation_process(self, input_file, output_file):
        """Excel数据整合处理（后台线程）"""
        try:
            # 清空结果
            self.root.after(0, self.clear_excel_results)
            
            # 显示开始信息
            self.root.after(0, lambda: self.append_result('excel_processor', 
                f"开始处理文件: {input_file}\n"))
            self.root.after(0, lambda: self.append_result('excel_processor', 
                f"输出文件: {output_file}\n"))
            self.root.after(0, lambda: self.append_result('excel_processor', 
                "-" * 50 + "\n"))
            
            # 获取选项
            group_column = self.excel_group_column_var.get().strip() or None
            include_summary = self.excel_include_summary_var.get()
            sheet_prefix = self.excel_sheet_prefix_var.get().strip()
            
            # 执行处理
            success = self.excel_processor.process_file(
                input_path=input_file,
                output_folder=os.path.dirname(output_file),
                output_filename=os.path.basename(output_file),
                group_column=group_column,
                include_summary=include_summary,
                sheet_prefix=sheet_prefix
            )
            
            # 显示结果
            if success:
                self.root.after(0, self._show_excel_success_result)
            else:
                self.root.after(0, self._show_excel_error_result, "处理失败")
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}"
            self.root.after(0, self._show_excel_error_result, error_msg)
    
    def _show_excel_success_result(self):
        """显示Excel整合成功结果"""
        report = self.excel_processor.get_process_report()
        self.append_result('excel_processor', report)
        self.append_result('excel_processor', "\n\n✅ Excel数据处理完成！")
        
        self.excel_process_button.config(state="normal")
        self.excel_preview_button.config(state="normal")
        self.status_var.set("Excel处理完成")
        
        messagebox.showinfo("成功", "Excel数据处理完成！请点击查看结果按钮查看详细报告")
    
    def _show_excel_error_result(self, error_msg):
        """显示Excel处理错误结果"""
        self.append_result('excel_processor', f"❌ {error_msg}\n")
        
        self.excel_process_button.config(state="normal")
        self.excel_preview_button.config(state="normal")
        self.status_var.set("Excel处理失败")
        
        messagebox.showerror("错误", error_msg)
    
    def preview_excel_data(self):
        """预览Excel数据"""
        input_file = self.excel_input_var.get().strip()
        
        if not input_file:
            messagebox.showerror("错误", "请先选择输入文件")
            return
        
        if not os.path.exists(input_file):
            messagebox.showerror("错误", "输入文件不存在")
            return
        
        try:
            # 读取文件
            df = self.excel_processor.read_excel_file(input_file)
            
            # 显示预览信息
            preview_text = f"文件预览: {os.path.basename(input_file)}\n"
            preview_text += f"总行数: {len(df)}\n"
            preview_text += f"总列数: {len(df.columns)}\n"
            preview_text += f"列名: {list(df.columns)}\n\n"
            
            # 显示前几行数据
            preview_text += "前5行数据:\n"
            preview_text += df.head().to_string()
            
            # 显示A列的唯一值
            if len(df) > 0:
                first_col = df.columns[0]
                unique_values = df[first_col].unique()
                preview_text += f"\n\n第一列 '{first_col}' 的唯一值:\n"
                for i, value in enumerate(unique_values[:10]):  # 只显示前10个
                    preview_text += f"{i+1}. {value}\n"
                if len(unique_values) > 10:
                    preview_text += f"... 还有 {len(unique_values) - 10} 个值\n"
            
            # 清空并显示预览
            self.clear_result('excel_processor')
            self.append_result('excel_processor', preview_text)
            messagebox.showinfo("预览", "预览数据加载完成！请点击查看结果按钮查看")
            
        except Exception as e:
            messagebox.showerror("错误", f"预览数据失败: {str(e)}")
    
    def clear_excel_results(self):
        """清空Excel整合结果"""
        self.clear_result('excel_processor')
    
    # Excel文本提取器相关方法 - 多语言版本
    def browse_extractor_language_path(self, language_name):
        """
        浏览特定语言的文件路径
        
        Args:
            language_name: 语言名称
        """
        choice = messagebox.askquestion("选择类型", 
                                       f"为 {language_name} 选择：\n\n是(Y) = 选择Excel文件\n否(N) = 选择目录",
                                       icon='question')
        
        if choice == 'yes':
            # 选择文件
            file_path = filedialog.askopenfilename(
                title=f"选择 {language_name} 的Excel文件",
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
            if file_path:
                self.extractor_language_paths[language_name].set(file_path)
                self.status_var.set(f"已设置 {language_name} 路径")
        else:
            # 选择目录
            directory = filedialog.askdirectory(title=f"选择 {language_name} 的目录")
            if directory:
                self.extractor_language_paths[language_name].set(directory)
                self.status_var.set(f"已设置 {language_name} 路径")
        
        # 自动设置输出目录（如果还未设置）
        if not self.extractor_output_var.get():
            for lang_var in self.extractor_language_paths.values():
                path = lang_var.get().strip()
                if path:
                    if os.path.isfile(path):
                        self.extractor_output_var.set(os.path.dirname(path))
                    else:
                        self.extractor_output_var.set(path)
                    break
    
    def clear_extractor_language_path(self, language_name):
        """
        清除特定语言的路径
        
        Args:
            language_name: 语言名称
        """
        self.extractor_language_paths[language_name].set("")
        self.status_var.set(f"已清除 {language_name} 路径")
    
    def browse_extractor_output_directory(self):
        """浏览文本提取器输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.extractor_output_var.set(directory)
    
    def start_text_extraction(self):
        """开始文本提取 - 多语言版本"""
        output_dir = self.extractor_output_var.get().strip()
        
        # 检查是否至少配置了一种语言
        active_languages = {}
        for lang_name, lang_var in self.extractor_language_paths.items():
            path = lang_var.get().strip()
            if path:
                if not os.path.exists(path):
                    messagebox.showerror("错误", f"{lang_name} 路径不存在:\n{path}")
                    return
                active_languages[lang_name] = path
        
        if not active_languages:
            messagebox.showerror("错误", "请至少配置一种语言的文件路径")
            return
        
        # 设置输出目录
        if not output_dir:
            first_path = list(active_languages.values())[0]
            if os.path.isfile(first_path):
                output_dir = os.path.dirname(first_path)
            else:
                output_dir = first_path
        
        # 在新线程中执行提取
        self.extractor_process_button.config(state="disabled")
        self.status_var.set("正在提取文本...")
        
        # 重置进度条
        self.extractor_progress_var.set(0)
        self.extractor_progress_text_var.set("准备开始...")
        
        thread = threading.Thread(target=self._text_extraction_multi_language, 
                                 args=(active_languages, output_dir))
        thread.daemon = True
        thread.start()
    
    def update_extractor_progress(self, current: int, total: int, filename: str, message: str):
        """更新提取进度"""
        if total > 0:
            percentage = (current / total) * 100
            self.extractor_progress_var.set(percentage)
            progress_text = f"[{current}/{total}] ({percentage:.1f}%) {filename}: {message}"
        else:
            self.extractor_progress_var.set(0)
            progress_text = f"{filename}: {message}"
        
        self.extractor_progress_text_var.set(progress_text)
        
        # 在日志窗口显示
        timestamp = self._get_timestamp()
        if "成功" in message or "完成" in message:
            log_message = f"✅ [{timestamp}] {progress_text}\n"
        elif "失败" in message or "错误" in message:
            log_message = f"❌ [{timestamp}] {progress_text}\n"
        elif "跳过" in message:
            log_message = f"⏭️ [{timestamp}] {progress_text}\n"
        elif "开始" in message:
            log_message = f"🚀 [{timestamp}] {progress_text}\n"
        else:
            log_message = f"ℹ️ [{timestamp}] {progress_text}\n"
        
        self.append_result('text_extractor', log_message)
        self.root.update_idletasks()
    
    def _sanitize_folder_name(self, name):
        """清理文件夹名称"""
        clean_name = name.replace('版', '')
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        return clean_name
    
    def _text_extraction_multi_language(self, active_languages, output_dir):
        """多语言文本提取（后台线程）- 生成汇总Excel表格"""
        try:
            # 清空结果
            self.root.after(0, self.clear_extractor_results)
            
            # 显示开始信息
            timestamp = self._get_timestamp()
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"🚀 [{timestamp}] 开始多语言翻译提取任务\n"))
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"📁 输出目录: {output_dir}\n"))
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"🌐 配置语言数: {len(active_languages)} 种\n"))
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"📋 输出格式: name | num | cn | vn | en | th\n"))
            self.root.after(0, lambda: self.append_result('text_extractor', 
                "=" * 60 + "\n\n"))
            
            # 语言映射
            lang_map = {
                '中文版': 'cn',
                '越南文版': 'vn',
                '英文版': 'en',
                '泰文版': 'th'
            }
            
            # 收集所有语言的文本数据
            all_lang_data = {}  # {lang_code: {file_name: extracted_data}}
            
            for idx, (lang_name, lang_path) in enumerate(active_languages.items(), 1):
                lang_code = lang_map[lang_name]
                self.root.after(0, lambda n=lang_name, i=idx, t=len(active_languages): 
                              self.append_result('text_extractor', 
                                  f"\n📚 提取 [{i}/{t}]: {n}\n"))
                
                try:
                    lang_extractor = ExcelTextExtractor(progress_callback=self.update_extractor_progress)
                    file_data = {}
                    
                    if os.path.isfile(lang_path):
                        # 单个文件
                        file_name = os.path.splitext(os.path.basename(lang_path))[0]
                        self.root.after(0, lambda f=file_name: 
                                      self.append_result('text_extractor', f"  📄 {f}\n"))
                        extracted_data = lang_extractor.extract_text_from_excel(lang_path, 1, 1)
                        if extracted_data:
                            file_data[file_name] = extracted_data
                    else:
                        # 目录批量处理
                        excel_files = lang_extractor.scan_directory(lang_path)
                        self.root.after(0, lambda c=len(excel_files): 
                                      self.append_result('text_extractor', f"  📁 找到 {c} 个文件\n"))
                        
                        for file_idx, file_path in enumerate(excel_files, 1):
                            file_name = os.path.splitext(os.path.basename(file_path))[0]
                            extracted_data = lang_extractor.extract_text_from_excel(file_path, file_idx, len(excel_files))
                            if extracted_data:
                                file_data[file_name] = extracted_data
                    
                    all_lang_data[lang_code] = file_data
                    self.root.after(0, lambda c=len(file_data): 
                                  self.append_result('text_extractor', f"  ✅ 成功提取 {c} 个文件\n"))
                    
                except Exception as e:
                    error_msg = f"  ❌ 提取 {lang_name} 时出错: {str(e)}\n"
                    self.root.after(0, lambda m=error_msg: self.append_result('text_extractor', m))
            
            # 生成汇总Excel表格
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"\n{'='*60}\n"))
            self.root.after(0, lambda: self.append_result('text_extractor', 
                f"📊 生成汇总Excel表格\n"))
            
            output_file = os.path.join(output_dir, "翻译提取汇总.xlsx")
            success = self._create_extractor_summary_excel(all_lang_data, output_file)
            
            if success:
                self.root.after(0, lambda: self.append_result('text_extractor', 
                    f"✅ 汇总表格已生成: {output_file}\n"))
                self.root.after(0, lambda: self.append_result('text_extractor', 
                    f"{'='*60}\n\n"))
                self.root.after(0, self._show_extractor_multi_lang_success)
            else:
                self.root.after(0, lambda: self.append_result('text_extractor', 
                    f"❌ 生成汇总表格失败\n"))
                self.root.after(0, self._show_extractor_error_result, "生成汇总表格失败")
            
        except Exception as e:
            error_msg = f"多语言提取过程中发生错误: {str(e)}"
            self.root.after(0, self._show_extractor_error_result, error_msg)
    
    def _create_extractor_summary_excel(self, all_lang_data, output_file):
        """创建汇总Excel表格"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 选择第一个有数据的语言作为基准
            base_lang = None
            base_data = None
            lang_priority = ['cn', 'vn', 'en', 'th']  # 优先级顺序
            
            for lang in lang_priority:
                if lang in all_lang_data and all_lang_data[lang]:
                    base_lang = lang
                    base_data = all_lang_data[lang]
                    logging.info(f"使用 {lang} 作为基准语言")
                    break
            
            if not base_lang or not base_data:
                logging.error("未找到任何语言数据，无法生成汇总表格")
                return False
            
            # 构建汇总数据
            summary_data = []
            
            for file_name, file_data in base_data.items():
                for sheet_name, sheet_data in file_data.items():
                    if not sheet_data or 'items' not in sheet_data:
                        continue
                    
                    for item in sheet_data['items']:
                        excel_pos = item.get('excel_row_ref', '')
                        base_text = item.get('text', '')
                        
                        # 查找所有语言对应位置的文本
                        cn_text = base_text if base_lang == 'cn' else self._find_extractor_text_at_position(all_lang_data.get('cn', {}), file_name, sheet_name, excel_pos)
                        vn_text = base_text if base_lang == 'vn' else self._find_extractor_text_at_position(all_lang_data.get('vn', {}), file_name, sheet_name, excel_pos)
                        en_text = base_text if base_lang == 'en' else self._find_extractor_text_at_position(all_lang_data.get('en', {}), file_name, sheet_name, excel_pos)
                        th_text = base_text if base_lang == 'th' else self._find_extractor_text_at_position(all_lang_data.get('th', {}), file_name, sheet_name, excel_pos)
                        
                        summary_data.append({
                            'name': file_name,
                            'num': excel_pos,
                            'cn': cn_text,
                            'vn': vn_text,
                            'en': en_text,
                            'th': th_text
                        })
            
            # 创建DataFrame
            df = pd.DataFrame(summary_data)
            
            # 保存到Excel
            df.to_excel(output_file, index=False, sheet_name='汇总')
            
            # 美化Excel格式
            wb = load_workbook(output_file)
            ws = wb['汇总']
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            ws.column_dimensions['A'].width = 25  # name
            ws.column_dimensions['B'].width = 10  # num
            ws.column_dimensions['C'].width = 40  # cn
            ws.column_dimensions['D'].width = 40  # vn
            ws.column_dimensions['E'].width = 40  # en
            ws.column_dimensions['F'].width = 40  # th
            
            # 保存
            wb.save(output_file)
            
            logging.info(f"汇总Excel表格创建成功: {output_file}, 总计 {len(summary_data)} 条记录")
            
            return True
            
        except Exception as e:
            logging.error(f"创建汇总Excel失败: {str(e)}")
            return False
    
    def _find_extractor_text_at_position(self, lang_data, file_name, sheet_name, excel_pos):
        """在指定语言数据中查找特定位置的文本"""
        if not lang_data or file_name not in lang_data:
            return ""
        
        file_data = lang_data[file_name]
        if sheet_name not in file_data:
            return ""
        
        sheet_data = file_data[sheet_name]
        if not sheet_data or 'items' not in sheet_data:
            return ""
        
        for item in sheet_data['items']:
            if item.get('excel_row_ref') == excel_pos:
                return item.get('text', '')
        
        return ""
    
    def _show_extractor_multi_lang_success(self):
        """显示多语言提取成功结果"""
        timestamp = self._get_timestamp()
        self.append_result('text_extractor', f"✅ [{timestamp}] 多语言文本提取完成！\n")
        
        self.extractor_process_button.config(state="normal")
        self.status_var.set("多语言提取完成")
        
        messagebox.showinfo("成功", "多语言Excel文本提取完成！\n\n汇总Excel表格已生成，格式：name | num | cn | vn | en | th\n\n请点击查看结果按钮查看详细日志")
    
    def _show_extractor_error_result(self, error_msg):
        """显示文本提取错误结果"""
        timestamp = self._get_timestamp()
        self.append_result('text_extractor', f"❌ [{timestamp}] {error_msg}\n")
        
        self.extractor_process_button.config(state="normal")
        self.status_var.set("文本提取失败")
        
        messagebox.showerror("错误", error_msg)
    
    def preview_extractor_files(self):
        """预览各语言的Excel文件 - 多语言版本"""
        # 收集已配置的语言路径
        active_languages = {}
        for lang_name, lang_var in self.extractor_language_paths.items():
            path = lang_var.get().strip()
            if path and os.path.exists(path):
                active_languages[lang_name] = path
        
        if not active_languages:
            messagebox.showerror("错误", "请先配置至少一种语言的文件路径")
            return
        
        try:
            # 清空并显示预览
            self.extractor_result_text.delete(1.0, tk.END)
            
            preview_text = "=" * 60 + "\n"
            preview_text += "📋 多语言文件预览\n"
            preview_text += "=" * 60 + "\n\n"
            
            total_files = 0
            
            for lang_name, lang_path in active_languages.items():
                preview_text += f"🌍 {lang_name}\n"
                preview_text += f"   路径: {lang_path}\n"
                
                if os.path.isfile(lang_path):
                    # 单个文件
                    preview_text += f"   类型: 单个Excel文件\n"
                    preview_text += f"   文件名: {os.path.basename(lang_path)}\n"
                    total_files += 1
                else:
                    # 目录
                    excel_files = self.text_extractor.scan_directory(lang_path)
                    preview_text += f"   类型: 目录\n"
                    preview_text += f"   找到文件: {len(excel_files)} 个\n"
                    
                    if excel_files:
                        preview_text += f"   文件列表:\n"
                        for i, file_path in enumerate(excel_files[:10]):
                            preview_text += f"      {i+1}. {os.path.basename(file_path)}\n"
                        if len(excel_files) > 10:
                            preview_text += f"      ... 还有 {len(excel_files) - 10} 个文件\n"
                    
                    total_files += len(excel_files)
                
                preview_text += "\n"
            
            preview_text += "=" * 60 + "\n"
            preview_text += f"📊 总计: {len(active_languages)} 种语言，{total_files} 个文件\n"
            preview_text += "=" * 60 + "\n"
            
            self.clear_result('text_extractor')
            self.append_result('text_extractor', preview_text)
            messagebox.showinfo("预览", "文件预览加载完成！请点击查看结果按钮查看")
            
        except Exception as e:
            messagebox.showerror("错误", f"预览文件失败: {str(e)}")
    
    def clear_extractor_results(self):
        """清空文本提取结果"""
        self.clear_result('text_extractor')
    
    # ==================== 跨项目翻译对应相关方法 ====================
    
    def browse_cpt_mapping_file(self):
        """浏览映射文件"""
        file_path = filedialog.askopenfilename(
            title="选择映射文件",
            filetypes=[
                ("Excel文件", "*.xlsx *.xls"),
                ("所有文件", "*.*")
            ]
        )
        if file_path:
            self.cpt_mapping_file_var.set(file_path)
            # 自动设置输出文件名
            if not self.cpt_output_file_var.get():
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = os.path.join(os.path.dirname(file_path), f"{base_name}_翻译对应结果.xlsx")
                self.cpt_output_file_var.set(output_path)
    
    def browse_cpt_project_directory(self):
        """浏览项目目录"""
        dir_path = filedialog.askdirectory(title="选择项目目录")
        if dir_path:
            self.cpt_project_dir_var.set(dir_path)
    
    def browse_cpt_output_file(self):
        """浏览输出文件"""
        file_path = filedialog.asksaveasfilename(
            title="选择输出文件",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("所有文件", "*.*")
            ]
        )
        if file_path:
            self.cpt_output_file_var.set(file_path)
    
    def start_cross_project_translation(self):
        """开始跨项目翻译对应"""
        mapping_file = self.cpt_mapping_file_var.get().strip()
        project_dir = self.cpt_project_dir_var.get().strip()
        output_file = self.cpt_output_file_var.get().strip()
        
        # 验证输入
        if not mapping_file:
            messagebox.showerror("错误", "请选择映射文件")
            return
        
        if not project_dir:
            messagebox.showerror("错误", "请选择项目目录")
            return
        
        if not output_file:
            messagebox.showerror("错误", "请设置输出文件")
            return
        
        if not os.path.exists(mapping_file):
            messagebox.showerror("错误", "映射文件不存在")
            return
        
        if not os.path.exists(project_dir):
            messagebox.showerror("错误", "项目目录不存在")
            return
        
        # 在新线程中执行翻译对应
        self.cpt_process_button.config(state="disabled")
        self.status_var.set("正在处理翻译对应...")
        
        thread = threading.Thread(target=self._cross_project_translation, 
                                 args=(mapping_file, project_dir, output_file))
        thread.daemon = True
        thread.start()
    
    def _cross_project_translation(self, mapping_file, project_dir, output_file):
        """跨项目翻译对应（后台线程）"""
        try:
            # 清空结果
            self.root.after(0, self.clear_cpt_results)
            
            # 开始处理
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"开始处理翻译对应...\n"))
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"映射文件: {mapping_file}\n"))
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"项目目录: {project_dir}\n"))
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"输出文件: {output_file}\n"))
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"{'='*60}\n"))
            
            # 处理翻译映射
            results = self.cross_project_translator.process_translation_mapping(
                mapping_file, project_dir)
            
            if results:
                # 显示处理报告
                report = self.cross_project_translator.get_processing_report()
                self.root.after(0, lambda: self.append_result('cross_project_translator', 
                    f"{report}\n"))
                
                # 导出结果
                if self.cross_project_translator.export_results(output_file):
                    self.root.after(0, lambda: self.append_result('cross_project_translator', 
                        f"结果已导出到: {output_file}\n"))
                    # 启用导出按钮
                    self.root.after(0, lambda: self.cpt_export_button.config(state="normal"))
                else:
                    self.root.after(0, lambda: self.append_result('cross_project_translator', 
                        f"导出失败！\n"))
                
                # 显示详细结果（前20条）
                self.root.after(0, lambda: self.append_result('cross_project_translator', 
                    f"\n详细结果（前20条）:\n"))
                self.root.after(0, lambda: self.append_result('cross_project_translator', 
                    f"{'='*60}\n"))
                
                for i, result in enumerate(results[:20]):
                    status_icon = "✅" if result['status'] == 'success' else "❌"
                    self.root.after(0, lambda r=result, icon=status_icon: 
                        self.append_result('cross_project_translator', 
                            f"{icon} 第{r['index']}行: {r['file_name']} -> {r['content'][:50]}...\n"))
                
                if len(results) > 20:
                    self.root.after(0, lambda: self.append_result('cross_project_translator', 
                        f"... 还有 {len(results) - 20} 条结果，请查看导出的Excel文件\n"))
                
            else:
                self.root.after(0, lambda: self.append_result('cross_project_translator', 
                    f"处理失败，没有生成结果\n"))
            
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"\n处理完成！\n"))
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}"
            self.root.after(0, lambda: self.append_result('cross_project_translator', 
                f"❌ {error_msg}\n"))
        
        # 恢复按钮状态
        self.root.after(0, lambda: self.cpt_process_button.config(state="normal"))
        self.root.after(0, lambda: self.status_var.set("翻译对应完成"))
        self.root.after(0, lambda: messagebox.showinfo("完成", "翻译对应完成！请点击查看结果按钮查看详细报告"))
    
    def clear_cpt_results(self):
        """清空跨项目翻译对应结果"""
        self.clear_result('cross_project_translator')
        self.cpt_export_button.config(state="disabled")
    
    def export_cpt_results(self):
        """导出跨项目翻译对应结果"""
        if not self.cross_project_translator.translation_results:
            messagebox.showwarning("警告", "没有结果可导出")
            return
        
        # 选择导出文件
        file_path = filedialog.asksaveasfilename(
            title="导出翻译对应结果",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            if self.cross_project_translator.export_results(file_path):
                messagebox.showinfo("成功", f"结果已导出到:\n{file_path}")
            else:
                messagebox.showerror("错误", "导出失败")
    
    # ==================== 表字段导出相关方法 ====================
    
    def browse_field_scan_directory(self):
        """浏览字段提取扫描目录"""
        dir_path = filedialog.askdirectory(title="选择扫描目录")
        if dir_path:
            self.field_scan_dir_var.set(dir_path)
            # 如果输出目录为空，自动设置为扫描目录
            if not self.field_output_dir_var.get():
                self.field_output_dir_var.set(dir_path)
    
    def browse_field_output_directory(self):
        """浏览字段提取输出目录"""
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.field_output_dir_var.set(dir_path)
    
    def start_field_extraction(self):
        """开始字段提取"""
        scan_dir = self.field_scan_dir_var.get().strip()
        output_dir = self.field_output_dir_var.get().strip()
        
        # 验证输入
        if not scan_dir:
            messagebox.showerror("错误", "请选择扫描目录")
            return
        
        if not os.path.exists(scan_dir):
            messagebox.showerror("错误", "扫描目录不存在")
            return
        
        if not output_dir:
            output_dir = scan_dir
            self.field_output_dir_var.set(output_dir)
        
        # 在新线程中执行提取
        self.field_extract_button.config(state="disabled")
        self.status_var.set("正在提取表字段...")
        
        thread = threading.Thread(target=self._field_extraction_thread, 
                                 args=(scan_dir, output_dir))
        thread.daemon = True
        thread.start()
    
    def _field_extraction_thread(self, scan_dir, output_dir):
        """字段提取线程"""
        try:
            # 清空结果显示
            self.root.after(0, lambda: self.field_results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self._log_field_result("=" * 60))
            self.root.after(0, lambda: self._log_field_result("开始提取表字段信息..."))
            self.root.after(0, lambda: self._log_field_result("=" * 60))
            self.root.after(0, lambda: self._log_field_result(f"扫描目录: {scan_dir}"))
            self.root.after(0, lambda: self._log_field_result(f"输出目录: {output_dir}"))
            self.root.after(0, lambda: self._log_field_result(f"输出格式: {self.field_output_format_var.get().upper()}"))
            self.root.after(0, lambda: self._log_field_result(f"递归扫描: {'是' if self.field_recursive_var.get() else '否'}"))
            self.root.after(0, lambda: self._log_field_result(""))
            
            # 执行提取
            stats = self.field_extractor.process_directory(
                directory_path=scan_dir,
                output_folder=output_dir,
                output_format=self.field_output_format_var.get(),
                recursive=self.field_recursive_var.get()
            )
            
            # 保存输出文件路径和结果数据
            self.results_storage['field_extractor'] = stats.get('output_file', '')
            self.field_extraction_results = stats.get('results', [])
            
            # 显示统计信息
            self.root.after(0, lambda: self._log_field_result(""))
            self.root.after(0, lambda: self._log_field_result("=" * 60))
            self.root.after(0, lambda: self._log_field_result("提取完成!"))
            self.root.after(0, lambda: self._log_field_result("=" * 60))
            self.root.after(0, lambda: self._log_field_result(f"扫描文件数: {stats['total_files']}"))
            self.root.after(0, lambda: self._log_field_result(f"工作表数: {stats['total_sheets']}"))
            self.root.after(0, lambda: self._log_field_result(f"提取字段数: {stats['total_fields']}"))
            self.root.after(0, lambda: self._log_field_result(f"输出文件: {stats['output_file']}"))
            self.root.after(0, lambda: self._log_field_result(""))
            
            # 如果是JSON格式，显示JSON预览
            if self.field_output_format_var.get() == 'json' and self.field_extraction_results:
                self.root.after(0, lambda: self._log_field_result("JSON结果预览:"))
                self.root.after(0, lambda: self._log_field_result("-" * 60))
                import json
                json_data = [{
                    "table_name": r['excel_file'],
                    "sheet_name": r['sheet_name'],
                    "fields": r['fields'],
                    "field_count": r['field_count']
                } for r in self.field_extraction_results]
                json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
                self.root.after(0, lambda: self._log_field_result(json_str))
                self.root.after(0, lambda: self._log_field_result("-" * 60))
                self.root.after(0, lambda: self._log_field_result(""))
            
            # 显示完成消息
            self.root.after(0, lambda: self.status_var.set("字段提取完成"))
            self.root.after(0, lambda: messagebox.showinfo(
                "完成",
                f"字段提取完成!\n\n"
                f"扫描文件数: {stats['total_files']}\n"
                f"工作表数: {stats['total_sheets']}\n"
                f"提取字段数: {stats['total_fields']}\n\n"
                f"结果已保存到:\n{stats['output_file']}"
            ))
            
        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            self.root.after(0, lambda: self._log_field_result(f"\n错误: {str(e)}"))
            self.root.after(0, lambda: self._log_field_result(error_msg))
            self.root.after(0, lambda: self.status_var.set("字段提取失败"))
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理失败:\n{str(e)}"))
        
        finally:
            self.root.after(0, lambda: self.field_extract_button.config(state="normal"))
    
    def _log_field_result(self, message):
        """记录字段提取结果"""
        self.field_results_text.insert(tk.END, message + "\n")
        self.field_results_text.see(tk.END)
    
    def clear_field_results(self):
        """清空字段提取结果"""
        self.field_results_text.delete(1.0, tk.END)
        self.results_storage['field_extractor'] = ''
        self.field_extraction_results = None
        # 清除提取器的日志
        self.field_extractor.clear_logs()
    
    def show_field_error_logs(self):
        """显示字段提取的错误和警告日志"""
        logs = self.field_extractor.get_all_logs()
        errors = logs['errors']
        warnings = logs['warnings']
        
        if not errors and not warnings:
            messagebox.showinfo("日志信息", "没有错误或警告日志")
            return
        
        # 创建新窗口显示日志
        log_window = tk.Toplevel(self.root)
        log_window.title("字段提取 - 错误与警告日志")
        log_window.geometry("900x600")
        
        # 创建笔记本（标签页）
        notebook = ttk.Notebook(log_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 错误日志标签页
        error_frame = ttk.Frame(notebook)
        notebook.add(error_frame, text=f"错误日志 ({len(errors)})")
        
        error_text = scrolledtext.ScrolledText(error_frame, 
                                               wrap=tk.WORD,
                                               font=('Consolas', 9))
        error_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        if errors:
            for i, error in enumerate(errors, 1):
                error_text.insert(tk.END, f"{i}. {error}\n\n")
        else:
            error_text.insert(tk.END, "无错误日志")
        
        error_text.config(state='disabled')
        
        # 警告日志标签页
        warning_frame = ttk.Frame(notebook)
        notebook.add(warning_frame, text=f"警告日志 ({len(warnings)})")
        
        warning_text = scrolledtext.ScrolledText(warning_frame,
                                                 wrap=tk.WORD,
                                                 font=('Consolas', 9))
        warning_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        if warnings:
            for i, warning in enumerate(warnings, 1):
                warning_text.insert(tk.END, f"{i}. {warning}\n\n")
        else:
            warning_text.insert(tk.END, "无警告日志")
        
        warning_text.config(state='disabled')
        
        # 底部按钮
        button_frame = ttk.Frame(log_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 保存日志按钮
        def save_logs():
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                title="保存日志",
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
            )
            if file_path:
                if self.field_extractor.save_logs_to_file(Path(file_path)):
                    messagebox.showinfo("成功", f"日志已保存到:\n{file_path}")
        
        ttk.Button(button_frame, text="保存日志", command=save_logs).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="关闭", command=log_window.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 统计信息
        stats_label = ttk.Label(button_frame, 
                               text=f"总计: {len(errors)} 个错误, {len(warnings)} 个警告",
                               foreground='#7f8c8d')
        stats_label.pack(side=tk.LEFT, padx=20)
    
    def copy_field_json_result(self):
        """复制字段提取的JSON结果到剪贴板"""
        if not self.field_extraction_results:
            messagebox.showwarning("警告", "没有可复制的结果，请先执行字段提取")
            return
        
        try:
            import json
            # 构建JSON数据
            json_data = [{
                "table_name": r['excel_file'],
                "sheet_name": r['sheet_name'],
                "fields_with_examples": r.get('fields_with_examples', []),
                "field_count": r['field_count']
            } for r in self.field_extraction_results]
            
            json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
            
            # 复制到剪贴板
            self.root.clipboard_clear()
            self.root.clipboard_append(json_str)
            self.root.update()
            
            messagebox.showinfo("成功", f"JSON结果已复制到剪贴板\n共 {len(json_data)} 条记录")
        except Exception as e:
            messagebox.showerror("错误", f"复制失败:\n{str(e)}")
    
    # ==================== 多语言翻译提取相关方法 ====================
    
    def browse_trt_json_file(self):
        """浏览JSON配置文件"""
        file_path = filedialog.askopenfilename(
            title="选择JSON配置文件",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
        )
        if file_path:
            self.trt_json_var.set(file_path)
    
    def browse_trt_vn_directory(self):
        """浏览越南文文件目录"""
        dir_path = filedialog.askdirectory(title="选择越南文Excel文件目录")
        if dir_path:
            self.trt_vn_dir_var.set(dir_path)
    
    def browse_trt_zh_directory(self):
        """浏览中文文件目录"""
        dir_path = filedialog.askdirectory(title="选择中文Excel文件目录（_zh后缀）")
        if dir_path:
            self.trt_zh_dir_var.set(dir_path)
    
    def browse_trt_th_directory(self):
        """浏览泰文文件目录"""
        dir_path = filedialog.askdirectory(title="选择泰文Excel文件目录（_th后缀）")
        if dir_path:
            self.trt_th_dir_var.set(dir_path)
    
    def browse_trt_output_file(self):
        """浏览输出文件位置"""
        file_path = filedialog.asksaveasfilename(
            title="保存翻译总表",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.trt_output_var.set(file_path)
    
    def start_table_range_translation(self):
        """开始多语言翻译提取"""
        json_path = self.trt_json_var.get().strip()
        vn_dir = self.trt_vn_dir_var.get().strip()
        zh_dir = self.trt_zh_dir_var.get().strip()
        th_dir = self.trt_th_dir_var.get().strip()
        output_file = self.trt_output_var.get().strip()
        
        # 验证输入
        if not json_path:
            messagebox.showerror("错误", "请选择JSON配置文件")
            return
        
        if not vn_dir and not zh_dir and not th_dir:
            messagebox.showerror("错误", "请至少选择一个语言目录")
            return
        
        if not output_file:
            messagebox.showerror("错误", "请选择输出文件位置")
            return
        
        if not os.path.exists(json_path):
            messagebox.showerror("错误", "JSON配置文件不存在")
            return
        
        # 验证目录存在性
        if vn_dir and not os.path.exists(vn_dir):
            messagebox.showerror("错误", "越南文目录不存在")
            return
        
        if zh_dir and not os.path.exists(zh_dir):
            messagebox.showerror("错误", "中文目录不存在")
            return
        
        if th_dir and not os.path.exists(th_dir):
            messagebox.showerror("错误", "泰文目录不存在")
            return
        
        # 构建语言目录字典
        lang_dirs = {}
        if vn_dir:
            lang_dirs['vn'] = vn_dir
        if zh_dir:
            lang_dirs['zh'] = zh_dir
        if th_dir:
            lang_dirs['th'] = th_dir
        
        # 在新线程中执行提取
        self.trt_process_button.config(state="disabled")
        self.status_var.set("正在提取翻译内容...")
        
        thread = threading.Thread(target=self._table_range_translation_thread, 
                                 args=(json_path, lang_dirs, output_file))
        thread.daemon = True
        thread.start()
    
    def _table_range_translation_thread(self, json_path, lang_dirs, output_file):
        """多语言翻译提取线程"""
        try:
            # 清空结果
            self.root.after(0, self.clear_trt_results)
            
            # 开始处理
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                "=" * 70 + "\n"))
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                "开始多语言翻译提取...\n"))
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                "=" * 70 + "\n"))
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                f"JSON配置: {json_path}\n"))
            
            # 显示各语言目录
            for lang, dir_path in lang_dirs.items():
                lang_name = {'vn': '越南文', 'zh': '中文', 'th': '泰文'}.get(lang, lang)
                self.root.after(0, lambda ln=lang_name, dp=dir_path: 
                    self.append_result('table_range_translator', f"{ln}目录: {dp}\n"))
            
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                f"输出文件: {output_file}\n"))
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                "\n"))
            
            # 定义进度回调函数
            def progress_callback(msg):
                """进度回调，将消息显示到界面"""
                self.root.after(0, lambda m=msg: self.append_result('table_range_translator', m + "\n"))
            
            # 处理数据
            results = self.table_range_translator.process_with_json_config_multi_lang(
                json_path, lang_dirs, progress_callback=progress_callback)
            
            if results:
                self.root.after(0, lambda: self.append_result('table_range_translator', 
                    f"✓ 成功提取 {len(results)} 条数据\n\n"))
                
                # 生成翻译总表
                self.root.after(0, lambda: self.append_result('table_range_translator', 
                    "正在生成翻译总表...\n"))
                
                success = self.table_range_translator.generate_translation_master_table_multi_lang(
                    output_file)
                
                if success:
                    self.root.after(0, lambda: self.append_result('table_range_translator', 
                        f"✓ 翻译总表已生成: {output_file}\n\n"))
                    
                    # 显示处理报告
                    report = self.table_range_translator.get_processing_report()
                    self.root.after(0, lambda: self.append_result('table_range_translator', 
                        report + "\n"))
                    
                    # 显示成功消息
                    stats = self.table_range_translator.processing_stats
                    msg = (f"多语言翻译提取完成！\n\n"
                          f"处理表格: {stats['processed_tables']}/{stats['total_tables']}\n"
                          f"导出字段: {stats['exported_fields']} 个\n"
                          f"提取数据: {stats['total_rows']} 行\n\n"
                          f"翻译总表已生成:\n{output_file}")
                    self.root.after(0, lambda: messagebox.showinfo("完成", msg))
                else:
                    self.root.after(0, lambda: self.append_result('table_range_translator', 
                        "✗ 生成翻译总表失败\n"))
                    self.root.after(0, lambda: messagebox.showerror("错误", "生成翻译总表失败"))
            else:
                self.root.after(0, lambda: self.append_result('table_range_translator', 
                    "✗ 没有提取到数据\n"))
                self.root.after(0, lambda: messagebox.showwarning("警告", 
                    "没有提取到数据，请检查JSON配置和Excel文件"))
        
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}"
            self.root.after(0, lambda: self.append_result('table_range_translator', 
                f"\n✗ {error_msg}\n"))
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        
        finally:
            # 恢复按钮状态
            self.root.after(0, lambda: self.trt_process_button.config(state="normal"))
            self.root.after(0, lambda: self.status_var.set("就绪"))
    
    def clear_trt_results(self):
        """清空多语言翻译提取结果"""
        self.clear_result('table_range_translator')


def main():
    """主函数"""
    root = tk.Tk()
    app = GameToolsUnified(root)
    
    # 设置窗口关闭事件
    def on_closing():
        root.quit()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 启动主循环
    root.mainloop()


if __name__ == "__main__":
    main()