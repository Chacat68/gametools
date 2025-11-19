"""查看 Excel 文件内容"""
import openpyxl

wb = openpyxl.load_workbook('test_output/split_test_result.xlsx')
ws = wb.active

print('工作表名:', ws.title)
print('行数:', ws.max_row)
print('\n所有行内容:')

for i, row in enumerate(ws.iter_rows(), 1):
    values = [cell.value for cell in row]
    fill_color = ws[f'A{i}'].fill.start_color.rgb if ws[f'A{i}'].fill.start_color else 'None'
    print(f'行{i}: {values[:4]} | 背景色: {fill_color}')
