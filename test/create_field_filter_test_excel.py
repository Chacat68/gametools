#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建测试Excel文件，用于验证字段过滤功能
"""

import openpyxl
from pathlib import Path


def create_test_excel():
    """创建测试Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "角色数据"
    
    # 添加测试数据
    # 第1-4行：其他数据
    ws.append(["游戏角色配置表", "", "", "", "", "", ""])
    ws.append(["版本：1.0", "", "", "", "", "", ""])
    ws.append(["更新时间：2025-01-01", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    
    # 第5行：字段名（这一行会被提取）
    ws.append([
        "ID",                    # 纯数字ID - 应该忽略
        "角色名称",               # 中文 - 应该提取
        "Tên nhân vật",         # 越南文 - 应该提取
        "player_level",         # 纯英文代码 - 应该忽略
        "等级",                  # 中文 - 应该提取
        "HP_MAX",               # 英文常量 - 应该忽略
        "ชื่อ"                  # 泰文 - 应该提取
    ])
    
    # 第6行及以后：数据内容
    ws.append([1, "张三", "Nguyễn Văn A", 10, "初级", 100, "ชื่อหนึ่ง"])
    ws.append([2, "李四", "Trần Thị B", 20, "中级", 200, "ชื่อสอง"])
    ws.append([3, "王五", "Lê Văn C", 30, "高级", 300, "ชื่อสาม"])
    
    # 创建第二个工作表：道具配置
    ws2 = wb.create_sheet("道具数据")
    ws2.append(["道具配置表", "", "", "", ""])
    ws2.append(["", "", "", "", ""])
    ws2.append(["", "", "", "", ""])
    ws2.append(["", "", "", "", ""])
    ws2.append([
        "item_id",              # 纯英文 - 应该忽略
        "道具名称",              # 中文 - 应该提取
        "Tên vật phẩm",        # 越南文 - 应该提取
        "price",               # 纯英文 - 应该忽略
        "描述"                  # 中文 - 应该提取
    ])
    ws2.append([1001, "生命药水", "Thuốc HP", 50, "恢复100点生命值"])
    ws2.append([1002, "魔法药水", "Thuốc MP", 80, "恢复50点魔法值"])
    
    # 创建第三个工作表：纯代码配置（应该完全被忽略）
    ws3 = wb.create_sheet("SystemConfig")
    ws3.append(["Config Data", "", "", ""])
    ws3.append(["", "", "", ""])
    ws3.append(["", "", "", ""])
    ws3.append(["", "", "", ""])
    ws3.append([
        "config_key",
        "config_value",
        "config_type",
        "enabled"
    ])
    ws3.append(["MAX_PLAYER", "1000", "int", "true"])
    ws3.append(["SERVER_NAME", "Game01", "string", "true"])
    
    # 保存文件
    output_dir = Path(__file__).parent / "test_excel_files"
    output_dir.mkdir(exist_ok=True)
    
    output_file = output_dir / "field_filter_test.xlsx"
    wb.save(output_file)
    
    print(f"测试文件已创建: {output_file}")
    print("\n文件结构:")
    print("=" * 60)
    print("工作表1: 角色数据")
    print("  - 应该提取的字段: 角色名称, Tên nhân vật, 等级, ชื่อ")
    print("  - 应该忽略的字段: ID, player_level, HP_MAX")
    print("\n工作表2: 道具数据")
    print("  - 应该提取的字段: 道具名称, Tên vật phẩm, 描述")
    print("  - 应该忽略的字段: item_id, price")
    print("\n工作表3: SystemConfig")
    print("  - 应该完全忽略（纯英文配置）")
    print("=" * 60)


if __name__ == "__main__":
    create_test_excel()
