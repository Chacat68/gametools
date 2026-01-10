#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试数据生成工具 - 合并版
包含所有测试数据生成函数

使用方法:
    python create_test_data.py [选项]
    
选项:
    --all       生成所有测试数据
    --excel     生成基础Excel测试文件
    --field     生成字段提取测试文件
    --filter    生成过滤测试文件
    --mapping   生成映射文件
    --csv       生成CSV映射文件
    --range     生成表范围测试文件
"""

import sys
import os
from pathlib import Path
import argparse

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))


def create_basic_test_excel():
    """创建基础测试Excel文件"""
    import pandas as pd
    
    test_data = [
        ["act_20206_shilian_0.xlsx", "des", 1080601, "神兵天降", "Thần Binh Thiên Giáng"],
        ["act_20206_shilian_0.xlsx", "des", 1090601, "罗刹试炼", "Thí Luyện La Sát"],
        ["act_20206_shilian_0.xlsx", "des", 1090610, "凌虚试炼", "Thí Luyện Lăng Hư"],
        ["act_23201_rank_flower_list.xlsx", "task_name", 5010501, "合服狂欢-花语心愿", "Gộp Server - Hoa Tâm Nguyện"],
        ["act_23201_rank_flower_list.xlsx", "task_name", 5010501, "登录1次", "Đăng nhập 1 lần"],
    ]
    
    columns = ["文件名", "分类", "ID", "中文描述", "越南文描述"]
    df = pd.DataFrame(test_data, columns=columns)
    
    output_dir = Path("test_data")
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "test_data.xlsx"
    df.to_excel(output_file, index=False)
    
    print(f"✅ 基础测试Excel已创建: {output_file}")
    return str(output_file)


def create_field_extractor_test_files():
    """创建字段提取测试文件"""
    import openpyxl
    
    test_dir = Path(__file__).parent / "test_excel_files"
    test_dir.mkdir(parents=True, exist_ok=True)
    
    # 测试文件1：包含中文字段
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "测试表1"
    ws1['A5'], ws1['B5'], ws1['C5'], ws1['D5'], ws1['E5'] = 'ID', '名称', '描述', '数值', '类型'
    data_rows = [
        [1001, '物品A', '这是物品A的描述', 100, '装备'],
        [1002, '物品B', '这是物品B的描述', 200, '消耗品'],
    ]
    for i, row in enumerate(data_rows, start=6):
        for j, value in enumerate(row, start=1):
            ws1.cell(row=i, column=j, value=value)
    wb1.save(test_dir / "测试表1.xlsx")
    
    # 测试文件2：包含越南文字段
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Bảng thử nghiệm"
    ws2['A5'], ws2['B5'], ws2['C5'], ws2['D5'] = 'ID', 'Tên', 'Mô tả', 'Giá trị'
    data_rows2 = [
        [2001, 'Vật phẩm A', 'Đây là mô tả của vật phẩm A', 500],
    ]
    for i, row in enumerate(data_rows2, start=6):
        for j, value in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=value)
    wb2.save(test_dir / "越南文测试表.xlsx")
    
    # 测试文件3：混合类型
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "混合类型表"
    ws3['A5'], ws3['B5'], ws3['C5'], ws3['D5'], ws3['E5'] = '编号', '名称', '数量', '价格', '备注'
    data_rows3 = [['A001', '商品1', 10, 99.5, '热销']]
    for i, row in enumerate(data_rows3, start=6):
        for j, value in enumerate(row, start=1):
            ws3.cell(row=i, column=j, value=value)
    wb3.save(test_dir / "混合类型表.xlsx")
    
    # 测试文件4：多工作表
    wb4 = openpyxl.Workbook()
    ws4_1 = wb4.active
    ws4_1.title = "角色表"
    ws4_1['A5'], ws4_1['B5'], ws4_1['C5'] = '角色ID', '角色名', '等级'
    ws4_2 = wb4.create_sheet("技能表")
    ws4_2['A5'], ws4_2['B5'], ws4_2['C5'] = '技能ID', '技能名称', '消耗MP'
    wb4.save(test_dir / "多工作表测试.xlsx")
    
    print(f"✅ 字段提取测试文件已创建在: {test_dir}")


def create_field_type_test_files():
    """创建字段类型测试文件"""
    import openpyxl
    
    output_dir = Path("test_excel_files")
    output_dir.mkdir(exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试表"
    
    # 第5行：字段名
    ws['A5'], ws['B5'], ws['C5'], ws['D5'], ws['E5'], ws['F5'], ws['G5'] = \
        "c_", "des_cn", "des_vcn", "des_en", "name", "model", "c_"
    
    # 第6行：字段类型
    ws['B6'], ws['C6'], ws['D6'], ws['E6'], ws['F6'] = "策划", "前端", "后端", "前后端", "策划"
    
    # 数据行
    ws['B7'], ws['C7'], ws['D7'], ws['E7'], ws['F7'] = "张若", "Tiểu Long Nữ", "Little Dragon Girl", "npc_001", "0"
    
    wb.save(output_dir / "test_field_types.xlsx")
    print(f"✅ 字段类型测试文件已创建: {output_dir / 'test_field_types.xlsx'}")


def create_filter_test_excel():
    """创建过滤功能测试文件"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "测试表"
    
    # 第5行：字段名（包含需要过滤的字段）
    fields = ['c_', '序号', '索引1', '索引2', 'des_cn', 'des_vcn', 'des', 'name', 'model', 'id', 'c_']
    for idx, field in enumerate(fields, 1):
        ws.cell(row=5, column=idx, value=field)
    
    # 测试数据
    test_data = [
        [1, '张若', '小龙女', 'Tiểu Long Nữ', 'npc104_ui', 0, 1001],
        [2, '黄蓉', '黄蓉', 'Hoàng Dung', 'npc105_ui', 0, 1002],
    ]
    
    for row_idx, data in enumerate(test_data, 6):
        ws.cell(row=row_idx, column=2, value=data[0])
        ws.cell(row=row_idx, column=5, value=data[1])
        ws.cell(row=row_idx, column=6, value=data[2])
        ws.cell(row=row_idx, column=7, value=data[3])
        ws.cell(row=row_idx, column=8, value=data[4])
        ws.cell(row=row_idx, column=9, value=data[5])
        ws.cell(row=row_idx, column=10, value=data[6])
    
    output_path = Path('test_excel_files/test_field_filter.xlsx')
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ 过滤测试文件已创建: {output_path}")


def create_mapping_file():
    """创建映射文件测试数据"""
    import pandas as pd
    
    output_dir = Path("test_data")
    output_dir.mkdir(exist_ok=True)
    
    data = {
        '文件名': ['test1.xlsx', 'test1.xlsx', 'test2.xlsx'],
        '分类': ['des', 'name', 'des'],
        'ID': [1001, 1002, 2001],
        '中文': ['中文内容1', '中文内容2', '中文内容3'],
        '越南文': ['Nội dung 1', 'Nội dung 2', 'Nội dung 3'],
    }
    df = pd.DataFrame(data)
    output_file = output_dir / "mapping.xlsx"
    df.to_excel(output_file, index=False)
    
    print(f"✅ 映射文件已创建: {output_file}")


def create_csv_mapping():
    """创建CSV格式映射文件"""
    import pandas as pd
    
    output_dir = Path("test_data")
    output_dir.mkdir(exist_ok=True)
    
    data = {
        'Table': ['test.xlsx', 'test.xlsx', 'test.xlsx'],
        'Sheet': ['sheet1', 'sheet1', 'sheet1'],
        'Field': ['name', 'name', 'desc'],
        'Position': ['B7', 'B8', 'E7'],
        'ZH': ['测试1', '测试2', '描述1'],
        'VN': ['Test 1', 'Test 2', 'Description 1'],
    }
    df = pd.DataFrame(data)
    output_file = output_dir / "mapping.csv"
    df.to_csv(output_file, index=False, encoding='utf-8-sig')
    
    print(f"✅ CSV映射文件已创建: {output_file}")


def create_table_range_test_files():
    """创建表范围翻译测试文件"""
    import openpyxl
    import json
    
    output_dir = Path("test_table_range")
    output_dir.mkdir(exist_ok=True)
    
    # 创建配置JSON
    config = {
        "tables": [
            {"file": "角色配置.xlsx", "fields": ["name", "desc"]},
            {"file": "物品配置.xlsx", "fields": ["name", "desc"]},
        ]
    }
    with open(output_dir / "field_config.json", 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    # 创建测试Excel
    for table in config['tables']:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A5'], ws['B5'], ws['C5'] = 'ID', 'name', 'desc'
        ws['A6'], ws['B6'], ws['C6'] = 1, '测试名称', '测试描述'
        wb.save(output_dir / table['file'])
    
    print(f"✅ 表范围测试文件已创建在: {output_dir}")


def main():
    parser = argparse.ArgumentParser(description='测试数据生成工具')
    parser.add_argument('--all', action='store_true', help='生成所有测试数据')
    parser.add_argument('--excel', action='store_true', help='生成基础Excel测试文件')
    parser.add_argument('--field', action='store_true', help='生成字段提取测试文件')
    parser.add_argument('--filter', action='store_true', help='生成过滤测试文件')
    parser.add_argument('--mapping', action='store_true', help='生成映射文件')
    parser.add_argument('--csv', action='store_true', help='生成CSV映射文件')
    parser.add_argument('--range', action='store_true', help='生成表范围测试文件')
    parser.add_argument('--types', action='store_true', help='生成字段类型测试文件')
    
    args = parser.parse_args()
    
    # 如果没有指定任何选项，默认生成全部
    if not any([args.all, args.excel, args.field, args.filter, args.mapping, args.csv, args.range, args.types]):
        args.all = True
    
    print("=" * 60)
    print("测试数据生成工具")
    print("=" * 60)
    
    if args.all or args.excel:
        create_basic_test_excel()
    if args.all or args.field:
        create_field_extractor_test_files()
    if args.all or args.types:
        create_field_type_test_files()
    if args.all or args.filter:
        create_filter_test_excel()
    if args.all or args.mapping:
        create_mapping_file()
    if args.all or args.csv:
        create_csv_mapping()
    if args.all or args.range:
        create_table_range_test_files()
    
    print("\n" + "=" * 60)
    print("✅ 测试数据生成完成")
    print("=" * 60)


if __name__ == "__main__":
    main()
