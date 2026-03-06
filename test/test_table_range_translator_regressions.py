#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试多语言翻译提取器的回归问题
"""

import json
import sys
import tempfile
from pathlib import Path

import openpyxl

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.table_range_translator import TableRangeTranslator, TableRangeTranslatorError


def create_translation_excel(file_path: Path, text_value: str):
    """创建最小可提取的测试Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=5, column=1, value="id")
    ws.cell(row=5, column=3, value="desc")
    ws.cell(row=6, column=1, value="int")
    ws.cell(row=6, column=3, value="前端")
    ws.cell(row=7, column=1, value=1001)
    ws.cell(row=7, column=3, value=text_value)
    wb.save(file_path)


def create_merged_json(file_path: Path):
    """创建最小合并JSON配置。"""
    payload = {
        "ZH": {
            "text_tables": [
                {
                    "table_name": "items.xlsx",
                    "sheet_name": "Sheet1",
                    "fields_with_examples": ["desc,前端,C"],
                    "field_column_letters": ["C"],
                }
            ],
            "no_text_tables": []
        },
        "VN": {
            "text_tables": [
                {
                    "table_name": "items.xlsx",
                    "sheet_name": "Sheet1",
                    "fields_with_examples": ["desc,前端,C"],
                    "field_column_letters": ["C"],
                }
            ],
            "no_text_tables": []
        }
    }
    file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def test_translator_state_resets_between_runs():
    """重复执行同一实例时，统计信息不应累加。"""
    print("=" * 60)
    print("测试翻译提取器状态重置")
    print("=" * 60)

    translator = TableRangeTranslator()

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        zh_dir = temp_path / "zh"
        vn_dir = temp_path / "vn"
        zh_dir.mkdir()
        vn_dir.mkdir()

        create_translation_excel(zh_dir / "items.xlsx", "中文内容")
        create_translation_excel(vn_dir / "items.xlsx", "Noi dung tieng Viet")

        merged_json = temp_path / "merged.json"
        create_merged_json(merged_json)

        lang_dirs = {'zh': str(zh_dir), 'vn': str(vn_dir)}

        first_results = translator.process_with_merged_json(str(merged_json), lang_dirs)
        second_results = translator.process_with_merged_json(str(merged_json), lang_dirs)

        assert len(first_results) == 1, f"第一次提取结果异常: {len(first_results)}"
        assert len(second_results) == 1, f"第二次提取结果异常: {len(second_results)}"
        assert translator.processing_stats['total_tables'] == 1, translator.processing_stats
        assert translator.processing_stats['processed_tables'] == 1, translator.processing_stats
        assert translator.processing_stats['exported_fields'] == 1, translator.processing_stats
        assert translator.processing_stats['total_rows'] == 1, translator.processing_stats

    print("✅ 重复执行后统计信息未累加")
    return True


def test_translator_raises_on_fatal_error():
    """致命错误应上抛，而不是伪装成空结果。"""
    print("\n" + "=" * 60)
    print("测试翻译提取器错误上抛")
    print("=" * 60)

    translator = TableRangeTranslator()

    with tempfile.TemporaryDirectory() as temp_dir:
        bad_json = Path(temp_dir) / "bad.json"
        bad_json.write_text("{not valid json", encoding="utf-8")

        try:
            translator.process_with_merged_json(str(bad_json), {'zh': temp_dir})
        except TableRangeTranslatorError as exc:
            print(f"✅ 正确抛出异常: {exc}")
            assert translator.error_logs, "异常发生后应记录错误日志"
            return True

    print("❌ 预期抛出 TableRangeTranslatorError")
    return False


def main():
    results = []
    results.append(("状态重置", test_translator_state_resets_between_runs()))
    results.append(("错误上抛", test_translator_raises_on_fatal_error()))

    print("\n" + "=" * 60)
    print("测试结果汇总")
    print("=" * 60)

    for name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"{status} - {name}")

    passed = sum(1 for _, result in results if result)
    return 0 if passed == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())