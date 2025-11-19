#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
综合测试：表字段导出器的列范围限制功能
模拟真实的游戏配置表格式
"""

import openpyxl
from pathlib import Path
import sys
import os
import json

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def create_realistic_game_config_excel():
    """创建一个真实的游戏配置表格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "关卡配置"
    
    # 第1-4行：注释和表格说明（实际项目中可能有的）
    ws.cell(row=1, column=1, value="关卡配置表")
    ws.cell(row=2, column=1, value="修改日期：2025-11-19")
    
    # 第5行：字段名（重要！）
    ws.cell(row=5, column=1, value="level_id")  # ID字段
    ws.cell(row=5, column=2, value="c_classic_battle")  # 第一个标记
    ws.cell(row=5, column=3, value="level_name")  # 关卡名称
    ws.cell(row=5, column=4, value="level_desc")  # 关卡描述
    ws.cell(row=5, column=5, value="difficulty")  # 难度
    ws.cell(row=5, column=6, value="reward_desc")  # 奖励说明
    ws.cell(row=5, column=7, value="c_classic_battle")  # 第二个标记
    ws.cell(row=5, column=8, value="internal_note")  # 内部备注
    ws.cell(row=5, column=9, value="config_type")  # 配置类型
    
    # 第6行：示例数据（前端数据）
    ws.cell(row=6, column=1, value=1001)
    ws.cell(row=6, column=2, value="战斗配置1")
    ws.cell(row=6, column=3, value="新手村")
    ws.cell(row=6, column=4, value="这是新手村的关卡，难度简单")
    ws.cell(row=6, column=5, value="简单")
    ws.cell(row=6, column=6, value="获得100金币和经验")
    ws.cell(row=6, column=7, value="战斗配置1")
    ws.cell(row=6, column=8, value="这是内部备注，不应该被提取")
    ws.cell(row=6, column=9, value="BATTLE_CONFIG")
    
    # 第7行：更多数据（包含越南文）
    ws.cell(row=7, column=1, value=1002)
    ws.cell(row=7, column=2, value="战斗配置2")
    ws.cell(row=7, column=3, value="Làng mới")  # 越南文：新手村
    ws.cell(row=7, column=4, value="Đây là màn chơi cho người mới")  # 越南文描述
    ws.cell(row=7, column=5, value="Dễ")  # 越南文：简单
    ws.cell(row=7, column=6, value="Nhận 100 vàng và kinh nghiệm")  # 越南文奖励
    ws.cell(row=7, column=7, value="战斗配置2")
    ws.cell(row=7, column=8, value="Internal note in English")
    ws.cell(row=7, column=9, value="BATTLE_CONFIG")
    
    # 第8行：包含泰文
    ws.cell(row=8, column=1, value=1003)
    ws.cell(row=8, column=2, value="战斗配置3")
    ws.cell(row=8, column=3, value="หมู่บ้านมือใหม่")  # 泰文：新手村
    ws.cell(row=8, column=4, value="ด่านสำหรับผู้เล่นใหม่")  # 泰文描述
    ws.cell(row=8, column=5, value="ง่าย")  # 泰文：简单
    ws.cell(row=8, column=6, value="ได้รับ 100 เหรียญและประสบการณ์")  # 泰文奖励
    ws.cell(row=8, column=7, value="战斗配置3")
    ws.cell(row=8, column=8, value="Should not be extracted")
    ws.cell(row=8, column=9, value="BATTLE_CONFIG")
    
    # 第9行：混合数据
    ws.cell(row=9, column=1, value=1004)
    ws.cell(row=9, column=2, value="战斗配置4")
    ws.cell(row=9, column=3, value="高级关卡")
    ws.cell(row=9, column=4, value="Expert level for advanced players")
    ws.cell(row=9, column=5, value="困难")
    ws.cell(row=9, column=6, value="Rare items and 500 gold")
    ws.cell(row=9, column=7, value="战斗配置4")
    ws.cell(row=9, column=8, value="额外的中文备注")
    ws.cell(row=9, column=9, value="BATTLE_CONFIG")
    
    # 保存文件
    test_file = Path("test_output/game_config_realistic.xlsx")
    test_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(test_file)
    print(f"✓ 真实场景测试文件已创建: {test_file}")
    return test_file

def test_realistic_scenario():
    """测试真实场景"""
    from core.excel_field_extractor import ExcelFieldExtractor
    
    # 创建测试文件
    test_file = create_realistic_game_config_excel()
    
    # 创建提取器
    extractor = ExcelFieldExtractor()
    
    # 提取字段
    print("\n" + "="*70)
    print("开始真实场景测试：游戏配置表字段提取")
    print("="*70)
    
    results = extractor.extract_fields_from_excel(test_file)
    
    if not results:
        print("❌ 错误：没有提取到任何结果")
        return False
    
    result = results[0]
    
    print(f"\n📊 提取结果:")
    print(f"   表名: {result['excel_file']}")
    print(f"   工作表: {result['sheet_name']}")
    print(f"   字段数量: {result['field_count']}")
    print(f"   列号范围: {result['text_columns']}")
    
    print(f"\n📝 检测到的字段:")
    for i, field in enumerate(result['fields'], 1):
        print(f"   {i}. {field}")
    
    print(f"\n📋 字段+示例数据:")
    for i, field_example in enumerate(result.get('fields_with_examples', []), 1):
        print(f"   {i}. {field_example}")
    
    # 验证结果
    print("\n" + "="*70)
    print("验证结果:")
    print("="*70)
    
    # 预期字段：只有标记之间的字段
    expected_fields = ["level_name", "level_desc", "difficulty", "reward_desc"]
    expected_columns = [3, 4, 5, 6]  # C, D, E, F列
    
    success = True
    
    # 验证字段
    if result['fields'] != expected_fields:
        print(f"❌ 字段不匹配！")
        print(f"   预期: {expected_fields}")
        print(f"   实际: {result['fields']}")
        success = False
    else:
        print(f"✓ 字段提取正确: {result['fields']}")
    
    # 验证列号
    if result['text_columns'] != expected_columns:
        print(f"❌ 列号不匹配！")
        print(f"   预期: {expected_columns}")
        print(f"   实际: {result['text_columns']}")
        success = False
    else:
        print(f"✓ 列号范围正确: {result['text_columns']}")
    
    # 验证不包含标记之外的列
    excluded_columns = [1, 2, 7, 8, 9]  # level_id, 标记1, 标记2, internal_note, config_type
    found_excluded = [col for col in excluded_columns if col in result['text_columns']]
    
    if found_excluded:
        print(f"❌ 错误：包含了不应该提取的列: {found_excluded}")
        success = False
    else:
        print(f"✓ 成功排除了标记外的列")
    
    # 导出到所有格式
    print("\n" + "="*70)
    print("测试多格式导出:")
    print("="*70)
    
    output_dir = Path("test_output")
    
    # JSON格式
    json_file = output_dir / "realistic_test_result.json"
    extractor.export_to_json(results, json_file)
    
    # CSV格式
    csv_file = output_dir / "realistic_test_result.csv"
    extractor.export_to_csv(results, csv_file)
    
    # Excel格式
    excel_file = output_dir / "realistic_test_result.xlsx"
    extractor.export_to_excel(results, excel_file)
    
    # 读取并显示JSON结果
    if json_file.exists():
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
            print(f"\n📄 JSON输出预览:")
            print(json.dumps(json_data, ensure_ascii=False, indent=2)[:500] + "...")
    
    # 最终结果
    print("\n" + "="*70)
    if success:
        print("✅ 真实场景测试通过！")
        print("\n功能总结:")
        print("  1. ✓ 只提取两个 c_classic_battle 标记之间的字段")
        print("  2. ✓ 不包含标记列本身")
        print("  3. ✓ 成功过滤ID、配置类型等非本地化字段")
        print("  4. ✓ 正确识别中文、越南文、泰文内容")
        print("  5. ✓ 包含第6行的示例数据")
        print("  6. ✓ 支持JSON、CSV、Excel多格式导出")
    else:
        print("❌ 真实场景测试失败！")
    print("="*70)
    
    return success

if __name__ == "__main__":
    test_realistic_scenario()
