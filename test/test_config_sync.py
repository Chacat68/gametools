#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel配置同步工具测试脚本
"""

import os
import sys
import json
import shutil
from pathlib import Path

# 添加模块路径
sys.path.insert(0, str(Path(__file__).parent))

from core.excel_config_sync import ExcelConfigSync


def create_test_directories():
    """创建测试目录和文件"""
    test_base = Path("test_config_sync")
    
    # 创建目录结构
    source_dir = test_base / "source"
    target1_dir = test_base / "target1"
    target2_dir = test_base / "target2"
    
    for d in [source_dir, target1_dir, target2_dir]:
        d.mkdir(parents=True, exist_ok=True)
    
    # 创建测试Excel文件
    import pandas as pd
    from openpyxl import Workbook
    
    # 创建带字段行的Excel文件（字段在第5行）
    def create_excel_with_fields(path, data, fields):
        wb = Workbook()
        ws = wb.active
        
        # 写入前4行空数据（模拟表头区域）
        for i in range(1, 5):
            ws.cell(row=i, column=1, value=f"Header{i}")
        
        # 第5行写入字段名
        for col, field in enumerate(fields, 1):
            ws.cell(row=5, column=col, value=field)
        
        # 第6行开始写入数据
        for row_idx, row_data in enumerate(data, 6):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(path)
    
    # 源文件数据
    source_data = [
        [1, '源-A', 100, '描述1'],
        [2, '源-B', 200, '描述2'],
        [3, '源-C', 300, '描述3']
    ]
    fields = ['ID', 'Name', 'Value', 'Description']
    
    create_excel_with_fields(source_dir / "test1.xlsx", source_data, fields)
    create_excel_with_fields(source_dir / "test2.xlsx", source_data, fields)
    
    # 目标文件1（旧数据）
    target1_data = [
        [1, '目标1-A', 10, '旧描述1'],
        [2, '目标1-B', 20, '旧描述2'],
        [3, '目标1-C', 30, '旧描述3']
    ]
    create_excel_with_fields(target1_dir / "test1.xlsx", target1_data, fields)
    create_excel_with_fields(target1_dir / "test2.xlsx", target1_data, fields)
    
    # 目标文件2（旧数据）
    target2_data = [
        [1, '目标2-A', 11, '旧描述2-1'],
        [2, '目标2-B', 22, '旧描述2-2'],
        [3, '目标2-C', 33, '旧描述2-3']
    ]
    create_excel_with_fields(target2_dir / "test1.xlsx", target2_data, fields)
    
    # 创建过滤配置文件
    filter_config = {
        "skip_fields": {
            "test1.xlsx": ["Name", "Description"],
            "test2.xlsx": ["Description"]
        }
    }
    
    filter_path = test_base / "filter_config.json"
    with open(filter_path, 'w', encoding='utf-8') as f:
        json.dump(filter_config, f, ensure_ascii=False, indent=2)
    
    print(f"测试目录已创建: {test_base}")
    print(f"  源目录: {source_dir}")
    print(f"  目标目录1: {target1_dir}")
    print(f"  目标目录2: {target2_dir}")
    print(f"  过滤配置: {filter_path}")
    
    return str(source_dir), str(target1_dir), str(target2_dir), str(filter_path)


def test_config_sync():
    """测试配置同步功能"""
    print("\n" + "=" * 60)
    print("Excel配置同步工具测试")
    print("=" * 60)
    
    # 创建测试数据
    source_dir, target1_dir, target2_dir, filter_path = create_test_directories()
    
    # 创建同步器
    syncer = ExcelConfigSync()
    
    # 设置进度回调
    def progress_callback(msg, percentage=None):
        if percentage is not None:
            print(f"[{percentage:.1f}%] {msg}")
        else:
            print(f"[...] {msg}")
    
    syncer.set_progress_callback(progress_callback)
    
    # 测试文件匹配
    print("\n1. 测试文件匹配...")
    matching = syncer.find_matching_files(source_dir, [target1_dir, target2_dir])
    print(f"   匹配的文件: {list(matching.keys())}")
    
    # 测试过滤配置加载
    print("\n2. 加载过滤配置...")
    syncer.load_filter_config(filter_path)
    print(f"   过滤规则数: {len(syncer.skip_fields)}")
    for table, fields in syncer.skip_fields.items():
        if '.xlsx' in table:  # 只显示完整表名
            print(f"   - {table}: 跳过字段 {fields}")
    
    # 执行同步（带过滤）
    print("\n3. 执行同步（带过滤）...")
    stats = syncer.sync_directories(
        source_dir=source_dir,
        target_dir1=target1_dir,
        target_dir2=target2_dir
    )
    
    # 显示统计
    print("\n4. 同步统计:")
    print(syncer.get_stats_summary())
    
    # 生成报告
    report_path = "test_config_sync/sync_report.xlsx"
    print(f"\n5. 生成报告: {report_path}")
    syncer.generate_sync_report(report_path)
    
    # 验证同步结果
    print("\n6. 验证同步结果...")
    from openpyxl import load_workbook
    
    # test1.xlsx 应该跳过 Name 和 Description 字段
    wb = load_workbook(os.path.join(target1_dir, "test1.xlsx"))
    ws = wb.active
    
    # 第6行是第一行数据，第2列是Name，第4列是Description
    name_value = ws.cell(row=6, column=2).value
    value_value = ws.cell(row=6, column=3).value
    desc_value = ws.cell(row=6, column=4).value
    
    print(f"   test1.xlsx 第一行:")
    print(f"     Name (应保持原值): {name_value}")
    print(f"     Value (应同步): {value_value}")
    print(f"     Description (应保持原值): {desc_value}")
    
    # 验证：Name应该保持原值，Value应该被同步
    if name_value == '目标1-A' and value_value == 100:
        print("   ✓ 过滤同步成功！Name保持原值，Value被同步")
    else:
        print("   ✗ 同步结果不符合预期！")
    
    wb.close()
    
    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)


def cleanup_test_data():
    """清理测试数据"""
    test_base = Path("test_config_sync")
    if test_base.exists():
        shutil.rmtree(test_base)
        print("测试目录已清理")


if __name__ == "__main__":
    try:
        test_config_sync()
    finally:
        # 可选：清理测试数据
        # cleanup_test_data()
        pass
