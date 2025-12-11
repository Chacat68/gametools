#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建测试Excel文件用于验证字段类型提取
"""

import openpyxl
from pathlib import Path

def create_test_excel():
    """创建包含字段类型的测试Excel文件"""
    
    output_dir = Path("test_excel_files")
    output_dir.mkdir(exist_ok=True)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试表"
    
    # 第1行：表头说明
    ws['A1'] = "序号"
    ws['B1'] = "配置说明"
    
    # 第5行：字段名（前后有 c_ 标记）
    ws['A5'] = "c_"
    ws['B5'] = "des_cn"
    ws['C5'] = "des_vcn"
    ws['D5'] = "des_en"
    ws['E5'] = "name"
    ws['F5'] = "model"
    ws['G5'] = "c_"
    
    # 第6行：字段类型
    ws['A6'] = ""
    ws['B6'] = "策划"
    ws['C6'] = "前端"
    ws['D6'] = "后端"
    ws['E6'] = "前后端"
    ws['F6'] = "策划"
    ws['G6'] = ""
    
    # 第7行及以后：数据
    ws['A7'] = "1"
    ws['B7'] = "张若"
    ws['C7'] = "Tiểu Long Nữ"
    ws['D7'] = "Little Dragon Girl"
    ws['E7'] = "npc_001"
    ws['F7'] = "0"
    ws['G7'] = ""
    
    ws['A8'] = "2"
    ws['B8'] = "黄蓉"
    ws['C8'] = "Hoàng Dung"
    ws['D8'] = "Huang Rong"
    ws['E8'] = "npc_002"
    ws['F8'] = "1"
    ws['G8'] = ""
    
    # 保存文件
    output_file = output_dir / "test_field_types.xlsx"
    wb.save(output_file)
    print(f"✅ 创建测试文件: {output_file}")
    
    # 创建第二个测试文件（无字段类型）
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "无类型表"
    
    ws2['A5'] = "c_"
    ws2['B5'] = "desc"
    ws2['C5'] = "c_"
    
    # 第6行为空（测试缺失字段类型的情况）
    ws2['A6'] = ""
    ws2['B6'] = ""
    ws2['C6'] = ""
    
    ws2['A7'] = "1"
    ws2['B7'] = "测试描述"
    ws2['C7'] = ""
    
    output_file2 = output_dir / "test_no_types.xlsx"
    wb2.save(output_file2)
    print(f"✅ 创建测试文件: {output_file2}")
    
    print("\n📋 测试文件结构:")
    print("  - test_field_types.xlsx: 包含完整的字段类型")
    print("  - test_no_types.xlsx: 缺失字段类型（用于测试警告）")

if __name__ == "__main__":
    create_test_excel()
