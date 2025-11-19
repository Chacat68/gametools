#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建测试Excel文件用于字段导出测试
"""

import pandas as pd
import openpyxl
from pathlib import Path


def create_test_excel_files():
    """创建测试Excel文件"""
    test_dir = Path(__file__).parent.parent / "test_excel_files"
    test_dir.mkdir(parents=True, exist_ok=True)
    
    # 测试文件1：包含中文字段
    print("创建测试文件1：中文字段...")
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "测试表1"
    
    # 写入数据（第5行是字段名）
    ws1['A1'] = '表说明'
    ws1['A2'] = '策划：张三'
    ws1['A3'] = '版本：1.0'
    ws1['A4'] = ''
    ws1['A5'] = 'ID'
    ws1['B5'] = '名称'
    ws1['C5'] = '描述'
    ws1['D5'] = '数值'
    ws1['E5'] = '类型'
    
    # 填充一些数据行
    data_rows = [
        [1001, '物品A', '这是物品A的描述', 100, '装备'],
        [1002, '物品B', '这是物品B的描述', 200, '消耗品'],
        [1003, '物品C', '这是物品C的描述', 300, '材料'],
    ]
    
    for i, row in enumerate(data_rows, start=6):
        for j, value in enumerate(row, start=1):
            ws1.cell(row=i, column=j, value=value)
    
    wb1.save(test_dir / "测试表1.xlsx")
    print("  已创建: 测试表1.xlsx")
    
    # 测试文件2：包含越南文字段
    print("创建测试文件2：越南文字段...")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Bảng thử nghiệm"
    
    ws2['A5'] = 'ID'
    ws2['B5'] = 'Tên'
    ws2['C5'] = 'Mô tả'
    ws2['D5'] = 'Giá trị'
    
    data_rows2 = [
        [2001, 'Vật phẩm A', 'Đây là mô tả của vật phẩm A', 500],
        [2002, 'Vật phẩm B', 'Đây là mô tả của vật phẩm B', 600],
    ]
    
    for i, row in enumerate(data_rows2, start=6):
        for j, value in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=value)
    
    wb2.save(test_dir / "越南文测试表.xlsx")
    print("  已创建: 越南文测试表.xlsx")
    
    # 测试文件3：混合文本和纯数字列
    print("创建测试文件3：混合类型...")
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "混合类型表"
    
    ws3['A5'] = '编号'
    ws3['B5'] = '名称'
    ws3['C5'] = '数量'  # 纯数字列
    ws3['D5'] = '价格'  # 纯数字列
    ws3['E5'] = '备注'
    
    data_rows3 = [
        ['A001', '商品1', 10, 99.5, '热销'],
        ['A002', '商品2', 20, 199.0, '新品'],
        ['A003', '商品3', 30, 299.5, '促销中'],
    ]
    
    for i, row in enumerate(data_rows3, start=6):
        for j, value in enumerate(row, start=1):
            ws3.cell(row=i, column=j, value=value)
    
    wb3.save(test_dir / "混合类型表.xlsx")
    print("  已创建: 混合类型表.xlsx")
    
    # 测试文件4：多工作表
    print("创建测试文件4：多工作表...")
    wb4 = openpyxl.Workbook()
    
    # 第一个工作表
    ws4_1 = wb4.active
    ws4_1.title = "角色表"
    ws4_1['A5'] = '角色ID'
    ws4_1['B5'] = '角色名'
    ws4_1['C5'] = '等级'
    ws4_1['D5'] = '职业'
    
    # 第二个工作表
    ws4_2 = wb4.create_sheet("技能表")
    ws4_2['A5'] = '技能ID'
    ws4_2['B5'] = '技能名称'
    ws4_2['C5'] = '消耗MP'
    ws4_2['D5'] = '冷却时间'
    
    # 第三个工作表
    ws4_3 = wb4.create_sheet("装备表")
    ws4_3['A5'] = '装备ID'
    ws4_3['B5'] = '装备名称'
    ws4_3['C5'] = '品质'
    ws4_3['D5'] = '部位'
    
    wb4.save(test_dir / "多工作表测试.xlsx")
    print("  已创建: 多工作表测试.xlsx")
    
    print(f"\n所有测试文件已创建在: {test_dir}")
    print("\n现在可以运行测试:")
    print("  python test/test_field_extractor.py")
    print("或启动GUI:")
    print("  python gui/excel_field_extractor_gui.py")


if __name__ == "__main__":
    create_test_excel_files()
