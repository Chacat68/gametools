#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试新的JSON格式输出
验证字段名+字段类型的提取
"""

import sys
import json
import tempfile
from pathlib import Path

import openpyxl

# 添加项目根目录到路径（确保可导入 core 包）
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.excel_field_extractor import ExcelFieldExtractor


def create_test_excel_files(base_dir: Path) -> Path:
    """创建最小可复现的 JSON 导出测试数据。"""
    base_dir.mkdir(parents=True, exist_ok=True)

    text_file = base_dir / "dialog.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=5, column=2, value="c_")
    ws.cell(row=5, column=3, value="title")
    ws.cell(row=5, column=4, value="desc")
    ws.cell(row=5, column=5, value="c_")
    ws.cell(row=6, column=3, value="前端")
    ws.cell(row=6, column=4, value="后端")
    ws.cell(row=7, column=3, value="你好")
    ws.cell(row=7, column=4, value="欢迎来到游戏")
    ws.cell(row=8, column=3, value="再见")
    ws.cell(row=8, column=4, value="下次再来")
    wb.save(text_file)

    no_text_file = base_dir / "config.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=5, column=2, value="c_")
    ws.cell(row=5, column=3, value="code")
    ws.cell(row=5, column=4, value="value")
    ws.cell(row=5, column=5, value="c_")
    ws.cell(row=6, column=3, value="后端")
    ws.cell(row=6, column=4, value="后端")
    ws.cell(row=7, column=3, value="CFG_001")
    ws.cell(row=7, column=4, value="100")
    wb.save(no_text_file)

    return base_dir

def test_json_format():
    """测试JSON格式输出"""
    print("=" * 60)
    print("测试表字段导出工具 - 新JSON格式")
    print("=" * 60)
    
    # 创建提取器实例
    extractor = ExcelFieldExtractor()

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        test_dir = create_test_excel_files(temp_path / "test_excel_files")
        output_dir = temp_path / "test_output"
        output_dir.mkdir(exist_ok=True)
        
        print(f"\n📁 扫描目录: {test_dir}")
        print(f"📤 输出目录: {output_dir}")
        print(f"📋 输出格式: JSON\n")
        
        stats = extractor.process_directory(
            directory_path=str(test_dir),
            output_folder=str(output_dir),
            output_format='json',
            recursive=True
        )
    
    print("\n" + "=" * 60)
    print("统计信息")
    print("=" * 60)
    print(f"✅ 扫描文件数: {stats['total_files']}")
    print(f"✅ 工作表数: {stats['total_sheets']}")
    print(f"✅ 字段总数: {stats['total_fields']}")
    print(f"✅ 输出文件: {stats['output_file']}")
    
    # 显示日志
    logs = extractor.get_all_logs()
    if logs['warnings']:
        print(f"\n⚠️ 警告数: {len(logs['warnings'])}")
    if logs['errors']:
        print(f"\n❌ 错误数: {len(logs['errors'])}")
    
        # 读取并显示JSON输出
        output_file = Path(stats['output_file'])
        if output_file.exists():
            print("\n" + "=" * 60)
            print("JSON输出预览")
            print("=" * 60)
            with open(output_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            print(f"\n📊 无文本表格数: {len(data['no_text_tables'])}")
            if data['no_text_tables']:
                print("前3个示例:")
                for item in data['no_text_tables'][:3]:
                    print(f"  - {item['table_name']} # {item['sheet_name']}")
            
            print(f"\n📊 包含文本表格数: {len(data['text_tables'])}")
            if data['text_tables']:
                print("前3个示例:")
                for item in data['text_tables'][:3]:
                    print(f"  - {item['table_name']} # {item['sheet_name']}")
                    print(f"    字段数: {item['field_count']}")
                    if item['fields_with_examples']:
                        print(f"    字段+类型: {item['fields_with_examples'][:2]}")

            assert any(item['table_name'] == 'dialog.xlsx' for item in data['text_tables'])
            assert any(item['table_name'] == 'config.xlsx' for item in data['no_text_tables'])
            dialog_table = next(item for item in data['text_tables'] if item['table_name'] == 'dialog.xlsx')
            assert dialog_table['field_count'] == 2
            assert any(entry.startswith('title,前端') for entry in dialog_table['fields_with_examples'])
        else:
            return False
    
    print("\n✅ 测试完成！")
    return True

if __name__ == "__main__":
    success = test_json_format()
    sys.exit(0 if success else 1)
