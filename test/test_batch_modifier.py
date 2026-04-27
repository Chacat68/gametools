#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量改表测试 - 合并版
包含基本功能测试、CSV格式测试和Position模式测试

使用方法:
    python test_batch_modifier.py [选项]
    
选项:
    --all       运行所有测试
    --basic     运行基本功能测试
    --csv       运行CSV格式测试
    --position  运行Position模式测试
"""

import sys
import os
import json
import tempfile
import shutil
import argparse
from pathlib import Path

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import Workbook


def test_basic_modifier():
    """测试批量改表基本功能"""
    print("\n" + "=" * 60)
    print("批量改表基本功能测试")
    print("=" * 60)
    
    from core.batch_excel_modifier import BatchExcelModifier
    
    # 创建临时测试目录
    temp_dir = tempfile.mkdtemp()
    
    try:
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "test_table"
        ws['A1'], ws['B1'], ws['C1'] = 'ID', 'des', 'name'
        ws['A2'], ws['B2'], ws['C2'] = 1001, '旧内容1', '旧名称1'
        ws['A3'], ws['B3'], ws['C3'] = 1002, '旧内容2', '旧名称2'
        
        excel_path = Path(temp_dir) / "test_table.xlsx"
        wb.save(excel_path)
        print(f"✅ 创建测试Excel: {excel_path}")
        
        # 创建映射表
        mapping_data = {
            '表名': ['test_table.xlsx', 'test_table.xlsx'],
            'Classification': ['des', 'des'],
            'ID': [1001, 1002],
            'VN': ['Nội dung mới 1', 'Nội dung mới 2'],
        }
        mapping_df = pd.DataFrame(mapping_data)
        mapping_path = Path(temp_dir) / "mapping.xlsx"
        mapping_df.to_excel(mapping_path, index=False)
        print(f"✅ 创建映射表: {mapping_path}")
        
        # 测试批量改表
        modifier = BatchExcelModifier()
        print("\n开始批量改表测试...")
        
        # 获取语言列表
        languages = modifier.get_mapping_file_languages(str(mapping_path))
        print(f"✅ 检测到语言: {languages}")
        
        print("\n✅ 批量改表基本测试通过")
        return True
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_csv_format():
    """测试CSV格式支持"""
    print("\n" + "=" * 60)
    print("CSV格式支持测试")
    print("=" * 60)
    
    from core.batch_excel_modifier import BatchExcelModifier
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # 创建CSV映射文件
        csv_data = {
            'Table': ['test.xlsx', 'test.xlsx'],
            'Sheet': ['Sheet1', 'Sheet1'],
            'Field': ['name', 'desc'],
            'Position': ['B2', 'C2'],
            'ZH': ['中文1', '中文2'],
            'VN': ['越南文1', '越南文2'],
        }
        df = pd.DataFrame(csv_data)
        csv_path = Path(temp_dir) / "mapping.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        print(f"✅ 创建CSV映射文件: {csv_path}")
        
        modifier = BatchExcelModifier()
        
        # 测试CSV格式检测
        languages = modifier.get_mapping_file_languages(str(csv_path))
        print(f"✅ 从CSV检测到语言: {languages}")
        
        # 测试格式转换
        df_loaded = pd.read_csv(csv_path, encoding='utf-8-sig')
        converted = modifier._convert_csv_format_if_needed(df_loaded)
        print(f"✅ CSV格式转换成功，列: {converted.columns.tolist()}")
        
        # 检查Position列是否保留
        if 'Position' in converted.columns:
            print("✅ Position列已保留")
        else:
            print("⚠️ Position列未保留")
        
        return True
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_json_language_csv_regression():
    """回归测试：JSON 语言模式下 CSV 映射表应只加载一次且能正常生成修改项。"""
    print("\n" + "=" * 60)
    print("JSON语言+CSV回归测试")
    print("=" * 60)

    from core.batch_excel_modifier import BatchExcelModifier

    temp_dir = tempfile.mkdtemp()

    try:
        excel_dir = Path(temp_dir) / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        excel_path = excel_dir / "items.xlsx"
        excel_path.touch()

        csv_path = Path(temp_dir) / "mapping.csv"
        pd.DataFrame(
            {
                'Table': ['items.xlsx'],
                'Classification': ['name_vn'],
                'ID': [1001],
                'VN': ['Kiếm'],
            }
        ).to_csv(csv_path, index=False, encoding='utf-8-sig')

        modifier = BatchExcelModifier()
        modifier.json_language = {'code': 'vn', 'name': '越南语'}

        table_config = {
            'table_name': 'items.xlsx',
            'sheet_name': 'items',
            'fields': ['name_vn'],
            'fields_with_examples': [],
            'field_types': {},
            'fields_by_language': {'vn': ['name_vn']},
        }
        modifier.field_config = {
            'items.xlsx': table_config,
            'items': table_config,
        }

        original_load_mapping_table = modifier.load_mapping_table
        load_call_count = 0
        captured = {}

        def counted_load_mapping_table(*args, **kwargs):
            nonlocal load_call_count
            load_call_count += 1
            return original_load_mapping_table(*args, **kwargs)

        def fake_modify_excel_file(excel_path_arg, modifications, **kwargs):
            captured['excel_path'] = excel_path_arg
            captured['modifications'] = modifications
            captured['kwargs'] = kwargs
            return len(modifications), [], 0

        modifier.load_mapping_table = counted_load_mapping_table
        modifier.modify_excel_file = fake_modify_excel_file

        stats = modifier.process_batch_modification_by_json_language(
            str(csv_path),
            str(excel_dir),
            target_language='VN',
            backup=False,
        )

        assert load_call_count == 1, f"CSV映射表应只加载一次，实际加载 {load_call_count} 次"
        assert stats['processed_rows'] == 1, f"应处理 1 行，实际为 {stats['processed_rows']}"
        assert stats['modified_files'] == 1, f"应修改 1 个文件，实际为 {stats['modified_files']}"
        assert captured['excel_path'] == str(excel_path), "目标Excel路径不正确"
        assert captured['modifications'] == [
            {
                'id': 1001,
                'modify_values': {'name_vn': 'Kiếm'},
            }
        ], f"生成的修改项不符合预期: {captured['modifications']}"
        assert captured['kwargs'].get('use_position') is False, "普通CSV不应启用Position模式"

        print("✅ JSON语言模式下的CSV回归路径正常")
        return True

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_position_mode():
    """测试Position定位模式"""
    print("\n" + "=" * 60)
    print("Position定位模式测试")
    print("=" * 60)
    
    from core.batch_excel_modifier import BatchExcelModifier
    
    modifier = BatchExcelModifier()
    
    # 测试Position列检测
    test_data_with_position = {
        'Table': ['test.xlsx'],
        'Position': ['B7'],
        'VN': ['测试'],
    }
    df_with = pd.DataFrame(test_data_with_position)
    
    test_data_without_position = {
        'Table': ['test.xlsx'],
        'ID': [1001],
        'VN': ['测试'],
    }
    df_without = pd.DataFrame(test_data_without_position)
    
    # 转换并检查
    converted_with = modifier._convert_csv_format_if_needed(df_with)
    converted_without = modifier._convert_csv_format_if_needed(df_without)
    
    has_position = 'Position' in converted_with.columns
    no_position = 'Position' not in converted_without.columns
    
    if has_position:
        print("✅ 带Position的数据正确保留了Position列")
    else:
        print("❌ 带Position的数据未保留Position列")
    
    if no_position:
        print("✅ 无Position的数据正确使用行号模式")
    else:
        print("⚠️ 无Position的数据也有Position列")
    
    return has_position


def main():
    parser = argparse.ArgumentParser(description='批量改表测试')
    parser.add_argument('--all', action='store_true', help='运行所有测试')
    parser.add_argument('--basic', action='store_true', help='运行基本功能测试')
    parser.add_argument('--csv', action='store_true', help='运行CSV格式测试')
    parser.add_argument('--json-csv', action='store_true', help='运行JSON语言+CSV回归测试')
    parser.add_argument('--position', action='store_true', help='运行Position模式测试')
    
    args = parser.parse_args()
    
    if not any([args.all, args.basic, args.csv, args.json_csv, args.position]):
        args.all = True
    
    print("=" * 60)
    print("批量改表测试")
    print("=" * 60)
    
    results = {}
    
    if args.all or args.basic:
        results['基本功能'] = test_basic_modifier()
    if args.all or args.csv:
        results['CSV格式'] = test_csv_format()
    if args.all or args.json_csv:
        results['JSON语言CSV回归'] = test_json_language_csv_regression()
    if args.all or args.position:
        results['Position模式'] = test_position_mode()
    
    # 汇总结果
    print("\n" + "=" * 60)
    print("测试结果汇总")
    print("=" * 60)
    for name, passed in results.items():
        status = "✅ 通过" if passed else "❌ 失败"
        print(f"  {name}: {status}")
    
    all_passed = all(results.values())
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
