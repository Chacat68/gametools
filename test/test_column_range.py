#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试表字段导出器的列范围限制功能
验证是否只扫描两个 c_classic_battle 之间的列
"""

import openpyxl
from pathlib import Path
import sys
import os

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def create_test_excel():
    """创建测试用的Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试工作表"
    
    # 第5行：字段名
    ws.cell(row=5, column=1, value="id")  # 列A - 在第一个标记之前
    ws.cell(row=5, column=2, value="c_classic_battle")  # 列B - 第一个标记
    ws.cell(row=5, column=3, value="名称")  # 列C - 应该被扫描
    ws.cell(row=5, column=4, value="描述")  # 列D - 应该被扫描
    ws.cell(row=5, column=5, value="类型")  # 列E - 应该被扫描
    ws.cell(row=5, column=6, value="c_classic_battle")  # 列F - 第二个标记
    ws.cell(row=5, column=7, value="额外字段")  # 列G - 在第二个标记之后
    
    # 第6行：示例数据
    ws.cell(row=6, column=1, value=1001)  # 不应被扫描
    ws.cell(row=6, column=2, value="标记1")
    ws.cell(row=6, column=3, value="测试名称")  # 应该被检测到
    ws.cell(row=6, column=4, value="这是描述")  # 应该被检测到
    ws.cell(row=6, column=5, value="战斗类型")  # 应该被检测到
    ws.cell(row=6, column=6, value="标记2")
    ws.cell(row=6, column=7, value="不应被扫描")  # 不应被扫描
    
    # 第7行：更多数据
    ws.cell(row=7, column=1, value=1002)
    ws.cell(row=7, column=2, value="标记1")
    ws.cell(row=7, column=3, value="越南测试")  # 应该被检测到
    ws.cell(row=7, column=4, value="Mô tả tiếng Việt")  # 应该被检测到（越南文）
    ws.cell(row=7, column=5, value="泰文测试 ภาษาไทย")  # 应该被检测到（泰文）
    ws.cell(row=7, column=6, value="标记2")
    ws.cell(row=7, column=7, value="外部中文")  # 不应被扫描
    
    # 保存文件
    test_file = Path("test_output/test_column_range.xlsx")
    test_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(test_file)
    print(f"✓ 测试文件已创建: {test_file}")
    return test_file

def test_column_range():
    """测试列范围限制功能"""
    from core.excel_field_extractor import ExcelFieldExtractor
    
    # 创建测试文件
    test_file = create_test_excel()
    
    # 创建提取器
    extractor = ExcelFieldExtractor()
    
    # 提取字段
    print("\n开始测试字段提取...")
    results = extractor.extract_fields_from_excel(test_file)
    
    # 验证结果
    print("\n" + "="*60)
    print("测试结果:")
    print("="*60)
    
    if not results:
        print("❌ 错误：没有提取到任何结果")
        return False
    
    result = results[0]
    print(f"表名: {result['excel_file']}")
    print(f"工作表: {result['sheet_name']}")
    print(f"检测到的字段: {result['fields']}")
    print(f"字段数量: {result['field_count']}")
    print(f"列号范围: {result['text_columns']}")
    
    # 验证字段
    expected_fields = ["名称", "描述", "类型"]
    actual_fields = result['fields']
    
    print("\n预期字段:", expected_fields)
    print("实际字段:", actual_fields)
    
    # 检查是否只包含预期字段
    success = True
    
    if set(actual_fields) != set(expected_fields):
        print("❌ 字段不匹配！")
        success = False
    else:
        print("✓ 字段匹配！")
    
    # 检查列号范围（应该是3, 4, 5，即C, D, E列）
    expected_columns = [3, 4, 5]
    if result['text_columns'] != expected_columns:
        print(f"❌ 列号范围不正确！预期 {expected_columns}，实际 {result['text_columns']}")
        success = False
    else:
        print(f"✓ 列号范围正确！{result['text_columns']}")
    
    # 检查是否包含了标记之外的列
    if 1 in result['text_columns']:
        print("❌ 错误：包含了第一个标记之前的列（列1: id）")
        success = False
    
    if 7 in result['text_columns']:
        print("❌ 错误：包含了第二个标记之后的列（列7: 额外字段）")
        success = False
    
    if success:
        print("\n" + "="*60)
        print("✅ 所有测试通过！列范围限制功能正常工作。")
        print("="*60)
    else:
        print("\n" + "="*60)
        print("❌ 测试失败！")
        print("="*60)
    
    return success

if __name__ == "__main__":
    test_column_range()
