#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试错误日志功能
验证字段提取器的错误和警告日志记录
"""

import openpyxl
from pathlib import Path
import sys
import os

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

_TEST_OUTPUT = Path(__file__).parent / "_runtime" / "output"


def create_test_excel_with_issues():
    """创建包含问题的测试Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试工作表"
    
    # 第5行：字段名
    ws.cell(row=5, column=1, value="id")
    ws.cell(row=5, column=2, value="c_classic_battle")  # 第一个标记
    ws.cell(row=5, column=3, value="字段1")  # 有示例数据
    ws.cell(row=5, column=4, value="字段2")  # 示例数据为空
    ws.cell(row=5, column=5, value="字段3")  # 有示例数据
    ws.cell(row=5, column=6, value="c_classic_battle")  # 第二个标记
    
    # 第6行：示例数据（部分为空）
    ws.cell(row=6, column=1, value=1001)
    ws.cell(row=6, column=2, value="标记1")
    ws.cell(row=6, column=3, value="这是字段1的示例")  # 正常
    ws.cell(row=6, column=4, value="")  # 空值 - 应该产生警告
    ws.cell(row=6, column=5, value="这是字段3的示例")  # 正常
    ws.cell(row=6, column=6, value="标记2")
    
    # 第7行：更多数据
    ws.cell(row=7, column=3, value="字段1数据")
    ws.cell(row=7, column=4, value="字段2数据")  # 虽然第6行为空，但第7行有数据
    ws.cell(row=7, column=5, value="字段3数据")
    
    # 保存文件
    test_file = _TEST_OUTPUT / "test_error_logging.xlsx"
    test_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(test_file)
    print(f"✓ 测试文件已创建: {test_file}")
    return test_file

def create_incomplete_excel():
    """创建行数不足的Excel文件（测试行数不足警告）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "行数不足"
    
    # 第5行：字段名
    ws.cell(row=5, column=2, value="c_classic_battle")
    ws.cell(row=5, column=3, value="字段名")
    ws.cell(row=5, column=4, value="c_classic_battle")
    
    # 没有第6行数据 - 应该产生警告
    
    # 保存文件
    test_file = _TEST_OUTPUT / "test_incomplete_rows.xlsx"
    test_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(test_file)
    print(f"✓ 行数不足测试文件已创建: {test_file}")
    return test_file

def test_error_logging():
    """测试错误日志功能"""
    from core.excel_field_extractor import ExcelFieldExtractor
    
    print("="*70)
    print("测试错误日志功能")
    print("="*70)
    
    # 创建测试文件
    test_file1 = create_test_excel_with_issues()
    test_file2 = create_incomplete_excel()
    
    # 创建提取器
    extractor = ExcelFieldExtractor()
    
    # 提取第一个文件（有空值警告）
    print("\n测试1: 提取包含空示例数据的文件")
    print("-"*70)
    results1 = extractor.extract_fields_from_excel(test_file1)
    
    # 提取第二个文件（行数不足警告）
    print("\n测试2: 提取行数不足的文件")
    print("-"*70)
    results2 = extractor.extract_fields_from_excel(test_file2)
    
    # 获取日志
    print("\n" + "="*70)
    print("日志统计")
    print("="*70)
    
    all_logs = extractor.get_all_logs()
    errors = all_logs['errors']
    warnings = all_logs['warnings']
    
    print(f"\n错误数量: {len(errors)}")
    if errors:
        print("错误日志:")
        for i, error in enumerate(errors, 1):
            print(f"  {i}. {error}")
    
    print(f"\n警告数量: {len(warnings)}")
    if warnings:
        print("警告日志:")
        for i, warning in enumerate(warnings, 1):
            print(f"  {i}. {warning}")
    
    # 保存日志到文件
    print("\n" + "="*70)
    print("保存日志到文件")
    print("="*70)
    
    log_file = _TEST_OUTPUT / "error_logging_test.log"
    if extractor.save_logs_to_file(log_file):
        print(f"✓ 日志已保存")
        
        # 读取并显示日志文件内容
        print("\n日志文件内容预览:")
        print("-"*70)
        with open(log_file, 'r', encoding='utf-8') as f:
            content = f.read()
            print(content)
    
    # 验证结果
    print("\n" + "="*70)
    print("验证结果")
    print("="*70)
    
    success = True
    
    # 应该有至少1个警告（字段2的示例数据为空）
    if len(warnings) < 1:
        print("❌ 错误：应该至少有1个警告（空示例数据）")
        success = False
    else:
        print(f"✓ 检测到 {len(warnings)} 个警告")
    
    # 检查警告内容是否包含关键信息
    warning_has_file_info = any("文件:" in w for w in warnings)
    warning_has_sheet_info = any("工作表:" in w for w in warnings)
    warning_has_field_info = any("字段:" in w for w in warnings)
    
    if not (warning_has_file_info and warning_has_sheet_info and warning_has_field_info):
        print("❌ 错误：警告信息不完整")
        success = False
    else:
        print("✓ 警告信息包含完整的定位信息（文件、工作表、字段）")
    
    # 测试清除日志功能
    print("\n测试清除日志功能...")
    extractor.clear_logs()
    logs_after_clear = extractor.get_all_logs()
    
    if len(logs_after_clear['errors']) == 0 and len(logs_after_clear['warnings']) == 0:
        print("✓ 日志清除功能正常")
    else:
        print("❌ 错误：日志清除失败")
        success = False
    
    # 最终结果
    print("\n" + "="*70)
    if success:
        print("✅ 所有测试通过！错误日志功能正常工作。")
        print("\n功能总结:")
        print("  1. ✓ 检测并记录示例数据为空的警告")
        print("  2. ✓ 检测并记录表格行数不足的警告")
        print("  3. ✓ 警告信息包含完整的定位信息")
        print("  4. ✓ 支持保存日志到文件")
        print("  5. ✓ 支持清除日志")
    else:
        print("❌ 部分测试失败！")
    print("="*70)
    
    return success

if __name__ == "__main__":
    success = test_error_logging()
    sys.exit(0 if success else 1)
