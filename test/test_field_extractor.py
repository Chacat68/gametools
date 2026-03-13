#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
字段提取功能测试 - 合并版
包含CSV、JSON、Excel格式输出测试及字段过滤测试

使用方法:
    python test_field_extractor.py [选项]
    
选项:
    --all       运行所有测试
    --csv       测试CSV格式输出
    --json      测试JSON格式输出
    --excel     测试Excel格式输出
    --filter    测试字段过滤功能
"""

import sys
import json
import argparse
import tempfile
from pathlib import Path

import openpyxl

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.excel_field_extractor import ExcelFieldExtractor


def create_test_excel_files(base_dir: Path) -> Path:
    """创建字段提取测试数据，避免依赖外部夹具。"""
    base_dir.mkdir(parents=True, exist_ok=True)

    text_file = base_dir / "items.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=5, column=2, value="c_")
    ws.cell(row=5, column=3, value="name")
    ws.cell(row=5, column=4, value="title")
    ws.cell(row=5, column=5, value="desc")
    ws.cell(row=5, column=6, value="c_")
    ws.cell(row=6, column=3, value="前端")
    ws.cell(row=6, column=4, value="前端")
    ws.cell(row=6, column=5, value="后端")
    ws.cell(row=7, column=3, value="item_001")
    ws.cell(row=7, column=4, value="武器")
    ws.cell(row=7, column=5, value="强大的武器")
    ws.cell(row=8, column=3, value="item_002")
    ws.cell(row=8, column=4, value="护甲")
    ws.cell(row=8, column=5, value="坚固的护甲")
    wb.save(text_file)

    no_text_file = base_dir / "codes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=5, column=2, value="c_")
    ws.cell(row=5, column=3, value="code")
    ws.cell(row=5, column=4, value="model")
    ws.cell(row=5, column=5, value="type")
    ws.cell(row=5, column=6, value="c_")
    ws.cell(row=6, column=3, value="后端")
    ws.cell(row=6, column=4, value="后端")
    ws.cell(row=6, column=5, value="后端")
    ws.cell(row=7, column=3, value="ITEM_001")
    ws.cell(row=7, column=4, value="armor_a")
    ws.cell(row=7, column=5, value="1")
    ws.cell(row=8, column=3, value="ITEM_002")
    ws.cell(row=8, column=4, value="armor_b")
    ws.cell(row=8, column=5, value="2")
    wb.save(no_text_file)

    return base_dir


def test_csv_output():
    """测试CSV格式输出"""
    print("\n" + "=" * 60)
    print("测试CSV格式输出")
    print("=" * 60)
    
    extractor = ExcelFieldExtractor()
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        test_dir = create_test_excel_files(temp_path / "test_excel_files")
        output_dir = temp_path / "test_output"
        output_dir.mkdir(parents=True, exist_ok=True)

        stats = extractor.process_directory(
            directory_path=str(test_dir),
            output_folder=str(output_dir),
            output_format='csv',
            recursive=True
        )
        
        print(f"✅ 扫描文件数: {stats['total_files']}")
        print(f"✅ 工作表数: {stats['total_sheets']}")
        print(f"✅ 提取字段数: {stats['total_fields']}")
        print(f"✅ 输出文件: {stats['output_file']}")

        csv_file = Path(stats['output_file'])
        content = csv_file.read_text(encoding='utf-8-sig')
        return csv_file.exists() and "items.xlsx#Sheet1" in content and "codes.xlsx#Sheet1" in content


def test_json_output():
    """测试JSON格式输出"""
    print("\n" + "=" * 60)
    print("测试JSON格式输出")
    print("=" * 60)
    
    extractor = ExcelFieldExtractor()
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        test_dir = create_test_excel_files(temp_path / "test_excel_files")
        output_dir = temp_path / "test_output"
        output_dir.mkdir(parents=True, exist_ok=True)

        stats = extractor.process_directory(
            directory_path=str(test_dir),
            output_folder=str(output_dir),
            output_format='json',
            recursive=True
        )
        
        print(f"✅ 扫描文件数: {stats['total_files']}")
        print(f"✅ 输出文件: {stats['output_file']}")
        
        # 验证JSON文件
        json_file = Path(stats['output_file'])
        if json_file.exists():
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            print(f"✅ JSON文件验证成功，共 {len(json_data)} 条记录")
            if json_data:
                print("\nJSON格式示例:")
                preview = {
                    'no_text_tables': json_data.get('no_text_tables', [])[:1],
                    'text_tables': json_data.get('text_tables', [])[:1]
                }
                print(json.dumps(preview, ensure_ascii=False, indent=2))

            text_tables = json_data.get('text_tables', [])
            no_text_tables = json_data.get('no_text_tables', [])
            return any(item['table_name'] == 'items.xlsx' for item in text_tables) and \
                any(item['table_name'] == 'codes.xlsx' for item in no_text_tables)

        return False


def test_excel_output():
    """测试Excel格式输出"""
    print("\n" + "=" * 60)
    print("测试Excel格式输出")
    print("=" * 60)
    
    extractor = ExcelFieldExtractor()
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        test_dir = create_test_excel_files(temp_path / "test_excel_files")
        output_dir = temp_path / "test_output"
        output_dir.mkdir(parents=True, exist_ok=True)

        stats = extractor.process_directory(
            directory_path=str(test_dir),
            output_folder=str(output_dir),
            output_format='excel',
            recursive=True
        )
        
        print(f"✅ 扫描文件数: {stats['total_files']}")
        print(f"✅ 工作表数: {stats['total_sheets']}")
        print(f"✅ 提取字段数: {stats['total_fields']}")
        print(f"✅ 输出文件: {stats['output_file']}")

        excel_file = Path(stats['output_file'])
        return excel_file.exists()


def test_field_filter():
    """测试字段过滤功能"""
    print("\n" + "=" * 60)
    print("测试字段过滤功能")
    print("=" * 60)
    
    extractor = ExcelFieldExtractor()
    print(f"过滤的字段名: {getattr(extractor, 'excluded_field_names', set())}")

    with tempfile.TemporaryDirectory() as temp_dir:
        test_dir = create_test_excel_files(Path(temp_dir) / "test_excel_files")
        results = extractor.scan_directory(test_dir, recursive=True)
    
        print(f"\n扫描到 {len(results)} 个工作表")
        
        all_passed = True
        for idx, result in enumerate(results, 1):
            print(f"\n[{idx}] {result['excel_file']} - {result['sheet_name']}")
            print(f"    是否包含文本: {result['has_text']}")
            
            if result['has_text']:
                print(f"    字段数量: {result['field_count']}")
                print(f"    字段列表: {', '.join(result['fields'])}")
                
                filtered_fields = [f for f in result['fields'] 
                                 if f.lower() in ['name', 'model', 'id', 'code', 'type']]
                if filtered_fields:
                    print(f"    ⚠️ 警告: 以下字段应该被过滤但未被过滤: {filtered_fields}")
                    all_passed = False
                else:
                    print(f"    ✅ 过滤成功")
        
        if extractor.error_logs:
            print("\n错误日志:")
            for error in extractor.error_logs:
                print(f"  ❌ {error}")
        
        return all_passed


def main():
    parser = argparse.ArgumentParser(description='字段提取功能测试')
    parser.add_argument('--all', action='store_true', help='运行所有测试')
    parser.add_argument('--csv', action='store_true', help='测试CSV格式输出')
    parser.add_argument('--json', action='store_true', help='测试JSON格式输出')
    parser.add_argument('--excel', action='store_true', help='测试Excel格式输出')
    parser.add_argument('--filter', action='store_true', help='测试字段过滤功能')
    
    args = parser.parse_args()
    
    # 如果没有指定任何选项，默认运行全部
    if not any([args.all, args.csv, args.json, args.excel, args.filter]):
        args.all = True
    
    print("=" * 60)
    print("字段提取功能测试")
    print("=" * 60)
    
    results = {}
    
    if args.all or args.csv:
        results['CSV输出'] = test_csv_output()
    if args.all or args.json:
        results['JSON输出'] = test_json_output()
    if args.all or args.excel:
        results['Excel输出'] = test_excel_output()
    if args.all or args.filter:
        results['字段过滤'] = test_field_filter()
    
    # 汇总结果
    print("\n" + "=" * 60)
    print("测试结果汇总")
    print("=" * 60)
    for name, passed in results.items():
        status = "✅ 通过" if passed else "❌ 失败"
        print(f"  {name}: {status}")
    
    all_passed = all(results.values())
    print(f"\n总体结果: {'✅ 全部通过' if all_passed else '❌ 存在失败'}")
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
