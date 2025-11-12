#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文本提取器 - 图形界面版本
检测目录中的Excel文件，提取文本内容并创建同名的新Excel文件
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import sys
from pathlib import Path
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 添加模块路径
sys.path.append(str(Path(__file__).parent.parent))

from tools.excel_text_extractor import ExcelTextExtractor


class ExcelTextExtractorGUI:
    """Excel文本提取器图形界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("翻译提取 - 多语言版本")
        self.root.geometry("900x750")
        self.root.minsize(800, 650)
        
        # 语言路径配置（中文为准，默认越南文，额外支持英文、泰文）
        self.language_paths = {
            '中文版': tk.StringVar(),
            '越南文版': tk.StringVar(),
            '英文版': tk.StringVar(),
            '泰文版': tk.StringVar(),
        }
        
        # 设置样式
        self.setup_styles()
        
        # 创建界面
        self.create_widgets()
        
        # 初始化提取器
        self.text_extractor = ExcelTextExtractor(progress_callback=self.update_progress)
    
    def setup_styles(self):
        """设置界面样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置样式
        style.configure('Title.TLabel', font=('Microsoft YaHei', 16, 'bold'))
        style.configure('Heading.TLabel', font=('Microsoft YaHei', 12, 'bold'))
        style.configure('Info.TLabel', font=('Microsoft YaHei', 10))
        style.configure('Success.TLabel', font=('Microsoft YaHei', 10), foreground='green')
        style.configure('Error.TLabel', font=('Microsoft YaHei', 10), foreground='red')
        style.configure('Accent.TButton', font=('Microsoft YaHei', 10, 'bold'))
    
    def create_widgets(self):
        """创建界面控件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # 标题
        title_label = ttk.Label(main_frame, text="翻译提取 - 多语言版本", 
                               style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 10))
        
        # 说明标签
        info_label = ttk.Label(main_frame, 
                              text="💡 为每种语言版本指定独立的Excel文件路径，系统将分别提取文本内容", 
                              style='Info.TLabel', foreground='blue')
        info_label.grid(row=1, column=0, pady=(0, 15))
        
        # 多语言路径选择框架
        lang_frame = ttk.LabelFrame(main_frame, text="多语言文件路径配置", padding="10")
        lang_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        lang_frame.columnconfigure(1, weight=1)
        
        # 为每种语言创建路径选择行
        row_idx = 0
        for lang_name, lang_var in self.language_paths.items():
            # 语言标签
            lang_label = ttk.Label(lang_frame, text=f"{lang_name}:", style='Info.TLabel')
            lang_label.grid(row=row_idx, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 5))
            
            # 路径输入框
            lang_entry = ttk.Entry(lang_frame, textvariable=lang_var, width=50)
            lang_entry.grid(row=row_idx, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(5, 5))
            
            # 浏览按钮
            browse_btn = ttk.Button(lang_frame, text="浏览", 
                                   command=lambda l=lang_name: self.browse_language_path(l))
            browse_btn.grid(row=row_idx, column=2, pady=(5, 5))
            
            # 清除按钮
            clear_btn = ttk.Button(lang_frame, text="✕", width=3,
                                  command=lambda l=lang_name: self.clear_language_path(l))
            clear_btn.grid(row=row_idx, column=3, padx=(5, 0), pady=(5, 5))
            
            row_idx += 1
        
        # 通用输出目录框架
        output_frame = ttk.LabelFrame(main_frame, text="输出配置", padding="10")
        output_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)
        
        # 输出目录
        ttk.Label(output_frame, text="输出目录:", style='Info.TLabel').grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.output_var = tk.StringVar()
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=50)
        self.output_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        self.output_browse_button = ttk.Button(output_frame, text="浏览", 
                                             command=self.browse_output_directory)
        self.output_browse_button.grid(row=0, column=2)
        
        # 输出说明
        output_info = ttk.Label(output_frame, 
                               text="提示：每种语言的提取结果将保存在输出目录的对应子文件夹中", 
                               style='Info.TLabel', foreground='gray')
        output_info.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # 选项设置框架
        options_frame = ttk.LabelFrame(main_frame, text="提取选项", padding="10")
        options_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        options_frame.columnconfigure(1, weight=1)
        
        # 递归扫描选项
        self.recursive_var = tk.BooleanVar(value=True)
        self.recursive_check = ttk.Checkbutton(options_frame, text="递归扫描子目录", 
                                             variable=self.recursive_var)
        self.recursive_check.grid(row=0, column=0, columnspan=3, sticky=tk.W)
        
        # 文本类型过滤
        ttk.Label(options_frame, text="文本类型:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        self.text_type_var = tk.StringVar(value="全部")
        text_type_combo = ttk.Combobox(options_frame, textvariable=self.text_type_var, 
                                      values=["全部", "中文", "英文", "中英混合"], state="readonly", width=15)
        text_type_combo.grid(row=1, column=1, sticky=tk.W, pady=(10, 0))
        
        # 策划检测说明
        planner_info = ttk.Label(options_frame, 
                                text="💡 自动检测：如果Excel文件第6行包含'策划'，将跳过该文件的文本提取", 
                                style='Info.TLabel', foreground='blue')
        planner_info.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
        
        # 语言支持说明
        language_info = ttk.Label(options_frame, 
                                 text="🌐 支持语言：中文、越南文（跳过纯英文），从第7行开始检测，同时提取A列内容", 
                                 style='Info.TLabel', foreground='green')
        language_info.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # 进度显示框架
        progress_frame = ttk.LabelFrame(main_frame, text="处理进度", padding="10")
        progress_frame.grid(row=5, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                          maximum=100, length=400)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # 进度文本
        self.progress_text_var = tk.StringVar(value="就绪")
        self.progress_text = ttk.Label(progress_frame, textvariable=self.progress_text_var, 
                                     style='Info.TLabel')
        self.progress_text.grid(row=1, column=0, sticky=tk.W)
        
        # 控制按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=6, column=0, pady=(0, 10))
        
        self.process_button = ttk.Button(button_frame, text="开始提取", 
                                        command=self.start_extraction, 
                                        style='Accent.TButton')
        self.process_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.clear_button = ttk.Button(button_frame, text="清空结果", 
                                     command=self.clear_results)
        self.clear_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.preview_button = ttk.Button(button_frame, text="预览文件", 
                                        command=self.preview_files,
                                        state="disabled")
        self.preview_button.pack(side=tk.LEFT)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(main_frame, text="提取结果", padding="10")
        result_frame.grid(row=7, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        
        self.result_text = scrolledtext.ScrolledText(result_frame, 
                                                    wrap=tk.WORD, 
                                                    font=("Consolas", 10),
                                                    height=15)
        self.result_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪 - 请配置语言文件路径")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, 
                              relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=8, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def update_progress(self, current: int, total: int, filename: str, message: str):
        """
        更新进度显示
        
        Args:
            current: 当前处理的文件索引
            total: 总文件数
            filename: 当前处理的文件名
            message: 处理消息
        """
        if total > 0:
            percentage = (current / total) * 100
            self.progress_var.set(percentage)
            progress_text = f"[{current}/{total}] ({percentage:.1f}%) {filename}: {message}"
        else:
            self.progress_var.set(0)
            progress_text = f"{filename}: {message}"
        
        # 更新进度文本显示
        self.progress_text_var.set(progress_text)
        
        # 在日志窗口显示详细信息
        timestamp = self._get_timestamp()
        log_message = f"[{timestamp}] {progress_text}"
        
        # 添加带颜色的日志（根据消息类型）
        if "成功" in message or "完成" in message:
            log_message = f"✅ {log_message}"
        elif "失败" in message or "错误" in message:
            log_message = f"❌ {log_message}"
        elif "跳过" in message:
            log_message = f"⏭️ {log_message}"
        elif "开始" in message:
            log_message = f"🚀 {log_message}"
        else:
            log_message = f"ℹ️ {log_message}"
        
        # 更新结果文本区域
        self.result_text.insert(tk.END, f"{log_message}\n")
        self.result_text.see(tk.END)
        
        # 强制更新界面
        self.root.update_idletasks()
    
    def _get_timestamp(self):
        """获取当前时间戳"""
        import datetime
        return datetime.datetime.now().strftime("%H:%M:%S")
    
    def browse_language_path(self, language_name):
        """
        浏览特定语言的文件路径
        
        Args:
            language_name: 语言名称
        """
        # 可以选择目录或文件
        choice = messagebox.askquestion("选择类型", 
                                       f"为 {language_name} 选择：\n\n是(Y) = 选择Excel文件\n否(N) = 选择目录",
                                       icon='question')
        
        if choice == 'yes':
            # 选择文件
            file_path = filedialog.askopenfilename(
                title=f"选择 {language_name} 的Excel文件",
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
            if file_path:
                self.language_paths[language_name].set(file_path)
                self.status_var.set(f"已设置 {language_name} 路径")
        else:
            # 选择目录
            directory = filedialog.askdirectory(title=f"选择 {language_name} 的目录")
            if directory:
                self.language_paths[language_name].set(directory)
                self.status_var.set(f"已设置 {language_name} 路径")
        
        # 自动设置输出目录（如果还未设置）
        if not self.output_var.get():
            # 找到第一个已设置的路径作为参考
            for lang_var in self.language_paths.values():
                path = lang_var.get().strip()
                if path:
                    if os.path.isfile(path):
                        self.output_var.set(os.path.dirname(path))
                    else:
                        self.output_var.set(path)
                    break
    
    def clear_language_path(self, language_name):
        """
        清除特定语言的路径
        
        Args:
            language_name: 语言名称
        """
        self.language_paths[language_name].set("")
        self.status_var.set(f"已清除 {language_name} 路径")
    
    def browse_output_directory(self):
        """浏览输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_var.set(directory)
    
    def start_extraction(self):
        """开始文本提取"""
        output_dir = self.output_var.get().strip()
        
        # 检查是否至少配置了一种语言
        active_languages = {}
        for lang_name, lang_var in self.language_paths.items():
            path = lang_var.get().strip()
            if path:
                if not os.path.exists(path):
                    messagebox.showerror("错误", f"{lang_name} 路径不存在:\n{path}")
                    return
                active_languages[lang_name] = path
        
        if not active_languages:
            messagebox.showerror("错误", "请至少配置一种语言的文件路径")
            return
        
        # 设置输出目录
        if not output_dir:
            # 使用第一个语言路径的目录作为默认输出
            first_path = list(active_languages.values())[0]
            if os.path.isfile(first_path):
                output_dir = os.path.dirname(first_path)
            else:
                output_dir = first_path
        
        # 在新线程中执行提取
        self.process_button.config(state="disabled")
        self.status_var.set("正在提取文本...")
        
        # 重置进度条
        self.progress_var.set(0)
        self.progress_text_var.set("准备开始...")
        
        thread = threading.Thread(target=self._extraction_multi_language, 
                                 args=(active_languages, output_dir))
        thread.daemon = True
        thread.start()
    
    def _extraction_multi_language(self, active_languages, output_dir):
        """
        多语言文本提取（后台线程）- 生成汇总Excel表格
        
        Args:
            active_languages: 字典，键为语言名称，值为文件/目录路径
            output_dir: 输出目录
        """
        try:
            # 清空结果
            self.root.after(0, self.clear_results)
            
            # 显示开始信息
            timestamp = self._get_timestamp()
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"🚀 [{timestamp}] 开始多语言翻译提取任务\n"))
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"📁 输出目录: {output_dir}\n"))
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"🌐 配置语言数: {len(active_languages)} 种\n"))
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"📋 输出格式: name | num | cn | vn | en | th\n"))
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                "=" * 80 + "\n\n"))
            
            # 语言映射
            lang_map = {
                '中文版': 'cn',
                '越南文版': 'vn',
                '英文版': 'en',
                '泰文版': 'th'
            }
            
            # 收集所有语言的文本数据
            all_lang_data = {}  # {lang_code: {file_name: extracted_data}}
            
            for idx, (lang_name, lang_path) in enumerate(active_languages.items(), 1):
                lang_code = lang_map[lang_name]
                self.root.after(0, lambda n=lang_name, i=idx, t=len(active_languages): 
                              self.result_text.insert(tk.END, 
                                  f"\n📚 提取 [{i}/{t}]: {n}\n"))
                
                try:
                    lang_extractor = ExcelTextExtractor(progress_callback=self.update_progress)
                    file_data = {}
                    
                    if os.path.isfile(lang_path):
                        # 单个文件
                        file_name = os.path.splitext(os.path.basename(lang_path))[0]
                        self.root.after(0, lambda f=file_name: 
                                      self.result_text.insert(tk.END, f"  📄 {f}\n"))
                        extracted_data = lang_extractor.extract_text_from_excel(lang_path, 1, 1)
                        if extracted_data:
                            file_data[file_name] = extracted_data
                    else:
                        # 目录批量处理
                        excel_files = lang_extractor.scan_directory(lang_path)
                        self.root.after(0, lambda c=len(excel_files): 
                                      self.result_text.insert(tk.END, f"  📁 找到 {c} 个文件\n"))
                        
                        for file_idx, file_path in enumerate(excel_files, 1):
                            file_name = os.path.splitext(os.path.basename(file_path))[0]
                            extracted_data = lang_extractor.extract_text_from_excel(file_path, file_idx, len(excel_files))
                            if extracted_data:
                                file_data[file_name] = extracted_data
                    
                    all_lang_data[lang_code] = file_data
                    self.root.after(0, lambda c=len(file_data): 
                                  self.result_text.insert(tk.END, f"  ✅ 成功提取 {c} 个文件\n"))
                    
                except Exception as e:
                    error_msg = f"  ❌ 提取 {lang_name} 时出错: {str(e)}\n"
                    self.root.after(0, lambda m=error_msg: self.result_text.insert(tk.END, m))
            
            # 生成汇总Excel表格
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"\n{'='*80}\n"))
            self.root.after(0, lambda: self.result_text.insert(tk.END, 
                f"📊 生成汇总Excel表格\n"))
            
            output_file = os.path.join(output_dir, "翻译提取汇总.xlsx")
            success = self._create_summary_excel(all_lang_data, output_file)
            
            if success:
                self.root.after(0, lambda: self.result_text.insert(tk.END, 
                    f"✅ 汇总表格已生成: {output_file}\n"))
                self.root.after(0, lambda: self.result_text.insert(tk.END, 
                    f"{'='*80}\n\n"))
                self.root.after(0, self._show_multi_lang_success)
            else:
                self.root.after(0, lambda: self.result_text.insert(tk.END, 
                    f"❌ 生成汇总表格失败\n"))
                self.root.after(0, self._show_error_result, "生成汇总表格失败")
            
        except Exception as e:
            error_msg = f"多语言提取过程中发生错误: {str(e)}"
            self.root.after(0, self._show_error_result, error_msg)
    
    def _create_summary_excel(self, all_lang_data, output_file):
        """
        创建汇总Excel表格
        
        Args:
            all_lang_data: {lang_code: {file_name: extracted_data}}
            output_file: 输出文件路径
            
        Returns:
            是否成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 以中文版为基准
            if 'cn' not in all_lang_data or not all_lang_data['cn']:
                logger.error("未找到中文版数据，无法生成汇总表格")
                return False
            
            # 构建汇总数据
            summary_data = []
            
            for file_name, cn_data in all_lang_data['cn'].items():
                # 遍历中文版的每个工作表
                for sheet_name, sheet_data in cn_data.items():
                    if not sheet_data or 'items' not in sheet_data:
                        continue
                    
                    # 遍历每个提取项
                    for item in sheet_data['items']:
                        excel_pos = item.get('excel_row_ref', '')
                        cn_text = item.get('text', '')
                        
                        # 查找其他语言对应位置的文本
                        vn_text = self._find_text_at_position(all_lang_data.get('vn', {}), file_name, sheet_name, excel_pos)
                        en_text = self._find_text_at_position(all_lang_data.get('en', {}), file_name, sheet_name, excel_pos)
                        th_text = self._find_text_at_position(all_lang_data.get('th', {}), file_name, sheet_name, excel_pos)
                        
                        summary_data.append({
                            'name': file_name,
                            'num': excel_pos,
                            'cn': cn_text,
                            'vn': vn_text,
                            'en': en_text,
                            'th': th_text
                        })
            
            # 创建DataFrame
            df = pd.DataFrame(summary_data)
            
            # 保存到Excel
            df.to_excel(output_file, index=False, sheet_name='汇总')
            
            # 美化Excel格式
            wb = load_workbook(output_file)
            ws = wb['汇总']
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            ws.column_dimensions['A'].width = 25  # name
            ws.column_dimensions['B'].width = 10  # num
            ws.column_dimensions['C'].width = 40  # cn
            ws.column_dimensions['D'].width = 40  # vn
            ws.column_dimensions['E'].width = 40  # en
            ws.column_dimensions['F'].width = 40  # th
            
            # 保存
            wb.save(output_file)
            
            logger.info(f"汇总Excel表格创建成功: {output_file}")
            logger.info(f"总计 {len(summary_data)} 条记录")
            
            return True
            
        except Exception as e:
            logger.error(f"创建汇总Excel失败: {str(e)}")
            return False
    
    def _find_text_at_position(self, lang_data, file_name, sheet_name, excel_pos):
        """
        在指定语言数据中查找特定位置的文本
        
        Args:
            lang_data: 语言数据字典
            file_name: 文件名
            sheet_name: 工作表名
            excel_pos: Excel位置（如F7）
            
        Returns:
            找到的文本，未找到返回空字符串
        """
        if not lang_data or file_name not in lang_data:
            return ""
        
        file_data = lang_data[file_name]
        if sheet_name not in file_data:
            return ""
        
        sheet_data = file_data[sheet_name]
        if not sheet_data or 'items' not in sheet_data:
            return ""
        
        for item in sheet_data['items']:
            if item.get('excel_row_ref') == excel_pos:
                return item.get('text', '')
        
        return ""
    
    def _sanitize_folder_name(self, name):
        """
        清理文件夹名称，移除不合法字符
        
        Args:
            name: 原始名称
            
        Returns:
            清理后的名称
        """
        # 移除版本后缀
        clean_name = name.replace('版', '')
        # 移除不合法字符
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        return clean_name
    
    def _show_success_result(self):
        """显示提取成功结果"""
        timestamp = self._get_timestamp()
        self.result_text.insert(tk.END, "=" * 60 + "\n")
        self.result_text.insert(tk.END, f"✅ [{timestamp}] Excel文本提取完成！\n")
        
        report = self.text_extractor.get_processing_report()
        self.result_text.insert(tk.END, report)
        self.result_text.insert(tk.END, "\n")
        
        self.result_text.see(tk.END)
        
        self.process_button.config(state="normal")
        self.preview_button.config(state="normal")
        self.status_var.set("文本提取完成")
        
        messagebox.showinfo("成功", "Excel文本提取完成！")
    
    def _show_multi_lang_success(self):
        """显示多语言提取成功结果"""
        timestamp = self._get_timestamp()
        self.result_text.insert(tk.END, f"✅ [{timestamp}] 多语言文本提取完成！\n")
        self.result_text.see(tk.END)
        
        self.process_button.config(state="normal")
        self.preview_button.config(state="normal")
        self.status_var.set("多语言提取完成")
        
        messagebox.showinfo("成功", "多语言Excel文本提取完成！\n\n汇总Excel表格已生成，格式：name | num | cn | vn | en | th")
    
    def _show_error_result(self, error_msg):
        """显示提取错误结果"""
        timestamp = self._get_timestamp()
        self.result_text.insert(tk.END, "=" * 60 + "\n")
        self.result_text.insert(tk.END, f"❌ [{timestamp}] {error_msg}\n")
        self.result_text.see(tk.END)
        
        self.process_button.config(state="normal")
        self.preview_button.config(state="normal")
        self.status_var.set("文本提取失败")
        
        messagebox.showerror("错误", error_msg)
    
    def preview_files(self):
        """预览各语言的Excel文件"""
        # 收集已配置的语言路径
        active_languages = {}
        for lang_name, lang_var in self.language_paths.items():
            path = lang_var.get().strip()
            if path and os.path.exists(path):
                active_languages[lang_name] = path
        
        if not active_languages:
            messagebox.showerror("错误", "请先配置至少一种语言的文件路径")
            return
        
        try:
            # 清空并显示预览
            self.result_text.delete(1.0, tk.END)
            
            preview_text = "=" * 60 + "\n"
            preview_text += "📋 多语言文件预览\n"
            preview_text += "=" * 60 + "\n\n"
            
            total_files = 0
            
            for lang_name, lang_path in active_languages.items():
                preview_text += f"🌍 {lang_name}\n"
                preview_text += f"   路径: {lang_path}\n"
                
                if os.path.isfile(lang_path):
                    # 单个文件
                    preview_text += f"   类型: 单个Excel文件\n"
                    preview_text += f"   文件名: {os.path.basename(lang_path)}\n"
                    total_files += 1
                else:
                    # 目录
                    excel_files = self.text_extractor.scan_directory(lang_path)
                    preview_text += f"   类型: 目录\n"
                    preview_text += f"   找到文件: {len(excel_files)} 个\n"
                    
                    if excel_files:
                        preview_text += f"   文件列表:\n"
                        for i, file_path in enumerate(excel_files[:10]):  # 只显示前10个
                            preview_text += f"      {i+1}. {os.path.basename(file_path)}\n"
                        if len(excel_files) > 10:
                            preview_text += f"      ... 还有 {len(excel_files) - 10} 个文件\n"
                    
                    total_files += len(excel_files)
                
                preview_text += "\n"
            
            preview_text += "=" * 60 + "\n"
            preview_text += f"📊 总计: {len(active_languages)} 种语言，{total_files} 个文件\n"
            preview_text += "=" * 60 + "\n"
            
            self.result_text.insert(1.0, preview_text)
            
        except Exception as e:
            messagebox.showerror("错误", f"预览文件失败: {str(e)}")
    
    def clear_results(self):
        """清空提取结果"""
        self.result_text.delete(1.0, tk.END)


def main():
    """主函数"""
    root = tk.Tk()
    app = ExcelTextExtractorGUI(root)
    
    # 设置窗口关闭事件
    def on_closing():
        root.quit()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 启动主循环
    root.mainloop()


if __name__ == "__main__":
    main()
