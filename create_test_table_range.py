#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试表范围翻译提取器
创建测试用的Excel和JSON文件
"""

import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_test_excel_files():
    """创建测试Excel文件"""
    
    # 创建测试目录
    test_dir = "test_table_range"
    os.makedirs(test_dir, exist_ok=True)
    
    print("创建测试Excel文件...")
    
    # 测试文件1: 角色配置表
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "角色列表"
    
    # 第1-4行：说明行
    ws1['A1'] = "角色配置表"
    ws1['A2'] = "包含角色的基本信息"
    ws1['A3'] = ""
    ws1['A4'] = ""
    
    # 第5行：字段名
    ws1['A5'] = 'c_'
    ws1['B5'] = 'id'
    ws1['C5'] = 'name_cn'
    ws1['D5'] = 'name_vn'
    ws1['E5'] = 'name_th'
    ws1['F5'] = 'desc_cn'
    ws1['G5'] = 'model'
    ws1['H5'] = 'c_'
    
    # 第6行：字段类型
    ws1['A6'] = ''
    ws1['B6'] = '策划'
    ws1['C6'] = '前端'
    ws1['D6'] = '前端'
    ws1['E6'] = '前端'
    ws1['F6'] = '后端'
    ws1['G6'] = '策划'
    ws1['H6'] = ''
    
    # 第7行开始：数据
    data_rows = [
        ['', 1001, '张三', 'Tiểu Tam', 'จางซาน', '一个普通的角色', 'model_001', ''],
        ['', 1002, '李四', 'Lý Tứ', 'หลีซื่อ', '另一个角色', 'model_002', ''],
        ['', 1003, '王五', 'Vương Ngũ', 'หวังอู๋', '第三个角色', 'model_003', ''],
    ]
    
    for row in data_rows:
        ws1.append(row)
    
    wb1.save(os.path.join(test_dir, "角色配置.xlsx"))
    print("  ✓ 角色配置.xlsx")
    
    # 测试文件2: 物品配置表
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "物品列表"
    
    # 第1-4行：说明行
    ws2['A1'] = "物品配置表"
    ws2['A2'] = ""
    ws2['A3'] = ""
    ws2['A4'] = ""
    
    # 第5行：字段名
    ws2['A5'] = 'c_'
    ws2['B5'] = 'item_id'
    ws2['C5'] = 'item_name_cn'
    ws2['D5'] = 'item_name_vn'
    ws2['E5'] = 'item_desc'
    ws2['F5'] = 'price'
    ws2['G5'] = 'c_'
    
    # 第6行：字段类型
    ws2['A6'] = ''
    ws2['B6'] = '策划'
    ws2['C6'] = '前端'
    ws2['D6'] = '前端'
    ws2['E6'] = '前后端'
    ws2['F6'] = '后端'
    ws2['G6'] = ''
    
    # 第7行开始：数据
    item_data = [
        ['', 2001, '宝剑', 'Kiếm', '锋利的宝剑', 1000, ''],
        ['', 2002, '盾牌', 'Khiên', '坚固的盾牌', 800, ''],
        ['', 2003, '药水', 'Thuốc', '恢复生命的药水', 50, ''],
    ]
    
    for row in item_data:
        ws2.append(row)
    
    wb2.save(os.path.join(test_dir, "物品配置.xlsx"))
    print("  ✓ 物品配置.xlsx")
    
    # 测试文件3: 任务配置表
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "任务列表"
    
    # 第1-4行：说明行
    ws3['A1'] = "任务配置表"
    ws3['A2'] = ""
    ws3['A3'] = ""
    ws3['A4'] = ""
    
    # 第5行：字段名
    ws3['A5'] = 'c_'
    ws3['B5'] = 'quest_id'
    ws3['C5'] = 'quest_title'
    ws3['D5'] = 'quest_desc'
    ws3['E5'] = 'reward'
    ws3['F5'] = 'c_'
    
    # 第6行：字段类型
    ws3['A6'] = ''
    ws3['B6'] = '策划'
    ws3['C6'] = '前端'
    ws3['D6'] = '前后端'
    ws3['E6'] = '策划'
    ws3['F6'] = ''
    
    # 第7行开始：数据
    quest_data = [
        ['', 3001, '新手任务', 'ภารกิจสำหรับผู้เริ่มต้น', 100, ''],
        ['', 3002, '寻找宝藏', 'Tìm kho báu bị mất', 500, ''],
    ]
    
    for row in quest_data:
        ws3.append(row)
    
    wb3.save(os.path.join(test_dir, "任务配置.xlsx"))
    print("  ✓ 任务配置.xlsx")
    
    return test_dir


def create_test_json_config(test_dir):
    """创建测试JSON配置文件"""
    
    print("\n创建测试JSON配置...")
    
    json_config = {
        "no_text_tables": [
            {
                "table_name": "系统配置.xlsx",
                "sheet_name": "配置项"
            }
        ],
        "text_tables": [
            {
                "table_name": "角色配置.xlsx",
                "sheet_name": "角色列表",
                "fields_with_examples": [
                    "id,策划",
                    "name_cn,前端",
                    "name_vn,前端",
                    "name_th,前端",
                    "desc_cn,后端",
                    "model,策划"
                ],
                "field_count": 6
            },
            {
                "table_name": "物品配置.xlsx",
                "sheet_name": "物品列表",
                "fields_with_examples": [
                    "item_id,策划",
                    "item_name_cn,前端",
                    "item_name_vn,前端",
                    "item_desc,前后端",
                    "price,后端"
                ],
                "field_count": 5
            },
            {
                "table_name": "任务配置.xlsx",
                "sheet_name": "任务列表",
                "fields_with_examples": [
                    "quest_id,策划",
                    "quest_title,前端",
                    "quest_desc,前后端",
                    "reward,策划"
                ],
                "field_count": 4
            }
        ]
    }
    
    json_path = os.path.join(test_dir, "field_config.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_config, f, ensure_ascii=False, indent=2)
    
    print(f"  ✓ field_config.json")
    
    return json_path


def create_readme(test_dir):
    """创建测试说明文档"""
    
    readme_content = """# 表范围翻译提取器测试文件

## 测试文件说明

### Excel测试文件

1. **角色配置.xlsx**
   - 工作表: 角色列表
   - 字段: id(策划), name_cn(前端), name_vn(前端), name_th(前端), desc_cn(后端), model(策划)
   - 预期导出: name_cn, name_vn, name_th, desc_cn（跳过id和model）

2. **物品配置.xlsx**
   - 工作表: 物品列表
   - 字段: item_id(策划), item_name_cn(前端), item_name_vn(前端), item_desc(前后端), price(后端)
   - 预期导出: item_name_cn, item_name_vn, item_desc, price（跳过item_id）

3. **任务配置.xlsx**
   - 工作表: 任务列表
   - 字段: quest_id(策划), quest_title(前端), quest_desc(前后端), reward(策划)
   - 预期导出: quest_title, quest_desc（跳过quest_id和reward）

### JSON配置文件

**field_config.json**
- 包含 no_text_tables（跳过）和 text_tables（处理）
- 每个表格的字段名和字段类型配置

## 使用方法

### 命令行测试
```bash
python core/table_range_translator.py field_config.json test_table_range --output 翻译总表.xlsx
```

### GUI测试
1. 启动GUI: `python gui/gametools_unified.py`
2. 选择"表范围翻译提取"页签
3. 选择JSON配置: test_table_range/field_config.json
4. 选择Excel目录: test_table_range/
5. 选择输出文件: 翻译总表.xlsx
6. 点击"开始提取"

## 预期结果

生成的翻译总表应包含3个工作表:
- 角色配置
- 物品配置
- 任务配置

每个工作表包含列:
- 字段名
- 字段类型
- ID
- 行号
- 中文内容
- 越南文
- 泰文
- 语言类型

## 功能特点

✓ 自动跳过 no_text_tables
✓ 只导出前端、后端、前后端字段
✓ 忽略策划字段
✓ 按表格分工作表
✓ 支持中文、越南文、泰文识别
"""
    
    readme_path = os.path.join(test_dir, "README.md")
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print(f"  ✓ README.md")


def main():
    """主函数"""
    print("=" * 60)
    print("表范围翻译提取器 - 测试文件生成器")
    print("=" * 60)
    print()
    
    # 创建测试Excel文件
    test_dir = create_test_excel_files()
    
    # 创建JSON配置
    json_path = create_test_json_config(test_dir)
    
    # 创建说明文档
    create_readme(test_dir)
    
    print()
    print("=" * 60)
    print("✓ 测试文件创建完成！")
    print("=" * 60)
    print(f"\n测试目录: {test_dir}/")
    print(f"JSON配置: {json_path}")
    print("\n测试命令:")
    print(f"  python core/table_range_translator.py {json_path} {test_dir} --output 翻译总表.xlsx")
    print("\n或使用GUI进行测试")


if __name__ == "__main__":
    main()
