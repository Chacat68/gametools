#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建包含 name 和 model 字段的测试Excel文件
用于验证字段过滤功能
"""

import openpyxl
from openpyxl import Workbook
from pathlib import Path

def create_test_excel_with_filtered_fields():
    """创建包含需要过滤字段的测试Excel文件"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试表"
    
    # 第1-4行：表头信息
    ws['A1'] = '%%% -*- coding: utf-8 -*-'
    ws['A2'] = 'latin-1 -*-'
    ws['A3'] = '%%! -filetype'
    ws['A4'] = 'record consult'
    
    # 第5行：字段名（包含需要过滤的字段）
    fields = ['c_', '序号', '索引1', '索引2', 'des_cn', 'des_vcn', 'des', 'name', 'model', 'id', 'c_']
    for idx, field in enumerate(fields, 1):
        ws.cell(row=5, column=idx, value=field)
    
    # 第6行及以后：数据
    # 添加测试数据
    test_data = [
        [1, '张若', '小龙女', 'Tiểu Long Nữ', 'npc104_ui', 0, 1001],
        [2, '黄蓉', '黄蓉', 'Hoàng Dung', 'npc105_ui', 0, 1002],
        [3, '刘音', '活泼子', 'Tiểu Đạo Tử', 'npc201_ui', 0, 1003],
        [4, '汪宣亲', '五岳子', 'Võ Nhai Tử', 'npc208_ui', 0, 1004],
        [5, '徐观龙', '太极宗师张三丰', 'Thái Cực Tông Sư - Trương Tam Phong', 'npc400_ui', 0, 1005],
    ]
    
    for row_idx, data in enumerate(test_data, 6):
        # 跳过第一列（c_标记列）
        ws.cell(row=row_idx, column=2, value=data[0])  # 序号
        ws.cell(row=row_idx, column=3, value='string')  # 索引1
        ws.cell(row=row_idx, column=4, value='string')  # 索引2
        ws.cell(row=row_idx, column=5, value=data[1])  # des_cn (中文)
        ws.cell(row=row_idx, column=6, value=data[2])  # des_vcn (越南文)
        ws.cell(row=row_idx, column=7, value=data[3])  # des (越南文)
        ws.cell(row=row_idx, column=8, value=data[4])  # name (代码标识符，应该被过滤)
        ws.cell(row=row_idx, column=9, value=data[5])  # model (数字，应该被过滤)
        ws.cell(row=row_idx, column=10, value=data[6]) # id (数字，应该被过滤)
    
    # 保存文件
    output_path = Path('test_excel_files/test_field_filter.xlsx')
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ 测试文件已创建: {output_path}")
    print(f"   - 包含字段: {', '.join(fields)}")
    print(f"   - 应该被过滤的字段: name, model, id")
    print(f"   - 应该被保留的字段: des_cn, des_vcn, des")

if __name__ == "__main__":
    create_test_excel_with_filtered_fields()
